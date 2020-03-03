import TSD_Checker_V8_5
import time
from PyQt5 import QtGui
import xlwt
from xlutils.copy import copy
import openpyxl
from openpyxl.styles import Color, Font

def TestReturn(criticity, testName, message, localisation, workBook, TSDApp):

    return_dict = {}
    return_dict["criticity"] = criticity
    return_dict["testName"] = testName
    return_dict["message"] = message
    return_dict["localisation"] = localisation
    TSDApp.return_list.append(return_dict)

    if criticity.casefold() == "blocking":
        TSDApp.criticity_blocking += 1
    else:
        if criticity.casefold() == "warning":
            TSDApp.criticity_warning += 1
        else:
            TSDApp.criticity_information += 1

    tempString = str()
    if localisation is None or localisation == "":
        tempString = "OK"
        if criticity.casefold() == "blocking":
            TSDApp.criticity_blocking_passed += 1
        else:
            if criticity.casefold() == "warning":
                TSDApp.criticity_warning_passed += 1
            else:
                TSDApp.criticity_information_passed += 1
    else:
        tempString = "NOK"

    textBoxText = TSDApp.tab1.textbox.toPlainText()
    textBoxText = textBoxText + "\n" + testName + " " + tempString
    TSDApp.tab1.textbox.setText(textBoxText)
    TSDApp.tab1.textbox.moveCursor(QtGui.QTextCursor.End)

    TSDApp.IncrementProgressBar()


def deleteSheet(TSDApp, workbook, sheet_name1, sheet_name2):
    index_test_report = -1
    index_info_report = -1
    for sheetname in TSDApp.WorkbookStats.sheetNames:
        if sheetname == 'test report':
            index_test_report = TSDApp.WorkbookStats.sheetNames.index('test report')
        if sheetname == 'report information':
            index_info_report = TSDApp.WorkbookStats.sheetNames.index('report information')

    new_wb = copy(workbook)
    if index_info_report != -1:
        new_wb._Workbook__worksheets = [worksheet for worksheet in new_wb._Workbook__worksheets if worksheet.name.casefold() != sheet_name1]
    if index_test_report != -1:
        new_wb._Workbook__worksheets = [worksheet for worksheet in new_wb._Workbook__worksheets if worksheet.name.casefold() != sheet_name2]

    if index_info_report != -1:
        if new_wb._Workbook__worksheet_idx_from_name['report information'] > -1:
            del new_wb._Workbook__worksheet_idx_from_name['report information']
    if index_test_report != -1:
        if new_wb._Workbook__worksheet_idx_from_name['test report'] > -1:
            del new_wb._Workbook__worksheet_idx_from_name['test report']
    return new_wb


def column_string(n):
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string

def ExcelWrite_del_information(return_list, path, TSDApp, workBook):

    DOC3 = workBook
    new_wb = deleteSheet(TSDApp, DOC3, "report information", "test report")

    for link in TSDApp.links:
        if 'http' not in link[4]:
            sheet_to_check = new_wb._Workbook__worksheets[link[0]]
            sheet_to_check.write(link[2], link[3], xlwt.Formula(link[4]))
        else:
            sheet_to_check = new_wb._Workbook__worksheets[link[0]]
            sheet_to_check.write(link[2], link[3], xlwt.Formula('HYPERLINK("%s";"%s")' % (link[4], link[1])))

    if TSDApp.convergence != "":
        try:
            workSheet = new_wb.get_sheet("tableau")
        except:
            workSheet = new_wb.get_sheet("Table")

        if TSDApp.refSignature == -1:
            try:
                workSheet.write(TSDApp.tableHeaderRow, TSDApp.WorkbookStats.tableLastCol, 'Unique Test Signature')
                for element in TSDApp.unique_items:
                    if TSDApp.unique_list.count(element['value']) == 1:
                        workSheet.write(element['localisation'], TSDApp.WorkbookStats.tableLastCol, '1')
                        # NbUniqueSignatureTests += 1
                    else:
                        for elem in TSDApp.unique_items:
                            if element['value'] == elem['value']:
                                workSheet.write(elem['localisation'], TSDApp.WorkbookStats.tableLastCol, '0')
            except:
                text = TSDApp.tab1.textbox.toPlainText()
                TSDApp.tab1.textbox.setText(text + '\n' + "Warning: Only 256 first columns filled in 'tableau' sheet (xls format limitation)")
        else:
            for element in TSDApp.unique_items:
                if TSDApp.unique_list.count(element['value']) == 1:
                    workSheet.write(element['localisation'], TSDApp.refSignature, '1')
                    # NbUniqueSignatureTests += 1
                else:
                    for elem in TSDApp.unique_items:
                        if element['value'] == elem['value']:
                            workSheet.write(elem['localisation'], TSDApp.refSignature, '0')

    workSheet_info_report = new_wb.add_sheet('Report information', cell_overwrite_ok=True)

    col1 = workSheet_info_report.col(0)
    col1.width = 256 * 35
    col2 = workSheet_info_report.col(1)
    col2.width = 256 * 120
    col3 = workSheet_info_report.col(2)
    col3.width = 256 * 10

    workSheet_info_report.write(0, 0, "Tool version:")
    workSheet_info_report.write(0, 1, TSD_Checker_V8_5.appName)

    workSheet_info_report.write(2, 0, "Criticity configuration file:")
    workSheet_info_report.write(2, 1, TSDApp.DOC9Path)
    workSheet_info_report.write(2, 2, TSDApp.version_criticity_file)

    workSheet_info_report.write(3, 0, "Extract CESARE file:")
    workSheet_info_report.write(3, 1, TSDApp.DOC8Path)
    workSheet_info_report.write(3, 2, TSDApp.version_cesare_file)

    workSheet_info_report.write(4, 0, "Customer effects file:")
    workSheet_info_report.write(4, 1, TSDApp.DOC7Path)
    workSheet_info_report.write(4, 2, TSDApp.version_cutomer_effect)

    workSheet_info_report.write(5, 0, "Diversity management file:")
    workSheet_info_report.write(5, 1, TSDApp.DOC13Path)
    workSheet_info_report.write(5, 2, TSDApp.version_diversity_file)

    workSheet_info_report.write(6, 0, "CESARE file reference:")
    if TSDApp.tab2.myTextBox7.toPlainText() == "":
        workSheet_info_report.write(6, 1, TSDApp.DOC8Link.split("/")[-3])
    else:
        workSheet_info_report.write(6, 1, TSDApp.tab2.myTextBox7.toPlainText())

    workSheet_info_report.write(7, 0, "Criticity configuration file reference:")
    if TSDApp.tab2.myTextBox8.toPlainText() == "":
        workSheet_info_report.write(7, 1, TSDApp.DOC9Link.split("/")[-3])
    else:
        workSheet_info_report.write(7, 1, TSDApp.tab2.myTextBox8.toPlainText())

    workSheet_info_report.write(8, 0, "Customer effect file reference:")
    if TSDApp.tab2.myTextBox9.toPlainText() == "":
        workSheet_info_report.write(8, 1, TSDApp.DOC7Link.split("/")[-3])
    else:
        workSheet_info_report.write(8, 1, TSDApp.tab2.myTextBox9.toPlainText())

    workSheet_info_report.write(9, 0, "Diversity management file reference:")
    if TSDApp.tab2.myTextBox10.toPlainText() == "":
        workSheet_info_report.write(9, 1, TSDApp.DOC13Link.split("/")[-3])
    else:
        workSheet_info_report.write(9, 1, TSDApp.tab2.myTextBox10.toPlainText())

    workSheet_info_report.write(10, 0, "FSE TSD template:")
    if TSDApp.DOC3Exists and not TSDApp.DOC4Exists and not TSDApp.DOC5Exists:
        workSheet_info_report.write(10, 1, TSDApp.DOC3Name)
    else:
        workSheet_info_report.write(10, 1, "not used")

    workSheet_info_report.write(11, 0, "TSD Vehicle Funtion template:")
    if not TSDApp.DOC3Exists and TSDApp.DOC4Exists:
        workSheet_info_report.write(11, 1, TSDApp.DOC4Name)
    else:
        workSheet_info_report.write(11, 1, "not used")

    workSheet_info_report.write(12, 0, "TSD System template:")
    if not TSDApp.DOC3Exists and not TSDApp.DOC4Exists and TSDApp.DOC5Exists:
        workSheet_info_report.write(12, 1, TSDApp.DOC5Name)
    else:
        workSheet_info_report.write(12, 1, "not used")

    if TSDApp.DOC3Exists and not TSDApp.DOC4Exists and not TSDApp.DOC5Exists:
        workSheet_info_report.write(13, 0, "FSE TSD template reference:")
        if TSDApp.tab2.myTextBox11.toPlainText() == "":
            workSheet_info_report.write(13, 1, TSDApp.DOC3Link.split("/")[-3])
        else:
            workSheet_info_report.write(13, 1, TSDApp.tab2.myTextBox11.toPlainText())
    elif not TSDApp.DOC3Exists and TSDApp.DOC4Exists:
        workSheet_info_report.write(13, 0, "TSD Vehicle Function template reference:")
        if TSDApp.tab2.myTextBox12.toPlainText() == "":
            workSheet_info_report.write(13, 1, TSDApp.DOC4Link.split("/")[-3])
        else:
            workSheet_info_report.write(13, 1, TSDApp.tab2.myTextBox12.toPlainText())
    if not TSDApp.DOC3Exists and not TSDApp.DOC4Exists and TSDApp.DOC5Exists:
        workSheet_info_report.write(13, 0, "TSD System template reference:")
        if TSDApp.tab2.myTextBox13.toPlainText() == "":
            workSheet_info_report.write(13, 1, TSDApp.DOC5Link.split("/")[-3])
        else:
            workSheet_info_report.write(13, 1, TSDApp.tab2.myTextBox13.toPlainText())

    workSheet_info_report.write(14, 0, "Check level:")
    workSheet_info_report.write(14, 1, TSDApp.checkLevel)

    workSheet_info_report.write(16, 0, "Date of the test:")
    workSheet_info_report.write(16, 1, time.strftime("%d/%m/%Y"))

    workSheet_info_report.write(17, 0, "Time of the test:")
    workSheet_info_report.write(17, 1, time.strftime("%X"))

    workSheet_info_report.write(18, 0, "Test duration:")
    workSheet_info_report.write(18, 1, time.strftime('%H:%M:%S', time.gmtime(TSDApp.end_time - TSDApp.start_time)))

    workSheet_info_report.write(19, 0, "Opening duration:")
    workSheet_info_report.write(19, 1, time.strftime('%H:%M:%S', time.gmtime(TSDApp.opening_time - TSDApp.start_time)))

    workSheet_info_report.write(21, 0, "TSD file checked:")
    workSheet_info_report.write(21, 1, TSDApp.DOC3Path)

    workSheet_info_report.write(22, 0, "TSD function file checked:")
    workSheet_info_report.write(22, 1, TSDApp.DOC4Path)

    workSheet_info_report.write(23, 0, "TSD system file checked:")
    workSheet_info_report.write(23, 1, TSDApp.DOC5Path)

    workSheet_info_report.write(25, 0, "AMDEC:")
    workSheet_info_report.write(25, 1, TSDApp.tab1.myTextBox4.toPlainText())

    workSheet_info_report.write(26, 0, "Export MedialecMatrice:")
    workSheet_info_report.write(26, 1, TSDApp.tab1.myTextBox5.toPlainText())

    workSheet_info_report.write(27, 0, "Diagnostic Messagerie (ODX):")
    workSheet_info_report.write(27, 1, TSDApp.tab1.myTextBox6.toPlainText())

    workSheet_info_report.write(28, 0, "SubFamily:")
    workSheet_info_report.write(28, 1, TSDApp.tab1.myTextBox61.toPlainText())

    workSheet_info_report.write(30, 0, "Architecture type:")
    workSheet_info_report.write(30, 1, TSDApp.tab1.combo2.currentText())

    workSheet_info_report.write(31, 0, "Diversity Management:")
    workSheet_info_report.write(31, 1, TSDApp.tab1.combo3.currentText())

    workSheet_info_report.write(32, 0, "Project name:")
    workSheet_info_report.write(32, 1, TSDApp.tab1.combo1.currentText())

    workSheet_info_report.write(34, 0, "Status:")
    workSheet_info_report.write(34, 1, str(TSDApp.status))

    workSheet_info_report.write(35, 0, "Coverage Indicator:")
    workSheet_info_report.write(35, 1, str(TSDApp.coverage)[0:4] + "%")
    if str(TSDApp.coverage)[0:4] + "%" == "0.00%":
        workSheet_info_report.write(35, 2, "WARNING: The coverage indicator will not be calculated because at least one of its parameters is missing.")

    workSheet_info_report.write(36, 0, "Convergence Indicator:")
    workSheet_info_report.write(36, 1, str(TSDApp.convergence)[0:4] + "%")
    if str(TSDApp.convergence)[0:4] + "%" == "0.00%":
        workSheet_info_report.write(36, 2, "WARNING: The convergence indicator will not be calculated because at least one of its parameters is missing.")

    workSheet_info_report.write(38, 0, "Blocking Points Failed")
    workSheet_info_report.write(38, 1, str(TSDApp.criticity_blocking - TSDApp.criticity_blocking_passed))

    workSheet_info_report.write(39, 0, "Warning Points Failed")
    workSheet_info_report.write(39, 1, str(TSDApp.criticity_warning - TSDApp.criticity_warning_passed))

    workSheet_info_report.write(40, 0, "Information Points Failed")
    workSheet_info_report.write(40, 1, str(TSDApp.criticity_information - TSDApp.criticity_information_passed))

    workSheet_info_report.write(41, 0, "Total number of tests performed")
    workSheet_info_report.write(41, 1, str(TSDApp.criticity_blocking + TSDApp.criticity_warning + TSDApp.criticity_information))

    workSheet_test_report = new_wb.add_sheet('Test report', cell_overwrite_ok=True)

    lastRow = 0
    workSheet_test_report.write(lastRow, 0, 'Criticity')
    workSheet_test_report.write(lastRow, 1, 'Requirements')
    workSheet_test_report.write(lastRow, 2, 'Message')
    workSheet_test_report.write(lastRow, 3, 'Localisation')

    col1 = workSheet_test_report.col(0)
    col1.width = 256 * 15
    col2 = workSheet_test_report.col(1)
    col2.width = 256 * 45
    col3 = workSheet_test_report.col(2)
    col3.width = 256 * 50
    col4 = workSheet_test_report.col(3)
    col4.width = 256 * 25


    lastRow += 1
    blocking_style = xlwt.easyxf('pattern: pattern solid, fore_colour red;')
    warning_style = xlwt.easyxf('pattern: pattern solid, fore_colour yellow;')
    text_style = xlwt.easyxf('font: colour white, bold False;')

    for elem in return_list:

        if elem["criticity"].casefold().strip() == "blocking":
            workSheet_test_report.write(lastRow, 0, elem["criticity"], blocking_style)
        elif elem["criticity"].casefold().strip() == "warning":
            workSheet_test_report.write(lastRow, 0, elem["criticity"], warning_style)
        else:
            workSheet_test_report.write(lastRow, 0, elem["criticity"])

        workSheet_test_report.write(lastRow, 1, elem["testName"])

        if elem["localisation"] is None or elem["localisation"] == "":
            workSheet_test_report.write(lastRow, 2, "OK")
        else:
            workSheet_test_report.write(lastRow, 2, elem["message"])

        if elem["localisation"] is None or elem["localisation"] == "":
            workSheet_test_report.write(lastRow, 3, elem["localisation"])
            lastRow += 1

        if elem["testName"] != "Test_02043_18_04939_COH_2121":
            try:
                if elem["localisation"] is not None and elem["localisation"] != "":
                    if isinstance(elem["localisation"][0], str):
                        for index, element in enumerate(elem["localisation"]):
                            workSheet_test_report.write(lastRow + index, 3, element)

                        for index in range(1, len(elem["localisation"]) + 1):
                            workSheet_test_report.write(lastRow + index, 0, elem["criticity"], text_style)
                            workSheet_test_report.write(lastRow + index, 1, elem["testName"], text_style)

                        lastRow += index
                    else:
                        for index, element in enumerate(elem["localisation"]):
                            index_coloana = element[2]
                            link = "HYPERLINK(\"#\'" + str(element[0]) + "\'!$" + column_string(index_coloana + 1) + "$" + str(element[1] + 1) + "\",\"$" + column_string(index_coloana + 1) + "$" + str(element[1] + 1) + "\")"
                            workSheet_test_report.write(lastRow + index, 3, xlwt.Formula(link))

                        for index in range(1, len(elem["localisation"]) + 1):
                            workSheet_test_report.write(lastRow + index, 0, elem["criticity"], text_style)
                            workSheet_test_report.write(lastRow + index, 1, elem["testName"], text_style)

                        lastRow = lastRow + index
            except:
                TSDApp.tab1.textbox.setText("Warning: Only 65536 first rows filled in Test report (xls format limitation)")
                new_wb.save(path)
                return
        else:
            try:
                if elem["localisation"] is not None and elem["localisation"] != "":
                    for index, element in enumerate(elem["localisation"]):
                        index_coloana = element[2]
                        workSheet_test_report.write(lastRow + index, 3,"The cel " + column_string(index_coloana + 1) + str(element[1] + 1) + " from the sheet " + element[0] + " form the file " + TSDApp.DOC4Path.split("/")[-1])

                    for index in range(1, len(elem["localisation"]) + 1):
                        workSheet_test_report.write(lastRow + index, 0, elem["criticity"], text_style)
                        workSheet_test_report.write(lastRow + index, 1, elem["testName"], text_style)

                    lastRow = lastRow + index
            except:
                TSDApp.tab1.textbox.setText("Warning: Only 65536 first rows filled in Test report (xls format limitation)")
                new_wb.save(path)
                return

    new_wb.save(path)


def ExcelWrite2(return_list, workBook, TSDApp, path):

    if path.split('.')[-1] == 'xlsm':
        try:
            wb = openpyxl.load_workbook(path, keep_vba=True)
        except:
            return
    else:
        wb = openpyxl.load_workbook(path, keep_vba=False)

    index_test_report = -1
    index_info_report = -1
    for sheetname in TSDApp.WorkbookStats.sheetNames:
        if sheetname == 'test report':
            index_test_report = TSDApp.WorkbookStats.sheetNames.index('test report')
        if sheetname == 'report information':
            index_info_report = TSDApp.WorkbookStats.sheetNames.index('report information')

    if index_info_report == -1:
        workSheet_info_report = wb.create_sheet("Report information")

        workSheet_info_report['A1'] = "Tool version:"
        workSheet_info_report['B1'] = TSD_Checker_V8_5.appName

        workSheet_info_report['A3'] = "Criticity configuration file:"
        workSheet_info_report['B3'] = TSDApp.DOC9Path
        workSheet_info_report['C3'] = TSDApp.version_cesare_file

        workSheet_info_report['A4'] = "Extract CESARE file:"
        workSheet_info_report['B4'] = TSDApp.DOC8Path
        workSheet_info_report['C4'] = TSDApp.version_criticity_file

        workSheet_info_report['A5'] = "Customer effects file:"
        workSheet_info_report['B5'] = TSDApp.DOC7Path
        workSheet_info_report['C5'] = TSDApp.version_cutomer_effect

        workSheet_info_report['A6'] = "Diversity management file:"
        workSheet_info_report['B6'] = TSDApp.DOC13Path
        workSheet_info_report['C6'] = TSDApp.version_diversity_file

        workSheet_info_report['A7'] = "CESARE file reference:"
        if TSDApp.tab2.myTextBox7.toPlainText() == "":
            workSheet_info_report['B7'] = TSDApp.DOC8Link.split("/")[-3]
        else:
            workSheet_info_report['B7'] = TSDApp.tab2.myTextBox7.toPlainText()

        workSheet_info_report['A8'] = "Criticity configuration file reference:"
        if TSDApp.tab2.myTextBox8.toPlainText() == "":
            workSheet_info_report['B8'] = TSDApp.DOC9Link.split("/")[-3]
        else:
            workSheet_info_report['B8'] = TSDApp.tab2.myTextBox8.toPlainText()

        workSheet_info_report['A9'] = "Customer effect file reference:"
        if TSDApp.tab2.myTextBox9.toPlainText() == "":
            workSheet_info_report['B9'] = TSDApp.DOC7Link.split("/")[-3]
        else:
            workSheet_info_report['B9'] = TSDApp.tab2.myTextBox9.toPlainText()

        workSheet_info_report['A10'] = "Diversity management file reference:"
        if TSDApp.tab2.myTextBox10.toPlainText() == "":
            workSheet_info_report['B10'] = TSDApp.DOC13Link.split("/")[-3]
        else:
            workSheet_info_report['B10'] = TSDApp.tab2.myTextBox10.toPlainText()

        workSheet_info_report['A11'] = "FSE TSD template:"
        if TSDApp.DOC3Exists and not TSDApp.DOC4Exists and not TSDApp.DOC5Exists:
            workSheet_info_report['B11'] = TSDApp.DOC3Name
        else:
            workSheet_info_report['B11'] = "not used"

        workSheet_info_report['A12'] = "TSD Vehicle Funtion template:"
        if not TSDApp.DOC3Exists and TSDApp.DOC4Exists:
            workSheet_info_report['B12'] = TSDApp.DOC4Name
        else:
            workSheet_info_report['B12'] = "not used"

        workSheet_info_report['A13'] = "TSD System template:"
        if not TSDApp.DOC3Exists and not TSDApp.DOC4Exists and TSDApp.DOC5Exists:
            workSheet_info_report['B13'] = TSDApp.DOC5Name
        else:
            workSheet_info_report['B13'] = "not used"

        if TSDApp.DOC3Exists and not TSDApp.DOC4Exists and not TSDApp.DOC5Exists:
            workSheet_info_report['A14'] = "FSE TSD template reference:"
            if TSDApp.tab2.myTextBox11.toPlainText() == "":
                workSheet_info_report['B14'] = TSDApp.DOC3Link.split("/")[-3]
            else:
                workSheet_info_report['B14'] = TSDApp.tab2.myTextBox11.toPlainText()
        elif not TSDApp.DOC3Exists and TSDApp.DOC4Exists:
            workSheet_info_report['A14'] = "TSD Vehicle Function template reference:"
            if TSDApp.tab2.myTextBox12.toPlainText() == "":
                workSheet_info_report['B14'] = TSDApp.DOC4Link.split("/")[-3]
            else:
                workSheet_info_report['B14'] = TSDApp.tab2.myTextBox12.toPlainText()
        if not TSDApp.DOC3Exists and not TSDApp.DOC4Exists and TSDApp.DOC5Exists:
            workSheet_info_report['A14'] = "TSD System template reference:"
            if TSDApp.tab2.myTextBox13.toPlainText() == "":
                workSheet_info_report['B14'] = TSDApp.DOC5Link.split("/")[-3]
            else:
                workSheet_info_report['B14'] = TSDApp.tab2.myTextBox13.toPlainText()

        workSheet_info_report['A15'] = "Check level:"
        workSheet_info_report['B15'] = TSDApp.checkLevel

        workSheet_info_report['A17'] = "Date of the test:"
        workSheet_info_report['B17'] = time.strftime("%d/%m/%Y")

        workSheet_info_report['A18'] = "Time of the test:"
        workSheet_info_report['B18'] = time.strftime("%X")

        workSheet_info_report['A19'] = "Test duration:"
        workSheet_info_report['B19'] = time.strftime('%H:%M:%S', time.gmtime(TSDApp.end_time - TSDApp.start_time))

        workSheet_info_report['A20'] = "Opening duration:"
        workSheet_info_report['B20'] = time.strftime('%H:%M:%S', time.gmtime(TSDApp.opening_time - TSDApp.start_time))

        workSheet_info_report['A22'] = "TSD file checked:"
        workSheet_info_report['B22'] = TSDApp.DOC3Path

        workSheet_info_report['A23'] = "TSD function file checked:"
        workSheet_info_report['b23'] = TSDApp.DOC4Path

        workSheet_info_report['A24'] = "TSD system file checked:"
        workSheet_info_report['B24'] = TSDApp.DOC5Path

        workSheet_info_report['A26'] = "AMDEC:"
        workSheet_info_report['B26'] = TSDApp.tab1.myTextBox4.toPlainText()

        workSheet_info_report['A27'] = "Export MedialecMatrice:"
        workSheet_info_report['B27'] = TSDApp.tab1.myTextBox5.toPlainText()

        workSheet_info_report['A28'] = "Diagnostic Messagerie (ODX):"
        workSheet_info_report['B28'] = TSDApp.tab1.myTextBox6.toPlainText()

        workSheet_info_report['A29'] = "SubFamily:"
        workSheet_info_report['B29'] = TSDApp.tab1.myTextBox61.toPlainText()

        workSheet_info_report['A31'] = "Architecture type:"
        workSheet_info_report['B31'] = TSDApp.tab1.combo2.currentText()

        workSheet_info_report['A32'] = "Diversity Management:"
        workSheet_info_report['B32'] = TSDApp.tab1.combo3.currentText()

        workSheet_info_report['A33'] = "Project name:"
        workSheet_info_report['B33'] = TSDApp.tab1.combo1.currentText()

        workSheet_info_report['A35'] = "Status:"
        workSheet_info_report['B35'] = str(TSDApp.status)

        workSheet_info_report['A36'] = "Coverage Indicator:"
        workSheet_info_report['B36'] = str(TSDApp.coverage)[0:4] + "%"
        if str(TSDApp.coverage)[0:4] + "%" == "0.00%":
            workSheet_info_report['C36'] = "WARNING: The coverage indicator will not be calculated because at least one of its parameters is missing."

        workSheet_info_report['A37'] = "Convergence Indicator:"
        workSheet_info_report['B37'] = str(TSDApp.convergence)[0:4] + "%"
        if str(TSDApp.convergence)[0:4] + "%" == "0.00%":
            workSheet_info_report['C37'] = "WARNING: The convergence indicator will not be calculated because at least one of its parameters is missing."

        workSheet_info_report['A39'] = "Blocking Points Failed"
        workSheet_info_report['B39'] = str(TSDApp.criticity_blocking - TSDApp.criticity_blocking_passed)

        workSheet_info_report['A40'] = "Warning Points Failed"
        workSheet_info_report['B40'] = str(TSDApp.criticity_warning - TSDApp.criticity_warning_passed)

        workSheet_info_report['A41'] = "Information Points Failed"
        workSheet_info_report['B41'] = str(TSDApp.criticity_information - TSDApp.criticity_information_passed)

        workSheet_info_report['A42'] = "Total number of tests performed"
        workSheet_info_report['B42'] = str(TSDApp.criticity_blocking + TSDApp.criticity_warning + TSDApp.criticity_information)

    else:
        workSheet_info_report = wb.get_sheet_by_name("Report information")
        wb.remove_sheet(workSheet_info_report)
        workSheet_info_report = wb.create_sheet("Report information")

        workSheet_info_report['A1'] = "Tool version:"
        workSheet_info_report['B1'] = TSD_Checker_V8_5.appName

        workSheet_info_report['A3'] = "Criticity configuration file:"
        workSheet_info_report['B3'] = TSDApp.DOC9Path
        workSheet_info_report['C3'] = TSDApp.version_cesare_file

        workSheet_info_report['A4'] = "Extract CESARE file:"
        workSheet_info_report['B4'] = TSDApp.DOC8Path
        workSheet_info_report['C4'] = TSDApp.version_criticity_file

        workSheet_info_report['A5'] = "Customer effects file:"
        workSheet_info_report['B5'] = TSDApp.DOC7Path
        workSheet_info_report['C5'] = TSDApp.version_cutomer_effect

        workSheet_info_report['A6'] = "Diversity management file:"
        workSheet_info_report['B6'] = TSDApp.DOC13Path
        workSheet_info_report['C6'] = TSDApp.version_diversity_file

        workSheet_info_report['A7'] = "CESARE file reference:"
        if TSDApp.tab2.myTextBox7.toPlainText() == "":
            workSheet_info_report['B7'] = TSDApp.DOC8Link.split("/")[-3]
        else:
            workSheet_info_report['B7'] = TSDApp.tab2.myTextBox7.toPlainText()

        workSheet_info_report['A8'] = "Criticity configuration file reference:"
        if TSDApp.tab2.myTextBox8.toPlainText() == "":
            workSheet_info_report['B8'] = TSDApp.DOC9Link.split("/")[-3]
        else:
            workSheet_info_report['B8'] = TSDApp.tab2.myTextBox8.toPlainText()

        workSheet_info_report['A9'] = "Customer effect file reference:"
        if TSDApp.tab2.myTextBox9.toPlainText() == "":
            workSheet_info_report['B9'] = TSDApp.DOC7Link.split("/")[-3]
        else:
            workSheet_info_report['B9'] = TSDApp.tab2.myTextBox9.toPlainText()

        workSheet_info_report['A10'] = "Diversity management file reference:"
        if TSDApp.tab2.myTextBox10.toPlainText() == "":
            workSheet_info_report['B10'] = TSDApp.DOC13Link.split("/")[-3]
        else:
            workSheet_info_report['B10'] = TSDApp.tab2.myTextBox10.toPlainText()

        workSheet_info_report['A11'] = "FSE TSD template:"
        if TSDApp.DOC3Exists and not TSDApp.DOC4Exists and not TSDApp.DOC5Exists:
            workSheet_info_report['B11'] = TSDApp.DOC3Name
        else:
            workSheet_info_report['B11'] = "not used"

        workSheet_info_report['A12'] = "TSD Vehicle Funtion template:"
        if not TSDApp.DOC3Exists and TSDApp.DOC4Exists:
            workSheet_info_report['B12'] = TSDApp.DOC4Name
        else:
            workSheet_info_report['B12'] = "not used"

        workSheet_info_report['A13'] = "TSD System template:"
        if not TSDApp.DOC3Exists and not TSDApp.DOC4Exists and TSDApp.DOC5Exists:
            workSheet_info_report['B13'] = TSDApp.DOC5Name
        else:
            workSheet_info_report['B13'] = "not used"

        if TSDApp.DOC3Exists and not TSDApp.DOC4Exists and not TSDApp.DOC5Exists:
            workSheet_info_report['A14'] = "FSE TSD template reference:"
            if TSDApp.tab2.myTextBox11.toPlainText() == "":
                workSheet_info_report['B14'] = TSDApp.DOC3Link.split("/")[-3]
            else:
                workSheet_info_report['B14'] = TSDApp.tab2.myTextBox11.toPlainText()
        elif not TSDApp.DOC3Exists and TSDApp.DOC4Exists:
            workSheet_info_report['A14'] = "TSD Vehicle Function template reference:"
            if TSDApp.tab2.myTextBox12.toPlainText() == "":
                workSheet_info_report['B14'] = TSDApp.DOC4Link.split("/")[-3]
            else:
                workSheet_info_report['B14'] = TSDApp.tab2.myTextBox12.toPlainText()
        if not TSDApp.DOC3Exists and not TSDApp.DOC4Exists and TSDApp.DOC5Exists:
            workSheet_info_report['A14'] = "TSD System template reference:"
            if TSDApp.tab2.myTextBox13.toPlainText() == "":
                workSheet_info_report['B14'] = TSDApp.DOC5Link.split("/")[-3]
            else:
                workSheet_info_report['B14'] = TSDApp.tab2.myTextBox13.toPlainText()

        workSheet_info_report['A15'] = "Check level:"
        workSheet_info_report['B15'] = TSDApp.checkLevel

        workSheet_info_report['A17'] = "Date of the test:"
        workSheet_info_report['B17'] = time.strftime("%d/%m/%Y")

        workSheet_info_report['A18'] = "Time of the test:"
        workSheet_info_report['B18'] = time.strftime("%X")

        workSheet_info_report['A19'] = "Test duration:"
        workSheet_info_report['B19'] = time.strftime('%H:%M:%S', time.gmtime(TSDApp.end_time - TSDApp.start_time))

        workSheet_info_report['A20'] = "Opening duration:"
        workSheet_info_report['B20'] = time.strftime('%H:%M:%S', time.gmtime(TSDApp.opening_time - TSDApp.start_time))

        workSheet_info_report['A22'] = "TSD file checked:"
        workSheet_info_report['B22'] = TSDApp.DOC3Path

        workSheet_info_report['A23'] = "TSD function file checked:"
        workSheet_info_report['b23'] = TSDApp.DOC4Path

        workSheet_info_report['A24'] = "TSD system file checked:"
        workSheet_info_report['B24'] = TSDApp.DOC5Path

        workSheet_info_report['A26'] = "AMDEC:"
        workSheet_info_report['B26'] = TSDApp.tab1.myTextBox4.toPlainText()

        workSheet_info_report['A27'] = "Export MedialecMatrice:"
        workSheet_info_report['B27'] = TSDApp.tab1.myTextBox5.toPlainText()

        workSheet_info_report['A28'] = "Diagnostic Messagerie (ODX):"
        workSheet_info_report['B28'] = TSDApp.tab1.myTextBox6.toPlainText()

        workSheet_info_report['A29'] = "SubFamily:"
        workSheet_info_report['B29'] = TSDApp.tab1.myTextBox61.toPlainText()

        workSheet_info_report['A31'] = "Architecture type:"
        workSheet_info_report['B31'] = TSDApp.tab1.combo2.currentText()

        workSheet_info_report['A32'] = "Diversity Management:"
        workSheet_info_report['B32'] = TSDApp.tab1.combo3.currentText()

        workSheet_info_report['A33'] = "Project name:"
        workSheet_info_report['B33'] = TSDApp.tab1.combo1.currentText()

        workSheet_info_report['A35'] = "Status:"
        workSheet_info_report['B35'] = str(TSDApp.status)

        workSheet_info_report['A36'] = "Coverage Indicator:"
        workSheet_info_report['B36'] = str(TSDApp.coverage)[0:4] + "%"
        if str(TSDApp.coverage)[0:4] + "%" == "0.00%":
            workSheet_info_report['C36'] = "WARNING: The coverage indicator will not be calculated because at least one of its parameters is missing."

        workSheet_info_report['A37'] = "Convergence Indicator:"
        workSheet_info_report['B37'] = str(TSDApp.convergence)[0:4] + "%"
        if str(TSDApp.convergence)[0:4] + "%" == "0.00%":
            workSheet_info_report['C37'] = "WARNING: The convergence indicator will not be calculated because at least one of its parameters is missing."

        workSheet_info_report['A39'] = "Blocking Points Failed"
        workSheet_info_report['B39'] = str(TSDApp.criticity_blocking - TSDApp.criticity_blocking_passed)

        workSheet_info_report['A40'] = "Warning Points Failed"
        workSheet_info_report['B40'] = str(TSDApp.criticity_warning - TSDApp.criticity_warning_passed)

        workSheet_info_report['A41'] = "Information Points Failed"
        workSheet_info_report['B41'] = str(TSDApp.criticity_information - TSDApp.criticity_information_passed)

        workSheet_info_report['A42'] = "Total number of tests performed"
        workSheet_info_report['B42'] = str(TSDApp.criticity_blocking + TSDApp.criticity_warning + TSDApp.criticity_information)

    workSheet_info_report.column_dimensions['A'].width = 40
    workSheet_info_report.column_dimensions['B'].width = 140


    if index_test_report == -1:
        workSheet_test_report = wb.create_sheet("Test report")

        lastRow = 1
        workSheet_test_report['A1'] = "Criticity"
        workSheet_test_report['B1'] = "Requirements"
        workSheet_test_report['C1'] = "Message"
        workSheet_test_report['D1'] = "Localisation"

        lastRow += 1

        my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
        blocking_style = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_red)
        my_yellow = openpyxl.styles.colors.Color(rgb='00FFFF00')
        warning_style = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_yellow)
        text_style = Font(color='FFFFFFFF')

        for elem in return_list:

            if elem["criticity"].casefold().strip() == "blocking":
                workSheet_test_report.cell(lastRow, 1, elem["criticity"]).fill = blocking_style
            elif elem["criticity"].casefold().strip() == "warning":
                workSheet_test_report.cell(lastRow, 1, elem["criticity"]).fill = warning_style
            else:
                workSheet_test_report.cell(lastRow, 1, elem["criticity"])

            workSheet_test_report.cell(lastRow, 2, elem["testName"])

            if elem["localisation"] is None or elem["localisation"] == "":
                workSheet_test_report.cell(lastRow, 3, "OK")
            else:
                workSheet_test_report.cell(lastRow, 3, elem["message"])

            if elem["localisation"] is None or elem["localisation"] == "":
                workSheet_test_report.cell(lastRow, 4, elem["localisation"])
                lastRow += 1

            if elem["testName"] != "Test_02043_18_04939_COH_2121":
                if elem["localisation"] is not None and elem["localisation"] != "":
                    if isinstance(elem["localisation"][0], str):
                        for index, element in enumerate(elem["localisation"]):
                            workSheet_test_report.cell(lastRow + index, 4, element)

                        if len(elem['localisation']) > 1:
                            for index in range(1, len(elem["localisation"])):
                                workSheet_test_report.cell(lastRow + index, 1, elem["criticity"]).font = text_style
                                workSheet_test_report.cell(lastRow + index, 2, elem["testName"]).font = text_style

                        lastRow += index + 1
                    else:
                        for index, element in enumerate(elem["localisation"]):
                            index_coloana = element[2]
                            workSheet_test_report.cell(lastRow + index, 4).value = '$' + column_string(index_coloana + 1) + '$' + str(element[1] + 1)
                            workSheet_test_report.cell(lastRow + index, 4).hyperlink = '#%s!%s' % ("'" + str(element[0]) + "'", column_string(index_coloana + 1) + str(element[1] + 1))

                        if len(elem['localisation']) > 1:
                            for index in range(1, len(elem["localisation"])):
                                workSheet_test_report.cell(lastRow + index, 1, elem["criticity"]).font = text_style
                                workSheet_test_report.cell(lastRow + index, 2, elem["testName"]).font = text_style

                        lastRow += index + 1
            else:
                if elem["localisation"] is not None and elem["localisation"] != "":
                    for index, element in enumerate(elem["localisation"]):
                        index_coloana = element[2]
                        workSheet_test_report.cell(lastRow + index, 4).value = 'The cel ' + column_string(index_coloana + 1) + str(element[1] + 1) + " from the sheet " + element[0] + " from the file " + TSDApp.DOC4Path.split("/")[-1]
                        # workSheet_test_report.cell(lastRow + index, 4).hyperlink = '#%s!%s' % ("'" + str(element[0]) + "'", column_string(index_coloana + 1) + str(element[1] + 1))

                    if len(elem['localisation']) > 1:
                        for index in range(1, len(elem["localisation"])):
                            workSheet_test_report.cell(lastRow + index, 1, elem["criticity"]).font = text_style
                            workSheet_test_report.cell(lastRow + index, 2, elem["testName"]).font = text_style

                    lastRow += index + 1
    else:
        workSheet_test_report = wb.get_sheet_by_name("Test report")
        wb.remove_sheet(workSheet_test_report)
        workSheet_test_report = wb.create_sheet("Test report")

        lastRow = 1
        workSheet_test_report['A1'] = "Criticity"
        workSheet_test_report['B1'] = "Requirements"
        workSheet_test_report['C1'] = "Message"
        workSheet_test_report['D1'] = "Localisation"

        lastRow += 1

        my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
        blocking_style = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_red)
        my_yellow = openpyxl.styles.colors.Color(rgb='00FFFF00')
        warning_style = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_yellow)
        text_style = Font(color='FFFFFFFF')

        for elem in return_list:

            if elem["criticity"].casefold().strip() == "blocking":
                workSheet_test_report.cell(lastRow, 1, elem["criticity"]).fill = blocking_style
            elif elem["criticity"].casefold().strip() == "warning":
                workSheet_test_report.cell(lastRow, 1, elem["criticity"]).fill = warning_style
            else:
                workSheet_test_report.cell(lastRow, 1, elem["criticity"])

            workSheet_test_report.cell(lastRow, 2, elem["testName"])

            if elem["localisation"] is None or elem["localisation"] == "":
                workSheet_test_report.cell(lastRow, 3, "OK")
            else:
                workSheet_test_report.cell(lastRow, 3, elem["message"])

            if elem["localisation"] is None or elem["localisation"] == "":
                workSheet_test_report.cell(lastRow, 4, elem["localisation"])
                lastRow += 1

            if elem["testName"] != "Test_02043_18_04939_COH_2121":
                if elem["localisation"] is not None and elem["localisation"] != "":
                    if isinstance(elem["localisation"][0], str):
                        for index, element in enumerate(elem["localisation"]):
                            workSheet_test_report.cell(lastRow + index, 4, element)

                        if len(elem['localisation']) > 1:
                            for index in range(1, len(elem["localisation"])):
                                workSheet_test_report.cell(lastRow + index, 1, elem["criticity"]).font = text_style
                                workSheet_test_report.cell(lastRow + index, 2, elem["testName"]).font = text_style

                        lastRow += index + 1
                    else:
                        for index, element in enumerate(elem["localisation"]):
                            index_coloana = element[2]
                            workSheet_test_report.cell(lastRow + index, 4).value = '$' + column_string(index_coloana + 1) + '$' + str(element[1] + 1)
                            workSheet_test_report.cell(lastRow + index, 4).hyperlink = '#%s!%s' % ("'" + str(element[0]) + "'", column_string(index_coloana + 1) + str(element[1] + 1))

                        if len(elem['localisation']) > 1:
                            for index in range(1, len(elem["localisation"])):
                                workSheet_test_report.cell(lastRow + index, 1, elem["criticity"]).font = text_style
                                workSheet_test_report.cell(lastRow + index, 2, elem["testName"]).font = text_style

                        lastRow += index + 1
            else:
                if elem["localisation"] is not None and elem["localisation"] != "":
                    for index, element in enumerate(elem["localisation"]):
                        index_coloana = element[2]
                        workSheet_test_report.cell(lastRow + index, 4).value = 'The cel ' + column_string(index_coloana + 1) + str(element[1] + 1) + " from the sheet " + element[0] + " from the file " + TSDApp.DOC4Path.split("/")[-1]
                        # workSheet_test_report.cell(lastRow + index, 4).hyperlink = '#%s!%s' % ("'" + str(element[0]) + "'", column_string(index_coloana + 1) + str(element[1] + 1))

                    if len(elem['localisation']) > 1:
                        for index in range(1, len(elem["localisation"])):
                            workSheet_test_report.cell(lastRow + index, 1, elem["criticity"]).font = text_style
                            workSheet_test_report.cell(lastRow + index, 2, elem["testName"]).font = text_style

                    lastRow += index + 1

    workSheet_test_report.column_dimensions['A'].width = 20
    workSheet_test_report.column_dimensions['B'].width = 40
    workSheet_test_report.column_dimensions['C'].width = 80
    workSheet_test_report.column_dimensions['D'].width = 20

    workSheet_test_report.auto_filter.ref = workSheet_test_report.dimensions

    wb.save(workBook)
    #wb.save("C:\\Users\\msnecula\\Downloads\\documente_TSD\\aaa.xlsx")

def TestReturnName(criticity, testName, message, name, workBook, TSDApp):
    return_dict = {}
    return_dict["criticity"] = criticity
    return_dict["testName"] = testName
    return_dict["message"] = message
    return_dict["localisation"] = name
    TSDApp.return_list.append(return_dict)

    if criticity.casefold() == "blocking":
        TSDApp.criticity_blocking += 1
    else:
        if criticity.casefold() == "warning":
            TSDApp.criticity_warning += 1
        else:
            TSDApp.criticity_information += 1

    tempString = str()
    if name is None or name == "":
        tempString = "OK"
        if criticity.casefold() == "blocking":
            TSDApp.criticity_blocking_passed += 1
        else:
            if criticity.casefold() == "warning":
                TSDApp.criticity_warning_passed += 1
            else:
                TSDApp.criticity_information_passed += 1
    else:
        tempString = "NOK"

    textBoxText = TSDApp.tab1.textbox.toPlainText()
    textBoxText = textBoxText + "\n" + testName + " " + tempString
    TSDApp.tab1.textbox.setText(textBoxText)
    TSDApp.tab1.textbox.moveCursor(QtGui.QTextCursor.End)

    TSDApp.IncrementProgressBar()
