import TSD_Checker_V8_6
from ExcelEdit import TestReturnName as show
from ErrorMessages import errorMessagesDict as error
import openpyxl


def coverageIndicator(workBook, TSDApp):
    testName = "Test_02043_18_04939_IND_6030"
    index = 0
    for sheetname in TSDApp.WorkbookStats.sheetNames:
        if sheetname == 'tableau':
            index = TSDApp.WorkbookStats.sheetNames.index('tableau')
            break
        if sheetname == 'table':
            index = TSDApp.WorkbookStats.sheetNames.index('table')
            break

    workSheet = workBook.sheet_by_index(index)
    nrCols = workSheet.ncols
    nrRows = workSheet.nrows

    refColBase = -1
    refColDTC = -1
    refCelParam = -1
    refCelDiag = -1
    refRowBase = -1

    for index1 in range(0, TSDApp.WorkbookStats.tableLastRow):
        for index2 in range(0, TSDApp.WorkbookStats.tableLastCol):
            cel = str(workSheet.cell(index1, index2).value).casefold().strip().replace("\n","")
            if cel == "Constituant défaillant détecté".casefold() or cel == "Defective part".casefold():
                refColBase = index2
                refRowBase = index1
            if cel == "Code défaut".casefold() or cel == "Data Trouble code".casefold():
                refColDTC = index2
            if cel == "mesures et commandes (Mesure Parametre et Test Actionneur) / Tests de cohérence".casefold() or cel == "Read data or I/O control".casefold():
                refCelParam = index2
            if cel == "DIAGNOSTIC DEBARQUE".casefold() or cel == "Non-embedded diagnosis".casefold():
                refCelDiag = index2


    NbComponentsOfTheFunction = 0
    NbComponentWithDiagPossible = 0

    name = []
    if refColBase == -1:
        name.append("Constituant défaillant détecté/Defective part")
    if refColDTC == -1:
        name.append("Code défaut/Data Trouble code")
    if refCelParam == -1:
        name.append("mesures et commandes (Mesure Parametre et Test Actionneur) / Tests de cohérence/Read data or I/O control")
    if refCelDiag == -1:
        name.append("DIAGNOSTIC DEBARQUE/Non-embedded diagnosis")
    if not name:
        name = None


    if refColBase != -1 and refColDTC != -1 and refCelParam != -1 and refCelDiag != -1:
        for index in range(refRowBase + 2, nrRows):
            if str(workSheet.cell(index, refColBase).value) is not None and str(workSheet.cell(index,refColBase).value) != "":
                NbComponentsOfTheFunction += 1
                if  (str(workSheet.cell(index, refColDTC).value) is not None and str(workSheet.cell(index, refColDTC).value) !="" and str(workSheet.cell(index,refColDTC).value) != "NO DTC") or (
                        str(workSheet.cell(index, refCelParam).value) is not None and str(workSheet.cell(index, refCelParam).value) != "" and str(workSheet.cell(index,refCelParam).value) != "N/A") or (
                        str(workSheet.cell(index, refCelDiag).value) is not None and str(workSheet.cell(index, refCelDiag).value) != "" and str(workSheet.cell(index,refCelDiag).value) != "N/A"):
                    NbComponentWithDiagPossible += 1
        show("", testName, error[testName], name, workBook, TSDApp)
        try:
            return (NbComponentWithDiagPossible / NbComponentsOfTheFunction)
        except:
            text = TSDApp.tab1.textbox.toPlainText()
            TSDApp.tab1.textbox.setText(text + '\n' + "Warning: The coverage indicator will not be calculated because there are no records!")
            return str(0.00000)
    else:
        # warning = "WARNING: The coverage indicator will not be calculated because at least one of its parameters is missing."
        # textBoxText = TSDApp.tab1.textbox.toPlainText()
        # textBoxText = textBoxText + "\n" + warning
        # TSDApp.tab1.textbox.setText(textBoxText)
        show("", testName, error[testName], name, workBook, TSDApp)
        return str(0.00000)


def convergenceIndicator(workBook, TSDApp, path):
    testName = "Test_02043_18_04939_IND_6140"

    index = -1
    for sheetname in TSDApp.WorkbookStats.sheetNames:
        if sheetname == 'tableau':
            index = TSDApp.WorkbookStats.sheetNames.index('tableau')
            break
        if sheetname == 'table':
            index = TSDApp.WorkbookStats.sheetNames.index('table')
            break

    rb_sheet = workBook.sheet_by_index(index)
    nrCols = rb_sheet.ncols
    nrRows = rb_sheet.nrows

    refColBase = -1
    refColDTC = -1
    refCelParam = -1
    refCelDiag = -1
    refCelEff = -1
    refCelVoyant = -1

    TSDApp.refSignature = -1
    refCritere = -1

    for index1 in range(0, nrRows):
        for index2 in range(0, nrCols):
            cel = str(rb_sheet.cell(index1, index2).value).casefold().strip().replace("\n", "")
            if cel == "Critère de decision".casefold() or cel == "decision criterion":
                refCritere = index2
                refRowIndex = index1
            if cel == "Unique Test Signature".casefold():
                TSDApp.refSignature = index2
                refRowIndex = index1
        if refCritere != -1 and TSDApp.refSignature != -1:
            break

    for index1 in range(0, nrRows):
        for index2 in range(0, nrCols):
            cel = str(rb_sheet.cell(index1, index2).value).casefold().strip().replace("\n","")
            if cel == "Constituant défaillant détecté".casefold() or cel == "Defective part".casefold():
                refColBase = index2
                refRowIndex = index1
            if cel == "Code défaut".casefold() or cel == "Data Trouble code".casefold():
                refColDTC = index2
            if cel == "mesures et commandes (Mesure Parametre et Test Actionneur) / Tests de cohérence".casefold() or cel == "Read data or I/O control".casefold():
                refCelParam = index2
            if cel == "DIAGNOSTIC DEBARQUE".casefold() or cel == "Non-embedded diagnosis".casefold():
                refCelDiag = index2
            if cel == "Effet(s) client(s)".casefold() or cel == "Technical effect".casefold():
                refCelEff = index2
            if cel == "Voyant(s) ou message(s)".casefold() or cel == "HMI(Indicator lights/messages)".casefold():
                refCelVoyant = index2
        if refColBase != -1 or refColDTC != -1 or refCelParam != -1 or refCelDiag != -1 or refCelEff != -1 or refCelVoyant != -1:
            break

    name = []
    if refColBase == -1:
        name.append("Constituant défaillant détecté/Defective part")
    if refColDTC == -1:
        name.append("Code défaut/Data Trouble code")
    if refCelParam == -1:
        name.append("mesures et commandes (Mesure Parametre et Test Actionneur) / Tests de cohérence/Read data or I/O control")
    if refCelDiag == -1:
        name.append("DIAGNOSTIC DEBARQUE/Non-embedded diagnosis")
    if refCelEff == -1:
        name.append("Effet(s) client(s)/Technical effect")
    if refCelVoyant == -1:
        name.append("Voyant(s) ou message(s)/HMI(Indicator lights/messages)")
    if not name:
        name = None

    if refColBase == -1 or refColDTC == -1 or refCelParam == -1 or refCelDiag == -1 or refCelEff == -1 or refCelVoyant == -1:
        # warning = "WARNING: The convergence indicator will not be calculated because at least one of its parameters is missing."
        # textBoxText = TSDApp.tab1.textbox.toPlainText()
        # textBoxText = textBoxText + "\n" + warning
        # TSDApp.tab1.textbox.setText(textBoxText)
        show("", testName, error[testName], name, workBook, TSDApp)
        return str(0.00000)
    else:
        if path.split('.')[-1] == 'xls':
            NbUniqueSignatureTests = 0
            NbAMDECLine = 0

            for index in range(TSDApp.tableFirstInfoRow, nrRows):
                if rb_sheet.cell(index, refColBase).value != "":
                    NbAMDECLine += 1
                    dict = {}
                    dict['value'] = [rb_sheet.cell(index, refColDTC).value, rb_sheet.cell(index, refCelParam).value,rb_sheet.cell(index, refCelDiag).value, rb_sheet.cell(index, refCelEff).value, rb_sheet.cell(index, refCelVoyant).value]
                    dict['localisation'] = index
                    TSDApp.unique_items.append(dict)
                    TSDApp.unique_list.append([rb_sheet.cell(index, refColDTC).value, rb_sheet.cell(index, refCelParam).value,rb_sheet.cell(index, refCelDiag).value, rb_sheet.cell(index, refCelEff).value, rb_sheet.cell(index, refCelVoyant).value])

            for element in TSDApp.unique_items:
                if TSDApp.unique_list.count(element['value']) == 1:
                    NbUniqueSignatureTests += 1

        else:
            if path.split('.')[-1] == 'xlsm':
                wb = openpyxl.load_workbook(path, keep_vba=True)
            else:
                wb = openpyxl.load_workbook(path, keep_vba=False)

            if TSDApp.refSignature == -1:
                if "tableau" in wb.sheetnames:
                    workSheet = wb.get_sheet_by_name("tableau")
                elif "Table" in wb.sheetnames:
                    workSheet = wb.get_sheet_by_name("Table")

                workSheet.cell(TSDApp.tableHeaderRow + 1, TSDApp.WorkbookStats.tableLastCol + 1, "Unique Test Signature")
                NbUniqueSignatureTests = 0
                NbAMDECLine = 0
                unique_items = []
                unique_list = []

                for index in range(TSDApp.tableFirstInfoRow + 1, nrRows + 1):
                    if workSheet.cell(index, refColBase + 1).value != "":
                        NbAMDECLine += 1
                        dict = {}
                        dict['value'] = [workSheet.cell(index, refColDTC + 1).value, workSheet.cell(index, refCelParam + 1).value,workSheet.cell(index, refCelDiag + 1).value, workSheet.cell(index, refCelEff + 1).value, workSheet.cell(index, refCelVoyant + 1).value]
                        dict['localisation'] = index
                        unique_items.append(dict)
                        unique_list.append([workSheet.cell(index, refColDTC + 1).value, workSheet.cell(index, refCelParam + 1).value,workSheet.cell(index, refCelDiag + 1).value, workSheet.cell(index, refCelEff + 1).value, workSheet.cell(index, refCelVoyant + 1).value])

                for element in unique_items:
                    if unique_list.count(element['value']) == 1:
                        workSheet.cell(element['localisation'], TSDApp.WorkbookStats.tableLastCol + 1,'1')
                        NbUniqueSignatureTests += 1
                    else:
                        for elem in unique_items:
                            if element['value'] == elem['value']:
                                workSheet.cell(element['localisation'], TSDApp.WorkbookStats.tableLastCol + 1, '0')
                wb.save(path)

            else:
                if "tableau" in wb.sheetnames:
                    workSheet = wb.get_sheet_by_name("tableau")
                elif "Table" in wb.sheetnames:
                    workSheet = wb.get_sheet_by_name("Table")

                NbUniqueSignatureTests = 0
                NbAMDECLine = 0
                unique_items = []
                unique_list = []

                for index in range(TSDApp.tableFirstInfoRow + 1, nrRows + 1):
                    if workSheet.cell(index, refColBase + 1).value != "":
                        NbAMDECLine += 1
                        dict = {}
                        dict['value'] = [workSheet.cell(index, refColDTC + 1).value,
                                         workSheet.cell(index, refCelParam + 1).value,
                                         workSheet.cell(index, refCelDiag + 1).value,
                                         workSheet.cell(index, refCelEff + 1).value,
                                         workSheet.cell(index, refCelVoyant + 1).value]
                        dict['localisation'] = index
                        unique_items.append(dict)
                        unique_list.append(
                            [workSheet.cell(index, refColDTC + 1).value, workSheet.cell(index, refCelParam + 1).value,
                             workSheet.cell(index, refCelDiag + 1).value, workSheet.cell(index, refCelEff + 1).value, workSheet.cell(index, refCelVoyant + 1).value])

                for element in unique_items:
                    if unique_list.count(element['value']) == 1:
                        workSheet.cell(element['localisation'], TSDApp.refSignature + 1, '1')
                        NbUniqueSignatureTests += 1
                    else:
                        for elem in unique_items:
                            if element['value'] == elem['value']:
                                workSheet.cell(element['localisation'], TSDApp.refSignature + 1, '0')
                wb.save(path)

        show("", testName, error[testName], name, workBook, TSDApp)
        try:
            return (NbUniqueSignatureTests / NbAMDECLine)
        except:
            text = TSDApp.tab1.textbox.toPlainText()
            TSDApp.tab1.textbox.setText(text + '\n' + "Warning: The covergence indicator will not be calculated because there are no records!")
            return str(0.00000)
