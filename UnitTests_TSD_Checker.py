import unittest
import TSD_Checker
import xlrd
import openpyxl


#Requirements for General structure
class Test_02043_18_04939_STRUCT_0000(unittest.TestCase):


    def setUp(self):

        self.app = TSD_Checker.QApplication(TSD_Checker.sys.argv)
        self.TSD_checker = TSD_Checker.Test()

    def test_ok_xls(self):
        self.workbook = xlrd.open_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0000/doc4_ok_xls.xls", formatting_info=True)
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0000_XLS(self.workbook), 1)

    def test_ok_uppercase_xls(self):
        self.workbook = xlrd.open_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0000/doc4_ok_uppercase_xls.xls", formatting_info=True)
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0000_XLS(self.workbook), 1)

    def test_ok_french_xls(self):
        self.workbook = xlrd.open_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0000/doc4_ok_french_xls.xls", formatting_info=True)
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0000_XLS(self.workbook), 1)

    def test_notok_xls(self):
        self.workbook = xlrd.open_workbook(
            "C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0000/doc4_notok_xls.xls",
            formatting_info=True)
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0000_XLS(self.workbook), False)

    def test_notok_french_xls(self):
        self.workbook = xlrd.open_workbook(
            "C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0000/doc4_notok_french_xls.xls",
            formatting_info=True)
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0000_XLS(self.workbook), False)




    def test_ok_xlsx(self):
        self.workbook = openpyxl.load_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0000/doc4_ok_xlsx.xlsx")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0000_XLSX_XLSM(self.workbook), 1)

    def test_ok_uppercase_xlsx(self):
        self.workbook = openpyxl.load_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0000/doc4_ok_uppercase_xlsx.xlsx")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0000_XLSX_XLSM(self.workbook), 1)

    def test_ok_french_xlsx(self):
        self.workbook = openpyxl.load_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0000/doc4_ok_french_xlsx.xlsx")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0000_XLSX_XLSM(self.workbook), 1)

    def test_notok_xlsx(self):
        self.workbook = openpyxl.load_workbook(
            "C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0000/doc4_notok_xlsx.xlsx")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0000_XLSX_XLSM(self.workbook), False)

    def test_notok_french_xlsx(self):
        self.workbook = openpyxl.load_workbook(
            "C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0000/doc4_notok_french_xlsx.xlsx")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0000_XLSX_XLSM(self.workbook), False)



    def test_ok_xlsm(self):
        self.workbook = openpyxl.load_workbook(
            "C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0000/doc4_ok_xlsm.xlsm")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0000_XLSX_XLSM(self.workbook), 1)

    def test_ok_uppercase_xlsm(self):
        self.workbook = openpyxl.load_workbook(
            "C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0000/doc4_ok_uppercase_xlsm.xlsm")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0000_XLSX_XLSM(self.workbook), 1)

    def test_ok_french_xlsm(self):
        self.workbook = openpyxl.load_workbook(
            "C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0000/doc4_ok_french_xlsm.xlsm")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0000_XLSX_XLSM(self.workbook), 1)

    def test_notok_xlsm(self):
        self.workbook = openpyxl.load_workbook(
            "C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0000/doc4_notok_xlsm.xlsm")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0000_XLSX_XLSM(self.workbook), False)

    def test_notok_french_xlsm(self):
        self.workbook = openpyxl.load_workbook(
            "C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0000/doc4_notok_french_xlsm.xlsm")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0000_XLSX_XLSM(self.workbook), False)


class Test_02043_18_04939_STRUCT_0005(unittest.TestCase):
    def setUp(self):
        self.app = TSD_Checker.QApplication(TSD_Checker.sys.argv)
        self.TSD_checker = TSD_Checker.Test()

    def test_notok_formula_xls(self):

        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0005_XLS("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0005/doc4_notok_formula_xls.xls"), 0)

    def test_ok_notformula_xls(self):

        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0005_XLS("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0005/doc4_ok_notformula_xls.xls"), 1)

    def test_ok_empty_xls(self):
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0005_XLS("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0005/doc4_ok_empty_xls.xls"), 1)



    def test_notok_formula_xlsx(self):
        self.workbook = openpyxl.load_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0005/doc4_notok_formula_xlsx.xlsx")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0005_XLSX_XLSM(self.workbook), 1)
    def test_ok_notformula_xlsx(self):
        self.workbook = openpyxl.load_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0005/doc4_ok_empty_xlsx.xlsx")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0005_XLSX_XLSM(self.workbook), 1)

    def test_ok_empty_xlsx(self):
        self.workbook = openpyxl.load_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0005/doc4_ok_empty_xlsx.xlsx")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0005_XLSX_XLSM(self.workbook), 1)



    def test_notok_formula_xlsm(self):
        self.workbook = openpyxl.load_workbook( "C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0005/doc4_notok_formula_xlsm.xlsm")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0005_XLSX_XLSM(self.workbook), 1)

    def test_ok_notformula_xlsm(self):
        self.workbook = openpyxl.load_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0005/doc4_ok_empty_xlsm.xlsm")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0005_XLSX_XLSM(self.workbook), 1)

    def test_ok_empty_xlsm(self):
        self.workbook = openpyxl.load_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0005/doc4_ok_empty_xlsm.xlsm")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0005_XLSX_XLSM(self.workbook), 1)


class Test_02043_18_04939_STRUCT_0010(unittest.TestCase):
    def setUp(self):
        self.app = TSD_Checker.QApplication(TSD_Checker.sys.argv)
        self.TSD_checker = TSD_Checker.Test()



    def test_ok_IsReference_xls(self):
        self.workbook = xlrd.open_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0010/doc4_ok_IsReference_xls.xls", formatting_info=True)
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0010_XLS(self.workbook), 1)

    def test_notok_IsEmpty_xls(self):
        self.workbook = xlrd.open_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0010/doc4_notok_IsEmpty_xls.xls", formatting_info=True)
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0010_XLS(self.workbook), 0)


    def test_ok_IsReference_xlsx(self):
        self.workbook = openpyxl.load_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0010/doc4_ok_IsReference_xlsx.xlsx")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0010_XLSX_XLSM(self.workbook), 1)


    def test_notok_IsEmpty_xlsx(self):
        self.workbook = openpyxl.load_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0010/doc4_notok_IsEmpty_xlsx.xlsx")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0010_XLSX_XLSM(self.workbook), 0)


    def test_ok_IsReference_xlsm(self):
        self.workbook = openpyxl.load_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0010/doc4_ok_IsReference_xlsm.xlsm")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0010_XLSX_XLSM(self.workbook), 1)

    def test_notok_IsEmpty_xlsm(self):
        self.workbook = openpyxl.load_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0010/doc4_notok_IsEmpty_xlsm.xlsm")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0010_XLSX_XLSM(self.workbook), 0)


class Test_02043_18_04939_STRUCT_0011(unittest.TestCase):
    def setUp(self):
        self.app = TSD_Checker.QApplication(TSD_Checker.sys.argv)
        self.TSD_checker = TSD_Checker.Test()

    def test_ok_equal_xls(self):
        self.workbook = xlrd.open_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0011/doc4_ok_equal_xls.xls")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0011_XLS(self.workbook), 1)

    def test_notok_notequal_xls(self):
        self.workbook = xlrd.open_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0011/doc4_notok_notequal_xls.xls")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0011_XLS(self.workbook), 0)

    def test_ok_equal_xlsx(self):
        self.workbook = openpyxl.load_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0011/doc4_ok_equal_xlsx.xlsx")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0011_XLSX_XLSM(self.workbook), 1)

    def test_notok_notequal_xlsx(self):
        self.workbook = openpyxl.load_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0011/doc4_notok_notequal_xlsx.xlsx")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0011_XLSX_XLSM(self.workbook), 0)


class Test_02043_18_04939_STRUCT_0020(unittest.TestCase):
    def setUp(self):
        self.app = TSD_Checker.QApplication(TSD_Checker.sys.argv)
        self.TSD_checker = TSD_Checker.Test()

    def test_ok_xls(self):
        self.workbook = xlrd.open_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0020/doc4_ok_xls.xls")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0020_XLS(self.workbook), 1)

    def test_ok_upper_xls(self):
        self.workbook = xlrd.open_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0020/doc4_ok_upper_xls.xls")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0020_XLS(self.workbook), 1)

    def test_notok_xls(self):
        self.workbook = xlrd.open_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0020/doc4_notok_xls.xls")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0020_XLS(self.workbook), 0)

    def test_ok_xlsx(self):
        self.workbook = openpyxl.load_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0020/doc4_ok_xlsx.xlsx")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0020_XLSX_XLSM(self.workbook), 1)

    def test_notok_xlsx(self):
        self.workbook = openpyxl.load_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0020/doc4_notok_xlsx.xlsx")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0020_XLSX_XLSM(self.workbook), 0)


class Test_02043_18_04939_STRUCT_0025_0030_0035_0040(unittest.TestCase):
    def setUp(self):
        self.app = TSD_Checker.QApplication(TSD_Checker.sys.argv)
        self.TSD_checker = TSD_Checker.Test()

    def test_ok_0025_xls(self):
        self.workbook = xlrd.open_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0025_0030_0035_0040/doc4_ok_xls.xls")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0025_XLS(self.workbook), 1)

    def test_ok_french_0025_xls(self):
        self.workbook = xlrd.open_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0025_0030_0035_0040/doc4_ok_french_xls.xls")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0025_XLS(self.workbook), 1)
        
    def test_notok_0025_xls(self):
        self.workbook = xlrd.open_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0025_0030_0035_0040/doc4_notok_xls.xls")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0025_XLS(self.workbook), 0)


    def test_ok_0025_xlsx(self):
        self.workbook =  openpyxl.load_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0025_0030_0035_0040/doc4_ok_xlsx.xlsx")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0025_XLSX_XLSM(self.workbook), 1)


    def test_ok_french_0025_xlsx(self):
        self.workbook =  openpyxl.load_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0025_0030_0035_0040/doc4_ok_french_xlsx.xlsx")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0025_XLSX_XLSM(self.workbook), 1)

    def test_notok_0025_xlsx(self):
        self.workbook = openpyxl.load_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0025_0030_0035_0040/doc4_notok_xlsx.xlsx")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0025_XLSX_XLSM(self.workbook), 0)


    def test_ok_0030_xls(self):
        self.workbook = xlrd.open_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0025_0030_0035_0040/doc4_ok_xls.xls")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0030_XLS(self.workbook), 1)


    def test_ok_french_0030_xls(self):
        self.workbook = xlrd.open_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0025_0030_0035_0040/doc4_ok_french_xls.xls")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0030_XLS(self.workbook), 1)


    def test_notok_0030_xls(self):
        self.workbook = xlrd.open_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0025_0030_0035_0040/doc4_notok_xls.xls")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0030_XLS(self.workbook), 0)

    def test_ok_0030_xlsx(self):
        self.workbook = openpyxl.load_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0025_0030_0035_0040/doc4_ok_xlsx.xlsx")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0030_XLSX_XLSM(self.workbook), 1)


    def test_ok_french_0030_xlsx(self):
        self.workbook = openpyxl.load_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0025_0030_0035_0040/doc4_ok_french_xlsx.xlsx")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0030_XLSX_XLSM(self.workbook), 1)


    def test_notok_0030_xlsx(self):
        self.workbook = openpyxl.load_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0025_0030_0035_0040/doc4_notok_xlsx.xlsx")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0030_XLSX_XLSM(self.workbook), 0)


    def test_ok_0035_xls(self):
        self.workbook = xlrd.open_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0025_0030_0035_0040/doc4_ok_xls.xls")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0035_XLS(self.workbook), 1)

    def test_ok_french_0035_xls(self):
        self.workbook = xlrd.open_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0025_0030_0035_0040/doc4_ok_french_xls.xls")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0035_XLS(self.workbook), 1)

    def test_notok_0035_xls(self):
        self.workbook = xlrd.open_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0025_0030_0035_0040/doc4_notok_xls.xls")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0035_XLS(self.workbook), 0)

    def test_ok_0035_xlsx(self):
        self.workbook = openpyxl.load_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0025_0030_0035_0040/doc4_ok_xlsx.xlsx")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0035_XLSX_XLSM(self.workbook), 1)

    def test_ok_french_0035_xlsx(self):
        self.workbook = openpyxl.load_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0025_0030_0035_0040/doc4_ok_french_xlsx.xlsx")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0035_XLSX_XLSM(self.workbook), 1)

    def test_notok_0035_xlsx(self):
        self.workbook = openpyxl.load_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0025_0030_0035_0040/doc4_notok_xlsx.xlsx")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0035_XLSX_XLSM(self.workbook), 0)



    def test_ok_0040_xls(self):
        self.workbook = xlrd.open_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0025_0030_0035_0040/doc4_ok_xls.xls")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0040_XLS(self.workbook), 1)

    def test_ok_french_0040_xls(self):
        self.workbook = xlrd.open_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0025_0030_0035_0040/doc4_ok_french_xls.xls")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0040_XLS(self.workbook), 1)


    def test_notok_0040_xls(self):
        self.workbook = xlrd.open_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0025_0030_0035_0040/doc4_notok_xls.xls")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0040_XLS(self.workbook), 0)


    def test_ok_0040_xlsx(self):
        self.workbook =  openpyxl.load_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0025_0030_0035_0040/doc4_ok_xlsx.xlsx")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0040_XLSX_XLSM(self.workbook), 1)

    def test_ok_french_0040_xlsx(self):
        self.workbook =  openpyxl.load_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0025_0030_0035_0040/doc4_ok_french_xlsx.xlsx")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0040_XLSX_XLSM(self.workbook), 1)

    def test_notok_0040_xlsx(self):
        self.workbook =  openpyxl.load_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0025_0030_0035_0040/doc4_notok_xlsx.xlsx")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0040_XLSX_XLSM(self.workbook), 0)


class Test_02043_18_04939_STRUCT_0051(unittest.TestCase):
    def setUp(self):
        self.app = TSD_Checker.QApplication(TSD_Checker.sys.argv)
        self.TSD_checker = TSD_Checker.Test()

#0051
    def test_ok_0051_xls(self):
        self.workbook = xlrd.open_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0051_52_53_..._60/doc_ok_xls.xls")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0051_XLS(self.workbook), 1)

    def test_ok_french_0051_xls(self):
        self.workbook = xlrd.open_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0051_52_53_..._60/doc_ok_french_xls.xls")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0051_XLS(self.workbook), 1)

    def test_notok_0051_xls(self):
        self.workbook = xlrd.open_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0051_52_53_..._60/doc_notok_xls.xls")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0051_XLS(self.workbook), 0)


    def test_ok_0051_xlsx(self):
        self.workbook = openpyxl.load_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0051_52_53_..._60/doc_ok_xlsx.xlsx")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0051_XLSX_XLSM(self.workbook), 1)

    def test_ok_french_0051_xlsx(self):
        self.workbook = openpyxl.load_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0051_52_53_..._60/doc_ok_french_xlsx.xlsx")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0051_XLSX_XLSM(self.workbook), 1)

    def test_notok_0051_xlsx(self):
        self.workbook = openpyxl.load_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0051_52_53_..._60/doc_notok_xlsx.xlsx")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0051_XLSX_XLSM(self.workbook), 0)

#0052

    def test_ok_0052_xls(self):
        self.workbook = xlrd.open_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0051_52_53_..._60/doc_ok_xls.xls")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0052_XLS(self.workbook), 1)

    def test_ok_french_0052_xls(self):
        self.workbook = xlrd.open_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0051_52_53_..._60/doc_ok_french_xls.xls")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0052_XLS(self.workbook), 1)

    def test_notok_0052_xls(self):
        self.workbook = xlrd.open_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0051_52_53_..._60/doc_notok_xls.xls")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0052_XLS(self.workbook), 0)


    def test_ok_0052_xlsx(self):
        self.workbook = openpyxl.load_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0051_52_53_..._60/doc_ok_xlsx.xlsx")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0052_XLSX_XLSM(self.workbook), 1)

    def test_ok_french_0052_xlsx(self):
        self.workbook = openpyxl.load_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0051_52_53_..._60/doc_ok_french_xlsx.xlsx")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0052_XLSX_XLSM(self.workbook), 1)

    def test_notok_0052_xlsx(self):
        self.workbook = openpyxl.load_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0051_52_53_..._60/doc_notok_xlsx.xlsx")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0052_XLSX_XLSM(self.workbook), 0)


#0053

    def test_ok_0053_xls(self):
        self.workbook = xlrd.open_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0051_52_53_..._60/doc_ok_xls.xls")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0053_XLS(self.workbook), 1)


    def test_ok_french_0053_xls(self):
        self.workbook = xlrd.open_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0051_52_53_..._60/doc_ok_french_xls.xls")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0053_XLS(self.workbook), 1)

    def test_notok_0053_xls(self):
        self.workbook = xlrd.open_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0051_52_53_..._60/doc_notok_xls.xls")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0053_XLS(self.workbook), 0)


    def test_ok_0053_xlsx(self):
        self.workbook = openpyxl.load_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0051_52_53_..._60/doc_ok_xlsx.xlsx")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0053_XLSX_XLSM(self.workbook), 1)


    def test_ok_french_0053_xlsx(self):
        self.workbook = openpyxl.load_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0051_52_53_..._60/doc_ok_french_xlsx.xlsx")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0053_XLSX_XLSM(self.workbook), 1)

    def test_notok_0053_xlsx(self):
        self.workbook = openpyxl.load_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0051_52_53_..._60/doc_notok_xlsx.xlsx")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0053_XLSX_XLSM(self.workbook), 0)

#0054
    def test_ok_0054_xls(self):
        self.workbook = xlrd.open_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0051_52_53_..._60/doc_ok_xls.xls")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0054_XLS(self.workbook), 1)


    def test_ok_french_0054_xls(self):
        self.workbook = xlrd.open_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0051_52_53_..._60/doc_ok_french_xls.xls")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0054_XLS(self.workbook), 1)

    def test_notok_0054_xls(self):
        self.workbook = xlrd.open_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0051_52_53_..._60/doc_notok_xls.xls")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0054_XLS(self.workbook), 0)

    def test_ok_0054_xlsx(self):
        self.workbook = openpyxl.load_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0051_52_53_..._60/doc_ok_xlsx.xlsx")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0054_XLSX_XLSM(self.workbook), 1)

    def test_ok_french_0054_xlsx(self):
        self.workbook = openpyxl.load_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0051_52_53_..._60/doc_ok_french_xlsx.xlsx")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0054_XLSX_XLSM(self.workbook), 1)

    def test_notok_0054_xlsx(self):
        self.workbook = openpyxl.load_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0051_52_53_..._60/doc_notok_xlsx.xlsx")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0054_XLSX_XLSM(self.workbook), 0)



#0055

    def test_ok_0055_xls(self):
        self.workbook = xlrd.open_workbook(
            "C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0051_52_53_..._60/doc_ok_xls.xls")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0055_XLS(self.workbook), 1)

    def test_notok_0055_xls(self):
        self.workbook = xlrd.open_workbook(
            "C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0051_52_53_..._60/doc_notok_xls.xls")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0055_XLS(self.workbook), 0)

    def test_ok_0055_xlsx(self):
        self.workbook = openpyxl.load_workbook(
            "C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0051_52_53_..._60/doc_ok_xlsx.xlsx")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0055_XLSX_XLSM(self.workbook), 1)

    def test_notok_0055_xlsx(self):
        self.workbook = openpyxl.load_workbook(
            "C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0051_52_53_..._60/doc_notok_xlsx.xlsx")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0055_XLSX_XLSM(self.workbook), 0)

#0056

    def test_ok_0056_xls(self):
        self.workbook = xlrd.open_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0051_52_53_..._60/doc_ok_xls.xls")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0056_XLS(self.workbook), 1)


    def test_notok_0056_xls(self):
        self.workbook = xlrd.open_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0051_52_53_..._60/doc_notok_xls.xls")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0056_XLS(self.workbook), 0)


    def test_ok_0056_xlsx(self):
         self.workbook = openpyxl.load_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0051_52_53_..._60/doc_ok_xlsx.xlsx")
         self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0056_XLSX_XLSM(self.workbook), 1)


    def test_notok_0056_xlsx(self):
         self.workbook = openpyxl.load_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0051_52_53_..._60/doc_notok_xlsx.xlsx")
         self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0056_XLSX_XLSM(self.workbook), 0)


#0057

    def test_ok_0057_xls(self):
        self.workbook = xlrd.open_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0051_52_53_..._60/doc_ok_xls.xls")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0057_XLS(self.workbook), 1)

    def test_notok_0057_xls(self):
        self.workbook = xlrd.open_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0051_52_53_..._60/doc_notok_xls.xls")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0057_XLS(self.workbook), 0)


    def test_ok_0057_xlsx(self):
        self.workbook = openpyxl.load_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0051_52_53_..._60/doc_ok_xlsx.xlsx")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0057_XLSX_XLSM(self.workbook), 1)

    def test_notok_0057_xlsx(self):
        self.workbook = openpyxl.load_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0051_52_53_..._60/doc_notok_xlsx.xlsx")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0057_XLSX_XLSM(self.workbook), 0)


#0058
    def test_ok_0058_xls(self):
        self.workbook = xlrd.open_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0051_52_53_..._60/doc_ok_xls.xls")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0058_XLS(self.workbook), 1)

    def test_notok_0058_xls(self):
        self.workbook = xlrd.open_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0051_52_53_..._60/doc_notok_xls.xls")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0058_XLS(self.workbook), 0)


    def test_ok_0058_xlsx(self):
        self.workbook = openpyxl.load_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0051_52_53_..._60/doc_ok_xlsx.xlsx")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0058_XLSX_XLSM(self.workbook), 1)

    def test_notok_0058_xlsx(self):
        self.workbook = openpyxl.load_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0051_52_53_..._60/doc_notok_xlsx.xlsx")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0058_XLSX_XLSM(self.workbook), 0)

#0059

    def test_ok_0059_xls(self):
        self.workbook = xlrd.open_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0051_52_53_..._60/doc_ok_xls.xls")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0059_XLS(self.workbook), 1)

    def test_notok_0059_xls(self):
        self.workbook = xlrd.open_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0051_52_53_..._60/doc_notok_xls.xls")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0059_XLS(self.workbook), 0)


    def test_ok_0059_xlsx(self):
        self.workbook = openpyxl.load_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0051_52_53_..._60/doc_ok_xlsx.xlsx")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0059_XLSX_XLSM(self.workbook), 1)

    def test_notok_0059_xlsx(self):
        self.workbook = openpyxl.load_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0051_52_53_..._60/doc_notok_xlsx.xlsx")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0059_XLSX_XLSM(self.workbook), 0)

#0060

    def test_ok_0060_xls(self):
        self.workbook = xlrd.open_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0051_52_53_..._60/doc_ok_xls.xls")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0060_XLS(self.workbook), 1)

    def test_notok_0060_xls(self):
        self.workbook = xlrd.open_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0051_52_53_..._60/doc_notok_xls.xls")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0060_XLS(self.workbook), 0)


    def test_ok_0060_xlsx(self):
        self.workbook = openpyxl.load_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0051_52_53_..._60/doc_ok_xlsx.xlsx")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0060_XLSX_XLSM(self.workbook), 1)

    def test_notok_0060_xlsx(self):
        self.workbook = openpyxl.load_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0051_52_53_..._60/doc_notok_xlsx.xlsx")
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0060_XLSX_XLSM(self.workbook), 0)

#Requirements for [DOC4]

class Test_02043_18_04939_STRUCT_0400(unittest.TestCase):
    def setUp(self):
        self.app = TSD_Checker.QApplication(TSD_Checker.sys.argv)
        self.TSD_checker = TSD_Checker.Test()

    def test_ok_xls(self):
        self.workbook = xlrd.open_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0400/doc4_ok_xls.xls", formatting_info=True)
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0400_XLS(self.workbook), 1)

    def test_notok_xls(self):
        self.workbook = xlrd.open_workbook("C:/Users/admacesanu/Desktop/EXCEL_TEST/02043_18_04939_STRUCT_0400/doc4_notok_xls.xls", formatting_info=True)
        self.assertEqual(self.TSD_checker.Test_02043_18_04939_STRUCT_0400_XLS(self.workbook), 0)



if __name__ == '__main__':
    unittest.main()