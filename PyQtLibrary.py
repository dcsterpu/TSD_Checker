import sys
from PyQt5.QtWidgets import QWidget, QPushButton, QApplication, QComboBox, QLabel, QLineEdit,  QTabWidget, QVBoxLayout, QProgressBar
from PyQt5 import QtCore, QtWidgets
import openpyxl
import xlrd
import win32com.client as win32
import requests
import os


class Application(QWidget):

    def __init__(self):
        super().__init__()
        self.left = 200
        self.top = 200
        self.width = 1000
        self.height = 450
        self.tabs = QTabWidget()
        self.tab1 = QWidget()
        self.tab2 = QWidget()
        self.tabs.addTab(self.tab1, "TSD Checker")
        self.tabs.addTab(self.tab2, "Options")
        self.initUI(self.tab1)
        self.initUIOptions(self.tab2)
        self.layout = QVBoxLayout()
        self.layout.addWidget(self.tabs)
        self.setLayout(self.layout)

    def openFileNameDialog1(self):
        fileName1, _filter = QtWidgets.QFileDialog.getOpenFileName(self.tab1, 'Open File', QtCore.QDir.rootPath(), '*.*')
        self.tab1.myTextBox1.setText(fileName1)
        self.tab1.textbox.setText("next file")

    def openFileNameDialog2(self):
        fileName2, _filter = QtWidgets.QFileDialog.getOpenFileName(self.tab, 'Open File', QtCore.QDir.rootPath(), '*.*')
        self.tab.myTextBox2.setText(fileName2)



    def openFileNameDialog3(self):
        fileName3, _filter = QtWidgets.QFileDialog.getOpenFileName(self.tab, 'Open File', QtCore.QDir.rootPath(), '*.*')
        self.tab.myTextBox3.setText(fileName3)


    def openFileNameDialog4(self):
        fileName4, _filter = QtWidgets.QFileDialog.getOpenFileName(self.tab, 'Open File', QtCore.QDir.rootPath(), '*.*')
        self.tab.myTextBox4.setText(fileName4)


    def openFileNameDialog5(self):
        fileName5, _filter = QtWidgets.QFileDialog.getOpenFileName(self.tab, 'Open File', QtCore.QDir.rootPath(), '*.*')
        self.tab.myTextBox5.setText(fileName5)


    def openFileNameDialog6(self):
        fileName6, _filter = QtWidgets.QFileDialog.getOpenFileName(self.tab, 'Open File', QtCore.QDir.rootPath(), '*.*')
        self.tab.myTextBox6.setText(fileName6)


    def openFileNameDialog7(self):
        fileName7, _filter = QtWidgets.QFileDialog.getOpenFileName(self.tab, 'Open File', QtCore.QDir.rootPath(), '*.*')
        self.tab.myTextBox7.setText(fileName7)


    def openFileNameDialog8(self):
        fileName8, _filter = QtWidgets.QFileDialog.getOpenFileName(self.tab, 'Open File', QtCore.QDir.rootPath(), '*.*')
        self.tab.myTextBox8.setText(fileName8)

    def openFileNameDialog9(self):
        fileName9, _filter = QtWidgets.QFileDialog.getOpenFileName(self.tab, 'Open File', QtCore.QDir.rootPath(), '*.*')
        self.tab.myTextBox9.setText(fileName9)

    def openFileNameDialog10(self):
        fileName10, _filter = QtWidgets.QFileDialog.getOpenFileName(self.tab, 'Open File', QtCore.QDir.rootPath(), '*.*')
        self.tab.myTextBox10.setText(fileName10)


    def ButtonReportClick(self):
        self.popUp = QWidget()
        self.popUp.setWindowTitle("ERRROR")
        self.popUp.Label = QLabel(self.popUp)
        self.popUp.Label.setText("No Report")
        self.popUp.Label.setAlignment(QtCore.Qt.AlignCenter)
        self.popUp.setGeometry(550, 550, 200, 50)
        self.popUp.show()

    def buttonClicked(self):
        return

    def colorButton(self):
        return

    def download_file(self, url):
        user = self.tab2.TextBoxUser.text()
        user = str(user)
        password = self.tab2.TextBoxPass.text()
        password = str(password)
        out_path = "C:/Users/" + "admacesanu" + "/AppData/Local/Temp/TSD_Checker/"
        user_path = "C:/Users/" + "admacesanu"
        if not user or not password:
            self.errorPopUp = QWidget()
            self.errorPopUp.setWindowTitle("ERRROR")
            self.errorPopUp.Label = QLabel(self.errorPopUp)
            self.errorPopUp.Label.setText("Missing Username or Password")
            self.errorPopUp.Label.setAlignment(QtCore.Qt.AlignCenter)
            self.errorPopUp.setGeometry(550, 550, 200, 50)
            self.errorPopUp.show()
            return
        ''' try:
            os.stat(user_path)
          except:
            self.errorPopUp = QWidget()
            self.errorPopUp.setWindowTitle("ERRROR")
            self.errorPopUp.Label = QLabel(self.errorPopUp)
            self.errorPopUp.Label.setText("Username or Password Incorrect")
            self.errorPopUp.Label.setAlignment(QtCore.Qt.AlignCenter)
            self.errorPopUp.setGeometry(550, 550, 200, 50)
            self.errorPopUp.show()
            return '''
        try:
            os.stat(out_path)
        except:
            os.mkdir(out_path)

        response = requests.get(url, stream=True, auth=(user, password))
        status = response.status_code
        if status == 401:
            self.errorPopUp = QWidget()
            self.errorPopUp.setWindowTitle("ERRROR")
            self.errorPopUp.Label = QLabel(self.errorPopUp)
            self.errorPopUp.Label.setText("Username or Password Incorrect")
            self.errorPopUp.Label.setAlignment(QtCore.Qt.AlignCenter)
            self.errorPopUp.setGeometry(550, 550, 200, 50)
            self.errorPopUp.show()
            return

        FileName = response.headers['Content-Disposition'].split('"')[1]
        FilePath = out_path + "/" + FileName
        print("Saving file to location:" + FilePath)
        with open(FilePath, 'wb') as f:
            for chuck in response.iter_content(chunk_size=128):
                f.write(chuck)
                return FilePath



    def initUI(self, tab):

    # Create a textbox
        tab.message = "message"
        tab.textbox = QtWidgets.QTextEdit(self.tab1)
        tab.textbox.setText(tab.message)
        tab.textbox.move(10, 260)
        tab.textbox.resize(700, 40)
        tab.textbox.setReadOnly(True)

        tab.pbar = QProgressBar(self.tab1)
        tab.pbar.setGeometry(10, 310, 700, 20)
        tab.pbar.setAlignment(QtCore.Qt.AlignCenter)
        tab.pbar.setValue(0)

    #Create a color textbox1
        tab.colorTextBox1 = QtWidgets.QTextEdit(self.tab1)
        tab.colorTextBox1.setStyleSheet(  " background-color: grey ")
        tab.colorTextBox1.resize(20, 20)
        tab.colorTextBox1.move(710, 10)


    # Create a color textbox2
        tab.colorTextBox2 = QtWidgets.QTextEdit(self.tab1)
        tab.colorTextBox2.setStyleSheet(" background-color: grey ")
        tab.colorTextBox2.resize(20, 20)
        tab.colorTextBox2.move(710, 40)

    # Create a color textbox3
        tab.colorTextBox3 = QtWidgets.QTextEdit(self.tab1)
        tab.colorTextBox3.setStyleSheet(" background-color: grey ")
        tab.colorTextBox3.resize(20, 20)
        tab.colorTextBox3.move(710, 70)

    # Create a color textbox4
        tab.colorTextBox4 = QtWidgets.QTextEdit(self.tab1)
        tab.colorTextBox4.setStyleSheet(" background-color: grey ")
        tab.colorTextBox4.resize(20, 20)
        tab.colorTextBox4.move(710, 100)

    # Create a color textbox5
        tab.colorTextBox5 = QtWidgets.QTextEdit(self.tab1)
        tab.colorTextBox5.setStyleSheet(" background-color: grey ")
        tab.colorTextBox5.resize(20, 20)
        tab.colorTextBox5.move(710, 130)

    # Create a color textbox6
        tab.colorTextBox6 = QtWidgets.QTextEdit(self.tab1)
        tab.colorTextBox6.setStyleSheet(" background-color: grey ")
        tab.colorTextBox6.resize(20, 20)
        tab.colorTextBox6.move(710, 160)

    #Create a drop down list
        tab.lbl = QLabel("Check level", tab)

        tab.combo = QComboBox(tab)
        tab.combo.addItem("   Option1   ")
        tab.combo.addItem("   Option2   ")
        tab.combo.addItem("   Option3   ")
        tab.combo.addItem("   Option4   ")
        tab.combo.addItem("   Option5   ")
        tab.combo.resize(508, 20.4)  #rezise the drop down list
        tab.combo.move(200, 190)
        tab.lbl.move(5, 195)
        tab.combo.activated[str].connect(self.onActivated)


    # Create a drop down list
        tab.lbl1 = QLabel("Project name", tab)

        tab.combo1 = QComboBox(tab)
        tab.combo1.addItem("   Generic   ")
        tab.combo1.addItem("   All   ")
        tab.combo1.resize(378, 20.4)  # rezise the drop down list
        tab.combo1.move(200, 220)
        tab.lbl1.move(5, 225)
        tab.combo1.activated[str].connect(self.onActivated)

        self.setGeometry(self.left, self.top, self.width, self.height)
        self.setWindowTitle('TSD Checker')

        tab.importNames = QPushButton(tab)
        tab.importNames.setText("Import Project Names")
        tab.importNames.resize(120, 20.4)
        tab.importNames.move(585, 220)


        #File Selectiom Dialog1
        tab.lbl2 = QLabel("TSD File:", tab)
        tab.lbl2.move(5,15)
        tab.myTextBox1 = QtWidgets.QTextEdit(tab)
        tab.myTextBox1.resize(460, 25)
        tab.myTextBox1.move(200, 10)
        tab.myTextBox1.setReadOnly(True)
        tab.button1 = QPushButton('...',tab)
        tab.button1.clicked.connect(self.openFileNameDialog1)
        tab.button1.move(660, 10)
        tab.button1.resize(45,22)

    # File Selectiom Dialog2
        tab.lbl3 = QLabel("TSD vehicle Function file:", tab)
        tab.lbl3.move(5, 45)
        tab.myTextBox2 = QtWidgets.QTextEdit(tab)
        tab.myTextBox2.resize(460, 25)
        tab.myTextBox2.move(200, 40)
        tab.myTextBox2.setReadOnly(True)
        tab.button2 = QPushButton('...', tab)
        tab.button2.clicked.connect(self.openFileNameDialog2)
        tab.button2.move(660, 40)
        tab.button2.resize(45, 22)

    # File Selectiom Dialog3
        tab.lbl4 = QLabel("TSD system file:", tab)
        tab.lbl4.move(5, 75)
        tab.myTextBox3 = QtWidgets.QTextEdit(tab)
        tab.myTextBox3.resize(460, 25)
        tab.myTextBox3.move(200, 70)
        tab.myTextBox3.setReadOnly(True)
        tab.button3 = QPushButton('...', tab)
        tab.button3.clicked.connect(self.openFileNameDialog3)
        tab.button3.move(660, 70)
        tab.button3.resize(45, 22)



    # File Selectiom Dialog7
        tab.lbl8 = QLabel("AMDEC:", tab)
        tab.lbl8.move(5, 105)
        tab.myTextBox7 = QtWidgets.QTextEdit(tab)
        tab.myTextBox7.resize(460, 25)
        tab.myTextBox7.move(200, 100)

        tab.myTextBox7.setReadOnly(True)
        tab.button7 = QPushButton('...', tab)
        tab.button7.clicked.connect(self.openFileNameDialog7)
        tab.button7.move(660, 100)
        tab.button7.resize(45, 22)

    # File Selectiom Dialog8
        tab.lbl9 = QLabel("export MedialecMatrice:", tab)
        tab.lbl9.move(5, 135)
        tab.myTextBox8 = QtWidgets.QTextEdit(tab)
        tab.myTextBox8.resize(460, 25)
        tab.myTextBox8.move(200, 130)
        tab.myTextBox8.setReadOnly(True)
        tab.button8 = QPushButton('...', tab)
        tab.button8.clicked.connect(self.openFileNameDialog8)
        tab.button8.move(660, 130)
        tab.button8.resize(45, 22)



    # File Selectiom Dialog10
        tab.lbl11 = QLabel("Diagnostic matrix file:", tab)
        tab.lbl11.move(5, 165)
        tab.myTextBox10 = QtWidgets.QTextEdit(tab)
        tab.myTextBox10.resize(460, 25)
        tab.myTextBox10.move(200, 160)
        tab.myTextBox10.setReadOnly(True)
        tab.button10 = QPushButton('...', tab)
        tab.button10.clicked.connect(self.openFileNameDialog10)
        tab.button10.move(660, 160)
        tab.button10.resize(45, 22)



    # Check button
        button = QPushButton('Check', tab)
        button.move(390, 360)
        button.resize(90,25)
        button.clicked.connect(self.buttonClicked)
        button.setStyleSheet('QPushButton {background-color: white; color: black;}')
        buttonStruct = QPushButton("Check Structure", tab)
        buttonStruct.move(270, 360)
        buttonStruct.resize(90, 25)
        buttonStruct.clicked.connect(self.colorButton)

        buttonNew = QPushButton("Open \nReport", tab)
        buttonNew.resize(80, 40)
        buttonNew.move(710, 260)
        buttonNew.clicked.connect(self.ButtonReportClick)



        self.show()

    def initUIOptions(self, tab):

        hdd_path = " C:/Users/u409465/AppData/Local/Temp/TSD_Checker/ "
        file_url = "https://docinfogroupe.psa-peugeot-citroen.com/ead/doc/ref.01272_18_00096/v.vc/pj"

        tab.lblUser = QLabel("USER:", tab)
        tab.lblUser.move(155,25)
        tab.TextBoxUser = QtWidgets.QLineEdit(tab)
        tab.TextBoxUser.resize(200,25)
        tab.TextBoxUser.move(200, 20)


        tab.lblPass = QLabel("PASSWORD:", tab)
        tab.lblPass.move(420,25)
        tab.TextBoxPass = QtWidgets.QLineEdit(tab)
        tab.TextBoxPass.resize(180,25)
        tab.TextBoxPass.move(520, 20)
        tab.TextBoxPass.setEchoMode((QLineEdit.Password))



        # File Selectiom Dialog5
        tab.lbl6 = QLabel("Famille/Sous-Famille list export(CESARE):", tab)
        tab.lbl6.move(5, 85)
        tab.myTextBox5 = QtWidgets.QTextEdit(tab)
        tab.myTextBox5.resize(460, 25)
        tab.myTextBox5.move(200, 80)
        tab.myTextBox5.setReadOnly(True)

        tab.link2 = QLabel('''<a href='https://docinfogroupe.psa-peugeot-citroen.com/ead/doc/ref.02043_18_05471/v.vc/pj'>DocInfo Reference: 02043_18_05471</a>''',tab)
        tab.link2.setOpenExternalLinks(True)
        tab.link2.move(720, 85)


        tab.button5 = QPushButton('...', tab)
        tab.button5.move(660, 80)
        tab.button5.resize(45, 22)



        # File Selectiom Dialog4
        tab.lbl5 = QLabel("TSD configuration file:", tab)
        tab.lbl5.move(5,145)
        tab.myTextBox4 = QtWidgets.QTextEdit(tab)
        tab.myTextBox4.resize(460, 25)
        tab.myTextBox4.move(200, 140)
        tab.myTextBox4.setReadOnly(True)

        tab.link1 = QLabel('''<a href='https://docinfogroupe.psa-peugeot-citroen.com/ead/doc/ref.02043_18_05472/v.vc/pj'>DocInfo Reference: 02043_18_05472</a>''', tab)
        tab.link1.setOpenExternalLinks(True)
        tab.link1.move(720, 145)

        tab.button4 = QPushButton('...', tab)
        tab.button4.clicked.connect(self.openFileNameDialog4)
        tab.button4.move(660, 140)
        tab.button4.resize(45, 22)



        # File Selectiom Dialog6
        tab.lbl7 = QLabel("Customer effect file:", tab)
        tab.lbl7.move(5, 205)
        tab.myTextBox6 = QtWidgets.QTextEdit(tab)
        tab.myTextBox6.resize(460, 25)
        tab.myTextBox6.move(200, 200)
        tab.myTextBox6.setReadOnly(True)

        tab.link3 = QLabel('''<a href='https://docinfogroupe.psa-peugeot-citroen.com/ead/doc/ref.02043_18_05499/v.vc/pj'>DocInfo Reference: 02043_18_05499</a>''', tab)
        tab.link3.setOpenExternalLinks(True)
        tab.link3.move(720, 205)



        tab.button6 = QPushButton('...', tab)
        tab.button6.clicked.connect(self.openFileNameDialog6)
        tab.button6.move(660, 200)
        tab.button6.resize(45, 22)

        # File Selectiom Dialog9
        tab.lbl10 = QLabel("Diversity management file:", tab)
        tab.lbl10.move(5, 265)
        tab.myTextBox9 = QtWidgets.QTextEdit(tab)
        tab.myTextBox9.resize(460, 25)
        tab.myTextBox9.move(200,260)
        tab.myTextBox9.setReadOnly(True)

        tab.link4 = QLabel('''<a href='https://docinfogroupe.psa-peugeot-citroen.com/ead/doc/ref.02016_11_04964/v.vc/pj'>DocInfo Reference: 02016_11_04964</a>''',tab)
        tab.link4.setOpenExternalLinks(True)
        tab.link4.move(720, 265)


        tab.button9 = QPushButton('...', tab)
        tab.button9.clicked.connect(self.openFileNameDialog9)
        tab.button9.move(660, 260)
        tab.button9.resize(45, 22)


    def onActivated(self):
        return


class Test(Application):
    Application.tsdFileExtension = str()
    Application.tsdVehicleFunctionFileExtension = str()
    Application.tsdSystemFileExtension = str()
    Application.amdecFileExtension = str()
    Application.exportMedialecMatriceFileExtension = str()
    Application.diagnosticMatrixFileExtension = str()

    def GetTsdFileExtension(self):
        fileName = self.tab1.myTextBox1.toPlainText()
        tokens = fileName.split(".")
        self.tsdFileExtension = tokens[-1]

    def GetTsdFileWorkbook(self):

        fileName = self.tab1.myTextBox1.toPlainText()
        if self.tsdFileExtension is "xls":
            return xlrd.open_workbook(fileName,  formatting_info=True)
        elif self.tsdFileExtension is "xlsx":
            return openpyxl.load_workbook(fileName)
        elif self.tsdFileExtension is "xlsm":
            return openpyxl.load_workbook(fileName, keep_vba=True)

    def GetTsdVehicleFunctionFileExtension(self):
        fileName = self.tab1.myTextBox2.toPlainText()
        tokens = fileName.split(".")
        self.tsdVehicleFucntionFileExtension = tokens[-1]

    def GetTsdVehicleFunctionFileWorkbook(self):

        fileName = self.tab1.myTextBox2.toPlainText()
        if self.tsdVehicleFunctionFileExtension is "xls":
            return xlrd.open_workbook(fileName,  formatting_info=True)
        elif self.tsdVehicleFunctionFileExtension is "xlsx":
            return openpyxl.load_workbook(fileName)
        elif self.tsdVehicleFunctionFileExtension is "xlsm":
            return openpyxl.load_workbook(fileName, keep_vba=True)

    def GetTsdSystemFileExtension(self):
        fileName = self.tab1.myTextBox3.toPlainText()
        tokens = fileName.split(".")
        self.tsdSystemFileExtension = tokens[-1]

    def GetTsdSystemFileWorkbook(self):

        fileName = self.tab1.myTextBox3.toPlainText()
        if self.tsdSystemFileExtension is "xls":
            return xlrd.open_workbook(fileName,  formatting_info=True)
        elif self.tsdSystemFileExtension is "xlsx":
            return openpyxl.load_workbook(fileName)
        elif self.tsdSystemFileExtension is "xlsm":
            return openpyxl.load_workbook(fileName, keep_vba=True)

    def GetAmdecFileExtension(self):
        fileName = self.tab1.myTextBox7.toPlainText()
        tokens = fileName.split(".")
        self.amdecFileExtension = tokens[-1]

    def GetAmdecFileWorkbook(self):

        fileName = self.tab1.myTextBox7.toPlainText()
        if self.amdecFileExtension is "xls":
            return xlrd.open_workbook(fileName,  formatting_info=True)
        elif self.amdecFileExtension is "xlsx":
            return openpyxl.load_workbook(fileName)
        elif self.amdecFileExtension is "xlsm":
            return openpyxl.load_workbook(fileName, keep_vba=True)

    def GetExportMedialecMatriceFileExtension(self):
        fileName = self.tab1.myTextBox8.toPlainText()
        tokens = fileName.split(".")
        self.exportMedialecMatriceFileExtension = tokens[-1]

    def GetExportMedialecMatriceFileWorkbook(self):

        fileName = self.tab1.myTextBox8.toPlainText()
        if self.exportMedialecMatriceFileExtension is "xls":
            return xlrd.open_workbook(fileName,  formatting_info=True)
        elif self.exportMedialecMatriceFileExtension is "xlsx":
            return openpyxl.load_workbook(fileName)
        elif self.exportMedialecMatriceFileExtension is "xlsm":
            return openpyxl.load_workbook(fileName, keep_vba=True)

    def GetDiagnosticMatrixFileExtension(self):
        fileName = self.tab1.myTextBox9.toPlainText()
        tokens = fileName.split(".")
        self.diagnosticMatrixFileExtension = tokens[-1]

    def GetDiagnosticMatrixFileWorkbook(self):

        fileName = self.tab1.myTextBox9.toPlainText()
        if self.diagnosticMatrixFileExtension is "xls":
            return xlrd.open_workbook(fileName,  formatting_info=True)
        elif self.diagnosticMatrixFileExtension is "xlsx":
            return openpyxl.load_workbook(fileName)
        elif self.diagnosticMatrixFileExtension is "xlsm":
            return openpyxl.load_workbook(fileName, keep_vba=True)

#Requirements for General structure
    def Test_02043_18_04939_STRUCT_0000_XLS(self, workBook):

        sheetNames = workBook.sheet_names()
        if sheetNames[0].casefold() in ["informations générales", "general information"]:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0000_XLSX_XLSM(self, workBook):

        sheetNames = workBook.sheetnames
        if sheetNames[0].casefold() in ["informations générales", "general information"]:
            return 1
        else:
            return 0


    def Test_02043_18_04939_STRUCT_0005_XLS(self, workBook):

        excel = win32.gencache.EnsureDispatch('Excel.Application')
        wb = excel.Workbooks.Open(workBook)
        ws = wb.Worksheets(1)
        if ws.Cells(52,2).HasFormula is False:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0005_XLSX_XLSM(self, workBook):

        workSheet = workBook.worksheets[0]
        if workSheet.cell(52, 2).value[0] is "=":
            return 0
        else:
            return 1

    def Test_02043_18_04939_STRUCT_0010_XLS(self, workBook):
        workSheet = workBook.worksheets[0]
        if isinstance(workSheet.cell_value(52,2), str) and workSheet.cell_value(52,2):
            return 1
        else:
            return 0


    def Test_02043_18_04939_STRUCT_0010_XLSX_XLSM(self, workBook):

        workSheet = workBook.worksheets[0]
        if isinstance(workSheet.cell(52, 2).value , str) and workSheet.cell(52, 2).value:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0011_XLS(self, workBook):
        workSheet = workBook.worksheets[0]
        if workSheet.cell_value(52,2) in {"AEEV_IAEE07_0033", "02043_12_01665", "02043_12_01666"}:
            return 1
        else:
            return 0



    def Test_02043_18_04939_STRUCT_0011_XLSX_XLSM(self, workBook):

        workSheet = workBook.worksheets[0]
        if workSheet.cell(52, 2).value in {"AEEV_IAEE07_0033", "02043_12_01665", "02043_12_01666"}:
            return 1
        else:
            return 0


    def Test_02043_18_04939_STRUCT_0020_XLS(self, workBook):

        if "suppression" in workBook.sheet_names().casefold():
            return 1
        else:
            return 0


    def  Test_02043_18_04939_STRUCT_0020_XLSX_XLSM(self, workBook):

         if "suppression" in workBook.sheetnames.casefold():
             return 1
         else:
             return 0


    def Test_02043_18_04939_STRUCT_0025_XLS(self, workBook):
        try:
            workSheet = workBook.sheet_by_name("Suppression")
        except:
            workSheet = workBook.sheet_by_name("suppression")

        row = workSheet.row(1)
        for cell in row:
            if cell.value.casefold() in {"sheet", "ongle"}:
                return 1
        return 0



    def Test_02043_18_04939_STRUCT_0025_XLSX_XLSM(self, workBook):
        try:
            workSheet = workBook.get_sheet_by_name("Suppression")
        except:
            workSheet = workBook.get_sheet_by_name("suppression")

        row = workSheet.iter_rows(min_col=1, min_row=1, max_row=1)
        for cell in row:
            if cell.value.casefold() in {"sheet", "onglet"}:
                return 1
        return 0


    def Test_02043_18_04939_STRUCT_0030_XLS(self, workBook):
        try:
            workSheet = workBook.sheet_by_name("Suppression")
        except:
            workSheet = workBook.sheet_by_name("suppression")

        row = workSheet.row(1)
        for cell in row:
            if cell.value.casefold() in {"référence de la ligne", "line number"}:
                return 1
        return 0


    def Test_02043_18_04939_STRUCT_0030_XLSX_XLSM(self, workBook):
        try:
            workSheet = workBook.get_sheet_by_name("Suppression")
        except:
            workSheet = workBook.get_sheet_by_name("suppression")

        row = workSheet.iter_rows(min_col=1, min_row=1, max_row=1)
        for cell in row:
            if cell.value.casefold() in {"référence de la ligne", "line number"}:
                return 1
        return 0


    def Test_02043_18_04939_STRUCT_0035_XLS(self, workBook):
        try:
            workSheet = workBook.sheet_by_name("Suppression")
        except:
            workSheet = workBook.sheet_by_name("suppression")

        row = workSheet.row(1)
        for cell in row:
            if cell.value.casefold() in {"version du tsd", "version of the document"}:
                return 1
        return 0

    def Test_02043_18_04939_STRUCT_0035_XLSX_XLSM(self, workBook):
        try:
            workSheet = workBook.get_sheet_by_name("Suppression")
        except:
            workSheet = workBook.get_sheet_by_name("suppression")

        row = workSheet.iter_rows(min_col=1, min_row=1, max_row=1)
        for cell in row:
            if cell.value.casefold() in {"version du tsd", "version of the document"}:
                return 1
        return 0


    def Test_02043_18_04939_STRUCT_0040_XLS(self, workBook):
        try:
            workSheet = workBook.sheet_by_name("Suppression")
        except:
            workSheet = workBook.sheet_by_name("suppression")

        row = workSheet.row(1)
        for cell in row:
            if cell.value.casefold() in {"justification de la modification", "change reason"}:
                return 1
        return 0

    def Test_02043_18_04939_STRUCT_0040_XLSX_XLSM(self, workBook):
        try:
            workSheet = workBook.get_sheet_by_name("Suppression")
        except:
            workSheet = workBook.get_sheet_by_name("suppression")

        row = workSheet.iter_rows(min_col=1, min_row=1, max_row=1)
        for cell in row:
            if cell.value.casefold() in {"justification de la modification", "change reason"}:
                return 1
        return 0

    def Test_02043_18_04939_STRUCT_0051_XLS(self, workBook):
        workSheet = workBook.worksheets[0]
        if isinstance(workSheet.cell_value(62, 5), str) and workSheet.cell_value(62, 5):
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0051_XLSX_XLSM(self, workBook):
        workSheet = workBook.worksheets[0]
        if isinstance(workSheet.cell(62, 5).value, str) and workSheet.cell(62, 5).value:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0052_XLS(self, workBook):
        workSheet = workBook.worksheets[0]
        if isinstance(workSheet.cell_value(63, 5), str) and workSheet.cell_value(63, 5):
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0052_XLSX_XLSM(self, workBook):
        workSheet = workBook.worksheets[0]
        if isinstance(workSheet.cell(63, 5).value, str) and workSheet.cell(63, 5).value:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0053_XLS(self, workBook):
        workSheet = workBook.worksheets[0]
        if isinstance(workSheet.cell_value(64, 5), str) and workSheet.cell_value(64, 5):
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0053_XLSX_XLSM(self, workBook):
        workSheet = workBook.worksheets[0]
        if isinstance(workSheet.cell(64, 5).value, str) and workSheet.cell(64, 5).value:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0054_XLS(self, workBook):
        workSheet = workBook.worksheets[0]
        if isinstance(workSheet.cell_value(65, 5), str) and workSheet.cell_value(65, 5):
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0054_XLSX_XLSM(self, workBook):
        workSheet = workBook.worksheets[0]
        if isinstance(workSheet.cell(65, 5).value, str) and workSheet.cell(65, 5).value:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0055_XLS(self, workBook):
        workSheet = workBook.worksheets[0]
        if isinstance(workSheet.cell_value(66, 5), str) and workSheet.cell_value(66, 5):
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0055_XLSX_XLSM(self, workBook):
        workSheet = workBook.worksheets[0]
        if isinstance(workSheet.cell(66, 5).value, str) and workSheet.cell(66, 5).value:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0056_XLS(self, workBook):
        workSheet = workBook.worksheets[0]
        if isinstance(workSheet.cell_value(67, 5), str) and workSheet.cell_value(67, 5):
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0056_XLSX_XLSM(self, workBook):
        workSheet = workBook.worksheets[0]
        if isinstance(workSheet.cell(67, 5).value, str) and workSheet.cell(67, 5).value:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0057_XLS(self, workBook):
        workSheet = workBook.worksheets[0]
        if isinstance(workSheet.cell_value(68, 5), str) and workSheet.cell_value(68, 5):
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0057_XLSX_XLSM(self, workBook):
        workSheet = workBook.worksheets[0]
        if isinstance(workSheet.cell(68, 5).value, str) and workSheet.cell(68, 5).value:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0058_XLS(self, workBook):
        workSheet = workBook.worksheets[0]
        if isinstance(workSheet.cell_value(69, 5), str) and workSheet.cell_value(69, 5):
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0058_XLSX_XLSM(self, workBook):
        workSheet = workBook.worksheets[0]
        if isinstance(workSheet.cell(69, 5).value, str) and workSheet.cell(69, 5).value:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0059_XLS(self, workBook):
        workSheet = workBook.worksheets[0]
        if isinstance(workSheet.cell_value(70, 5), str) and workSheet.cell_value(70, 5):
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0059_XLSX_XLSM(self, workBook):
        workSheet = workBook.worksheets[0]
        if isinstance(workSheet.cell(70, 5).value, str) and workSheet.cell(70, 5).value:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0060_XLS(self, workBook):
        workSheet = workBook.worksheets[0]
        if isinstance(workSheet.cell_value(71, 5), str) and workSheet.cell_value(71, 5):
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0060_XLSX_XLSM(self, workBook):
        workSheet = workBook.worksheets[0]
        if isinstance(workSheet.cell(71, 5).value, str) and workSheet.cell(71, 5).value:
            return 1
        else:
            return 0

#Requirements for [DOC4]

    def Test_02043_18_04939_STRUCT_0410_XLS(self, workBook):

        cellNamesRow3 = ["Version", "To diagnose", "Supplier system", "Logical flow", "Physical flow", "Client system",
                         "Type of connection", "Type", "Logical failure mode", "Physical failure mode", "Wiring harness cause",
                         "Other cause", "Operation situation / Scenario", "system effect", "Customer effect", "Comment",
                         "Feared event", "Severity", "Level", "target","Safety measure (G4) / Functional diagnostic(G3,G2,G1)",
                         "Type of failure", "Degraded mode /Safe state", "lead time", "Efficiency", "recovering mode",
                         "Requirement N° to the Design Document", "Requirement N° from Design document",
                         "research time allocated to the system (in minutes)", "HMI\n(Indicators/messages)","High level test",
                         "Diagnosis needs", "Comments"]

        sheetNames = workBook.sheet_names()
        sheetNames = [x.casefold() for x in sheetNames]
        index = sheetNames.index("table")
        workSheet = workBook.sheets[index]
        rowsIterator = workSheet.rows(2)
        row3CellValues = list()
        for cell in rowsIterator:
            row3CellValues.append(cell.value.casefold())
        row3NumbersOfValues = len(cellNamesRow3)
        trueCases = 0
        for value in cellNamesRow3:
            if value.casefold() in row3CellValues:
                if row3CellValues.count(value.casefold()) is 1:
                    trueCases = trueCases + 1
            if "reference" in row3CellValues:
                if row3CellValues.count("reference") is 2:
                    trueCases = trueCases + 1
            if not trueCases is row3NumbersOfValues + 1:
                return 0



    def Test_02043_18_04939_STRUCT_0410_XLSX_XLSM(self, workBook):

        cellNamesRow3 = ["Version", "To diagnose", "Supplier system", "Logical flow", "Physical flow", "Client system",
                         "Type of connection", "Type", "Logical failure mode", "Physical failure mode", "Wiring harness cause",
                         "Other cause", "Operation situation / Scenario", "system effect", "Customer effect", "Comment",
                         "Feared event", "Severity", "Level", "target", "Safety measure (G4) / Functional diagnostic(G3,G2,G1)",
                         "Type of failure", "Degraded mode /Safe state", "lead time", "Efficiency", "recovering mode",
                         "Requirement N° to the Design Document", "Requirement N° from Design document",
                         "research time allocated to the system (in minutes)", "HMI\n(Indicators/messages)", "High level test",
                         "Diagnosis needs", "Comments"]

        sheetNames = workBook.sheetnames()
        sheetNames = [x.casefold() for x in sheetNames]
        index = sheetNames.index("table")
        workSheet = workBook.sheets[index]
        rowsIterator = workSheet.iter_rows(min_row=3, max_row=3)
        row3CellValues = list()
        for row in rowsIterator:
            for cell in row:
                row3CellValues.append(cell.value.casefold())
        row3NumberOfValues = len(cellNamesRow3)
        trueCases = 0
        for value in cellNamesRow3:
            if value.casefold() in row3CellValues:
                if row3CellValues.count(value.casefold()) is 1:
                    trueCases = trueCases + 1
        if "reference" in row3CellValues:
            if row3CellValues.count("reference") is 2:
                trueCases = trueCases +1
        if not trueCases is row3NumberOfValues + 1:
            return 0



    def buttonClicked(self):
        if not self.tab2.myTextBox5.toPlainText():
            self.download_file("https://docinfogroupe.psa-peugeot-citroen.com/ead/doc/ref.01272_18_00096/v.vc/pj")
        return

    def __init__(self):
        super().__init__()



    def colorButton(self):
        self.tab1.colorTextBox1.setStyleSheet('background-color: green')
        self.tab1.colorTextBox2.setStyleSheet('background-color: green')
        self.tab1.colorTextBox3.setStyleSheet('background-color: green')
        self.tab1.colorTextBox4.setStyleSheet('background-color: green')
        self.tab1.colorTextBox5.setStyleSheet('background-color: green')
        self.tab1.colorTextBox6.setStyleSheet('background-color: green')


    def readExcel(self, fileName):
        if fileName.endswith(".xls"):
            import xlrd as reader
            def openExcel(filename):
                return reader.open_workbook(filename)

            def getSheetNames(workbook):
                return workbook.sheet_names()

            self.excelExtension = ".xls"

        elif fileName.endswith((".xlsx", ".xlsm")):
            import openpyxl.reader as reader
            from openpyxl import load_workbook
            def openExcel(filename):
                if filename.endswith(".xlsm"):
                    return load_workbook(filename, keep_vba=True)
                else:
                    return load_workbook(filename)

            def getSheetNames(workbook):
                return workbook.sheetnames

            self.excelExtension = fileName[-5:]

        else:
            self.tab1.textbox.setText("Invalid file format")
            return

        workBook = openExcel(fileName)
        self.excelFile = workBook
        sheetNamesList = getSheetNames(workBook)
        sheetNamesString = str()
        for sheetNames in sheetNamesList:
            sheetNamesString = sheetNamesString + " " + sheetNames

        self.tab1.textbox.setText(sheetNamesString)
        return





if __name__ == '__main__':
    app = QApplication(sys.argv)
    apel = Test()
    myQLabel = QLabel()
    sys.exit(app.exec_())



