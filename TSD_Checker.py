import sys
from PyQt5.QtWidgets import QWidget, QPushButton, QApplication, QComboBox, QLabel, QLineEdit,  QTabWidget, QVBoxLayout, QProgressBar, QRadioButton
from PyQt5 import QtCore, QtWidgets
import openpyxl
import xlrd
import win32com.client as win32
import requests
import os
import io
from win32ui import FindWindow
from ctypes import windll
import datetime
from openpyxl.styles import Color


def colnum_string(n):
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string

class Application(QWidget):

    def __init__(self):
        super().__init__()
        self.left = 200
        self.top = 200
        self.width = 900
        self.height = 550
        self.tabs = QTabWidget()
        self.tab1 = QWidget()
        self.tab2 = QWidget()
        self.CesareLink = '''https://docinfogroupe.psa-peugeot-citroen.com/ead/doc/ref.02043_18_05471/v.vc/pj'''
        self.TSDConfigLink = '''https://docinfogroupe.psa-peugeot-citroen.com/ead/doc/ref.02043_18_05472/v.vc/pj'''
        self.CustomerEffectLink = '''https://docinfogroupe.psa-peugeot-citroen.com/ead/doc/ref.02043_18_05499/v.vc/pj'''
        self.DiversityLink = '''https://docinfogroupe.psa-peugeot-citroen.com/ead/doc/ref.02016_11_04964/v.vc/pj'''
        self.DOC3Link = '''https://docinfogroupe.psa-peugeot-citroen.com/ead/doc/ref.AEEV_IAEE07_0033/v.vc/pj'''
        self.DOC4Link = '''https://docinfogroupe.psa-peugeot-citroen.com/ead/doc/ref.02043_12_01665/v.vc/pj'''
        self.DOC5Link = '''https://docinfogroupe.psa-peugeot-citroen.com/ead/doc/ref.02043_12_01666/v.vc/pj'''
        self.DOC9Link = "https://docinfogroupe.psa-peugeot-citroen.com/ead/doc/ref.02043_18_05474/v.vc/pj"
        self.DOC14Link = "https://docinfogroupe.psa-peugeot-citroen.com/ead/doc/ref.02043_19_00392/v.vc/pj"
        self.DOC8Link = "https://docinfogroupe.psa-peugeot-citroen.com/ead/doc/ref.02043_18_05474/v.vc/pj"
        self.tabs.addTab(self.tab1, "TSD Checker")
        self.tabs.addTab(self.tab2, "Options")
        self.initUI(self.tab1)
        self.initUIOptions(self.tab2)
        self.layout = QVBoxLayout()
        self.layout.addWidget(self.tabs)
        self.setLayout(self.layout)
        self.setWindowTitle("TSD Checker  V0.5")

    def ToggleLink(self):
        if self.tab2.RadioButtonInternet.isChecked() == True:
            self.CesareLink = '''https://docinfogroupe.psa-peugeot-citroen.com/ead/doc/ref.02043_18_05471/v.vc/pj'''
            self.TSDConfigLink = '''https://docinfogroupe.psa-peugeot-citroen.com/ead/doc/ref.02043_18_05472/v.vc/pj'''
            self.CustomerEffectLink = '''https://docinfogroupe.psa-peugeot-citroen.com/ead/doc/ref.02043_18_05499/v.vc/pj'''
            self.DiversityLink = '''https://docinfogroupe.psa-peugeot-citroen.com/ead/doc/ref.02016_11_04964/v.vc/pj'''
            self.DOC3Link = '''https://docinfogroupe.psa-peugeot-citroen.com/ead/doc/ref.AEEV_IAEE07_0033/v.vc/pj'''
            self.DOC4Link = '''https://docinfogroupe.psa-peugeot-citroen.com/ead/doc/ref.02043_12_01665/v.vc/pj'''
            self.DOC5Link = '''https://docinfogroupe.psa-peugeot-citroen.com/ead/doc/ref.02043_12_01666/v.vc/pj'''
            self.DOC9Link = "https://docinfogroupe.psa-peugeot-citroen.com/ead/doc/ref.02043_18_05474/v.vc/pj"
            self.DOC14Link = "https://docinfogroupe.psa-peugeot-citroen.com/ead/doc/ref.02043_19_00392/v.vc/pj"
            self.DOC8Link = "https://docinfogroupe.psa-peugeot-citroen.com/ead/doc/ref.02043_18_05474/v.vc/pj"
            self.tab2.link2.setText('''<a href=''' + self.CesareLink + '''>DocInfo Reference: 02043_18_05471</a>''')
            self.tab2.link1.setText('''<a href=''' + self.TSDConfigLink + '''>DocInfo Reference: 02043_18_05472</a>''')
            self.tab2.link3.setText('''<a href=''' + self.CustomerEffectLink + '''>DocInfo Reference: 02043_18_05499</a>''')
            self.tab2.link4.setText('''<a href=''' + self.DiversityLink + '''>DocInfo Reference: 02016_11_04964</a>''')
        elif self.tab2.RadioButtonIntranet.isChecked() == True:
            self.CesareLink = "http://docinfogroupe.inetpsa.com/ead/doc/ref.02043_18_05471/v.vc/pj"
            self.TSDConfigLink = "http://docinfogroupe.inetpsa.com/ead/doc/ref.02043_18_05472/v.vc/pj"
            self.CustomerEffectLink = "http://docinfogroupe.inetpsa.com/ead/doc/ref.02043_18_05499/v.vc/pj"
            self.DiversityLink = "http://docinfogroupe.inetpsa.com/ead/doc/ref.02016_11_04964/v.vc/pj"
            self.DOC3Link = '''http://docinfogroupe.inetpsa.com/ead/doc/ref.AEEV_IAEE07_0033/v.vc/pj'''
            self.DOC4Link = '''http://docinfogroupe.inetpsa.com/ead/doc/ref.02043_12_01665/v.vc/pj'''
            self.DOC5Link = '''http://docinfogroupe.inetpsa.com/ead/doc/ref.02043_12_01666/v.vc/pj'''
            self.DOC9Link = "http://docinfogroupe.inetpsa.com/ead/doc/ref.02043_18_05474/v.vc/pj"
            self.DOC14Link = "http://docinfogroupe.inetpsa.com/ead/doc/ref.02043_19_00392/v.vc/pj"
            self.DOC8Link = "http://docinfogroupe.inetpsa.com/ead/doc/ref.02043_18_05474/v.vc/pj"
            self.tab2.link2.setText('''<a href=''' + self.CesareLink + '''>DocInfo Reference: 02043_18_05471</a>''')
            self.tab2.link1.setText('''<a href=''' + self.TSDConfigLink + '''>DocInfo Reference: 02043_18_05472</a>''')
            self.tab2.link3.setText('''<a href=''' + self.CustomerEffectLink + '''>DocInfo Reference: 02043_18_05499</a>''')
            self.tab2.link4.setText('''<a href=''' + self.DiversityLink + '''>DocInfo Reference: 02016_11_04964</a>''')

        else:
            self.tab1.setText("ERROR: Incorrect network type")

    def openFileNameDialog1(self):
        fileName1, _filter = QtWidgets.QFileDialog.getOpenFileName(self.tab1, 'Open File', QtCore.QDir.rootPath(), '*.*')
        self.tab1.myTextBox1.setText(fileName1)
        #self.tab1.textbox.setText("next file")

    def openFileNameDialog2(self):
        fileName2, _filter = QtWidgets.QFileDialog.getOpenFileName(self.tab1, 'Open File', QtCore.QDir.rootPath(), '*.*')
        self.tab1.myTextBox2.setText(fileName2)

    def openFileNameDialog3(self):
        fileName3, _filter = QtWidgets.QFileDialog.getOpenFileName(self.tab1, 'Open File', QtCore.QDir.rootPath(), '*.*')
        self.tab1.myTextBox3.setText(fileName3)

    def openFileNameDialog4(self):
        fileName4, _filter = QtWidgets.QFileDialog.getOpenFileName(self.tab1, 'Open File', QtCore.QDir.rootPath(), '*.*')
        self.tab2.myTextBox4.setText(fileName4)

    def openFileNameDialog5(self):
        fileName5, _filter = QtWidgets.QFileDialog.getOpenFileName(self.tab1, 'Open File', QtCore.QDir.rootPath(), '*.*')
        self.tab2.myTextBox5.setText(fileName5)

    def openFileNameDialog6(self):
        fileName6, _filter = QtWidgets.QFileDialog.getOpenFileName(self.tab1, 'Open File', QtCore.QDir.rootPath(), '*.*')
        self.tab2.myTextBox6.setText(fileName6)

    def openFileNameDialog7(self):
        fileName7, _filter = QtWidgets.QFileDialog.getOpenFileName(self.tab1, 'Open File', QtCore.QDir.rootPath(), '*.*')
        self.tab1.myTextBox7.setText(fileName7)

    def openFileNameDialog8(self):
        fileName8, _filter = QtWidgets.QFileDialog.getOpenFileName(self.tab1, 'Open File', QtCore.QDir.rootPath(), '*.*')
        self.tab1.myTextBox8.setText(fileName8)

    def openFileNameDialog9(self):
        fileName9, _filter = QtWidgets.QFileDialog.getOpenFileName(self.tab1, 'Open File', QtCore.QDir.rootPath(), '*.*')
        self.tab2.myTextBox9.setText(fileName9)

    def openFileNameDialog10(self):
        fileName10, _filter = QtWidgets.QFileDialog.getOpenFileName(self.tab1, 'Open File', QtCore.QDir.rootPath(), '*.*')
        self.tab1.myTextBox10.setText(fileName10)

    def ButtonReportClick(self):
        if self.tab1.myTextBox1.toPlainText():
           fileName = self.tab1.myTextBox1.toPlainText()
           self.excel = win32.gencache.EnsureDispatch('Excel.Application')
           self.excel.Visible = True
           self.excel.Workbooks.Open(fileName)

        if self.tab1.myTextBox2.toPlainText():
           fileName = self.tab1.myTextBox2.toPlainText()
           self.excel = win32.gencache.EnsureDispatch('Excel.Application')
           self.excel.Visible = True
           self.excel.Workbooks.Open(fileName)

        if self.tab1.myTextBox3.toPlainText():
           fileName = self.tab1.myTextBox3.toPlainText()
           self.excel = win32.gencache.EnsureDispatch('Excel.Application')
           self.excel.Visible = True
           self.excel.Workbooks.Open(fileName)
        #return fileName

    def buttonClicked(self):
        return

    def download_doc9(self, url):

        user = self.tab2.TextBoxUser.text()
        user = str(user)
        password = self.tab2.TextBoxPass.text()
        password = str(password)
        username = os.environ['USERNAME']
        out_path = "C:/Users/" + username + "/AppData/Local/Temp/TSD_Checker/"
        if not user or not password:
            self.tab1.textbox.setText("Missing Username or Password")
            return "False"

        try:
            os.stat(out_path)
        except:
            os.mkdir(out_path)

        try:
           response = requests.get(url, stream=True, auth=(user, password))
        except:
            return "Error"

        status = response.status_code
        if status == 401:
            self.tab1.textbox.setText("Username or Password Incorrect")
            return "False"

        FileName = response.headers['Content-Disposition'].split('"')[1]
        FilePath = out_path + FileName
        print("Saving file to location:" + FilePath)
        with open(FilePath, 'wb') as f:
            for chuck in response.iter_content(chunk_size=128):
                f.write(chuck)
        return FilePath

    def download_DOC3(self, url):
        user = self.tab2.TextBoxUser.text()
        user = str(user)
        password = self.tab2.TextBoxPass.text()
        password = str(password)
        username = os.environ['USERNAME']
        out_path = "C:/Users/" + username + "/AppData/Local/Temp/TSD_Checker/"
        if not user or not password:
            if not user or not password:
                self.tab1.textbox.setText("Missing Username or Password")
                return "False"

        try:
            os.stat(out_path)
        except:
            os.mkdir(out_path)

        try:
           response = requests.get(url, stream=True, auth=(user, password))
        except:
            return "Error"

        status = response.status_code
        if status == 401:
            self.tab1.textbox.setText("Username or Password Incorrect")
            return "False"

        FileName = response.headers['Content-Disposition'].split('"')[1]

        if self.tab1.myTextBox1.toPlainText():
            FileName_new = "DOC3" + os.path.splitext(FileName)[1]

        # change file extension
        FilePath = out_path + os.path.splitext(FileName_new)[0] + ".xls"

        with open(FilePath, 'wb') as f:
            for chuck in response.iter_content(chunk_size=128):
                f.write(chuck)
        return FilePath

    def download_DOC4(self, url):
        user = self.tab2.TextBoxUser.text()
        user = str(user)
        password = self.tab2.TextBoxPass.text()
        password = str(password)
        username = os.environ['USERNAME']
        out_path = "C:/Users/" + username + "/AppData/Local/Temp/TSD_Checker/"
        if not user or not password:
            self.tab1.textbox.setText("Missing Username or Password")
            return "False"

        try:
            os.stat(out_path)
        except:
            os.mkdir(out_path)

        try:
           response = requests.get(url, stream=True, auth=(user, password))
        except:
            return "Error"

        status = response.status_code
        if status == 401:
            self.tab1.textbox.setText("Username or Password Incorrect")
            return "False"

        FileName = response.headers['Content-Disposition'].split('"')[1]

        if self.tab1.myTextBox2.toPlainText():
            FileName_new = "DOC4" + os.path.splitext(FileName)[1]

        # change file extension
        FilePath = out_path + os.path.splitext(FileName_new)[0] + ".xls"

        with open(FilePath, 'wb') as f:
            for chuck in response.iter_content(chunk_size=128):
                f.write(chuck)
        return FilePath



    def download_file(self, url):
        user = self.tab2.TextBoxUser.text()
        user = str(user)
        password = self.tab2.TextBoxPass.text()
        password = str(password)
        username = os.environ['USERNAME']
        out_path = "C:/Users/" + username + "/AppData/Local/Temp/TSD_Checker/"
        if not user or not password:
            self.tab1.textbox.setText("Missing Username or Password")
            return "False"
        try:
            os.stat(out_path)
        except:
            os.mkdir(out_path)
        try:
            response = requests.get(url, stream=True, auth=(user, password))
        except:
            return "Error"
        status = response.status_code
        if status == 401:
            self.tab1.textbox.setText("Username or Password Incorrect")
            return "False"

        FileName = response.headers['Content-Disposition'].split('"')[1]
        FilePath = out_path + FileName
        success_download = self.tab1.textbox.toPlainText()
        success_download = success_download + "\nfile " + FileName + " has been successfully downloaded\n=======================\n"
        print("Saving file to location:" + FilePath)
        self.tab1.textbox.setText(success_download)
        with open(FilePath, 'wb') as f:
            for chuck in response.iter_content(chunk_size=128):
                f.write(chuck)
        return FilePath

    def initUI(self, tab):

    # Create a textbox
        tab.message = ""
        tab.textbox = QtWidgets.QTextEdit(self.tab1)
        tab.textbox.setText(tab.message)
        tab.textbox.move(10, 270)
        tab.textbox.resize(700, 130)
        tab.textbox.setReadOnly(True)

        sb = tab.textbox.verticalScrollBar()
        sb.setValue(sb.minimum())

    #create a progress bar
        tab.pbar = QProgressBar(self.tab1)
        tab.pbar.setGeometry(10, 310, 700, 20)
        tab.pbar.setAlignment(QtCore.Qt.AlignCenter)
        tab.pbar.setValue(0)
        tab.pbar.move(10, 410)

    #Create a color textbox1
        tab.colorTextBox1 = QtWidgets.QTextEdit(self.tab1)
        tab.colorTextBox1.setStyleSheet(  " background-color: grey ")
        tab.colorTextBox1.resize(20, 20)
        tab.colorTextBox1.move(710, 10)


    # Create a color textbox2
        tab.colorTextBox2 = QtWidgets.QTextEdit(self.tab1)
        tab.colorTextBox2.setStyleSheet(" background-color: grey ")
        tab.colorTextBox2.resize(20, 20)
        tab.colorTextBox2.move(710, 40)

    # Create a color textbox3
        tab.colorTextBox3 = QtWidgets.QTextEdit(self.tab1)
        tab.colorTextBox3.setStyleSheet(" background-color: grey ")
        tab.colorTextBox3.resize(20, 20)
        tab.colorTextBox3.move(710, 70)

    # Create a color textbox4
        tab.colorTextBox4 = QtWidgets.QTextEdit(self.tab1)
        tab.colorTextBox4.setStyleSheet(" background-color: grey ")
        tab.colorTextBox4.resize(20, 20)
        tab.colorTextBox4.move(710, 100)

    # Create a color textbox5
        tab.colorTextBox5 = QtWidgets.QTextEdit(self.tab1)
        tab.colorTextBox5.setStyleSheet(" background-color: grey ")
        tab.colorTextBox5.resize(20, 20)
        tab.colorTextBox5.move(710, 130)

    # Create a color textbox6
        tab.colorTextBox6 = QtWidgets.QTextEdit(self.tab1)
        tab.colorTextBox6.setStyleSheet(" background-color: grey ")
        tab.colorTextBox6.resize(20, 20)
        tab.colorTextBox6.move(710, 160)

    #Create a drop down list
        tab.lbl = QLabel("Check level", tab)

        tab.combo = QComboBox(tab)
        tab.combo.addItem("Previsional")
        tab.combo.addItem("Consolidated")
        tab.combo.addItem("Validated")
        tab.combo.resize(508, 20.4)  #rezise the drop down list
        tab.combo.move(200, 200)
        tab.lbl.move(5, 205)
        tab.combo.activated[str].connect(self.onActivated)


    # Create a drop down list
        tab.lbl1 = QLabel("Project name", tab)

        tab.combo1 = QComboBox(tab)
        tab.combo1.addItem("   Generic   ")
        tab.combo1.addItem("   All   ")
        tab.combo1.resize(378, 20.4)  # rezise the drop down list
        tab.combo1.move(200, 230)
        tab.lbl1.move(5, 235)
        tab.combo1.activated[str].connect(self.onActivated)

        self.setGeometry(self.left, self.top, self.width, self.height)
        self.setWindowTitle('TSD Checker')

        tab.importNames = QPushButton(tab)
        tab.importNames.setText("Import Project Names")
        tab.importNames.resize(120, 20.4)
        tab.importNames.move(585, 230)


        #File Selectiom Dialog1
        tab.lbl2 = QLabel("TSD File:", tab)
        tab.lbl2.move(5,15)
        tab.myTextBox1 = QtWidgets.QTextEdit(tab)
        tab.myTextBox1.resize(460, 25)
        tab.myTextBox1.move(200, 10)
        tab.myTextBox1.setReadOnly(True)
        tab.myTextBox1.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        tab.button1 = QPushButton('...',tab)
        tab.button1.clicked.connect(self.openFileNameDialog1)
        tab.button1.move(660, 10)
        tab.button1.resize(45,22)

    # File Selectiom Dialog2
        tab.lbl3 = QLabel("TSD vehicle Function file:", tab)
        tab.lbl3.move(5, 45)
        tab.myTextBox2 = QtWidgets.QTextEdit(tab)
        tab.myTextBox2.resize(460, 25)
        tab.myTextBox2.move(200, 40)
        tab.myTextBox2.setReadOnly(True)
        tab.myTextBox2.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        tab.button2 = QPushButton('...', tab)
        tab.button2.clicked.connect(self.openFileNameDialog2)
        tab.button2.move(660, 40)
        tab.button2.resize(45, 22)

    # File Selectiom Dialog3
        tab.lbl4 = QLabel("TSD system file:", tab)
        tab.lbl4.move(5, 75)
        tab.myTextBox3 = QtWidgets.QTextEdit(tab)
        tab.myTextBox3.resize(460, 25)
        tab.myTextBox3.move(200, 70)
        tab.myTextBox3.setReadOnly(True)
        tab.myTextBox3.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        tab.button3 = QPushButton('...', tab)
        tab.button3.clicked.connect(self.openFileNameDialog3)
        tab.button3.move(660, 70)
        tab.button3.resize(45, 22)



    # File Selectiom Dialog7
        tab.lbl8 = QLabel("AMDEC:", tab)
        tab.lbl8.move(5, 105)
        tab.myTextBox7 = QtWidgets.QTextEdit(tab)
        tab.myTextBox7.resize(460, 25)
        tab.myTextBox7.move(200, 100)
        tab.myTextBox7.setReadOnly(True)
        tab.myTextBox7.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        tab.button7 = QPushButton('...', tab)
        tab.button7.clicked.connect(self.openFileNameDialog7)
        tab.button7.move(660, 100)
        tab.button7.resize(45, 22)

    # File Selectiom Dialog8
        tab.lbl9 = QLabel("export MedialecMatrice:", tab)
        tab.lbl9.move(5, 135)
        tab.myTextBox8 = QtWidgets.QTextEdit(tab)
        tab.myTextBox8.resize(460, 25)
        tab.myTextBox8.move(200, 130)
        tab.myTextBox8.setReadOnly(True)
        tab.button8 = QPushButton('...', tab)
        tab.button8.clicked.connect(self.openFileNameDialog8)
        tab.button8.move(660, 130)
        tab.button8.resize(45, 22)
        tab.myTextBox8.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)



    # File Selectiom Dialog10
        tab.lbl11 = QLabel("Diagnostic matrix file:", tab)
        tab.lbl11.move(5, 165)
        tab.myTextBox10 = QtWidgets.QTextEdit(tab)
        tab.myTextBox10.resize(460, 25)
        tab.myTextBox10.move(200, 160)
        tab.myTextBox10.setReadOnly(True)
        tab.button10 = QPushButton('...', tab)
        tab.button10.clicked.connect(self.openFileNameDialog10)
        tab.button10.move(660, 160)
        tab.button10.resize(45, 22)
        tab.myTextBox10.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)



    # Check button
        button = QPushButton('Check', tab)
        button.move(310, 470)
        button.resize(90,25)
        button.clicked.connect(self.buttonClicked)
        button.setStyleSheet('QPushButton {background-color: white; color: black;}')
        buttonNew = QPushButton("Open \nReport", tab)
        buttonNew.resize(90, 60)
        buttonNew.move(710, 310)
        buttonNew.clicked.connect(self.ButtonReportClick)


        self.show()

    def initUIOptions(self, tab):

        tab.lblUser = QLabel("USER:", tab)
        tab.lblUser.move(165,25)
        tab.TextBoxUser = QtWidgets.QLineEdit(tab)
        tab.TextBoxUser.resize(200,25)
        tab.TextBoxUser.move(200, 20)
        tab.TextBoxUser.setText("E518720")


        tab.lblPass = QLabel("PASSWORD:", tab)
        tab.lblPass.move(450,25)
        tab.TextBoxPass = QtWidgets.QLineEdit(tab)
        tab.TextBoxPass.resize(180,25)
        tab.TextBoxPass.move(520, 20)
        tab.TextBoxPass.setEchoMode((QLineEdit.Password))
        tab.TextBoxPass.setText("Cst78788")


        # File Selectiom Dialog5
        tab.lbl6 = QLabel("Famille/Sous-Famille list export(CESARE):", tab)
        tab.lbl6.move(5, 145)
        tab.myTextBox5 = QtWidgets.QTextEdit(tab)
        tab.myTextBox5.resize(460, 25)
        tab.myTextBox5.move(200, 140)
        tab.myTextBox5.setReadOnly(True)

        tab.link2 = QLabel('''<a href=''' + self.CesareLink + '''>DocInfo Reference: 02043_18_05471</a>''', tab)
        tab.link2.setOpenExternalLinks(True)
        tab.link2.move(720, 145)


        tab.button5 = QPushButton('...', tab)
        tab.button5.move(660, 140)
        tab.button5.resize(45, 22)



        # File Selectiom Dialog4
        tab.lbl5 = QLabel("TSD configuration file:", tab)
        tab.lbl5.move(5,185)
        tab.myTextBox4 = QtWidgets.QTextEdit(tab)
        tab.myTextBox4.resize(460, 25)
        tab.myTextBox4.move(200, 180)
        tab.myTextBox4.setReadOnly(True)

        tab.link1 = QLabel('''<a href='''+self.TSDConfigLink+'''>DocInfo Reference: 02043_18_05472</a>''', tab)
        tab.link1.setOpenExternalLinks(True)
        tab.link1.move(720, 185)

        tab.button4 = QPushButton('...', tab)
        tab.button4.clicked.connect(self.openFileNameDialog4)
        tab.button4.move(660, 180)
        tab.button4.resize(45, 22)



        # File Selectiom Dialog6
        tab.lbl7 = QLabel("Customer effect file:", tab)
        tab.lbl7.move(5, 225)
        tab.myTextBox6 = QtWidgets.QTextEdit(tab)
        tab.myTextBox6.resize(460, 25)
        tab.myTextBox6.move(200, 220)
        tab.myTextBox6.setReadOnly(True)

        tab.link3 = QLabel('''<a href=''' + self.CustomerEffectLink + '''>DocInfo Reference: 02043_18_05499</a>''', tab)
        tab.link3.setOpenExternalLinks(True)
        tab.link3.move(720, 225)



        tab.button6 = QPushButton('...', tab)
        tab.button6.clicked.connect(self.openFileNameDialog6)
        tab.button6.move(660, 220)
        tab.button6.resize(45, 22)

        # File Selectiom Dialog9
        tab.lbl10 = QLabel("Diversity management file:", tab)
        tab.lbl10.move(5, 265)
        tab.myTextBox9 = QtWidgets.QTextEdit(tab)
        tab.myTextBox9.resize(460, 25)
        tab.myTextBox9.move(200,260)
        tab.myTextBox9.setReadOnly(True)

        tab.link4 = QLabel('''<a href=''' + self.DiversityLink + '''>DocInfo Reference: 02016_11_04964</a>''', tab)
        tab.link4.setOpenExternalLinks(True)
        tab.link4.move(720, 265)


        tab.button9 = QPushButton('...', tab)
        tab.button9.clicked.connect(self.openFileNameDialog9)
        tab.button9.move(660, 260)
        tab.button9.resize(45, 22)

        tab.labelInternetAndIntranet = QLabel("Network Type:", tab)
        tab.labelInternetAndIntranet.move(130, 60)
        tab.RadioButtonInternet = QRadioButton(self.tab2)
        tab.RadioButtonInternet.setText("Internet link")
        tab.RadioButtonInternet.setChecked(True)
        tab.RadioButtonIntranet = QRadioButton(self.tab2)
        tab.RadioButtonIntranet.setText("Intranet link")
        tab.RadioButtonInternet.toggled.connect(self.ToggleLink)
        tab.RadioButtonIntranet.toggled.connect(self.ToggleLink)
        tab.RadioButtonInternet.move(210, 58)
        tab.RadioButtonIntranet.move(210, 90)

    def onActivated(self):
        return

class Test(Application):

    def __init__(self):
        super().__init__()
        self.tsdFileExtension = str()
        self.tsdVehicleFunctionFileExtension = str()
        self.tsdSystemFileExtension = str()
        self.amdecFileExtension = str()
        self.exportMedialecMatriceFileExtension = str()
        self.diagnosticMatrixFileExtension = str()
        self.pbvalue = 0
        username = os.environ['USERNAME']
        out_path = "C:/Users/" + username + "/AppData/Local/Temp/TSD_Checker/"

        try:
            os.stat(out_path)
            filelist = [file for file in os.listdir(out_path)]
            for filename in filelist:
                os.remove(out_path + filename)
        except:
            os.mkdir(out_path)

    def TestGeneralStructureXLS_DOC3(self, workBook, fileName, path_Cesare, path_effect):
        flag = 1
        testResult_0005 = self.Test_02043_18_04939_STRUCT_0005_XLS(fileName)
        fileName = self.tab1.myTextBox1.toPlainText()
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        reportWorkBook = excel.Workbooks.Open(fileName)
        reportWorkSheet1 = None
        reportWorkSheet2 = None
        for sheet in reportWorkBook.Worksheets:
            if sheet.Name == "Report Information":
                reportWorkSheet1 = sheet
        if not reportWorkSheet1:
            workSheetsNumber = reportWorkBook.Sheets.Count
            sheetAfter = reportWorkBook.Sheets(workSheetsNumber)
            reportWorkSheet1 = reportWorkBook.Worksheets.Add(None, sheetAfter)
            reportWorkSheet1.Name = "Report Information"

        reportInformationCol1StringList = ["Tool version:", "Criticity configuration file:", "", "Extract CESARE file:",
                                           "Customer effects file:", "check level:", "", "Date of the test:",
                                           "Time of the test:",
                                           "", "TSD file checked:", "TSD function file checked:",
                                           "TSD system file checked:",
                                           "", "AMDEC:", "export MedialecMatrice:", "", "Status:"]
        testReportRow1StringList = ["Criticity", "Requirements", "Message", "Localisation"]

        for i, name in enumerate(reportInformationCol1StringList):
            reportWorkSheet1.Cells(i + 1, 1).Value = name
        reportWorkSheet1.Columns.AutoFit()
        reportWorkSheet1.Columns.Font.Bold = True

        if self.tab1.myTextBox1.toPlainText():
            TSD_file_checked = self.tab1.myTextBox1.toPlainText()
        else:
            TSD_file_checked = "None"

        if self.tab1.myTextBox2.toPlainText():
            TSD_function_file_checked = self.tab1.myTextBox2.toPlainText()
        else:
            TSD_function_file_checked = "None"

        if self.tab1.myTextBox3.toPlainText():
            TSD_system_file_checked = self.tab1.myTextBox3.toPlainText()
        else:
            TSD_system_file_checked = "None"

        if self.tab1.myTextBox7.toPlainText():
            amdec = self.tab1.myTextBox7.toPlainText()
        else:
            amdec = "None"

        if self.tab1.myTextBox8.toPlainText():
            MedialecMatrice = self.tab1.myTextBox8.toPlainText()
        else:
            MedialecMatrice = "None"

        # if self.tab2.myTextBox5.toPlainText():
        #     cesare = self.tab2.myTextBox5.toPlainText()
        # else:
        #     cesare = self.download_file(self.CesareLink)
        #
        # if self.tab2.myTextBox6.toPlainText():
        #     CustomerEffect = self.tab2.myTextBox6.toPlainText()
        # else:
        #     CustomerEffect = self.download_file(self.CustomerEffectLink)

        if self.TestTsdFile is 1:
            status = "Pass"
        else:
            status = "Fail"

        check_level = str(self.tab1.combo.currentText())

        now = datetime.datetime.now()

        reportCol2 = ["V0.4", "02043_18_05474_2_1_Check_TSD_Criticity_Configuration_file.xlsx", "", path_Cesare.split("/")[-1],
                      path_effect.split("/")[-1], check_level,
                      "", now.strftime("%Y-%m-%d"), now.strftime("%H:%M"), "",
                      TSD_file_checked.split("/")[-1], TSD_function_file_checked.split("/")[-1], TSD_system_file_checked.split("/")[-1], "", amdec.split("/")[-1], MedialecMatrice.split("/")[-1],
                      "", status]
        for i, name in enumerate(reportCol2):
            reportWorkSheet1.Cells(i + 1, 2).Value = name

        for sheet in reportWorkBook.Worksheets:
            if sheet.Name == "Test Report":
                reportWorkSheet2 = sheet
        if not reportWorkSheet2:
            workSheetsNumber = reportWorkBook.Sheets.Count
            sheetAfter = reportWorkBook.Sheets(workSheetsNumber)
            reportWorkSheet2 = reportWorkBook.Worksheets.Add(None, sheetAfter)
            reportWorkSheet2.Name = "Test Report"

        for i, name in enumerate(testReportRow1StringList):
            reportWorkSheet2.Cells(1, i + 1).Value = name
        reportWorkSheet2.Columns.AutoFit()
        reportWorkSheet2.Columns.Font.Bold = True

        row = 2

        testResult = self.Test_02043_18_04939_STRUCT_0000_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0000 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0000", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0000: The sheet “Informations Générales” (or “General information”) is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0000"][self.checkLevel], "02043_18_04939_STRUCT_0000", "The sheet “Informations Générales” (or “General information”) is not present or not written correctly.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        if testResult_0005 == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0005 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0005", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0010: The field “REFERENCE” of the sheet “Informations Générales” (or “General information”)  in the line 52 shall be indicated."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0005"][self.checkLevel], "02043_18_04939_STRUCT_0005", "The field “REFERENCE” of the sheet “Informations Générales” (or “General information”)  in the line 52 shall be indicated.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0010_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0010 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0010", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0010: The information “Ref plan type” is missing in the sheet “Informations Générales” (or “General information”)."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0010"][self.checkLevel], "02043_18_04939_STRUCT_0010", "The information “Ref plan type” is missing in the sheet “Informations Générales” (or “General information”).", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)


        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0011_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0011 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0011", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0011: The document does not specify the template or the template reference is not indicated in the sheet “Informations Générales” (or “General information”). \nAs indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0011"][self.checkLevel], "02043_18_04939_STRUCT_0011", "The document does not specify the template or the template reference is not indicated in the sheet “Informations Générales” (or “General information”). \nAs indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)


        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0020_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0020 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0020", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0020: The sheet “Suppression ” (or “suppression ”) is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0020"][self.checkLevel], "02043_18_04939_STRUCT_0020", "The sheet “Suppression ” (or “suppression ”) is not present or not written correctly.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)


        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0025_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0025 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0025", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0025: The document does not follow the template, the column “Onglet” (or “sheet”) of the sheet “Suppression” (or “suppression”) is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0025"][self.checkLevel], "02043_18_04939_STRUCT_0025", "The document does not follow the template, the column “Onglet” (or “sheet”) of the sheet “Suppression” (or “suppression”) is not present or not written correctly.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)


        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0030_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0030 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0030", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0030: The document does not follow the template, the column “Version du TSD” (or “Version of the document”) of the sheet “Suppression” (or “suppression”) is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0030"][self.checkLevel], "02043_18_04939_STRUCT_0030", "The document does not follow the template, the column “Version du TSD” (or “Version of the document”) of the sheet “Suppression” (or “suppression”) is not present or not written correctly.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)


        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0035_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0035 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0035", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0035: The document does not follow the template, the column “Version du TSD” (or “Version of the document”) of the sheet “Suppression” (or “suppression”) is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0035"][self.checkLevel], "02043_18_04939_STRUCT_0030", "The document does not follow the template, the column “Version du TSD” (or “Version of the document”) of the sheet “Suppression” (or “suppression”) is not present or not written correctly.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)


        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0040_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0040 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0040", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0040: The document does not follow the template, the column “Justification de la modification” (or “Change reason”) of the sheet “Suppression” (or “suppression”) is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = False
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0040"][self.checkLevel], "02043_18_04939_STRUCT_0040", "The document does not follow the template, the column “Justification de la modification” (or “Change reason”) of the sheet “Suppression” (or “suppression”) is not present or not written correctly.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)


        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0051_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0051 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0051", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0051: The “Vehicle Architecture Schematic” document is not referenced. \nAs indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0051"][self.checkLevel], "02043_18_04939_STRUCT_0051", "The “Vehicle Architecture Schematic” document is not referenced. \nAs indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)


        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0052_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0052 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0052", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0052: The “Diagnostic Matrix” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0052"][self.checkLevel], "02043_18_04939_STRUCT_0052", "The “Diagnostic Matrix” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)


        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0053_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0053 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0053", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0053: The “Fault Tree” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0053"][self.checkLevel], "02043_18_04939_STRUCT_0053", "The “Fault Tree” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)


        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0054_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0054 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0054", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0054: The “ECU schematic” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0054"][self.checkLevel], "02043_18_04939_STRUCT_0054", "The “ECU schematic” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)


        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0055_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0055 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0055", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0055: The “STD” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0055"][self.checkLevel], "02043_18_04939_STRUCT_0055", "The “ECU schematic” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)


        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0056_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0056 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0056", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0056: The “Complexity Matrix (Decli EE)” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0056"][self.checkLevel], "02043_18_04939_STRUCT_0056", "The “Complexity Matrix (Decli EE)” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)


        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0057_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0057 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0057", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0057: The “Décli” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0057"][self.checkLevel], "02043_18_04939_STRUCT_0057", "The “Décli” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)


        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0058_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0058 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0058", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0058: The “DCEE” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0058"][self.checkLevel], "02043_18_04939_STRUCT_0058", "The “DCEE” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)


        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0059_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0059 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0059", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0059: The “EEAD” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0059"][self.checkLevel], "02043_18_04939_STRUCT_0059", "The “EEAD” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)


        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0060_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0060 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0060", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0060: The “TFD” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666"
            self.tab1.textbox.setText(text)
            flag = False
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0060"][self.checkLevel], "02043_18_04939_STRUCT_0060", "The “TFD” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        try:
            reportWorkBook.Save()
        except Exception as e:
            print(e)
        return flag

    def TestGeneralStructureXLSX_XLSM_DOC3(self, workBook,  path_Cesare, path_effect):
        flag = 1
        if self.tab1.myTextBox1.toPlainText():
           fileName = self.tab1.myTextBox1.toPlainText()
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        reportWorkBook = excel.Workbooks.Open(fileName)
        reportWorkSheet1 = None
        reportWorkSheet2 = None
        for sheet in reportWorkBook.Worksheets:
            if sheet.Name == "Report Information":
                reportWorkSheet1 = sheet
        if not reportWorkSheet1:
            workSheetsNumber = reportWorkBook.Sheets.Count
            sheetAfter = reportWorkBook.Sheets(workSheetsNumber)
            reportWorkSheet1 = reportWorkBook.Worksheets.Add(None, sheetAfter)
            reportWorkSheet1.Name = "Report Information"

        reportInformationCol1StringList = ["Tool version:", "Criticity configuration file:", "", "Extract CESARE file:",
                                           "Customer effects file:", "check level:", "", "Date of the test:",
                                           "Time of the test:",
                                           "", "TSD file checked:", "TSD function file checked:",
                                           "TSD system file checked:",
                                           "", "AMDEC:", "export MedialecMatrice:", "", "Status:"]
        testReportRow1StringList = ["Criticity", "Requirements", "Message", "Localisation"]

        for i, name in enumerate(reportInformationCol1StringList):
            reportWorkSheet1.Cells(i + 1, 1).Value = name
        reportWorkSheet1.Columns.AutoFit()
        reportWorkSheet1.Columns.Font.Bold = True

        if self.tab1.myTextBox1.toPlainText():
            TSD_file_checked = self.tab1.myTextBox1.toPlainText()
        else:
            TSD_file_checked = "None"

        if self.tab1.myTextBox2.toPlainText():
            TSD_function_file_checked = self.tab1.myTextBox2.toPlainText()
        else:
            TSD_function_file_checked = "None"

        if self.tab1.myTextBox3.toPlainText():
            TSD_system_file_checked = self.tab1.myTextBox3.toPlainText()
        else:
            TSD_system_file_checked = "None"

        if self.tab1.myTextBox7.toPlainText():
            amdec = self.tab1.myTextBox3.toPlainText()
        else:
            amdec = "None"

        if self.tab1.myTextBox8.toPlainText():
            MedialecMatrice = self.tab1.myTextBox8.toPlainText()
        else:
            MedialecMatrice = "None"

        # if self.tab2.myTextBox5.toPlainText():
        #     cesare = self.tab2.myTextBox5.toPlainText()
        # else:
        #     cesare = self.download_file(self.CesareLink)
        #
        # if self.tab2.myTextBox6.toPlainText():
        #     CustomerEffect = self.tab2.myTextBox6.toPlainText()
        # else:
        #     CustomerEffect = self.download_file(self.CustomerEffectLink)

        if self.TestTsdFile is 1:
            status = "Pass"
        else:
            status = "Fail"

        check_level = str(self.tab1.combo.currentText())

        now = datetime.datetime.now()

        reportCol2 = ["V0.4", "02043_18_05474_2_1_Check_TSD_Criticity_Configuration_file.xlsx", "", path_Cesare.split("/")[-1],
                      path_effect.split("/")[-1], check_level,
                      "", now.strftime("%Y-%m-%d"), now.strftime("%H:%M"), "",
                      TSD_file_checked.split("/")[-1], TSD_function_file_checked.split("/")[-1], TSD_system_file_checked.split("/")[-1], "", amdec, MedialecMatrice.split("/")[-1],
                      "", status]
        for i, name in enumerate(reportCol2):
            reportWorkSheet1.Cells(i + 1, 2).Value = name

        for sheet in reportWorkBook.Worksheets:
            if sheet.Name == "Test Report":
                reportWorkSheet2 = sheet
        if not reportWorkSheet2:
            workSheetsNumber = reportWorkBook.Sheets.Count
            sheetAfter = reportWorkBook.Sheets(workSheetsNumber)
            reportWorkSheet2 = reportWorkBook.Worksheets.Add(None, sheetAfter)
            reportWorkSheet2.Name = "Test Report"

        for i, name in enumerate(testReportRow1StringList):
            reportWorkSheet2.Cells(1, i + 1).Value = name
        reportWorkSheet2.Columns.AutoFit()
        reportWorkSheet2.Columns.Font.Bold = True

        row = 2

        testResult = self.Test_02043_18_04939_STRUCT_0000_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0000 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0000", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0000: The sheet “Informations Générales” (or “General information”) is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0000"][self.checkLevel], "02043_18_04939_STRUCT_0000", "The sheet “Informations Générales” (or “General information”) is not present or not written correctly.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0005_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0005 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0005", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0005: The field “REFERENCE” of the sheet “Informations Générales” (or “General information”)  in the line 52 shall be indicated."
            self.tab1.textbox.setText(text)
            flag = False
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0005"][self.checkLevel], "02043_18_04939_STRUCT_0005", "The field “REFERENCE” of the sheet “Informations Générales” (or “General information”)  in the line 52 shall be indicated.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)


        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0010_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0010 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0010", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0010: The information “Ref plan type” is missing in the sheet “Informations Générales” (or “General information”)."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0010"][self.checkLevel], "02043_18_04939_STRUCT_0010", "The information “Ref plan type” is missing in the sheet “Informations Générales” (or “General information”).", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0011_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0011 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0011", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0011: The document does not specify the template or the template reference is not indicated in the sheet “Informations Générales” (or “General information”). \nAs indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0011"][self.checkLevel], "02043_18_04939_STRUCT_0011", "The document does not specify the template or the template reference is not indicated in the sheet “Informations Générales” (or “General information”). \nAs indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0020_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0020 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0020", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0020: The sheet “Suppression ” (or “suppression ”) is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0020"][self.checkLevel], "02043_18_04939_STRUCT_0020", "The sheet “Suppression ” (or “suppression ”) is not present or not written correctly.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0025_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0025 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0025", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0025: The document does not follow the template, the column “Onglet” (or “sheet”) of the sheet “Suppression” (or “suppression”) is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0025"][self.checkLevel], "02043_18_04939_STRUCT_0025", "The document does not follow the template, the column “Onglet” (or “sheet”) of the sheet “Suppression” (or “suppression”) is not present or not written correctly.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0030_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0030 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0030", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0030: The document does not follow the template, the column “Version du TSD” (or “Version of the document”) of the sheet “Suppression” (or “suppression”) is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0030"][self.checkLevel], "02043_18_04939_STRUCT_0030", "The document does not follow the template, the column “Version du TSD” (or “Version of the document”) of the sheet “Suppression” (or “suppression”) is not present or not written correctly.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0035_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0035 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0035", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0035: The document does not follow the template, the column “Version du TSD” (or “Version of the document”) of the sheet “Suppression” (or “suppression”) is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0035"][self.checkLevel], "02043_18_04939_STRUCT_0030", "The document does not follow the template, the column “Version du TSD” (or “Version of the document”) of the sheet “Suppression” (or “suppression”) is not present or not written correctly.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0040_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0040 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0040", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0040: The document does not follow the template, the column “Justification de la modification” (or “Change reason”) of the sheet “Suppression” (or “suppression”) is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = False
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0040"][self.checkLevel], "02043_18_04939_STRUCT_0040", "The document does not follow the template, the column “Justification de la modification” (or “Change reason”) of the sheet “Suppression” (or “suppression”) is not present or not written correctly.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0051_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0051 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0051", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0051: The “Vehicle Architecture Schematic” document is not referenced. \nAs indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0051"][self.checkLevel], "02043_18_04939_STRUCT_0051", "The “Vehicle Architecture Schematic” document is not referenced. \nAs indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0052_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0052 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0052", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0052: The “Diagnostic Matrix” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0052"][self.checkLevel], "02043_18_04939_STRUCT_0052", "The “Diagnostic Matrix” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0053_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0053 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0053", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0053: The “Fault Tree” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0053"][self.checkLevel], "02043_18_04939_STRUCT_0053", "The “Fault Tree” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0054_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0054 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0054", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0054: The “ECU schematic” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0054"][self.checkLevel], "02043_18_04939_STRUCT_0054", "The “ECU schematic” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0055_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0055 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0055", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0055: The “STD” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0055"][self.checkLevel], "02043_18_04939_STRUCT_0055", "The “STD” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0056_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0056 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0056", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0056: The “Complexity Matrix (Decli EE)” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0056"][self.checkLevel], "02043_18_04939_STRUCT_0056", "The “Complexity Matrix (Decli EE)” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0057_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0057 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0057", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0057: The “Décli” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0057"][self.checkLevel], "02043_18_04939_STRUCT_0057", "The “Décli” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0058_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0058 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0058", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0058: The “DCEE” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0058"][self.checkLevel], "02043_18_04939_STRUCT_0058", "The “DCEE” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0059_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0059 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0059", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0059: The “EEAD” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0059"][self.checkLevel], "02043_18_04939_STRUCT_0059", "The “EEAD” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0060_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0060 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0060", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0060: The “TFD” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666"
            self.tab1.textbox.setText(text)
            flag = False
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0060"][self.checkLevel], "02043_18_04939_STRUCT_0060", "The “TFD” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        try:
            reportWorkBook.Save()
        except Exception as e:
            print(e)
        return flag

    def TestGeneralStructureXLS_DOC4(self, workBook, fileName, path_Cesare, path_effect):
        flag = 1
        testResult_0005 = self.Test_02043_18_04939_STRUCT_0005_XLS(fileName)
        fileName = self.tab1.myTextBox2.toPlainText()
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        reportWorkBook = excel.Workbooks.Open(fileName)
        reportWorkSheet1 = None
        reportWorkSheet2 = None
        for sheet in reportWorkBook.Worksheets:
            if sheet.Name == "Report Information":
                reportWorkSheet1 = sheet
        if not reportWorkSheet1:
            workSheetsNumber = reportWorkBook.Sheets.Count
            sheetAfter = reportWorkBook.Sheets(workSheetsNumber)
            reportWorkSheet1 = reportWorkBook.Worksheets.Add(None, sheetAfter)
            reportWorkSheet1.Name = "Report Information"

        reportInformationCol1StringList = ["Tool version:", "Criticity configuration file:", "", "Extract CESARE file:",
                                           "Customer effects file:", "check level:", "", "Date of the test:",
                                           "Time of the test:",
                                           "", "TSD file checked:", "TSD function file checked:",
                                           "TSD system file checked:",
                                           "", "AMDEC:", "export MedialecMatrice:", "", "Status:"]
        testReportRow1StringList = ["Criticity", "Requirements", "Message", "Localisation"]

        for i, name in enumerate(reportInformationCol1StringList):
            reportWorkSheet1.Cells(i + 1, 1).Value = name
        reportWorkSheet1.Columns.AutoFit()
        reportWorkSheet1.Columns.Font.Bold = True

        if self.tab1.myTextBox1.toPlainText():
            TSD_file_checked = self.tab1.myTextBox1.toPlainText()
        else:
            TSD_file_checked = "None"

        if self.tab1.myTextBox2.toPlainText():
            TSD_function_file_checked = self.tab1.myTextBox2.toPlainText()
        else:
            TSD_function_file_checked = "None"

        if self.tab1.myTextBox3.toPlainText():
            TSD_system_file_checked = self.tab1.myTextBox3.toPlainText()
        else:
            TSD_system_file_checked = "None"

        if self.tab1.myTextBox7.toPlainText():
            amdec = self.tab1.myTextBox3.toPlainText()
        else:
            amdec = "None"

        if self.tab1.myTextBox8.toPlainText():
            MedialecMatrice = self.tab1.myTextBox8.toPlainText()
        else:
            MedialecMatrice = "None"

        # if self.tab2.myTextBox5.toPlainText():
        #     cesare = self.tab2.myTextBox5.toPlainText()
        # else:
        #     cesare = self.download_file(self.CesareLink)
        #
        # if self.tab2.myTextBox6.toPlainText():
        #     CustomerEffect = self.tab2.myTextBox6.toPlainText()
        # else:
        #     CustomerEffect = self.download_file(self.CustomerEffectLink)

        if self.TestTsdFile is 1:
            status = "Pass"
        else:
            status = "Fail"

        check_level = str(self.tab1.combo.currentText())

        now = datetime.datetime.now()

        reportCol2 = ["V0.4", "02043_18_05474_2_1_Check_TSD_Criticity_Configuration_file.xlsx", "", path_Cesare.split("/")[-1],
                      path_effect.split("/")[-1], check_level,
                      "", now.strftime("%Y-%m-%d"), now.strftime("%H:%M"), "",
                      TSD_file_checked.split("/")[-1], TSD_function_file_checked.split("/")[-1], TSD_system_file_checked.split("/")[-1], "", amdec.split("/")[-1], MedialecMatrice.split("/")[-1],
                      "", status]
        for i, name in enumerate(reportCol2):
            reportWorkSheet1.Cells(i + 1, 2).Value = name

        for sheet in reportWorkBook.Worksheets:
            if sheet.Name == "Test Report":
                reportWorkSheet2 = sheet
        if not reportWorkSheet2:
            workSheetsNumber = reportWorkBook.Sheets.Count
            sheetAfter = reportWorkBook.Sheets(workSheetsNumber)
            reportWorkSheet2 = reportWorkBook.Worksheets.Add(None, sheetAfter)
            reportWorkSheet2.Name = "Test Report"

        for i, name in enumerate(testReportRow1StringList):
            reportWorkSheet2.Cells(1, i + 1).Value = name
        reportWorkSheet2.Columns.AutoFit()
        reportWorkSheet2.Columns.Font.Bold = True

        row = 2

        testResult = self.Test_02043_18_04939_STRUCT_0000_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0000 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0000", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0000: The sheet “Informations Générales” (or “General information”) is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0000"][self.checkLevel], "02043_18_04939_STRUCT_0000", "The sheet “Informations Générales” (or “General information”) is not present or not written correctly.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        if testResult_0005 == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0005 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0005", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0010: The field “REFERENCE” of the sheet “Informations Générales” (or “General information”)  in the line 52 shall be indicated."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0005"][self.checkLevel], "02043_18_04939_STRUCT_0005", "The field “REFERENCE” of the sheet “Informations Générales” (or “General information”)  in the line 52 shall be indicated.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0010_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0010 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0010", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0010: The information “Ref plan type” is missing in the sheet “Informations Générales” (or “General information”)."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0010"][self.checkLevel], "02043_18_04939_STRUCT_0010", "The information “Ref plan type” is missing in the sheet “Informations Générales” (or “General information”).", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0011_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0011 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0011", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0011: The document does not specify the template or the template reference is not indicated in the sheet “Informations Générales” (or “General information”). \nAs indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0011"][self.checkLevel], "02043_18_04939_STRUCT_0011", "The document does not specify the template or the template reference is not indicated in the sheet “Informations Générales” (or “General information”). \nAs indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0020_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0020 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0020", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0020: The sheet “Suppression ” (or “suppression ”) is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0020"][self.checkLevel], "02043_18_04939_STRUCT_0020", "The sheet “Suppression ” (or “suppression ”) is not present or not written correctly.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0025_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0025 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0025", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0025: The document does not follow the template, the column “Onglet” (or “sheet”) of the sheet “Suppression” (or “suppression”) is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0025"][self.checkLevel], "02043_18_04939_STRUCT_0025", "The document does not follow the template, the column “Onglet” (or “sheet”) of the sheet “Suppression” (or “suppression”) is not present or not written correctly.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0030_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0030 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0030", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0030: The document does not follow the template, the column “Version du TSD” (or “Version of the document”) of the sheet “Suppression” (or “suppression”) is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0030"][self.checkLevel], "02043_18_04939_STRUCT_0030", "The document does not follow the template, the column “Version du TSD” (or “Version of the document”) of the sheet “Suppression” (or “suppression”) is not present or not written correctly.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0035_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0035 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0035", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0035: The document does not follow the template, the column “Version du TSD” (or “Version of the document”) of the sheet “Suppression” (or “suppression”) is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0035"][self.checkLevel], "02043_18_04939_STRUCT_0030", "The document does not follow the template, the column “Version du TSD” (or “Version of the document”) of the sheet “Suppression” (or “suppression”) is not present or not written correctly.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0040_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0040 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0040", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0040: The document does not follow the template, the column “Justification de la modification” (or “Change reason”) of the sheet “Suppression” (or “suppression”) is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = False
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0040"][self.checkLevel], "02043_18_04939_STRUCT_0040", "The document does not follow the template, the column “Justification de la modification” (or “Change reason”) of the sheet “Suppression” (or “suppression”) is not present or not written correctly.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0051_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0051 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0051", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0051: The “Vehicle Architecture Schematic” document is not referenced. \nAs indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0051"][self.checkLevel], "02043_18_04939_STRUCT_0051", "The “Vehicle Architecture Schematic” document is not referenced. \nAs indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0052_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0052 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0052", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0052: The “Diagnostic Matrix” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0052"][self.checkLevel], "02043_18_04939_STRUCT_0052", "The “Diagnostic Matrix” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0053_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0053 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0053", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0053: The “Fault Tree” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0053"][self.checkLevel], "02043_18_04939_STRUCT_0053", "The “Fault Tree” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0054_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0054 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0054", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0054: The “ECU schematic” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0054"][self.checkLevel], "02043_18_04939_STRUCT_0054", "The “ECU schematic” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0055_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0055 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0055", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0055: The “STD” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0055"][self.checkLevel], "02043_18_04939_STRUCT_0055", "The “ECU schematic” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0056_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0056 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0056", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0056: The “Complexity Matrix (Decli EE)” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0056"][self.checkLevel], "02043_18_04939_STRUCT_0056", "The “Complexity Matrix (Decli EE)” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0057_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0057 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0057", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0057: The “Décli” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0057"][self.checkLevel], "02043_18_04939_STRUCT_0057", "The “Décli” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0058_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0058 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0058", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0058: The “DCEE” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0058"][self.checkLevel], "02043_18_04939_STRUCT_0058", "The “DCEE” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0059_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0059 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0059", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0059: The “EEAD” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0059"][self.checkLevel], "02043_18_04939_STRUCT_0059", "The “EEAD” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0060_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0060 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0060", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0060: The “TFD” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666"
            self.tab1.textbox.setText(text)
            flag = False
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0060"][self.checkLevel], "02043_18_04939_STRUCT_0060", "The “TFD” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        try:
            reportWorkBook.Save()
        except Exception as e:
            print(e)
        return flag

    def TestGeneralStructureXLSX_XLSM_DOC4(self, workBook,  path_Cesare, path_effect):
        flag = 1
        if self.tab1.myTextBox2.toPlainText():
           fileName = self.tab1.myTextBox2.toPlainText()
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        reportWorkBook = excel.Workbooks.Open(fileName)
        reportWorkSheet1 = None
        reportWorkSheet2 = None
        for sheet in reportWorkBook.Worksheets:
            if sheet.Name == "Report Information":
                reportWorkSheet1 = sheet
        if not reportWorkSheet1:
            workSheetsNumber = reportWorkBook.Sheets.Count
            sheetAfter = reportWorkBook.Sheets(workSheetsNumber)
            reportWorkSheet1 = reportWorkBook.Worksheets.Add(None, sheetAfter)
            reportWorkSheet1.Name = "Report Information"

        reportInformationCol1StringList = ["Tool version:", "Criticity configuration file:", "", "Extract CESARE file:",
                                           "Customer effects file:", "check level:", "", "Date of the test:",
                                           "Time of the test:",
                                           "", "TSD file checked:", "TSD function file checked:",
                                           "TSD system file checked:",
                                           "", "AMDEC:", "export MedialecMatrice:", "", "Status:"]
        testReportRow1StringList = ["Criticity", "Requirements", "Message", "Localisation"]

        for i, name in enumerate(reportInformationCol1StringList):
            reportWorkSheet1.Cells(i + 1, 1).Value = name
        reportWorkSheet1.Columns.AutoFit()
        reportWorkSheet1.Columns.Font.Bold = True

        if self.tab1.myTextBox1.toPlainText():
            TSD_file_checked = self.tab1.myTextBox1.toPlainText()
        else:
            TSD_file_checked = "None"

        if self.tab1.myTextBox2.toPlainText():
            TSD_function_file_checked = self.tab1.myTextBox2.toPlainText()
        else:
            TSD_function_file_checked = "None"

        if self.tab1.myTextBox3.toPlainText():
            TSD_system_file_checked = self.tab1.myTextBox3.toPlainText()
        else:
            TSD_system_file_checked = "None"

        if self.tab1.myTextBox7.toPlainText():
            amdec = self.tab1.myTextBox3.toPlainText()
        else:
            amdec = "None"

        if self.tab1.myTextBox8.toPlainText():
            MedialecMatrice = self.tab1.myTextBox8.toPlainText()
        else:
            MedialecMatrice = "None"

        # if self.tab2.myTextBox5.toPlainText():
        #     cesare = self.tab2.myTextBox5.toPlainText()
        # else:
        #     cesare = self.download_file(self.CesareLink)
        #
        # if self.tab2.myTextBox6.toPlainText():
        #     CustomerEffect = self.tab2.myTextBox6.toPlainText()
        # else:
        #     CustomerEffect = self.download_file(self.CustomerEffectLink)

        if self.TestTsdFile is 1:
            status = "Pass"
        else:
            status = "Fail"

        check_level = str(self.tab1.combo.currentText())

        now = datetime.datetime.now()

        reportCol2 = ["V0.4", "02043_18_05474_2_1_Check_TSD_Criticity_Configuration_file.xlsx", "", path_Cesare.split("/")[-1],
                      path_effect.split("/")[-1], check_level,
                      "", now.strftime("%Y-%m-%d"), now.strftime("%H:%M"), "",
                      TSD_file_checked.split("/")[-1], TSD_function_file_checked.split("/")[-1], TSD_system_file_checked.split("/")[-1], "", amdec.split("/")[-1], MedialecMatrice.split("/")[-1],
                      "", status]
        for i, name in enumerate(reportCol2):
            reportWorkSheet1.Cells(i + 1, 2).Value = name

        for sheet in reportWorkBook.Worksheets:
            if sheet.Name == "Test Report":
                reportWorkSheet2 = sheet
        if not reportWorkSheet2:
            workSheetsNumber = reportWorkBook.Sheets.Count
            sheetAfter = reportWorkBook.Sheets(workSheetsNumber)
            reportWorkSheet2 = reportWorkBook.Worksheets.Add(None, sheetAfter)
            reportWorkSheet2.Name = "Test Report"

        for i, name in enumerate(testReportRow1StringList):
            reportWorkSheet2.Cells(1, i + 1).Value = name
        reportWorkSheet2.Columns.AutoFit()
        reportWorkSheet2.Columns.Font.Bold = True

        row = 2

        testResult = self.Test_02043_18_04939_STRUCT_0000_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0000 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0000", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0000: The sheet “Informations Générales” (or “General information”) is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0000"][self.checkLevel], "02043_18_04939_STRUCT_0000", "The sheet “Informations Générales” (or “General information”) is not present or not written correctly.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0005_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0005 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0005", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0005: The field “REFERENCE” of the sheet “Informations Générales” (or “General information”)  in the line 52 shall be indicated."
            self.tab1.textbox.setText(text)
            flag = False
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0005"][self.checkLevel], "02043_18_04939_STRUCT_0005", "The field “REFERENCE” of the sheet “Informations Générales” (or “General information”)  in the line 52 shall be indicated.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0010_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0010 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0010", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0010: The information “Ref plan type” is missing in the sheet “Informations Générales” (or “General information”)."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0010"][self.checkLevel], "02043_18_04939_STRUCT_0010", "The information “Ref plan type” is missing in the sheet “Informations Générales” (or “General information”).", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0011_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0011 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0011", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0011: The document does not specify the template or the template reference is not indicated in the sheet “Informations Générales” (or “General information”). \nAs indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0011"][self.checkLevel], "02043_18_04939_STRUCT_0011", "The document does not specify the template or the template reference is not indicated in the sheet “Informations Générales” (or “General information”). \nAs indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0020_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0020 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0020", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0020: The sheet “Suppression ” (or “suppression ”) is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0020"][self.checkLevel], "02043_18_04939_STRUCT_0020", "The sheet “Suppression ” (or “suppression ”) is not present or not written correctly.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0025_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0025 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0025", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0025: The document does not follow the template, the column “Onglet” (or “sheet”) of the sheet “Suppression” (or “suppression”) is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0025"][self.checkLevel], "02043_18_04939_STRUCT_0025", "The document does not follow the template, the column “Onglet” (or “sheet”) of the sheet “Suppression” (or “suppression”) is not present or not written correctly.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0030_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0030 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0030", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0030: The document does not follow the template, the column “Version du TSD” (or “Version of the document”) of the sheet “Suppression” (or “suppression”) is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0030"][self.checkLevel], "02043_18_04939_STRUCT_0030", "The document does not follow the template, the column “Version du TSD” (or “Version of the document”) of the sheet “Suppression” (or “suppression”) is not present or not written correctly.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0035_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0035 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0035", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0035: The document does not follow the template, the column “Version du TSD” (or “Version of the document”) of the sheet “Suppression” (or “suppression”) is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0035"][self.checkLevel], "02043_18_04939_STRUCT_0030", "The document does not follow the template, the column “Version du TSD” (or “Version of the document”) of the sheet “Suppression” (or “suppression”) is not present or not written correctly.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0040_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0040 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0040", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0040: The document does not follow the template, the column “Justification de la modification” (or “Change reason”) of the sheet “Suppression” (or “suppression”) is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = False
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0040"][self.checkLevel], "02043_18_04939_STRUCT_0040", "The document does not follow the template, the column “Justification de la modification” (or “Change reason”) of the sheet “Suppression” (or “suppression”) is not present or not written correctly.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0051_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0051 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0051", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0051: The “Vehicle Architecture Schematic” document is not referenced. \nAs indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0051"][self.checkLevel], "02043_18_04939_STRUCT_0051", "The “Vehicle Architecture Schematic” document is not referenced. \nAs indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0052_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0052 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0052", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0052: The “Diagnostic Matrix” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0052"][self.checkLevel], "02043_18_04939_STRUCT_0052", "The “Diagnostic Matrix” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0053_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0053 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0053", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0053: The “Fault Tree” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0053"][self.checkLevel], "02043_18_04939_STRUCT_0053", "The “Fault Tree” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0054_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0054 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0054", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0054: The “ECU schematic” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0054"][self.checkLevel], "02043_18_04939_STRUCT_0054", "The “ECU schematic” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0055_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0055 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0055", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0055: The “STD” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0055"][self.checkLevel], "02043_18_04939_STRUCT_0055", "The “STD” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0056_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0056 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0056", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0056: The “Complexity Matrix (Decli EE)” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0056"][self.checkLevel], "02043_18_04939_STRUCT_0056", "The “Complexity Matrix (Decli EE)” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0057_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0057 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0057", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0057: The “Décli” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0057"][self.checkLevel], "02043_18_04939_STRUCT_0057", "The “Décli” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0058_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0058 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0058", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0058: The “DCEE” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0058"][self.checkLevel], "02043_18_04939_STRUCT_0058", "The “DCEE” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666.",""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0059_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0059 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0059", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0059: The “EEAD” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0059"][self.checkLevel], "02043_18_04939_STRUCT_0059", "The “EEAD” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0060_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0060 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0060", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0060: The “TFD” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666"
            self.tab1.textbox.setText(text)
            flag = False
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0060"][self.checkLevel], "02043_18_04939_STRUCT_0060", "The “TFD” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        try:
            reportWorkBook.Save()
        except Exception as e:
            print(e)
        return flag

    def TestGeneralStructureXLS_DOC5(self, workBook, fileName,  path_Cesare, path_effect):
        flag = 1
        testResult_0005 = self.Test_02043_18_04939_STRUCT_0005_XLS(fileName)
        fileName = self.tab1.myTextBox3.toPlainText()
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        reportWorkBook = excel.Workbooks.Open(fileName)
        reportWorkSheet1 = None
        reportWorkSheet2 = None
        for sheet in reportWorkBook.Worksheets:
            if sheet.Name == "Report Information":
                reportWorkSheet1 = sheet
        if not reportWorkSheet1:
            workSheetsNumber = reportWorkBook.Sheets.Count
            sheetAfter = reportWorkBook.Sheets(workSheetsNumber)
            reportWorkSheet1 = reportWorkBook.Worksheets.Add(None, sheetAfter)
            reportWorkSheet1.Name = "Report Information"

        reportInformationCol1StringList = ["Tool version:", "Criticity configuration file:", "", "Extract CESARE file:",
                                           "Customer effects file:", "check level:", "", "Date of the test:",
                                           "Time of the test:",
                                           "", "TSD file checked:", "TSD function file checked:",
                                           "TSD system file checked:",
                                           "", "AMDEC:", "export MedialecMatrice:", "", "Status:"]
        testReportRow1StringList = ["Criticity", "Requirements", "Message", "Localisation"]

        for i, name in enumerate(reportInformationCol1StringList):
            reportWorkSheet1.Cells(i + 1, 1).Value = name
        reportWorkSheet1.Columns.AutoFit()
        reportWorkSheet1.Columns.Font.Bold = True

        if self.tab1.myTextBox1.toPlainText():
            TSD_file_checked = self.tab1.myTextBox1.toPlainText()
        else:
            TSD_file_checked = "None"

        if self.tab1.myTextBox2.toPlainText():
            TSD_function_file_checked = self.tab1.myTextBox2.toPlainText()
        else:
            TSD_function_file_checked = "None"

        if self.tab1.myTextBox3.toPlainText():
            TSD_system_file_checked = self.tab1.myTextBox3.toPlainText()
        else:
            TSD_system_file_checked = "None"

        if self.tab1.myTextBox7.toPlainText():
            amdec = self.tab1.myTextBox3.toPlainText()
        else:
            amdec = "None"

        if self.tab1.myTextBox8.toPlainText():
            MedialecMatrice = self.tab1.myTextBox8.toPlainText()
        else:
            MedialecMatrice = "None"

        # if self.tab2.myTextBox5.toPlainText():
        #     cesare = self.tab2.myTextBox5.toPlainText()
        # else:
        #     cesare = self.download_file(self.CesareLink)
        #
        # if self.tab2.myTextBox6.toPlainText():
        #     CustomerEffect = self.tab2.myTextBox6.toPlainText()
        # else:
        #     CustomerEffect = self.download_file(self.CustomerEffectLink)

        if self.TestTsdFile is 1:
            status = "Pass"
        else:
            status = "Fail"

        check_level = str(self.tab1.combo.currentText())

        now = datetime.datetime.now()

        reportCol2 = ["V0.4", "02043_18_05474_2_1_Check_TSD_Criticity_Configuration_file.xlsx", "", path_Cesare.split("/")[-1],
                      path_effect.split("/")[-1], check_level,
                      "", now.strftime("%Y-%m-%d"), now.strftime("%H:%M"), "",
                      TSD_file_checked.split("/")[-1], TSD_function_file_checked.split("/")[-1], TSD_system_file_checked.split("/")[-1], "", amdec.split("/")[-1], MedialecMatrice.split("/")[-1],
                      "", status]
        for i, name in enumerate(reportCol2):
            reportWorkSheet1.Cells(i + 1, 2).Value = name

        for sheet in reportWorkBook.Worksheets:
            if sheet.Name == "Test Report":
                reportWorkSheet2 = sheet
        if not reportWorkSheet2:
            workSheetsNumber = reportWorkBook.Sheets.Count
            sheetAfter = reportWorkBook.Sheets(workSheetsNumber)
            reportWorkSheet2 = reportWorkBook.Worksheets.Add(None, sheetAfter)
            reportWorkSheet2.Name = "Test Report"

        for i, name in enumerate(testReportRow1StringList):
            reportWorkSheet2.Cells(1, i + 1).Value = name
        reportWorkSheet2.Columns.AutoFit()
        reportWorkSheet2.Columns.Font.Bold = True

        row = 2

        testResult = self.Test_02043_18_04939_STRUCT_0000_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0000 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0000", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0000: The sheet “Informations Générales” (or “General information”) is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0000"][self.checkLevel], "02043_18_04939_STRUCT_0000", "The sheet “Informations Générales” (or “General information”) is not present or not written correctly.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        if testResult_0005 == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0005 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0005", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0010: The field “REFERENCE” of the sheet “Informations Générales” (or “General information”)  in the line 52 shall be indicated."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0005"][self.checkLevel], "02043_18_04939_STRUCT_0005", "The field “REFERENCE” of the sheet “Informations Générales” (or “General information”)  in the line 52 shall be indicated.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0010_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0010 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0010", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0010: The information “Ref plan type” is missing in the sheet “Informations Générales” (or “General information”)."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0010"][self.checkLevel], "02043_18_04939_STRUCT_0010", "The information “Ref plan type” is missing in the sheet “Informations Générales” (or “General information”).", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0011_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0011 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0011", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0011: The document does not specify the template or the template reference is not indicated in the sheet “Informations Générales” (or “General information”). \nAs indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0011"][self.checkLevel], "02043_18_04939_STRUCT_0011", "The document does not specify the template or the template reference is not indicated in the sheet “Informations Générales” (or “General information”). \nAs indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0020_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0020 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0020", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0020: The sheet “Suppression ” (or “suppression ”) is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0020"][self.checkLevel], "02043_18_04939_STRUCT_0020", "The sheet “Suppression ” (or “suppression ”) is not present or not written correctly.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0025_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0025 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0025", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0025: The document does not follow the template, the column “Onglet” (or “sheet”) of the sheet “Suppression” (or “suppression”) is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0025"][self.checkLevel], "02043_18_04939_STRUCT_0025", "The document does not follow the template, the column “Onglet” (or “sheet”) of the sheet “Suppression” (or “suppression”) is not present or not written correctly.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0030_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0030 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0030", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0030: The document does not follow the template, the column “Version du TSD” (or “Version of the document”) of the sheet “Suppression” (or “suppression”) is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0030"][self.checkLevel], "02043_18_04939_STRUCT_0030", "The document does not follow the template, the column “Version du TSD” (or “Version of the document”) of the sheet “Suppression” (or “suppression”) is not present or not written correctly.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0035_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0035 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0035", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0035: The document does not follow the template, the column “Version du TSD” (or “Version of the document”) of the sheet “Suppression” (or “suppression”) is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0035"][self.checkLevel], "02043_18_04939_STRUCT_0030", "The document does not follow the template, the column “Version du TSD” (or “Version of the document”) of the sheet “Suppression” (or “suppression”) is not present or not written correctly.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0040_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0040 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0040", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0040: The document does not follow the template, the column “Justification de la modification” (or “Change reason”) of the sheet “Suppression” (or “suppression”) is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = False
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0040"][self.checkLevel], "02043_18_04939_STRUCT_0040", "The document does not follow the template, the column “Justification de la modification” (or “Change reason”) of the sheet “Suppression” (or “suppression”) is not present or not written correctly.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0051_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0051 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0051", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0051: The “Vehicle Architecture Schematic” document is not referenced. \nAs indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0051"][self.checkLevel], "02043_18_04939_STRUCT_0051", "The “Vehicle Architecture Schematic” document is not referenced. \nAs indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0052_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0052 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0052", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0052: The “Diagnostic Matrix” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0052"][self.checkLevel], "02043_18_04939_STRUCT_0052", "The “Diagnostic Matrix” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0053_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0053 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0053", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0053: The “Fault Tree” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0053"][self.checkLevel], "02043_18_04939_STRUCT_0053", "The “Fault Tree” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0054_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0054 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0054", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0054: The “ECU schematic” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0054"][self.checkLevel], "02043_18_04939_STRUCT_0054", "The “ECU schematic” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0055_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0055 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0055", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0055: The “STD” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0055"][self.checkLevel], "02043_18_04939_STRUCT_0055", "The “STD” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0056_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0056 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0056", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0056: The “Complexity Matrix (Decli EE)” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0056"][self.checkLevel], "02043_18_04939_STRUCT_0056", "The “Complexity Matrix (Decli EE)” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0057_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0057 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0057", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0057: The “Décli” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0057"][self.checkLevel], "02043_18_04939_STRUCT_0057", "The “Décli” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0058_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0058 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0058", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0058: The “DCEE” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0058"][self.checkLevel], "02043_18_04939_STRUCT_0058", "The “DCEE” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0059_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0059 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0059", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0059: The “EEAD” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0059"][self.checkLevel], "02043_18_04939_STRUCT_0059", "The “EEAD” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0060_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0060 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0060", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0060: The “TFD” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666"
            self.tab1.textbox.setText(text)
            flag = False
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0060"][self.checkLevel], "02043_18_04939_STRUCT_0060","The “TFD” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        try:
            reportWorkBook.Save()
        except Exception as e:
            print(e)
        return flag

    def TestGeneralStructureXLSX_XLSM_DOC5(self, workBook, path_Cesare, path_effect):
        flag = 1
        if self.tab1.myTextBox3.toPlainText():
           fileName = self.tab1.myTextBox3.toPlainText()
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        reportWorkBook = excel.Workbooks.Open(fileName)
        reportWorkSheet1 = None
        reportWorkSheet2 = None
        for sheet in reportWorkBook.Worksheets:
            if sheet.Name == "Report Information":
                reportWorkSheet1 = sheet
        if not reportWorkSheet1:
            workSheetsNumber = reportWorkBook.Sheets.Count
            sheetAfter = reportWorkBook.Sheets(workSheetsNumber)
            reportWorkSheet1 = reportWorkBook.Worksheets.Add(None, sheetAfter)
            reportWorkSheet1.Name = "Report Information"

        reportInformationCol1StringList = ["Tool version:", "Criticity configuration file:", "", "Extract CESARE file:",
                                           "Customer effects file:", "check level:", "", "Date of the test:",
                                           "Time of the test:",
                                           "", "TSD file checked:", "TSD function file checked:",
                                           "TSD system file checked:",
                                           "", "AMDEC:", "export MedialecMatrice:", "", "Status:"]
        testReportRow1StringList = ["Criticity", "Requirements", "Message", "Localisation"]

        for i, name in enumerate(reportInformationCol1StringList):
            reportWorkSheet1.Cells(i + 1, 1).Value = name
        reportWorkSheet1.Columns.AutoFit()
        reportWorkSheet1.Columns.Font.Bold = True

        if self.tab1.myTextBox1.toPlainText():
            TSD_file_checked = self.tab1.myTextBox1.toPlainText()
        else:
            TSD_file_checked = "None"

        if self.tab1.myTextBox2.toPlainText():
            TSD_function_file_checked = self.tab1.myTextBox2.toPlainText()
        else:
            TSD_function_file_checked = "None"

        if self.tab1.myTextBox3.toPlainText():
            TSD_system_file_checked = self.tab1.myTextBox3.toPlainText()
        else:
            TSD_system_file_checked = "None"

        if self.tab1.myTextBox7.toPlainText():
            amdec = self.tab1.myTextBox3.toPlainText()
        else:
            amdec = "None"

        if self.tab1.myTextBox8.toPlainText():
            MedialecMatrice = self.tab1.myTextBox8.toPlainText()
        else:
            MedialecMatrice = "None"

        # if self.tab2.myTextBox5.toPlainText():
        #     cesare = self.tab2.myTextBox5.toPlainText()
        # else:
        #     cesare = self.download_file(self.CesareLink)
        #
        # if self.tab2.myTextBox6.toPlainText():
        #     CustomerEffect = self.tab2.myTextBox6.toPlainText()
        # else:
        #     CustomerEffect = self.download_file(self.CustomerEffectLink)

        if self.TestTsdFile is 1:
            status = "Pass"
        else:
            status = "Fail"

        check_level = str(self.tab1.combo.currentText())

        now = datetime.datetime.now()

        reportCol2 = ["V0.4", "02043_18_05474_2_1_Check_TSD_Criticity_Configuration_file.xlsx", "", path_Cesare.split("/")[-1],
                      path_effect.split("/")[-1], check_level,
                      "", now.strftime("%Y-%m-%d"), now.strftime("%H:%M"), "",
                      TSD_file_checked.split("/")[-1], TSD_function_file_checked.split("/")[-1], TSD_system_file_checked.split("/")[-1], "", amdec, MedialecMatrice.split("/")[-1],
                      "", status]
        for i, name in enumerate(reportCol2):
            reportWorkSheet1.Cells(i + 1, 2).Value = name

        for sheet in reportWorkBook.Worksheets:
            if sheet.Name == "Test Report":
                reportWorkSheet2 = sheet
        if not reportWorkSheet2:
            workSheetsNumber = reportWorkBook.Sheets.Count
            sheetAfter = reportWorkBook.Sheets(workSheetsNumber)
            reportWorkSheet2 = reportWorkBook.Worksheets.Add(None, sheetAfter)
            reportWorkSheet2.Name = "Test Report"

        for i, name in enumerate(testReportRow1StringList):
            reportWorkSheet2.Cells(1, i + 1).Value = name
        reportWorkSheet2.Columns.AutoFit()
        reportWorkSheet2.Columns.Font.Bold = True

        row = 2

        testResult = self.Test_02043_18_04939_STRUCT_0000_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0000 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0000", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0000: The sheet “Informations Générales” (or “General information”) is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0000"][self.checkLevel], "02043_18_04939_STRUCT_0000", "The sheet “Informations Générales” (or “General information”) is not present or not written correctly.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0005_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0005 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0005", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0005: The field “REFERENCE” of the sheet “Informations Générales” (or “General information”)  in the line 52 shall be indicated."
            self.tab1.textbox.setText(text)
            flag = False
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0005"][self.checkLevel], "02043_18_04939_STRUCT_0005", "The field “REFERENCE” of the sheet “Informations Générales” (or “General information”)  in the line 52 shall be indicated.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0010_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0010 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0010", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0010: The information “Ref plan type” is missing in the sheet “Informations Générales” (or “General information”)."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0010"][self.checkLevel], "02043_18_04939_STRUCT_0010", "The information “Ref plan type” is missing in the sheet “Informations Générales” (or “General information”).", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0011_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0011 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0011", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0011: The document does not specify the template or the template reference is not indicated in the sheet “Informations Générales” (or “General information”). \nAs indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0011"][self.checkLevel], "02043_18_04939_STRUCT_0011", "The document does not specify the template or the template reference is not indicated in the sheet “Informations Générales” (or “General information”). \nAs indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0020_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0020 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0020", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0020: The sheet “Suppression ” (or “suppression ”) is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0020"][self.checkLevel], "02043_18_04939_STRUCT_0020", "The sheet “Suppression ” (or “suppression ”) is not present or not written correctly.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0025_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0025 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0025", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0025: The document does not follow the template, the column “Onglet” (or “sheet”) of the sheet “Suppression” (or “suppression”) is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0025"][self.checkLevel], "02043_18_04939_STRUCT_0025", "The document does not follow the template, the column “Onglet” (or “sheet”) of the sheet “Suppression” (or “suppression”) is not present or not written correctly.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0030_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0030 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0030", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0030: The document does not follow the template, the column “Version du TSD” (or “Version of the document”) of the sheet “Suppression” (or “suppression”) is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0030"][self.checkLevel], "02043_18_04939_STRUCT_0030", "The document does not follow the template, the column “Version du TSD” (or “Version of the document”) of the sheet “Suppression” (or “suppression”) is not present or not written correctly.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0035_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0035 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0035", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0035: The document does not follow the template, the column “Version du TSD” (or “Version of the document”) of the sheet “Suppression” (or “suppression”) is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0035"][self.checkLevel], "02043_18_04939_STRUCT_0030", "The document does not follow the template, the column “Version du TSD” (or “Version of the document”) of the sheet “Suppression” (or “suppression”) is not present or not written correctly.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0040_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0040 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0040", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0040: The document does not follow the template, the column “Justification de la modification” (or “Change reason”) of the sheet “Suppression” (or “suppression”) is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = False
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0040"][self.checkLevel], "02043_18_04939_STRUCT_0040", "The document does not follow the template, the column “Justification de la modification” (or “Change reason”) of the sheet “Suppression” (or “suppression”) is not present or not written correctly.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0051_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0051 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0051", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0051: The “Vehicle Architecture Schematic” document is not referenced. \nAs indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0051"][self.checkLevel], "02043_18_04939_STRUCT_0051", "The “Vehicle Architecture Schematic” document is not referenced. \nAs indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0052_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0052 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0052", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0052: The “Diagnostic Matrix” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0052"][self.checkLevel], "02043_18_04939_STRUCT_0052", "The “Diagnostic Matrix” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0053_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0053 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0053", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0053: The “Fault Tree” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0053"][self.checkLevel], "02043_18_04939_STRUCT_0053", "The “Fault Tree” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0054_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0054 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0054", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0054: The “ECU schematic” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0054"][self.checkLevel], "02043_18_04939_STRUCT_0054", "The “ECU schematic” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0055_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0055 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0055", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0055: The “STD” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0055"][self.checkLevel], "02043_18_04939_STRUCT_0055", "The “STD” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666",  ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0056_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0056 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0056", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0056: The “Complexity Matrix (Decli EE)” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0056"][self.checkLevel], "02043_18_04939_STRUCT_0056", "The “Complexity Matrix (Decli EE)” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0057_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0057 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0057", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0057: The “Décli” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0057"][self.checkLevel], "02043_18_04939_STRUCT_0057", "The “Décli” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0058_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0058 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0058", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0058: The “DCEE” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0058"][self.checkLevel], "02043_18_04939_STRUCT_0058", "The “DCEE” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0059_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0059 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0059", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0059: The “EEAD” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0059"][self.checkLevel], "02043_18_04939_STRUCT_0059", "The “EEAD” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        row += 1
        testResult = self.Test_02043_18_04939_STRUCT_0060_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0060 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_0060", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0060: The “TFD” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666"
            self.tab1.textbox.setText(text)
            flag = False
            for i, name in enumerate([self.testReqDict["02043_18_04939_STRUCT_0060"][self.checkLevel], "02043_18_04939_STRUCT_0060", "The “TFD” document is not referenced. As indicated in to one of the 3 references AEEV_IAEE07_0033 or 02043_12_01665 or 02043_12_01666", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        self.pbvalue = self.pbvalue + 0.8772
        self.tab1.pbar.setValue(self.pbvalue)

        try:
            reportWorkBook.Save()
        except Exception as e:
            print(e)
        return flag

    def TestGeneralStructure_DOC3_XLS(self, workBook):

        flag = 1
        fileName = self.tab1.myTextBox1.toPlainText()
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        reportWorkBook = excel.Workbooks.Open(fileName)
        reportWorkSheet2 = None

        for sheet in reportWorkBook.Worksheets:
            if sheet.Name == "Test Report":
                reportWorkSheet2 = sheet
        if not reportWorkSheet2:
            workSheetsNumber = reportWorkBook.Sheets.Count
            sheetAfter = reportWorkBook.Sheets(workSheetsNumber)
            reportWorkSheet2 = reportWorkBook.Worksheets.Add(None, sheetAfter)
            reportWorkSheet2.Name = "Test Report"

        testReportRow1StringList = ["Criticity", "Requirements", "Message", "Localisation"]

        for i, name in enumerate(testReportRow1StringList):
            reportWorkSheet2.Cells(1, i + 1).Value = name
        reportWorkSheet2.Columns.AutoFit()
        reportWorkSheet2.Columns.Font.Bold = True

        row = 21
        str1 = "02043_18_04939_STRUCT_0"
        stringInt = 10 - row
        str2 = str(stringInt + row)
        str3 = "0"
        String = str1 + str2 + str3

        testResult = self.Test_02043_18_04939_STRUCT_0100_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0100 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0100: The sheet “tableau” is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String, " The sheet “tableau” is not present or not written correctly.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0110_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0110 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0110: In the sheet “tableau”, the column XXXX (to be indicated) is not present or not written correctly.."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String, "In the sheet “tableau”, the column XXXX (to be indicated) is not present or not written correctly.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0120_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0120 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0120: The sheet “codes défauts” is not present or not written correctly"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String, "The sheet “codes défauts” is not present or not written correctly", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0130_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0130 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0130: In the sheet “codes défauts”, the column XXXX (to be indicated) is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String, "In the sheet “codes défauts”, the column XXXX (to be indicated) is not present or not written correctly.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0140_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0140 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0140: The sheet “mesures et commandes” is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String, "The sheet “mesures et commandes” is not present or not written correctly", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0150_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0140 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0150: In the sheet “mesures et commandes”, the column XXXX (to be indicated) is not present or not written correctly. "
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String, " In the sheet “mesures et commandes”, the column XXXX (to be indicated) is not present or not written correctly. ", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0160_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0160 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0160: The sheet “Diagnostic débarqués” is not present or not written correctly"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String, "The sheet “Diagnostic débarqués” is not present or not written correctly", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0170_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0170 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0170: In the sheet “Diagnostic débarqués”, the column XXXX (to be indicated) is not present or not written correctly.."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String, " In the sheet “Diagnostic débarqués”, the column XXXX (to be indicated) is not present or not written correctly.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0180_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0180 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0180: The sheet “Effets clients” is not present or not written correctly"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String, "The sheet “Effets clients” is not present or not written correctly", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0190_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0190 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0190: Effets clients”, the column XXXX (to be indicated) is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String, "Effets clients”, the column XXXX (to be indicated) is not present or not written correctly", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0200_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0200 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0200: The sheet “ER” is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String, "  The sheet “ER” is not present or not written correctly", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0210_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0210 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0210: In the sheet “ER”, the column XXXX (to be indicated) is not present or not written correctly.."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String, "In the sheet “ER”, the column XXXX (to be indicated) is not present or not written correctly.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0220_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0220 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0220: The sheet “Constituants” is not present or not written correctly"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String, "The sheet “Constituants” is not present or not written correctly", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0230_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0230 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0230: In the sheet “Constituants”, the column XXXX (to be indicated) is not present or not written correctly.."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String, "In the sheet “Constituants”, the column XXXX (to be indicated) is not present or not written correctly.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0240_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0240 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0240: The sheet “situations de vie” is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String, "The sheet “situations de vie” is not present or not written correctly", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0250_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0250 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0250: In the sheet “situations de vie”, the column XXXX (to be indicated) is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String, "In the sheet “situations de vie”, the column XXXX (to be indicated) is not present or not written correctly.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0260_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0260 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0260: The sheet “Liste MDD” is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String, "The sheet “Liste MDD” is not present or not written correctly", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0270_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0270 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0270: In the sheet “Liste MDD”, the column XXXX (to be indicated) is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String, " In the sheet “Liste MDD”, the column XXXX (to be indicated) is not present or not written correctly.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(
                    ["Good", "02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        elif testResult == 2:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060 NOT OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060",
                                      "The sheet “tableau” (or “table”) is not in the file", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            for key, value in testResult.items():
                if key == "0":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060: In the sheet “tableau” (or “table”), the column “Référence” is not completed."
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1000"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1000 ",
                                              "In the sheet “tableau” (or “table”), the column “Référence” is not completed.",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1000 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1000 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-1":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060: In the sheet “tableau” (or “table”), the column “Version” is not completed."
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1001"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1001 ",
                                              "In the sheet “tableau” (or “table”), the column “Version” is not completed.",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1001 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1001 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-2":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060: In the sheet “tableau” (or “table”), the columns “Constituant défaillant détecté” is not completed."
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1180"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1180 ",
                                              "In the sheet “tableau” (or “table”), the columns “Constituant défaillant détecté” is not completed",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1180 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1180 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-3":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060: In the sheet “tableau” (or “tableau”),  the columns “Défaillance constituant” is not completed"
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1190"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1190 ",
                                              "In the sheet “tableau” (or “tableau”),  the columns “Défaillance constituant” is not completed",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1190 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1190 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-4":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060: In the sheet “tableau” (or “tableau”), the columns “Situation de vie client” is not completed"
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1200"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1200 ",
                                              "In the sheet “tableau” (or “tableau”), the columns “Situation de vie client” is not completed ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1200 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1200 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-5":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060: In the sheet “tableau” (or “tableau”), the columns “Effet(s) client(s)” is not completed."
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1210"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1210 ",
                                              "In the sheet “tableau” (or “tableau”), the columns “Effet(s) client(s)” is not completed.",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1210 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1210 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-6":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060: In the sheet “tableau” (or “tableau”), the columns “Code défaut” is not completed"
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1220"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1220 ",
                                              "In the sheet “tableau” (or “tableau”), the columns “Code défaut” is not completed",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1220 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1220 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-7":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060: The project impact is not specified for the project in the sheet “tableau "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1060"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1060 ",
                                              "The project impact is not specified for the project in the sheet “tableau ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1060 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1060 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_WHOLENESS_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(
                    ["Good", "02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061", "",
                     ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        elif testResult == 2:
            row += 1
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061: NOT OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(
                    ["", "02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061",
                     "The sheet “codes défauts” is not present in the file", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            for key, value in testResult.items():
                if key == "0":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061:  In the sheet “codes défauts”, the column “Référence” is not completed "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1010"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1010",
                                              "In the sheet “codes défauts”, the column “Référence” is not completed. ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1010 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1010 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-1":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061: In the sheet “codes défauts”, the column “Version” is not completed. "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1011"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1011 ",
                                              "In the sheet “codes défauts”, the column “Version” is not completed. ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1011 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1011 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-2":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061:In the sheet “codes défauts”, the column “Code défaut” is not completed. "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1080"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1080 ",
                                              "In the sheet “codes défauts”, the column “Code défaut” is not completed. ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1080 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1080 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-3":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061:In the sheet “codes défauts”, the column “supporté par constituant (s)” is not completed. "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1090"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1090 ",
                                              "In the sheet “codes défauts”, the column “supporté par constituant (s)” is not completed. ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1090 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1090 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-4":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061:In the sheet “codes défauts”, the columns “libellé (signification)” is not completed "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1110"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1110 ",
                                              "In the sheet “codes défauts”, the columns “libellé (signification)” is not completed ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1110 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1110 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-5":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061:In the sheet “codes défauts”, the columns “Description de la strategie pour détecter le défaut” is not completed "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1120"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1120 ",
                                              "In the sheet “codes défauts”, the columns “Description de la strategie pour détecter le défaut” is not completed ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1120 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1120 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-6":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061:In the sheet “codes défauts”, the columns “Seuil de détection  /  valeur  du défaut” is not completed. "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1130"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1130 ",
                                              "In the sheet “codes défauts”, the columns “Seuil de détection  /  valeur  du défaut” is not completed.",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1130 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1130 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-7":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061:In the sheet “codes défauts”, the columns “Temps de confirmation du défaut” is not completed. "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1140"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1140 ",
                                              "In the sheet “codes défauts”, the columns “Temps de confirmation du défaut” is not completed. ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1140 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1140 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-8":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061:In the sheet “codes défauts”, the columns “Description de la strategie de disparition du défaut / Procedure à effectuer pour vérifier la disparition du défaut” is not completed "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1150"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1150 ",
                                              "In the sheet “codes défauts”, the columns “Description de la strategie de disparition du défaut / Procedure à effectuer pour vérifier la disparition du défaut” is not completed ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1150 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1150 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-9":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061:In the sheet “codes défauts”, the columns “Mode dégradé” is not completed. "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1160"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1160 ",
                                              "In the sheet “codes défauts”, the columns “Mode dégradé” is not completed. ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1160 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1160 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-10":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061:In the sheet “codes défauts”, the columns “Voyant” is not completed. "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1170"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1170 ",
                                              "In the sheet “codes défauts”, the columns “Voyant” is not completed. ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1170 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1170 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-11":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061: The project impact is not specified for the project in the sheet “codes défauts "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1061"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1061 ",
                                              " The project impact is not specified for the project in the sheet “codes défauts ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1061 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1061 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_WHOLENESS_1020_1021_1100_1062_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_WHOLENESS_1020_1021_1100_1062 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1020_1021_1100_1062 OK", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        elif testResult == 2:
            row += 1
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_WHOLENESS_1020_1021_1100_1062 NOT OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["", "02043_18_04939_WHOLENESS_1020_1021_1100_1062",
                                      "The sheet “mesures et commandes” is not in the file", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        else:
            for key, value in testResult.items():
                if key == "0":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1020_1021_1100_1062: In the sheet “mesures et commandes”, the column “Référence” is not completed"
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1020"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1020 ",
                                              "In the sheet “mesures et commandes”, the column “Référence” is not completed ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1020 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1020 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-1":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1020_1021_1100_1062: In the sheet “mesures et commandes”, the column “Version” is not completed."
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1021"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1021 ",
                                              "In the sheet “mesures et commandes”, the column “Version” is not completed.",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENES_1021 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1021 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-2":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1020_1021_1100_1062: In the sheet “mesures et commandes”, the column “supporté par constituant (s)” is not completed."
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1100"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1100 ",
                                              "In the sheet “mesures et commandes”, the column “supporté par constituant (s)” is not completed. ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1100 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1100 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-3":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1020_1021_1100_1062: The project impact is not specified for the project in the sheet “mesures et commandes”"
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1062"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1062 ",
                                              "The project impact is not specified for the project in the sheet “mesures et commandes” ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1062 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1062 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_WHOLENESS_1030_1031_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_WHOLENESS_1030_1031 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1030_1031", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        elif testResult == 2:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_WHOLENESS_1030_1031 NOT OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(
                    ["", "02043_18_04939_WHOLENESS_1030_1031", "The sheet “Diagnostic débarqués” is not in the file",
                     ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        else:
            for key, value in testResult.items():
                if key == "0":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nIn the sheet “Diagnostic débarqués”, the column “Référence” is not completed."
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1030"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1030 ",
                                              "Diagnostic débarqués”, the column “Référence” is not completed. ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1030 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "002043_18_04939_WHOLENESS_1030 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-1":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nIn the sheet “Diagnostic débarqués”, the column “Version” is not completed"
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1031"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1031",
                                              "In the sheet “Diagnostic débarqués”, the column “Version” is not completed.",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1031 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1031 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_WHOLENESS_1040_1041_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_WHOLENESS_1040_1041 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1040_1041", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        elif testResult == 2:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_WHOLENESS_1040_1041 NOT OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(
                    ["Good", "02043_18_04939_WHOLENESS_1040_1041", "The sheet “Liste MDD” is not in the file", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            for key, value in testResult.items():
                if key == "0":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1040_1041: In the sheet “Liste MDD”, the column “Référence” is not completed."
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1040"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1040",
                                              "In the sheet “Liste MDD”, the column “Référence” is not completed. ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1040 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1040 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-1":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1040_1041: In the sheet “Liste MDD”, the column “Version” is not completed."
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1041"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1041",
                                              "In the sheet “Liste MDD”, the column “Version” is not completed. ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1041 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1041 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_WHOLENESS_1050_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_1050 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_1050", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        elif testResult == 2:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_1050 NOT OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_1050",
                                      "The “tableau”(or “table”) or “codes défauts” is not presetnt in the file", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_1050: The projects listed in the following sheets “tableau” and “codes défauts” are not identical."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1050"][self.checkLevel],
                                      "02043_18_04939_WHOLENESS_1050 ",
                                      " The projects listed in the following sheets “tableau” and “codes défauts” are not identical. ",
                                      ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_WHOLENESS_1055_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_1055 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_1055", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        elif testResult == 2:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_1055 NOT OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_1055",
                                      "The “tableau”(or “table”) or “mesures et commandes” is not presetnt in the file",
                                      ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_1055: The projects listed in the following sheets “tableau” and “mesures et commandes” are not identical. "
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS _1055"][self.checkLevel],
                                      "02043_18_04939_WHOLENESS _1055",
                                      "The projects listed in the following sheets “tableau” and “mesures et commandes” are not identical. ",
                                      ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_WHOLENESS_1240_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_WHOLENESS_1240 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1240", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_WHOLENESS_1240: The reviewers have to be indicated in the paraph  “Liste de diffusion” or “Mailing list (the taking part)” of the sheet “Informations Générales” (or “General information”)."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1240"][self.checkLevel], "02043_18_04939_WHOLENESS_1240", "The reviewers have to be indicated in the paraph  “Liste de diffusion” or “Mailing list (the taking part)” of the sheet “Informations Générales” (or “General information”).",""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name


        row += 1
        testResult = self.Test_02043_18_04939_COH_2000_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_COH_2000 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_COH_2000", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            if testResult == 2:
                text = self.tab1.textbox.toPlainText()
                text = text + "\nTest_02043_18_04939_COH_2005:   The sheet or column “Code défaut” or “tableau”(“table”) is missing"

            else:
                text = self.tab1.textbox.toPlainText()
                text = text + "\nTest_02043_18_04939_COH_2000: The information specified in the column “mesures et commandes (Mesure Parametre et Test Actionneur) / Tests de cohérence” of the sheet “tableau” is missing in the column “libellé (signification)” of the sheet “mesures et commandes"
                self.tab1.textbox.setText(text)
                flag = 0
                for i, name in enumerate([self.testReqDict["02043_18_04939_COH_2000"][self.checkLevel],
                                          "02043_18_04939_COH_2000",
                                          "The information specified in the column “mesures et commandes (Mesure Parametre et Test Actionneur) / Tests de cohérence” of the sheet “tableau” is missing in the column “libellé (signification)” of the sheet “mesures et commandes",
                                          ""]):
                    reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult, famillyList, localisation = self.Test_02043_18_04939_COH_2005_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_COH_2005 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_COH_2005", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            if testResult == 2:
                text = self.tab1.textbox.toPlainText()
                text = text + "\nTest_02043_18_04939_COH_2005:   The sheet or column “Code défaut” is missing"
            else:
                text = self.tab1.textbox.toPlainText()
                text = text + "\nTest_02043_18_04939_COH_2005:   The DTC specified in the cell XXXX of the “Code défaut” sheet shall respect the following format: SubFamillyName-DTCcodeNumber-Caracterisation (UDS format) OR SubFamillyName-DTCcodeNumber (KW format)."
                self.tab1.textbox.setText(text)
                flag = 0
                for i, name in enumerate([self.testReqDict["02043_18_04939_COH_2005"][self.checkLevel],
                                          "02043_18_04939_COH_2005",
                                          " The DTC specified in the cell " + localisation + " of the “Code défaut” sheet shall respect the following format: SubFamillyName-DTCcodeNumber-Caracterisation (UDS format) OR SubFamillyName-DTCcodeNumber (KW format)",
                                          localisation]):
                    reportWorkSheet2.Cells(row, i + 1).Value = name


        row += 1
        testResult, List = self.Test_02043_18_04939_COH_2006_XLS(workBook, famillyList)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_COH_2006 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_COH_2006", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            if testResult == 2:
                text = self.tab1.textbox.toPlainText()
                text = text + "\nTest_02043_18_04939_COH_2005:   The sheet or column “Code défaut” is missing"
            else:
                text = self.tab1.textbox.toPlainText()
                text = text + "\nTest_02043_18_04939_COH_2006: The DTC specified in the cell XXXX of the “Table” sheet shall be present in the column “Nom de la sous famille” of the sheet “sous familles Cesare” in the document “Extract from CESARE” [DOC8]."
                self.tab1.textbox.setText(text)
                flag = 0
                for i, name in enumerate([self.testReqDict["02043_18_04939_COH_2006"][self.checkLevel],
                                          "02043_18_04939_COH_2006",
                                          "The DTC specified in the cell XXXX of the “Table” sheet shall be present in the column “Nom de la sous famille” of the sheet “sous familles Cesare” in the document “Extract from CESARE” [DOC8].",
                                          str(List[0]['col'] + List[0]['row'])]):
                    reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult, List = self.Test_02043_18_04939_COH_2007_XLS(workBook, famillyList)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_COH_2007 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_COH_2007", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            if testResult == 2:
                text = self.tab1.textbox.toPlainText()
                text = text + "\nTest_02043_18_04939_COH_2007:   The sheet or column “Code défaut” is missing"
            else:
                text = self.tab1.textbox.toPlainText()
                text = text + "\nTest_02043_18_04939_COH_2007: The DTC specified in the cell XXXX of the “Table” sheet shall be present in the column “Data Trouble Code (DTC)” of the sheet “Matrix” of the “Diagnostic matrix extract from DOTI” [DOC14]"
                self.tab1.textbox.setText(text)
                flag = 0
                for i, name in enumerate([self.testReqDict["02043_18_04939_COH_2007"][self.checkLevel],
                                          "02043_18_04939_COH_2007",
                                          "The DTC specified in the cell XXXX of the “Table” sheet shall be present in the column “Data Trouble Code (DTC)” of the sheet “Matrix” of the “Diagnostic matrix extract from DOTI” [DOC14]",
                                          str(List[0]['col'] + List[0]['row'])]):
                    reportWorkSheet2.Cells(row, i + 1).Value = name


        row += 1
        testResult = self.Test_02043_18_04939_COH_2010_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_COH_2010 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_COH_2010", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            if testResult == 2:
                text = self.tab1.textbox.toPlainText()
                text = text + "\nTest_02043_18_04939_COH_2010:   The sheet or column “Code défaut” or “tableau”(“table”) is missing"

            else:
                text = self.tab1.textbox.toPlainText()
                text = text + "\nTest_02043_18_04939_COH_2010: The information specified in the column “Code défaut” of the sheet “tableau”, is missing in the column “Code défaut” of the sheet “codes défauts"
                self.tab1.textbox.setText(text)
                flag = 0
                for i, name in enumerate([self.testReqDict["02043_18_04939_COH_2010"][self.checkLevel],
                                          "02043_18_04939_COH_2010",
                                          "The information specified in the column “Code défaut” of the sheet “tableau”, is missing in the column “Code défaut” of the sheet “codes défauts",
                                          ""]):
                    reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_COH_2020_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_COH_2020 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_COH_2020", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            if testResult == 2:
                text = self.tab1.textbox.toPlainText()
                text = text + "\nTest_02043_18_04939_COH_2020: The information specified in the sheet “Constituant défaillant détecté” of the sheet “tableau”, is missing in the column “Noms” of the sheet “Constituants"

            else:
                text = self.tab1.textbox.toPlainText()
                text = text + "\nTest_02043_18_04939_COH_2020: The information specified in the column “Code défaut” of the sheet “tableau”, is missing in the column “Code défaut” of the sheet “codes défauts"
                self.tab1.textbox.setText(text)
                flag = 0
                for i, name in enumerate([self.testReqDict["02043_18_04939_COH_2020"][self.checkLevel],
                                          "02043_18_04939_COH_2020",
                                          "The information specified in the column “Constituant défaillant détecté” of the sheet “tableau”, is missing in the column “Noms” of the sheet “Constituants",
                                          ""]):
                    reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_COH_2030_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_COH_2030 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_COH_2030", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            if testResult == 2:
                text = self.tab1.textbox.toPlainText()
                text = text + "\nTest_02043_18_04939_COH_2030: The information specified in the sheet “Effets clients” of the sheet “tableau”, is missing in the column “Noms” of the sheet “Constituants”"

            else:
                text = self.tab1.textbox.toPlainText()
                text = text + "\nTest_02043_18_04939_COH_2030: The information specified in the column “Effet(s) client(s)” of the sheet “tableau” is missing in the  column “Noms” of the sheet “Effets clients” "
                self.tab1.textbox.setText(text)
                flag = 0
                for i, name in enumerate([self.testReqDict["02043_18_04939_COH_2030"][self.checkLevel],
                                          "02043_18_04939_COH_2030",
                                          "The information specified in the column “Effet(s) client(s)” of the sheet “tableau” is missing in the  column “Noms” of the sheet “Effets clients”",
                                          ""]):
                    reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_COH_2040_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_COH_2040 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_COH_2040", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            if testResult == 2:
                text = self.tab1.textbox.toPlainText()
                text = text + "\nTest_02043_18_04939_COH_2040: The information specified in the column “DIAGNOSTIC DEBARQUE” of the sheet “tableau” is missing in the column “libellé (signification)” of the sheet “Diagnostic débarqués” "

            else:
                text = self.tab1.textbox.toPlainText()
                text = text + "\nTest_02043_18_04939_COH_2040: The information specified in the column “DIAGNOSTIC DEBARQUE” of the sheet “tableau” is missing in the column “libellé (signification)” of the sheet “Diagnostic débarqués” "
                self.tab1.textbox.setText(text)
                flag = 0
                for i, name in enumerate([self.testReqDict["02043_18_04939_COH_2040"][self.checkLevel],
                                          "02043_18_04939_COH_2040",
                                          "The information specified in the column “DIAGNOSTIC DEBARQUE” of the sheet “tableau” is missing in the column “libellé (signification)” of the sheet “Diagnostic débarqués” ",
                                          ""]):
                    reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_COH_2050_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_COH_2050 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_COH_2050", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            if testResult == 2:
                text = self.tab1.textbox.toPlainText()
                text = text + "\nTest_02043_18_04939_COH_2050: The information specified in the sheet “ER” of the sheet “tableau”, is missing in the column “Noms” of the sheet “Constituants” "

            else:
                text = self.tab1.textbox.toPlainText()
                text = text + "\nTest_02043_18_04939_COH_2050:  The information specified in the column “Evenement(s) redouté(s) (ER)” of the sheet “tableau” is missing in the column “nom” of the sheet “ER”  "
                self.tab1.textbox.setText(text)
                flag = 0
                for i, name in enumerate([self.testReqDict["02043_18_04939_COH_2050"][self.checkLevel],
                                          "02043_18_04939_COH_2050",
                                          " The information specified in the column “Evenement(s) redouté(s) (ER)” of the sheet “tableau” is missing in the column “nom” of the sheet “ER” ",
                                          ""]):
                    reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_COH_2060_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_COH_2060 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_COH_2060", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            if testResult == 2:
                text = self.tab1.textbox.toPlainText()
                text = text + "\nTest_02043_18_04939_COH_2060: One of the sheets “Effets clients”, “FR” or “GB” are not in the file  "

            else:
                text = self.tab1.textbox.toPlainText()
                text = text + "\nTest_02043_18_04939_COH_2060: The information specified in the column “Noms” of the sheet “Effets clients” is missing in the document [DOC7] in French or English. Please check it or ask ADRD people to add this new customer effect  "
                self.tab1.textbox.setText(text)
                flag = 0
                for i, name in enumerate([self.testReqDict["02043_18_04939_COH_2060"][self.checkLevel],
                                          "02043_18_04939_COH_2060",
                                          " The information specified in the column “Noms” of the sheet “Effets clients” is missing in the document [DOC7] in French or English. Please check it or ask ADRD people to add this new customer effect",
                                          ""]):
                    reportWorkSheet2.Cells(row, i + 1).Value = name

        try:
            reportWorkBook.Save()
        except Exception as e:
            print(e)

    def TestGeneralStructure_DOC3_XLSX_XLSM(self, workBook):

        flag = 1
        fileName = self.tab1.myTextBox1.toPlainText()
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        reportWorkBook = excel.Workbooks.Open(fileName)
        reportWorkSheet2 = None

        for sheet in reportWorkBook.Worksheets:
            if sheet.Name == "Test Report":
                reportWorkSheet2 = sheet
        if not reportWorkSheet2:
            workSheetsNumber = reportWorkBook.Sheets.Count
            sheetAfter = reportWorkBook.Sheets(workSheetsNumber)
            reportWorkSheet2 = reportWorkBook.Worksheets.Add(None, sheetAfter)
            reportWorkSheet2.Name = "Test Report"

        testReportRow1StringList = ["Criticity", "Requirements", "Message", "Localisation"]

        for i, name in enumerate(testReportRow1StringList):
            reportWorkSheet2.Cells(1, i + 1).Value = name
        reportWorkSheet2.Columns.AutoFit()
        reportWorkSheet2.Columns.Font.Bold = True

        row = 21
        str1 = "02043_18_04939_STRUCT_0"
        stringInt = 10 - row
        str2 = str(stringInt + row)
        str3 = "0"
        String = str1 + str2 + str3

        testResult = self.Test_02043_18_04939_STRUCT_0100_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0100 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0100: The sheet “tableau” is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String,
                                      " The sheet “tableau” is not present or not written correctly.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0110_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0110 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0110: In the sheet “tableau”, the column XXXX (to be indicated) is not present or not written correctly.."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String,
                                      "In the sheet “tableau”, the column XXXX (to be indicated) is not present or not written correctly.",
                                      ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0120_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0120 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0120: The sheet “codes défauts” is not present or not written correctly"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String,
                                      "The sheet “codes défauts” is not present or not written correctly", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0130_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0130 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0130: In the sheet “codes défauts”, the column XXXX (to be indicated) is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String,
                                      "In the sheet “codes défauts”, the column XXXX (to be indicated) is not present or not written correctly.",
                                      ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0140_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0140 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0140: The sheet “mesures et commandes” is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String,
                                      "The sheet “mesures et commandes” is not present or not written correctly", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0150_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0140 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0150: In the sheet “mesures et commandes”, the column XXXX (to be indicated) is not present or not written correctly. "
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String,
                                      " In the sheet “mesures et commandes”, the column XXXX (to be indicated) is not present or not written correctly. ",
                                      ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0160_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0160 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0160: The sheet “Diagnostic débarqués” is not present or not written correctly"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String,
                                      "The sheet “Diagnostic débarqués” is not present or not written correctly", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0170_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0170 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0170: In the sheet “Diagnostic débarqués”, the column XXXX (to be indicated) is not present or not written correctly.."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String,
                                      " In the sheet “Diagnostic débarqués”, the column XXXX (to be indicated) is not present or not written correctly.",
                                      ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0180_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0180 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0180: The sheet “Effets clients” is not present or not written correctly"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String,
                                      "The sheet “Effets clients” is not present or not written correctly", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0190_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0190 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0190: Effets clients”, the column XXXX (to be indicated) is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String,
                                      "Effets clients”, the column XXXX (to be indicated) is not present or not written correctly",
                                      ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0200_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0200 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0200: The sheet “ER” is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String,
                                      "  The sheet “ER” is not present or not written correctly", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0210_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0210 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0210: In the sheet “ER”, the column XXXX (to be indicated) is not present or not written correctly.."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String,
                                      "In the sheet “ER”, the column XXXX (to be indicated) is not present or not written correctly.",
                                      ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0220_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0220 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0220: The sheet “Constituants” is not present or not written correctly"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String,
                                      "The sheet “Constituants” is not present or not written correctly", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0230_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0230 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0230: In the sheet “Constituants”, the column XXXX (to be indicated) is not present or not written correctly.."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String,
                                      "In the sheet “Constituants”, the column XXXX (to be indicated) is not present or not written correctly.",
                                      ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0240_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0240 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0240: The sheet “situations de vie” is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String,
                                      "The sheet “situations de vie” is not present or not written correctly", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0250_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0250 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0250: In the sheet “situations de vie”, the column XXXX (to be indicated) is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String,
                                      "In the sheet “situations de vie”, the column XXXX (to be indicated) is not present or not written correctly.",
                                      ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0260_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0260 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0260: The sheet “Liste MDD” is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String,
                                      "The sheet “Liste MDD” is not present or not written correctly", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0270_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0270 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0270: In the sheet “Liste MDD”, the column XXXX (to be indicated) is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String,
                                      " In the sheet “Liste MDD”, the column XXXX (to be indicated) is not present or not written correctly.",
                                      ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(
                    ["Good", "02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        elif testResult == 2:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060 NOT OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060",
                                      "The sheet “tableau” (or “table”) is not in the file", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            for key, value in testResult.items():
                if key == "0":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060: In the sheet “tableau” (or “table”), the column “Référence” is not completed."
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1000"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1000 ",
                                              "In the sheet “tableau” (or “table”), the column “Référence” is not completed.",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1000 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1000 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-1":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060: In the sheet “tableau” (or “table”), the column “Version” is not completed."
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1001"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1001 ",
                                              "In the sheet “tableau” (or “table”), the column “Version” is not completed.",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1001 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1001 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-2":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060: In the sheet “tableau” (or “table”), the columns “Constituant défaillant détecté” is not completed."
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1180"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1180 ",
                                              "In the sheet “tableau” (or “table”), the columns “Constituant défaillant détecté” is not completed",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1180 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1180 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-3":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060: In the sheet “tableau” (or “tableau”),  the columns “Défaillance constituant” is not completed"
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1190"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1190 ",
                                              "In the sheet “tableau” (or “tableau”),  the columns “Défaillance constituant” is not completed",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1190 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1190 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-4":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060: In the sheet “tableau” (or “tableau”), the columns “Situation de vie client” is not completed"
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1200"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1200 ",
                                              "In the sheet “tableau” (or “tableau”), the columns “Situation de vie client” is not completed ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1200 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1200 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-5":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060: In the sheet “tableau” (or “tableau”), the columns “Effet(s) client(s)” is not completed."
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1210"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1210 ",
                                              "In the sheet “tableau” (or “tableau”), the columns “Effet(s) client(s)” is not completed.",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1210 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1210 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-6":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060: In the sheet “tableau” (or “tableau”), the columns “Code défaut” is not completed"
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1220"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1220 ",
                                              "In the sheet “tableau” (or “tableau”), the columns “Code défaut” is not completed",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1220 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1220 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-7":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060: The project impact is not specified for the project in the sheet “tableau "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1060"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1060 ",
                                              "The project impact is not specified for the project in the sheet “tableau ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1060 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1060 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name

        '''row += 1
        testResult = self.Test_02043_18_04939_WHOLENESS_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061_XLSX_XLSM(
            workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(
                    ["Good", "02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061", "",
                     ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        elif testResult == 2:
            row += 1
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061: NOT OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(
                    ["", "02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061",
                     "The sheet “codes défauts” is not present in the file", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            for key, value in testResult.items():
                if key == "0":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061:  In the sheet “codes défauts”, the column “Référence” is not completed "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1010"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1010",
                                              "In the sheet “codes défauts”, the column “Référence” is not completed. ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1010 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1010 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-1":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061: In the sheet “codes défauts”, the column “Version” is not completed. "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1011"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1011 ",
                                              "In the sheet “codes défauts”, the column “Version” is not completed. ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1011 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1011 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-2":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061:In the sheet “codes défauts”, the column “Code défaut” is not completed. "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1080"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1080 ",
                                              "In the sheet “codes défauts”, the column “Code défaut” is not completed. ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1080 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1080 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-3":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061:In the sheet “codes défauts”, the column “supporté par constituant (s)” is not completed. "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1090"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1090 ",
                                              "In the sheet “codes défauts”, the column “supporté par constituant (s)” is not completed. ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1090 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1090 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-4":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061:In the sheet “codes défauts”, the columns “libellé (signification)” is not completed "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1110"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1110 ",
                                              "In the sheet “codes défauts”, the columns “libellé (signification)” is not completed ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1110 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1110 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-5":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061:In the sheet “codes défauts”, the columns “Description de la strategie pour détecter le défaut” is not completed "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1120"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1120 ",
                                              "In the sheet “codes défauts”, the columns “Description de la strategie pour détecter le défaut” is not completed ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1120 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1120 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-6":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061:In the sheet “codes défauts”, the columns “Seuil de détection  /  valeur  du défaut” is not completed. "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1130"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1130 ",
                                              "In the sheet “codes défauts”, the columns “Seuil de détection  /  valeur  du défaut” is not completed.",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1130 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1130 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-7":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061:In the sheet “codes défauts”, the columns “Temps de confirmation du défaut” is not completed. "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1140"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1140 ",
                                              "In the sheet “codes défauts”, the columns “Temps de confirmation du défaut” is not completed. ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1140 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1140 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-8":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061:In the sheet “codes défauts”, the columns “Description de la strategie de disparition du défaut / Procedure à effectuer pour vérifier la disparition du défaut” is not completed "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1150"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1150 ",
                                              "In the sheet “codes défauts”, the columns “Description de la strategie de disparition du défaut / Procedure à effectuer pour vérifier la disparition du défaut” is not completed ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1150 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1150 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-9":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061:In the sheet “codes défauts”, the columns “Mode dégradé” is not completed. "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1160"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1160 ",
                                              "In the sheet “codes défauts”, the columns “Mode dégradé” is not completed. ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1160 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1160 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-10":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061:In the sheet “codes défauts”, the columns “Voyant” is not completed. "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1170"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1170 ",
                                              "In the sheet “codes défauts”, the columns “Voyant” is not completed. ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1170 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1170 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-11":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061: The project impact is not specified for the project in the sheet “codes défauts "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1061"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1061 ",
                                              " The project impact is not specified for the project in the sheet “codes défauts ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1061 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1061 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name'''

        row += 1
        testResult = self.Test_02043_18_04939_WHOLENESS_1020_1021_1100_1062_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_WHOLENESS_1020_1021_1100_1062 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1020_1021_1100_1062 OK", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        elif testResult == 2:
            row += 1
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_WHOLENESS_1020_1021_1100_1062 NOT OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["", "02043_18_04939_WHOLENESS_1020_1021_1100_1062",
                                      "The sheet “mesures et commandes” is not in the file", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        else:
            for key, value in testResult.items():
                if key == "0":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1020_1021_1100_1062: In the sheet “mesures et commandes”, the column “Référence” is not completed"
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1020"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1020 ",
                                              "In the sheet “mesures et commandes”, the column “Référence” is not completed ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1020 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1020 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-1":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1020_1021_1100_1062: In the sheet “mesures et commandes”, the column “Version” is not completed."
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1021"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1021 ",
                                              "In the sheet “mesures et commandes”, the column “Version” is not completed.",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENES_1021 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1021 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-2":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1020_1021_1100_1062: In the sheet “mesures et commandes”, the column “supporté par constituant (s)” is not completed."
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1100"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1100 ",
                                              "In the sheet “mesures et commandes”, the column “supporté par constituant (s)” is not completed. ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1100 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1100 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-3":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1020_1021_1100_1062: The project impact is not specified for the project in the sheet “mesures et commandes”"
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1062"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1062 ",
                                              "The project impact is not specified for the project in the sheet “mesures et commandes” ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1062 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1062 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_WHOLENESS_1030_1031_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_WHOLENESS_1030_1031 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1030_1031", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        elif testResult == 2:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_WHOLENESS_1030_1031 NOT OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(
                    ["", "02043_18_04939_WHOLENESS_1030_1031", "The sheet “Diagnostic débarqués” is not in the file",
                     ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        else:
            for key, value in testResult.items():
                if key == "0":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nIn the sheet “Diagnostic débarqués”, the column “Référence” is not completed."
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1030"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1030 ",
                                              "Diagnostic débarqués”, the column “Référence” is not completed. ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1030 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "002043_18_04939_WHOLENESS_1030 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-1":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nIn the sheet “Diagnostic débarqués”, the column “Version” is not completed"
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1031"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1031",
                                              "In the sheet “Diagnostic débarqués”, the column “Version” is not completed.",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1031 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1031 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_WHOLENESS_1040_1041_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_WHOLENESS_1040_1041 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1040_1041", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        elif testResult == 2:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_WHOLENESS_1040_1041 NOT OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(
                    ["Good", "02043_18_04939_WHOLENESS_1040_1041", "The sheet “Liste MDD” is not in the file", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            for key, value in testResult.items():
                if key == "0":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1040_1041: In the sheet “Liste MDD”, the column “Référence” is not completed."
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1040"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1040",
                                              "In the sheet “Liste MDD”, the column “Référence” is not completed. ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1040 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1040 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-1":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1040_1041: In the sheet “Liste MDD”, the column “Version” is not completed."
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1041"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1041",
                                              "In the sheet “Liste MDD”, the column “Version” is not completed. ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1041 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1041 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_WHOLENESS_1050_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_1050 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_1050", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        elif testResult == 2:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_1050 NOT OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_1050",
                                      "The “tableau”(or “table”) or “codes défauts” is not presetnt in the file", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_1050: The projects listed in the following sheets “tableau” and “codes défauts” are not identical."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1050"][self.checkLevel],
                                      "02043_18_04939_WHOLENESS_1050 ",
                                      " The projects listed in the following sheets “tableau” and “codes défauts” are not identical. ",
                                      ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_WHOLENESS_1055_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_1055 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_1055", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        elif testResult == 2:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_1055 NOT OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_1055",
                                      "The “tableau”(or “table”) or “mesures et commandes” is not presetnt in the file",
                                      ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_1055: The projects listed in the following sheets “tableau” and “mesures et commandes” are not identical. "
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS _1055"][self.checkLevel],
                                      "02043_18_04939_WHOLENESS _1055",
                                      "The projects listed in the following sheets “tableau” and “mesures et commandes” are not identical. ",
                                      ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_WHOLENESS_1240_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_WHOLENESS_1240 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1240", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_WHOLENESS_1240: The reviewers have to be indicated in the paraph  “Liste de diffusion” or “Mailing list (the taking part)” of the sheet “Informations Générales” (or “General information”)."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1240"][self.checkLevel],
                                      "02043_18_04939_WHOLENESS_1240",
                                      "The reviewers have to be indicated in the paraph  “Liste de diffusion” or “Mailing list (the taking part)” of the sheet “Informations Générales” (or “General information”).",
                                      ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_COH_2000_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_COH_2000 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_COH_2000", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_COH_2000: The information specified in the column “mesures et commandes (Mesure Parametre et Test Actionneur) / Tests de cohérence” of the sheet “tableau” is missing in the column “libellé (signification)” of the sheet “mesures et commandes"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_COH_2000"][self.checkLevel],
                                      "02043_18_04939_COH_2000",
                                      "The information specified in the column “mesures et commandes (Mesure Parametre et Test Actionneur) / Tests de cohérence” of the sheet “tableau” is missing in the column “libellé (signification)” of the sheet “mesures et commandes",
                                      ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult, famillyList, localisation = self.Test_02043_18_04939_COH_2005_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_COH_2005 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_COH_2005", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            if testResult == 2:
                text = self.tab1.textbox.toPlainText()
                text = text + "\nTest_02043_18_04939_COH_2005:   The sheet or column “Code défaut” is missing"
            else:
                text = self.tab1.textbox.toPlainText()
                text = text + "\nTest_02043_18_04939_COH_2005:   The DTC specified in the cell XXXX of the “Code défaut” sheet shall respect the following format: SubFamillyName-DTCcodeNumber-Caracterisation (UDS format) OR SubFamillyName-DTCcodeNumber (KW format)."
                self.tab1.textbox.setText(text)
                flag = 0
                for i, name in enumerate([self.testReqDict["02043_18_04939_COH_2005"][self.checkLevel],
                                          "02043_18_04939_COH_2005",
                                          " The DTC specified in the cell XXXX of the “Code défaut” sheet shall respect the following format: SubFamillyName-DTCcodeNumber-Caracterisation (UDS format) OR SubFamillyName-DTCcodeNumber (KW format)",
                                          localisation]):
                    reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult, List = self.Test_02043_18_04939_COH_2006_XLSX_XLSM(workBook, famillyList)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_COH_2006 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_COH_2006", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            if testResult == 2:
                text = self.tab1.textbox.toPlainText()
                text = text + "\nTest_02043_18_04939_COH_2005:   The sheet or column “Code défaut” is missing"
            else:
                text = self.tab1.textbox.toPlainText()
                text = text + "\nTest_02043_18_04939_COH_2006: The DTC specified in the cell XXXX of the “Table” sheet shall be present in the column “Nom de la sous famille” of the sheet “sous familles Cesare” in the document “Extract from CESARE” [DOC8]."
                self.tab1.textbox.setText(text)
                flag = 0
                for i, name in enumerate([self.testReqDict["02043_18_04939_COH_2006"][self.checkLevel],
                                          "02043_18_04939_COH_2006",
                                          "The DTC specified in the cell XXXX of the “Table” sheet shall be present in the column “Nom de la sous famille” of the sheet “sous familles Cesare” in the document “Extract from CESARE” [DOC8].",
                                          str(List[0]['col'] + List[0]['row'])]):
                    reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult, List = self.Test_02043_18_04939_COH_2007_XLSX_XLSM(workBook, famillyList)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_COH_2007 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_COH_2007", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            if testResult == 2:
                text = self.tab1.textbox.toPlainText()
                text = text + "\nTest_02043_18_04939_COH_2007:   The sheet or column “Code défaut” is missing"
            else:
                text = self.tab1.textbox.toPlainText()
                text = text + "\nTest_02043_18_04939_COH_2007: The DTC specified in the cell XXXX of the “Table” sheet shall be present in the column “Data Trouble Code (DTC)” of the sheet “Matrix” of the “Diagnostic matrix extract from DOTI” [DOC14]"
                self.tab1.textbox.setText(text)
                flag = 0
                for i, name in enumerate([self.testReqDict["02043_18_04939_COH_2007"][self.checkLevel],
                                          "02043_18_04939_COH_2007",
                                          "The DTC specified in the cell XXXX of the “Table” sheet shall be present in the column “Data Trouble Code (DTC)” of the sheet “Matrix” of the “Diagnostic matrix extract from DOTI” [DOC14]",
                                          str(List[0]['col'] + List[0]['row'])]):
                    reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_COH_2010_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_COH_2010 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_COH_2010", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            if testResult == 2:
                text = self.tab1.textbox.toPlainText()
                text = text + "\nTest_02043_18_04939_COH_2010: The sheet or column “Code défaut” or “tableau”(“table”) is missing"

            else:
                text = self.tab1.textbox.toPlainText()
                text = text + "\nTest_02043_18_04939_COH_2010: The information specified in the column “Code défaut” of the sheet “tableau”, is missing in the column “Code défaut” of the sheet “codes défauts"
                self.tab1.textbox.setText(text)
                flag = 0
                for i, name in enumerate([self.testReqDict["02043_18_04939_COH_2010"][self.checkLevel],
                                          "02043_18_04939_COH_2010",
                                          "The information specified in the column “Code défaut” of the sheet “tableau”, is missing in the column “Code défaut” of the sheet “codes défauts",
                                          ""]):
                    reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_COH_2020_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_COH_2020 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_COH_2020", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            if testResult == 2:
                text = self.tab1.textbox.toPlainText()
                text = text + "\nTest_02043_18_04939_COH_2020: Tthe information specified in the column “Constituant défaillant détecté” of the sheet “tableau”, is missing in the column “Noms” of the sheet “Constituants"

            else:
                text = self.tab1.textbox.toPlainText()
                text = text + "\nTest_02043_18_04939_COH_2020: The information specified in the column “Code défaut” of the sheet “tableau”, is missing in the column “Code défaut” of the sheet “codes défauts"
                self.tab1.textbox.setText(text)
                flag = 0
                for i, name in enumerate([self.testReqDict["02043_18_04939_COH_2020"][self.checkLevel],
                                          "02043_18_04939_COH_2020",
                                          "Tthe information specified in the column “Constituant défaillant détecté” of the sheet “tableau”, is missing in the column “Noms” of the sheet “Constituants",
                                          ""]):
                    reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_COH_2030_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_COH_2030 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_COH_2030", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            if testResult == 2:
                text = self.tab1.textbox.toPlainText()
                text = text + "\nTest_02043_18_04939_COH_2030: The information specified in the sheet “Effets clients” of the sheet “tableau”, is missing in the column “Noms” of the sheet “Constituants”"

            else:
                text = self.tab1.textbox.toPlainText()
                text = text + "\nTest_02043_18_04939_COH_2030: The information specified in the column “Effet(s) client(s)” of the sheet “tableau” is missing in the  column “Noms” of the sheet “Effets clients” "
                self.tab1.textbox.setText(text)
                flag = 0
                for i, name in enumerate([self.testReqDict["02043_18_04939_COH_2030"][self.checkLevel],
                                          "02043_18_04939_COH_2030",
                                          "The information specified in the column “Effet(s) client(s)” of the sheet “tableau” is missing in the  column “Noms” of the sheet “Effets clients”",
                                          ""]):
                    reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_COH_2040_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_COH_2040 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_COH_2040", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            if testResult == 2:
                text = self.tab1.textbox.toPlainText()
                text = text + "\nTest_02043_18_04939_COH_2040: The information specified in the column “DIAGNOSTIC DEBARQUE” of the sheet “tableau” is missing in the column “libellé (signification)” of the sheet “Diagnostic débarqués” "

            else:
                text = self.tab1.textbox.toPlainText()
                text = text + "\nTest_02043_18_04939_COH_2040: The information specified in the column “DIAGNOSTIC DEBARQUE” of the sheet “tableau” is missing in the column “libellé (signification)” of the sheet “Diagnostic débarqués” "
                self.tab1.textbox.setText(text)
                flag = 0
                for i, name in enumerate([self.testReqDict["02043_18_04939_COH_2040"][self.checkLevel],
                                          "02043_18_04939_COH_2040",
                                          "The information specified in the column “DIAGNOSTIC DEBARQUE” of the sheet “tableau” is missing in the column “libellé (signification)” of the sheet “Diagnostic débarqués” ",
                                          ""]):
                    reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_COH_2050_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_COH_2050 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_COH_2050", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            if testResult == 2:
                text = self.tab1.textbox.toPlainText()
                text = text + "\nTest_02043_18_04939_COH_2050: The information specified in the sheet “ER” of the sheet “tableau”, is missing in the column “Noms” of the sheet “Constituants” "

            else:
                text = self.tab1.textbox.toPlainText()
                text = text + "\nTest_02043_18_04939_COH_2050:  The information specified in the column “Evenement(s) redouté(s) (ER)” of the sheet “tableau” is missing in the column “nom” of the sheet “ER”  "
                self.tab1.textbox.setText(text)
                flag = 0
                for i, name in enumerate([self.testReqDict["02043_18_04939_COH_2050"][self.checkLevel],
                                          "02043_18_04939_COH_2050",
                                          " The information specified in the column “Evenement(s) redouté(s) (ER)” of the sheet “tableau” is missing in the column “nom” of the sheet “ER” ",
                                          ""]):
                    reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_COH_2060_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_COH_2060 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_COH_2060", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            if testResult == 2:
                text = self.tab1.textbox.toPlainText()
                text = text + "\nTest_02043_18_04939_COH_2060: One of the sheets “Effets clients”, “FR” or “GB” are not in the file  "

            else:
                text = self.tab1.textbox.toPlainText()
                text = text + "\nTest_02043_18_04939_COH_2060: The information specified in the column “Noms” of the sheet “Effets clients” is missing in the document [DOC7] in French or English. Please check it or ask ADRD people to add this new customer effect  "
                self.tab1.textbox.setText(text)
                flag = 0
                for i, name in enumerate([self.testReqDict["02043_18_04939_COH_2060"][self.checkLevel],
                                          "02043_18_04939_COH_2060",
                                          " The information specified in the column “Noms” of the sheet “Effets clients” is missing in the document [DOC7] in French or English. Please check it or ask ADRD people to add this new customer effect",
                                          ""]):
                    reportWorkSheet2.Cells(row, i + 1).Value = name

        try:
            reportWorkBook.Save()
        except Exception as e:
            print(e)

    def TestGeneralStructure_DOC4_XLS(self, workBook):

        flag = 1
        fileName = self.tab1.myTextBox2.toPlainText()
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        reportWorkBook = excel.Workbooks.Open(fileName)
        reportWorkSheet2 = None

        for sheet in reportWorkBook.Worksheets:
            if sheet.Name == "Test Report":
                reportWorkSheet2 = sheet
        if not reportWorkSheet2:
            workSheetsNumber = reportWorkBook.Sheets.Count
            sheetAfter = reportWorkBook.Sheets(workSheetsNumber)
            reportWorkSheet2 = reportWorkBook.Worksheets.Add(None, sheetAfter)
            reportWorkSheet2.Name = "Test Report"

        testReportRow1StringList = ["Criticity", "Requirements", "Message", "Localisation"]

        for i, name in enumerate(testReportRow1StringList):
            reportWorkSheet2.Cells(1, i + 1).Value = name
        reportWorkSheet2.Columns.AutoFit()
        reportWorkSheet2.Columns.Font.Bold = True

        row = 21
        str1 = "02043_18_04939_STRUCT_0"
        stringInt = 40 - row
        str2 = str(stringInt + row)
        str3 = "0"
        String = str1 + str2 + str3

        testResult = self.Test_02043_18_04939_STRUCT_0400_XLS(workBook)
        if testResult == 1:
            for i, name in enumerate(["Good", String, "", ""]):
                 reportWorkSheet2.Cells(row, i+1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0400: The sheet “tableau” is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String, "The sheet “Table” (or “tableau”) is not present or not written correctly.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0410_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0410 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                 reportWorkSheet2.Cells(row, i+1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0410: In the sheet “tableau” (or “tableau”), the column XXXX (to be indicated) is not present or not written correctly as in the document [DOC3].."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String, "In the sheet “tableau” (or “tableau”), the column XXXX (to be indicated) is not present or not written correctly as in the document [DOC3].", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0420_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0420 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                 reportWorkSheet2.Cells(row, i+1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0420: The sheet “Diagnostic Needs” is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String, "The sheet “Diagnostic Needs” is not present or not written correctly", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0430_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0430 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                 reportWorkSheet2.Cells(row, i+1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0430: In the sheet “Diagnostic Needs”, the column XXXX (to be indicated) is not present or not written correctly.."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String, "In the sheet “Diagnostic Needs”, the column XXXX (to be indicated) is not present or not written correctly.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0440_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0440 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                 reportWorkSheet2.Cells(row, i+1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0440: The sheet “Customer Effects” is not present or not written correctly"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String, "The sheet “Customer Effects” is not present or not written correctly", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0450_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0450 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                 reportWorkSheet2.Cells(row, i+1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0450: In the sheet “Customer Effects”, the column XXXX (to be indicated) is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String, "In the sheet “Customer Effects”, the column XXXX (to be indicated) is not present or not written correctly.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0460_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0460 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                 reportWorkSheet2.Cells(row, i+1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0460: The sheet “Feared events” (or “ER”)  is not present or not written correctly"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String, "The sheet “Feared events” (or “ER”)  is not present or not written correctly", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0470_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0470 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                 reportWorkSheet2.Cells(row, i+1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0470: In the sheet “Feared events”, the column XXXX (to be indicated) is not present or not written correctly as in the document [DOC3].."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String, " In the sheet “Feared events”, the column XXXX (to be indicated) is not present or not written correctly as in the document [DOC3].", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0480_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0480 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                 reportWorkSheet2.Cells(row, i+1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0480: The sheet “System” (or “Système”)  is not present or not written correctly"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String, "The sheet “System” (or “Système”)  is not present or not written correctly", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0490_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0490 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                 reportWorkSheet2.Cells(row, i+1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0490: In the sheet “System” (or “Système”), the column XXXX (to be indicated) is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String, "In the sheet “System” (or “Système”), the column XXXX (to be indicated) is not present or not written correctly.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0500_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0500 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                 reportWorkSheet2.Cells(row, i+1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0500: The sheet “Operation situation” is not present or not written correctly"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String, "The sheet “Operation situation” is not present or not written correctly", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0510_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0510 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                 reportWorkSheet2.Cells(row, i+1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0510: In the sheet “Operation situation”, the column XXXX (to be indicated) is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String, "In the sheet “Operation situation”, the column XXXX (to be indicated) is not present or not written correctly. ", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0520_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0520 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                 reportWorkSheet2.Cells(row, i+1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0520: The sheet “Req. of tech. effects” is not present or not written correctly"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String, "The sheet “Req. of tech. effects” is not present or not written correctly", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0530_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0530 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                 reportWorkSheet2.Cells(row, i+1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0530: In the sheet “Req. of tech. effects”, the column XXXX (to be indicated) is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String, "In the sheet “Req. of tech. effects”, the column XXXX (to be indicated) is not present or not written correctly.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(
                    ["Good", "02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        elif testResult == 2:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060 NOT OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060",
                                      "The sheet “tableau” (or “table”) is not in the file", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            for key, value in testResult.items():
                if key == "0":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060: In the sheet “tableau” (or “table”), the column “Référence” is not completed."
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1000"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1000 ",
                                              "In the sheet “tableau” (or “table”), the column “Référence” is not completed.",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1000 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1000 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-1":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060: In the sheet “tableau” (or “table”), the column “Version” is not completed."
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1001"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1001 ",
                                              "In the sheet “tableau” (or “table”), the column “Version” is not completed.",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1001 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1001 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-2":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060: In the sheet “tableau” (or “table”), the columns “Constituant défaillant détecté” is not completed."
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1180"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1180 ",
                                              "In the sheet “tableau” (or “table”), the columns “Constituant défaillant détecté” is not completed",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1180 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1180 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-3":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060: In the sheet “tableau” (or “tableau”),  the columns “Défaillance constituant” is not completed"
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1190"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1190 ",
                                              "In the sheet “tableau” (or “tableau”),  the columns “Défaillance constituant” is not completed",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1190 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1190 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-4":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060: In the sheet “tableau” (or “tableau”), the columns “Situation de vie client” is not completed"
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1200"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1200 ",
                                              "In the sheet “tableau” (or “tableau”), the columns “Situation de vie client” is not completed ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1200 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1200 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-5":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060: In the sheet “tableau” (or “tableau”), the columns “Effet(s) client(s)” is not completed."
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1210"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1210 ",
                                              "In the sheet “tableau” (or “tableau”), the columns “Effet(s) client(s)” is not completed.",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1210 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1210 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-6":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060: In the sheet “tableau” (or “tableau”), the columns “Code défaut” is not completed"
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1220"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1220 ",
                                              "In the sheet “tableau” (or “tableau”), the columns “Code défaut” is not completed",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1220 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1220 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-7":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060: The project impact is not specified for the project in the sheet “tableau "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1060"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1060 ",
                                              "The project impact is not specified for the project in the sheet “tableau ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1060 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1060 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_WHOLENESS_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061_XLS(
            workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(
                    ["Good", "02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061", "",
                     ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        elif testResult == 2:
            row += 1
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061: NOT OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(
                    ["", "02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061",
                     "The sheet “codes défauts” is not present in the file", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            for key, value in testResult.items():
                if key == "0":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061:  In the sheet “codes défauts”, the column “Référence” is not completed "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1010"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1010",
                                              "In the sheet “codes défauts”, the column “Référence” is not completed. ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1010 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1010 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-1":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061: In the sheet “codes défauts”, the column “Version” is not completed. "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1011"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1011 ",
                                              "In the sheet “codes défauts”, the column “Version” is not completed. ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1011 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1011 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-2":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061:In the sheet “codes défauts”, the column “Code défaut” is not completed. "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1080"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1080 ",
                                              "In the sheet “codes défauts”, the column “Code défaut” is not completed. ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1080 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1080 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-3":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061:In the sheet “codes défauts”, the column “supporté par constituant (s)” is not completed. "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1090"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1090 ",
                                              "In the sheet “codes défauts”, the column “supporté par constituant (s)” is not completed. ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1090 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1090 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-4":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061:In the sheet “codes défauts”, the columns “libellé (signification)” is not completed "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1110"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1110 ",
                                              "In the sheet “codes défauts”, the columns “libellé (signification)” is not completed ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1110 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1110 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-5":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061:In the sheet “codes défauts”, the columns “Description de la strategie pour détecter le défaut” is not completed "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1120"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1120 ",
                                              "In the sheet “codes défauts”, the columns “Description de la strategie pour détecter le défaut” is not completed ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1120 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1120 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-6":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061:In the sheet “codes défauts”, the columns “Seuil de détection  /  valeur  du défaut” is not completed. "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1130"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1130 ",
                                              "In the sheet “codes défauts”, the columns “Seuil de détection  /  valeur  du défaut” is not completed.",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1130 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1130 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-7":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061:In the sheet “codes défauts”, the columns “Temps de confirmation du défaut” is not completed. "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1140"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1140 ",
                                              "In the sheet “codes défauts”, the columns “Temps de confirmation du défaut” is not completed. ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1140 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1140 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-8":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061:In the sheet “codes défauts”, the columns “Description de la strategie de disparition du défaut / Procedure à effectuer pour vérifier la disparition du défaut” is not completed "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1150"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1150 ",
                                              "In the sheet “codes défauts”, the columns “Description de la strategie de disparition du défaut / Procedure à effectuer pour vérifier la disparition du défaut” is not completed ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1150 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1150 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-9":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061:In the sheet “codes défauts”, the columns “Mode dégradé” is not completed. "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1160"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1160 ",
                                              "In the sheet “codes défauts”, the columns “Mode dégradé” is not completed. ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1160 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1160 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-10":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061:In the sheet “codes défauts”, the columns “Voyant” is not completed. "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1170"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1170 ",
                                              "In the sheet “codes défauts”, the columns “Voyant” is not completed. ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1170 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1170 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-11":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061: The project impact is not specified for the project in the sheet “codes défauts "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1061"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1061 ",
                                              " The project impact is not specified for the project in the sheet “codes défauts ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1061 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1061 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_WHOLENESS_1020_1021_1100_1062_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_WHOLENESS_1020_1021_1100_1062 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1020_1021_1100_1062 OK", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        elif testResult == 2:
            row += 1
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_WHOLENESS_1020_1021_1100_1062 NOT OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["", "02043_18_04939_WHOLENESS_1020_1021_1100_1062",
                                      "The sheet “mesures et commandes” is not in the file", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        else:
            for key, value in testResult.items():
                if key == "0":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1020_1021_1100_1062: In the sheet “mesures et commandes”, the column “Référence” is not completed"
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1020"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1020 ",
                                              "In the sheet “mesures et commandes”, the column “Référence” is not completed ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1020 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1020 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-1":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1020_1021_1100_1062: In the sheet “mesures et commandes”, the column “Version” is not completed."
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1021"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1021 ",
                                              "In the sheet “mesures et commandes”, the column “Version” is not completed.",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENES_1021 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1021 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-2":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1020_1021_1100_1062: In the sheet “mesures et commandes”, the column “supporté par constituant (s)” is not completed."
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1100"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1100 ",
                                              "In the sheet “mesures et commandes”, the column “supporté par constituant (s)” is not completed. ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1100 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1100 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-3":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1020_1021_1100_1062: The project impact is not specified for the project in the sheet “mesures et commandes”"
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1062"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1062 ",
                                              "The project impact is not specified for the project in the sheet “mesures et commandes” ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1062 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1062 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_WHOLENESS_1030_1031_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_WHOLENESS_1030_1031 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1030_1031", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        elif testResult == 2:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_WHOLENESS_1030_1031 NOT OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(
                    ["", "02043_18_04939_WHOLENESS_1030_1031", "The sheet “Diagnostic débarqués” is not in the file",
                     ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        else:
            for key, value in testResult.items():
                if key == "0":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nIn the sheet “Diagnostic débarqués”, the column “Référence” is not completed."
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1030"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1030 ",
                                              "Diagnostic débarqués”, the column “Référence” is not completed. ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1030 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "002043_18_04939_WHOLENESS_1030 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-1":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nIn the sheet “Diagnostic débarqués”, the column “Version” is not completed"
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1031"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1031",
                                              "In the sheet “Diagnostic débarqués”, the column “Version” is not completed.",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1031 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1031 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_WHOLENESS_1040_1041_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_WHOLENESS_1040_1041 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1040_1041", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        elif testResult == 2:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_WHOLENESS_1040_1041 NOT OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(
                    ["Good", "02043_18_04939_WHOLENESS_1040_1041", "The sheet “Liste MDD” is not in the file", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            for key, value in testResult.items():
                if key == "0":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1040_1041: In the sheet “Liste MDD”, the column “Référence” is not completed."
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1040"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1040",
                                              "In the sheet “Liste MDD”, the column “Référence” is not completed. ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1040 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1040 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-1":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1040_1041: In the sheet “Liste MDD”, the column “Version” is not completed."
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1041"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1041",
                                              "In the sheet “Liste MDD”, the column “Version” is not completed. ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1041 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1041 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_WHOLENESS_1050_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_1050 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_1050", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        elif testResult == 2:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_1050 NOT OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_1050",
                                      "The “tableau”(or “table”) or “codes défauts” is not presetnt in the file", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_1050: The projects listed in the following sheets “tableau” and “codes défauts” are not identical."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1050"][self.checkLevel],
                                      "02043_18_04939_WHOLENESS_1050 ",
                                      " The projects listed in the following sheets “tableau” and “codes défauts” are not identical. ",
                                      ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_WHOLENESS_1055_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_1055 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_1055", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        elif testResult == 2:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_1055 NOT OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_1055",
                                      "The “tableau”(or “table”) or “mesures et commandes” is not presetnt in the file",
                                      ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_1055: The projects listed in the following sheets “tableau” and “mesures et commandes” are not identical. "
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS _1055"][self.checkLevel],
                                      "02043_18_04939_WHOLENESS _1055",
                                      "The projects listed in the following sheets “tableau” and “mesures et commandes” are not identical. ",
                                      ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_WHOLENESS_1240_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_WHOLENESS_1240 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1240", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_WHOLENESS_1240: The reviewers have to be indicated in the paraph  “Liste de diffusion” or “Mailing list (the taking part)” of the sheet “Informations Générales” (or “General information”)."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1240"][self.checkLevel],
                                      "02043_18_04939_WHOLENESS_1240",
                                      "The reviewers have to be indicated in the paraph  “Liste de diffusion” or “Mailing list (the taking part)” of the sheet “Informations Générales” (or “General information”).",
                                      ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_COH_2070_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_COH_2070 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_COH_2070", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            if testResult == 2:
                text = self.tab1.textbox.toPlainText()
                text = text + "\nTest_02043_18_04939_COH_2070: One of the sheets “Customer Effects”, “FR” or “GB” are not in the file  "

            else:
                text = self.tab1.textbox.toPlainText()
                text = text + "\nTest_02043_18_04939_COH_2070: The information specified in the column “Name” of the “Customer Effects” sheet is missing in the document [DOC7] in French or English. Please check it or ask ADRD people to add this new customer effect."
                self.tab1.textbox.setText(text)
                flag = 0
                for i, name in enumerate([self.testReqDict["02043_18_04939_COH_2070"][self.checkLevel],
                                          "02043_18_04939_COH_2070",
                                          " The information specified in the column “Name” of the “Customer Effects” sheet is missing in the document [DOC7] in French or English. Please check it or ask ADRD people to add this new customer effect.",
                                          ""]):
                    reportWorkSheet2.Cells(row, i + 1).Value = name

        try:
            reportWorkBook.Save()
        except Exception as e:
            print(e)

    def TestGeneralStructure_DOC4_XLSX_XLSM(self, workBook):
        flag = 1
        fileName = self.tab1.myTextBox2.toPlainText()
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        reportWorkBook = excel.Workbooks.Open(fileName)
        reportWorkSheet2 = None

        for sheet in reportWorkBook.Worksheets:
            if sheet.Name == "Test Report":
                reportWorkSheet2 = sheet
        if not reportWorkSheet2:
            workSheetsNumber = reportWorkBook.Sheets.Count
            sheetAfter = reportWorkBook.Sheets(workSheetsNumber)
            reportWorkSheet2 = reportWorkBook.Worksheets.Add(None, sheetAfter)
            reportWorkSheet2.Name = "Test Report"

        testReportRow1StringList = ["Criticity", "Requirements", "Message", "Localisation"]

        for i, name in enumerate(testReportRow1StringList):
            reportWorkSheet2.Cells(1, i + 1).Value = name
        reportWorkSheet2.Columns.AutoFit()
        reportWorkSheet2.Columns.Font.Bold = True

        row = 21
        str1 = "02043_18_04939_STRUCT_0"
        stringInt = 40 - row
        str2 = str(stringInt + row)
        str3 = "0"
        String = str1 + str2 + str3

        testResult = self.Test_02043_18_04939_STRUCT_0400_XLSX_XLSM(workBook)
        if testResult == 1:
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0400: The sheet “tableau” is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String,
                                      "The sheet “Table” (or “tableau”) is not present or not written correctly.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0410_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0410 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0410: In the sheet “tableau” (or “tableau”), the column XXXX (to be indicated) is not present or not written correctly as in the document [DOC3].."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String,
                                      "In the sheet “tableau” (or “tableau”), the column XXXX (to be indicated) is not present or not written correctly as in the document [DOC3].",
                                      ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0420_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0420 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0420: The sheet “Diagnostic Needs” is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String,
                                      "The sheet “Diagnostic Needs” is not present or not written correctly", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0430_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0430 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0430: In the sheet “Diagnostic Needs”, the column XXXX (to be indicated) is not present or not written correctly.."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String,
                                      "In the sheet “Diagnostic Needs”, the column XXXX (to be indicated) is not present or not written correctly.",
                                      ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0440_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0440 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0440: The sheet “Customer Effects” is not present or not written correctly"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String,
                                      "The sheet “Customer Effects” is not present or not written correctly", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0450_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0450 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0450: In the sheet “Customer Effects”, the column XXXX (to be indicated) is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String,
                                      "In the sheet “Customer Effects”, the column XXXX (to be indicated) is not present or not written correctly.",
                                      ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0460_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0460 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0460: The sheet “Feared events” (or “ER”)  is not present or not written correctly"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String,
                                      "The sheet “Feared events” (or “ER”)  is not present or not written correctly",
                                      ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0470_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0470 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0470: In the sheet “Feared events”, the column XXXX (to be indicated) is not present or not written correctly as in the document [DOC3].."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String,
                                      " In the sheet “Feared events”, the column XXXX (to be indicated) is not present or not written correctly as in the document [DOC3].",
                                      ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0480_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0480 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0480: The sheet “System” (or “Système”)  is not present or not written correctly"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String,
                                      "The sheet “System” (or “Système”)  is not present or not written correctly",
                                      ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0490_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0490 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0490: In the sheet “System” (or “Système”), the column XXXX (to be indicated) is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String,
                                      "In the sheet “System” (or “Système”), the column XXXX (to be indicated) is not present or not written correctly.",
                                      ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0500_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0500 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0500: The sheet “Operation situation” is not present or not written correctly"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String,
                                      "The sheet “Operation situation” is not present or not written correctly", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0510_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0510 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0510: In the sheet “Operation situation”, the column XXXX (to be indicated) is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String,
                                      "In the sheet “Operation situation”, the column XXXX (to be indicated) is not present or not written correctly. ",
                                      ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0520_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0520 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0520: The sheet “Req. of tech. effects” is not present or not written correctly"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String,
                                      "The sheet “Req. of tech. effects” is not present or not written correctly", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0530_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0530 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0530: In the sheet “Req. of tech. effects”, the column XXXX (to be indicated) is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String,
                                      "In the sheet “Req. of tech. effects”, the column XXXX (to be indicated) is not present or not written correctly.",
                                      ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(
                    ["Good", "02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        elif testResult == 2:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060 NOT OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060",
                                      "The sheet “tableau” (or “table”) is not in the file", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            for key, value in testResult.items():
                if key == "0":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060: In the sheet “tableau” (or “table”), the column “Référence” is not completed."
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1000"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1000 ",
                                              "In the sheet “tableau” (or “table”), the column “Référence” is not completed.",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1000 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1000 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-1":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060: In the sheet “tableau” (or “table”), the column “Version” is not completed."
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1001"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1001 ",
                                              "In the sheet “tableau” (or “table”), the column “Version” is not completed.",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1001 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1001 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-2":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060: In the sheet “tableau” (or “table”), the columns “Constituant défaillant détecté” is not completed."
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1180"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1180 ",
                                              "In the sheet “tableau” (or “table”), the columns “Constituant défaillant détecté” is not completed",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1180 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1180 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-3":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060: In the sheet “tableau” (or “tableau”),  the columns “Défaillance constituant” is not completed"
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1190"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1190 ",
                                              "In the sheet “tableau” (or “tableau”),  the columns “Défaillance constituant” is not completed",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1190 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1190 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-4":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060: In the sheet “tableau” (or “tableau”), the columns “Situation de vie client” is not completed"
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1200"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1200 ",
                                              "In the sheet “tableau” (or “tableau”), the columns “Situation de vie client” is not completed ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1200 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1200 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-5":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060: In the sheet “tableau” (or “tableau”), the columns “Effet(s) client(s)” is not completed."
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1210"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1210 ",
                                              "In the sheet “tableau” (or “tableau”), the columns “Effet(s) client(s)” is not completed.",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1210 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1210 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-6":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060: In the sheet “tableau” (or “tableau”), the columns “Code défaut” is not completed"
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1220"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1220 ",
                                              "In the sheet “tableau” (or “tableau”), the columns “Code défaut” is not completed",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1220 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1220 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-7":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060: The project impact is not specified for the project in the sheet “tableau "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1060"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1060 ",
                                              "The project impact is not specified for the project in the sheet “tableau ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1060 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1060 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_WHOLENESS_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061_XLSX_XLSM(
            workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(
                    ["Good", "02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061", "",
                     ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        elif testResult == 2:
            row += 1
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061: NOT OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(
                    ["", "02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061",
                     "The sheet “codes défauts” is not present in the file", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            for key, value in testResult.items():
                if key == "0":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061:  In the sheet “codes défauts”, the column “Référence” is not completed "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1010"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1010",
                                              "In the sheet “codes défauts”, the column “Référence” is not completed. ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1010 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1010 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-1":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061: In the sheet “codes défauts”, the column “Version” is not completed. "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1011"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1011 ",
                                              "In the sheet “codes défauts”, the column “Version” is not completed. ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1011 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1011 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-2":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061:In the sheet “codes défauts”, the column “Code défaut” is not completed. "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1080"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1080 ",
                                              "In the sheet “codes défauts”, the column “Code défaut” is not completed. ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1080 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1080 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-3":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061:In the sheet “codes défauts”, the column “supporté par constituant (s)” is not completed. "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1090"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1090 ",
                                              "In the sheet “codes défauts”, the column “supporté par constituant (s)” is not completed. ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1090 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1090 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-4":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061:In the sheet “codes défauts”, the columns “libellé (signification)” is not completed "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1110"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1110 ",
                                              "In the sheet “codes défauts”, the columns “libellé (signification)” is not completed ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1110 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1110 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-5":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061:In the sheet “codes défauts”, the columns “Description de la strategie pour détecter le défaut” is not completed "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1120"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1120 ",
                                              "In the sheet “codes défauts”, the columns “Description de la strategie pour détecter le défaut” is not completed ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1120 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1120 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-6":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061:In the sheet “codes défauts”, the columns “Seuil de détection  /  valeur  du défaut” is not completed. "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1130"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1130 ",
                                              "In the sheet “codes défauts”, the columns “Seuil de détection  /  valeur  du défaut” is not completed.",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1130 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1130 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-7":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061:In the sheet “codes défauts”, the columns “Temps de confirmation du défaut” is not completed. "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1140"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1140 ",
                                              "In the sheet “codes défauts”, the columns “Temps de confirmation du défaut” is not completed. ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1140 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1140 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-8":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061:In the sheet “codes défauts”, the columns “Description de la strategie de disparition du défaut / Procedure à effectuer pour vérifier la disparition du défaut” is not completed "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1150"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1150 ",
                                              "In the sheet “codes défauts”, the columns “Description de la strategie de disparition du défaut / Procedure à effectuer pour vérifier la disparition du défaut” is not completed ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1150 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1150 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-9":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061:In the sheet “codes défauts”, the columns “Mode dégradé” is not completed. "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1160"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1160 ",
                                              "In the sheet “codes défauts”, the columns “Mode dégradé” is not completed. ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1160 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1160 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-10":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061:In the sheet “codes défauts”, the columns “Voyant” is not completed. "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1170"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1170 ",
                                              "In the sheet “codes défauts”, the columns “Voyant” is not completed. ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1170 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1170 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-11":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061: The project impact is not specified for the project in the sheet “codes défauts "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1061"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1061 ",
                                              " The project impact is not specified for the project in the sheet “codes défauts ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1061 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1061 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_WHOLENESS_1020_1021_1100_1062_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_WHOLENESS_1020_1021_1100_1062 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1020_1021_1100_1062 OK", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        elif testResult == 2:
            row += 1
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_WHOLENESS_1020_1021_1100_1062 NOT OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["", "02043_18_04939_WHOLENESS_1020_1021_1100_1062",
                                      "The sheet “mesures et commandes” is not in the file", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        else:
            for key, value in testResult.items():
                if key == "0":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1020_1021_1100_1062: In the sheet “mesures et commandes”, the column “Référence” is not completed"
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1020"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1020 ",
                                              "In the sheet “mesures et commandes”, the column “Référence” is not completed ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1020 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1020 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-1":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1020_1021_1100_1062: In the sheet “mesures et commandes”, the column “Version” is not completed."
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1021"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1021 ",
                                              "In the sheet “mesures et commandes”, the column “Version” is not completed.",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENES_1021 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1021 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-2":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1020_1021_1100_1062: In the sheet “mesures et commandes”, the column “supporté par constituant (s)” is not completed."
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1100"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1100 ",
                                              "In the sheet “mesures et commandes”, the column “supporté par constituant (s)” is not completed. ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1100 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1100 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-3":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1020_1021_1100_1062: The project impact is not specified for the project in the sheet “mesures et commandes”"
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1062"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1062 ",
                                              "The project impact is not specified for the project in the sheet “mesures et commandes” ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1062 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1062 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_WHOLENESS_1030_1031_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_WHOLENESS_1030_1031 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1030_1031", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        elif testResult == 2:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_WHOLENESS_1030_1031 NOT OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(
                    ["", "02043_18_04939_WHOLENESS_1030_1031", "The sheet “Diagnostic débarqués” is not in the file",
                     ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        else:
            for key, value in testResult.items():
                if key == "0":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nIn the sheet “Diagnostic débarqués”, the column “Référence” is not completed."
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1030"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1030 ",
                                              "Diagnostic débarqués”, the column “Référence” is not completed. ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1030 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "002043_18_04939_WHOLENESS_1030 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-1":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nIn the sheet “Diagnostic débarqués”, the column “Version” is not completed"
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1031"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1031",
                                              "In the sheet “Diagnostic débarqués”, the column “Version” is not completed.",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1031 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1031 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_WHOLENESS_1040_1041_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_WHOLENESS_1040_1041 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1040_1041", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        elif testResult == 2:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_WHOLENESS_1040_1041 NOT OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(
                    ["Good", "02043_18_04939_WHOLENESS_1040_1041", "The sheet “Liste MDD” is not in the file", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            for key, value in testResult.items():
                if key == "0":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1040_1041: In the sheet “Liste MDD”, the column “Référence” is not completed."
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1040"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1040",
                                              "In the sheet “Liste MDD”, the column “Référence” is not completed. ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1040 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1040 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-1":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1040_1041: In the sheet “Liste MDD”, the column “Version” is not completed."
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1041"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1041",
                                              "In the sheet “Liste MDD”, the column “Version” is not completed. ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1041 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1041 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_WHOLENESS_1050_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_1050 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_1050", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        elif testResult == 2:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_1050 NOT OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_1050",
                                      "The “tableau”(or “table”) or “codes défauts” is not presetnt in the file", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_1050: The projects listed in the following sheets “tableau” and “codes défauts” are not identical."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1050"][self.checkLevel],
                                      "02043_18_04939_WHOLENESS_1050 ",
                                      " The projects listed in the following sheets “tableau” and “codes défauts” are not identical. ",
                                      ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_WHOLENESS_1055_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_1055 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_1055", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        elif testResult == 2:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_1055 NOT OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_1055",
                                      "The “tableau”(or “table”) or “mesures et commandes” is not presetnt in the file",
                                      ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_1055: The projects listed in the following sheets “tableau” and “mesures et commandes” are not identical. "
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS _1055"][self.checkLevel],
                                      "02043_18_04939_WHOLENESS _1055",
                                      "The projects listed in the following sheets “tableau” and “mesures et commandes” are not identical. ",
                                      ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_WHOLENESS_1240_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_WHOLENESS_1240 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1240", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_WHOLENESS_1240: The reviewers have to be indicated in the paraph  “Liste de diffusion” or “Mailing list (the taking part)” of the sheet “Informations Générales” (or “General information”)."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1240"][self.checkLevel],
                                      "02043_18_04939_WHOLENESS_1240",
                                      "The reviewers have to be indicated in the paraph  “Liste de diffusion” or “Mailing list (the taking part)” of the sheet “Informations Générales” (or “General information”).",
                                      ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_COH_2070_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_COH_2070 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_COH_2070", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            if testResult == 2:
                text = self.tab1.textbox.toPlainText()
                text = text + "\nTest_02043_18_04939_COH_2070: One of the sheets “Customer Effects”, “FR” or “GB” are not in the file  "

            else:
                text = self.tab1.textbox.toPlainText()
                text = text + "\nTest_02043_18_04939_COH_2070: The information specified in the column “Name” of the “Customer Effects” sheet is missing in the document [DOC7] in French or English. Please check it or ask ADRD people to add this new customer effect."
                self.tab1.textbox.setText(text)
                flag = 0
                for i, name in enumerate([self.testReqDict["02043_18_04939_COH_2070"][self.checkLevel],
                                          "02043_18_04939_COH_2070",
                                          " The information specified in the column “Name” of the “Customer Effects” sheet is missing in the document [DOC7] in French or English. Please check it or ask ADRD people to add this new customer effect.",
                                          ""]):
                    reportWorkSheet2.Cells(row, i + 1).Value = name

        try:
            reportWorkBook.Save()
        except Exception as e:
            print(e)

    def TestGeneralStructure_DOC5_XLS(self, workBook):

        flag = 1
        fileName = self.tab1.myTextBox3.toPlainText()
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        reportWorkBook = excel.Workbooks.Open(fileName)
        reportWorkSheet2 = None

        for sheet in reportWorkBook.Worksheets:
            if sheet.Name == "Test Report":
                reportWorkSheet2 = sheet
        if not reportWorkSheet2:
            workSheetsNumber = reportWorkBook.Sheets.Count
            sheetAfter = reportWorkBook.Sheets(workSheetsNumber)
            reportWorkSheet2 = reportWorkBook.Worksheets.Add(None, sheetAfter)
            reportWorkSheet2.Name = "Test Report"

        testReportRow1StringList = ["Criticity", "Requirements", "Message", "Localisation"]

        for i, name in enumerate(testReportRow1StringList):
            reportWorkSheet2.Cells(1, i + 1).Value = name
        reportWorkSheet2.Columns.AutoFit()
        reportWorkSheet2.Columns.Font.Bold = True


        row = 21
        str1 = "02043_18_04939_STRUCT_0"
        stringInt = 70 - row
        str2 = str(stringInt + row)
        str3 = "0"
        String = str1 + str2 + str3

        testResult = self.Test_02043_18_04939_STRUCT_0700_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0700 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                 reportWorkSheet2.Cells(row, i+1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0400: The sheet “Table” (or “tableau”) is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String, "The sheet “Table” (or “tableau”) is not present or not written correctly.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0710_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0710 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                 reportWorkSheet2.Cells(row, i+1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0710: In the sheet “tableau” (or “tableau”), the column XXXX (to be indicated) is not present or not written correctly as in the document [DOC3]."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String, "In the sheet “tableau” (or “tableau”), the column XXXX (to be indicated) is not present or not written correctly as in the document [DOC3]. ", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name


        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0720_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0720 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                 reportWorkSheet2.Cells(row, i+1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0720:  The sheet “Data trouble codes” (or “codes défauts”) is not present or not written correctly"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String, " The sheet “Data trouble codes” (or “codes défauts”) is not present or not written correctly", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name


        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0730_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0730 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                 reportWorkSheet2.Cells(row, i+1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0730: In the sheet “Data trouble codes” (or “codes défauts”), the column XXXX (to be indicated) is not present or not written correctly as in the document [DOC3]."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String, "In the sheet “Data trouble codes” (or “codes défauts”), the column XXXX (to be indicated) is not present or not written correctly as in the document [DOC3].", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name


        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0740_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0740 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                 reportWorkSheet2.Cells(row, i+1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0740: The sheet “Read data and IO control” is not present or not written correctly"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String, "The sheet “Read data and IO control” is not present or not written correctly", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name


        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0750_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0750 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                 reportWorkSheet2.Cells(row, i+1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0750: In the sheet “Read data and IO control” (or “mesures et commandes”), the column XXXX (to be indicated) is not present or not written correctly as in the document [DOC3]. "
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String, " In the sheet “Read data and IO control” (or “mesures et commandes”), the column XXXX (to be indicated) is not present or not written correctly as in the document [DOC3]. ", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name


        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0760_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0760 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                 reportWorkSheet2.Cells(row, i+1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0760: The sheet “Not embedded diagnosis”  is not present or not written correctly"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String, " The sheet “Not embedded diagnosis”  is not present or not written correctly", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name


        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0770_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0770 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                 reportWorkSheet2.Cells(row, i+1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0770: In the sheet “Not embedded diagnosis” (or “Read data and IO control”), the column XXXX (to be indicated) is not present or not written correctly as in the document [DOC3]."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String, "In the sheet “Not embedded diagnosis” (or “Read data and IO control”), the column XXXX (to be indicated) is not present or not written correctly as in the document [DOC3].", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name


        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0780_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0780 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                 reportWorkSheet2.Cells(row, i+1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0780: The sheet “Customer effect” (or “Effets clients”) is not present or not written correctly"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String, "The sheet “Customer effect” (or “Effets clients”) is not present or not written correctly", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name


        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0790_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0790 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                 reportWorkSheet2.Cells(row, i+1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0790: In the sheet “Customer effect” (or “Effets clients”), the column XXXX (to be indicated) is not present or not written correctly as in the document [DOC3]"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String, "In the sheet “Customer effect” (or “Effets clients”), the column XXXX (to be indicated) is not present or not written correctly as in the document [DOC3].", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name


        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0800_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0800 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                 reportWorkSheet2.Cells(row, i+1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0800: The sheet “Feared events” (or “ER”)  is not present or not written correctly"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String, "The sheet “Feared events” (or “ER”)  is not present or not written correctly", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name


        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0810_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0810 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                 reportWorkSheet2.Cells(row, i+1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0810: In the sheet “Feared events”, the column XXXX (to be indicated) is not present or not written correctly as in the document [DOC3].."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String, "In the sheet “Feared events”, the column XXXX (to be indicated) is not present or not written correctly as in the document [DOC3].", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name


        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0820_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0820 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                 reportWorkSheet2.Cells(row, i+1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0820: The sheet “Parts” (or “Constituants”) is not present or not written correctly"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String, "The sheet “Parts” (or “Constituants”) is not present or not written correctly", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name


        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0830_XLS(workBook)
        if testResult == 1:
            for i, name in enumerate(["Good", String, "", ""]):
                 reportWorkSheet2.Cells(row, i+1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0830:  In the sheet “Parts” (or “Constituants”), the column XXXX (to be indicated) is not present or not written correctly as in the document [DOC3]."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String, " In the sheet “Parts” (or “Constituants”), the column XXXX (to be indicated) is not present or not written correctly as in the document [DOC3]", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0840_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0840 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                 reportWorkSheet2.Cells(row, i+1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0840: The sheet “Situation” (or “situations de vie”) is not present or not written correctly"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String, "The sheet “Situation” (or “situations de vie”) is not present or not written correctly", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name


        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0850_XLS(workBook)
        if testResult == 1:
            for i, name in enumerate(["Good", String, "", ""]):
                 reportWorkSheet2.Cells(row, i+1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0850: In the sheet “Situation” (or “situations de vie”), the column XXXX (to be indicated) is not present or not written correctly as in the document [DOC3]. "
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String, "In the sheet “Situation” (or “situations de vie”), the column XXXX (to be indicated) is not present or not written correctly as in the document [DOC3]. ", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name


        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0860_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0860 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                 reportWorkSheet2.Cells(row, i+1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0860: The sheet “Degraded mode” (or “Liste MDD”)  is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String, "The sheet “Degraded mode” (or “Liste MDD”)  is not present or not written correctly.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name


        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0870_XLS(workBook)
        if testResult == 1:
            for i, name in enumerate(["Good", String, "", ""]):
                 reportWorkSheet2.Cells(row, i+1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0870:In the sheet “Degraded mode” (or “Liste MDD”), the column XXXX (to be indicated) is not present or not written correctly as in the document [DOC3].."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String, "In the sheet “Degraded mode” (or “Liste MDD”), the column XXXX (to be indicated) is not present or not written correctly as in the document [DOC3].", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name


        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0880_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0880 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                 reportWorkSheet2.Cells(row, i+1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0880: The sheet “Technical effect “ (or “Effets techniques”) is not present or not written correctly"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String, "The sheet “Technical effect “ (or “Effets techniques”) is not present or not written correctly”.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name


        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0890_XLS(workBook)
        if testResult == 1:
            for i, name in enumerate(["Good", String, "", ""]):
                 reportWorkSheet2.Cells(row, i+1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0890: In the sheet “Technical effect “ (or “Effets techniques”), the column XXXX (to be indicated) is not present or not written correctly as in the document [DOC3]."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String, " In the sheet “Technical effect “ (or “Effets techniques”), the column XXXX (to be indicated) is not present or not written correctly as in the document [DOC3].", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name


        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0900_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0900 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                 reportWorkSheet2.Cells(row, i+1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0900: The sheet “Variant “ (or “Variantes”) is not present or not written correctly.."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String, "The sheet “Variant “ (or “Variantes”) is not present or not written correctly.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name


        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0910_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0910 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                 reportWorkSheet2.Cells(row, i+1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0910: In the sheet “Variant “ (or “Variantes”), the column XXXX (to be indicated) is not present or not written correctly as in the document [DOC3]."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String, "In the sheet “Variant “ (or “Variantes”), the column XXXX (to be indicated) is not present or not written correctly as in the document [DOC3]. ", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060", "", ""]):
               reportWorkSheet2.Cells(row, i + 1).Value = name
        elif testResult == 2:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060 NOT OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060", "The sheet “tableau” (or “table”) is not in the file", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            for key, value in testResult.items():
                if key == "0":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060: In the sheet “tableau” (or “table”), the column “Référence” is not completed."
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1000"][self.checkLevel],"02043_18_04939_WHOLENESS_1000 ","In the sheet “tableau” (or “table”), the column “Référence” is not completed.", str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row +=1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1000 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1000 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-1":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060: In the sheet “tableau” (or “table”), the column “Version” is not completed."
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1001"][self.checkLevel],"02043_18_04939_WHOLENESS_1001 ","In the sheet “tableau” (or “table”), the column “Version” is not completed.", str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row +=1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1001 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1001 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-2":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060: In the sheet “tableau” (or “table”), the columns “Constituant défaillant détecté” is not completed."
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1180"][self.checkLevel], "02043_18_04939_WHOLENESS_1180 ", "In the sheet “tableau” (or “table”), the columns “Constituant défaillant détecté” is not completed", str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row +=1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1180 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1180 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-3":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060: In the sheet “tableau” (or “tableau”),  the columns “Défaillance constituant” is not completed"
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1190"][self.checkLevel],  "02043_18_04939_WHOLENESS_1190 ",  "In the sheet “tableau” (or “tableau”),  the columns “Défaillance constituant” is not completed", str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row +=1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1190 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1190 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-4":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060: In the sheet “tableau” (or “tableau”), the columns “Situation de vie client” is not completed"
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1200"][self.checkLevel], "02043_18_04939_WHOLENESS_1200 ", "In the sheet “tableau” (or “tableau”), the columns “Situation de vie client” is not completed ",  str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row +=1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1200 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1200 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-5":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060: In the sheet “tableau” (or “tableau”), the columns “Effet(s) client(s)” is not completed."
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1210"][self.checkLevel], "02043_18_04939_WHOLENESS_1210 ", "In the sheet “tableau” (or “tableau”), the columns “Effet(s) client(s)” is not completed.",  str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row +=1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1210 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1210 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-6":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060: In the sheet “tableau” (or “tableau”), the columns “Code défaut” is not completed"
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1220"][self.checkLevel], "02043_18_04939_WHOLENESS_1220 ", "In the sheet “tableau” (or “tableau”), the columns “Code défaut” is not completed",  str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row +=1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1220 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1220 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-7":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060: The project impact is not specified for the project in the sheet “tableau "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1060"][self.checkLevel], "02043_18_04939_WHOLENESS_1060 ", "The project impact is not specified for the project in the sheet “tableau ",  str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row +=1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1060 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1060 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_WHOLENESS_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        elif testResult == 2:
            row += 1
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061: NOT OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["", "02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061", "The sheet “codes défauts” is not present in the file",  ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            for key, value in testResult.items():
                if key == "0":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061:  In the sheet “codes défauts”, the column “Référence” is not completed "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1010"][self.checkLevel], "02043_18_04939_WHOLENESS_1010", "In the sheet “codes défauts”, the column “Référence” is not completed. ",  str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row +=1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1010 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1010 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-1":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061: In the sheet “codes défauts”, the column “Version” is not completed. "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1011"][self.checkLevel], "02043_18_04939_WHOLENESS_1011 ", "In the sheet “codes défauts”, the column “Version” is not completed. ",  str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row +=1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1011 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1011 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-2":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061:In the sheet “codes défauts”, the column “Code défaut” is not completed. "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1080"][self.checkLevel], "02043_18_04939_WHOLENESS_1080 ", "In the sheet “codes défauts”, the column “Code défaut” is not completed. ",  str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row +=1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1080 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1080 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-3":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061:In the sheet “codes défauts”, the column “supporté par constituant (s)” is not completed. "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1090"][self.checkLevel], "02043_18_04939_WHOLENESS_1090 ", "In the sheet “codes défauts”, the column “supporté par constituant (s)” is not completed. ",  str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row +=1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1090 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1090 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-4":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061:In the sheet “codes défauts”, the columns “libellé (signification)” is not completed "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1110"][self.checkLevel], "02043_18_04939_WHOLENESS_1110 ", "In the sheet “codes défauts”, the columns “libellé (signification)” is not completed ",  str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row +=1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1110 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1110 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-5":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061:In the sheet “codes défauts”, the columns “Description de la strategie pour détecter le défaut” is not completed "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1120"][self.checkLevel], "02043_18_04939_WHOLENESS_1120 ", "In the sheet “codes défauts”, the columns “Description de la strategie pour détecter le défaut” is not completed ",  str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row +=1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1120 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1120 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-6":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061:In the sheet “codes défauts”, the columns “Seuil de détection  /  valeur  du défaut” is not completed. "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1130"][self.checkLevel], "02043_18_04939_WHOLENESS_1130 ", "In the sheet “codes défauts”, the columns “Seuil de détection  /  valeur  du défaut” is not completed.",  str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row +=1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1130 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1130 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-7":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061:In the sheet “codes défauts”, the columns “Temps de confirmation du défaut” is not completed. "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1140"][self.checkLevel], "02043_18_04939_WHOLENESS_1140 ", "In the sheet “codes défauts”, the columns “Temps de confirmation du défaut” is not completed. ",  str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row +=1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1140 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1140 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-8":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061:In the sheet “codes défauts”, the columns “Description de la strategie de disparition du défaut / Procedure à effectuer pour vérifier la disparition du défaut” is not completed "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1150"][self.checkLevel], "02043_18_04939_WHOLENESS_1150 ", "In the sheet “codes défauts”, the columns “Description de la strategie de disparition du défaut / Procedure à effectuer pour vérifier la disparition du défaut” is not completed ", str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row +=1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1150 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1150 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-9":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061:In the sheet “codes défauts”, the columns “Mode dégradé” is not completed. "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1160"][self.checkLevel], "02043_18_04939_WHOLENESS_1160 ", "In the sheet “codes défauts”, the columns “Mode dégradé” is not completed. ",  str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row +=1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1160 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1160 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-10":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061:In the sheet “codes défauts”, the columns “Voyant” is not completed. "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1170"][self.checkLevel], "02043_18_04939_WHOLENESS_1170 ", "In the sheet “codes défauts”, the columns “Voyant” is not completed. ",  str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row +=1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1170 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1170 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-11":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061: The project impact is not specified for the project in the sheet “codes défauts "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1061"][self.checkLevel], "02043_18_04939_WHOLENESS_1061 ", " The project impact is not specified for the project in the sheet “codes défauts ",  str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row +=1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1061 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1061 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_WHOLENESS_1020_1021_1100_1062_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_WHOLENESS_1020_1021_1100_1062 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1020_1021_1100_1062 OK", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        elif testResult == 2:
            row += 1
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_WHOLENESS_1020_1021_1100_1062 NOT OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["", "02043_18_04939_WHOLENESS_1020_1021_1100_1062", "The sheet “mesures et commandes” is not in the file", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        else:
            for key, value in testResult.items():
                if key == "0":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1020_1021_1100_1062: In the sheet “mesures et commandes”, the column “Référence” is not completed"
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1020"][self.checkLevel], "02043_18_04939_WHOLENESS_1020 ", "In the sheet “mesures et commandes”, the column “Référence” is not completed ", str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row +=1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1020 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1020 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-1":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1020_1021_1100_1062: In the sheet “mesures et commandes”, the column “Version” is not completed."
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1021"][self.checkLevel], "02043_18_04939_WHOLENESS_1021 ", "In the sheet “mesures et commandes”, the column “Version” is not completed.", str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row +=1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENES_1021 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1021 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-2":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1020_1021_1100_1062: In the sheet “mesures et commandes”, the column “supporté par constituant (s)” is not completed."
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1100"][self.checkLevel], "02043_18_04939_WHOLENESS_1100 ", "In the sheet “mesures et commandes”, the column “supporté par constituant (s)” is not completed. ", str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row +=1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1100 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1100 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-3":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1020_1021_1100_1062: The project impact is not specified for the project in the sheet “mesures et commandes”"
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1062"][self.checkLevel], "02043_18_04939_WHOLENESS_1062 ", "The project impact is not specified for the project in the sheet “mesures et commandes” ", str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row +=1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1062 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1062 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_WHOLENESS_1030_1031_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_WHOLENESS_1030_1031 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1030_1031", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        elif testResult == 2:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_WHOLENESS_1030_1031 NOT OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["", "02043_18_04939_WHOLENESS_1030_1031", "The sheet “Diagnostic débarqués” is not in the file", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        else:
            for key, value in testResult.items():
                if key == "0":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nIn the sheet “Diagnostic débarqués”, the column “Référence” is not completed."
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1030"][self.checkLevel], "02043_18_04939_WHOLENESS_1030 ", "Diagnostic débarqués”, the column “Référence” is not completed. ", str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row +=1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1030 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "002043_18_04939_WHOLENESS_1030 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-1":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nIn the sheet “Diagnostic débarqués”, the column “Version” is not completed"
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1031"][self.checkLevel], "02043_18_04939_WHOLENESS_1031", "In the sheet “Diagnostic débarqués”, the column “Version” is not completed.", str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row +=1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1031 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1031 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_WHOLENESS_1040_1041_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_WHOLENESS_1040_1041 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1040_1041", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        elif testResult == 2:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_WHOLENESS_1040_1041 NOT OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1040_1041", "The sheet “Liste MDD” is not in the file", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            for key, value in testResult.items():
                if key == "0":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1040_1041: In the sheet “Liste MDD”, the column “Référence” is not completed."
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1040"][self.checkLevel], "02043_18_04939_WHOLENESS_1040", "In the sheet “Liste MDD”, the column “Référence” is not completed. ", str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row +=1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1040 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1040 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-1":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1040_1041: In the sheet “Liste MDD”, the column “Version” is not completed."
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1041"][self.checkLevel], "02043_18_04939_WHOLENESS_1041", "In the sheet “Liste MDD”, the column “Version” is not completed. ", str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row +=1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1041 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1041 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_WHOLENESS_1050_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_1050 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_1050", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        elif testResult == 2:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_1050 NOT OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_1050", "The “tableau”(or “table”) or “codes défauts” is not presetnt in the file", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_1050: The projects listed in the following sheets “tableau” and “codes défauts” are not identical."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1050"][self.checkLevel], "02043_18_04939_WHOLENESS_1050 ", " The projects listed in the following sheets “tableau” and “codes défauts” are not identical. ", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_WHOLENESS_1055_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_1055 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_1055", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        elif testResult == 2:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_1055 NOT OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_1055", "The “tableau”(or “table”) or “mesures et commandes” is not presetnt in the file", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_1055: The projects listed in the following sheets “tableau” and “mesures et commandes” are not identical. "
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS _1055"][self.checkLevel],"02043_18_04939_WHOLENESS _1055",  "The projects listed in the following sheets “tableau” and “mesures et commandes” are not identical. ", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_WHOLENESS_1240_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_WHOLENESS_1240 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1240", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_WHOLENESS_1240: The reviewers have to be indicated in the paraph  “Liste de diffusion” or “Mailing list (the taking part)” of the sheet “Informations Générales” (or “General information”)."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1240"][self.checkLevel],
                                      "02043_18_04939_WHOLENESS_1240",
                                      "The reviewers have to be indicated in the paraph  “Liste de diffusion” or “Mailing list (the taking part)” of the sheet “Informations Générales” (or “General information”).",
                                      ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_COH_2080_XLS(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_COH_2080 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_COH_2080", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            if testResult == 2:
                text = self.tab1.textbox.toPlainText()
                text = text + "\nTest_02043_18_04939_COH_2080: One of the sheets “Effets clients”, “FR” or “GB” are not in the file  "

            else:
                text = self.tab1.textbox.toPlainText()
                text = text + "\nTest_02043_18_04939_COH_2080: The information specified in the column “Noms” of the “Effets clients” sheet is missing in the document [DOC7] in French or English. Please check it or ask ADRD people to add this new customer effect."
                self.tab1.textbox.setText(text)
                flag = 0
                for i, name in enumerate([self.testReqDict["02043_18_04939_COH_2080"][self.checkLevel],
                                          "02043_18_04939_COH_2080",
                                          " The information specified in the column “Noms” of the “Effets clients” sheet is missing in the document [DOC7] in French or English. Please check it or ask ADRD people to add this new customer effect.",
                                          ""]):
                    reportWorkSheet2.Cells(row, i + 1).Value = name

        reportWorkBook.Save()

    def TestGeneralStructure_DOC5_XLSX_XLSM(self, workBook):


        flag = 1
        fileName = self.tab1.myTextBox3.toPlainText()
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        reportWorkBook = excel.Workbooks.Open(fileName)
        reportWorkSheet2 = None

        for sheet in reportWorkBook.Worksheets:
            if sheet.Name == "Test Report":
                reportWorkSheet2 = sheet
        if not reportWorkSheet2:
            workSheetsNumber = reportWorkBook.Sheets.Count
            sheetAfter = reportWorkBook.Sheets(workSheetsNumber)
            reportWorkSheet2 = reportWorkBook.Worksheets.Add(None, sheetAfter)
            reportWorkSheet2.Name = "Test Report"

        testReportRow1StringList = ["Criticity", "Requirements", "Message", "Localisation"]

        for i, name in enumerate(testReportRow1StringList):
            reportWorkSheet2.Cells(1, i + 1).Value = name
        reportWorkSheet2.Columns.AutoFit()
        reportWorkSheet2.Columns.Font.Bold = True


        row = 21
        str1 = "02043_18_04939_STRUCT_0"
        stringInt = 70 - row
        str2 = str(stringInt + row)
        str3 = "0"
        String = str1 + str2 + str3


        testResult = self.Test_02043_18_04939_STRUCT_0700_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0700 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0400: The sheet “Table” (or “tableau”) is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String,
                                      "The sheet “Table” (or “tableau”) is not present or not written correctly.", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name



        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0710_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0710 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0710: In the sheet “tableau” (or “tableau”), the column XXXX (to be indicated) is not present or not written correctly as in the document [DOC3]."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String,
                                      "In the sheet “tableau” (or “tableau”), the column XXXX (to be indicated) is not present or not written correctly as in the document [DOC3]. ",
                                      ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0720_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0720 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0720:  The sheet “Data trouble codes” (or “codes défauts”) is not present or not written correctly"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String,
                                      " The sheet “Data trouble codes” (or “codes défauts”) is not present or not written correctly",
                                      ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0730_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0730 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0730: In the sheet “Data trouble codes” (or “codes défauts”), the column XXXX (to be indicated) is not present or not written correctly as in the document [DOC3]."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String,
                                      "In the sheet “Data trouble codes” (or “codes défauts”), the column XXXX (to be indicated) is not present or not written correctly as in the document [DOC3].",
                                      ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0740_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0740 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0740: The sheet “Read data and IO control” is not present or not written correctly"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String,
                                      "The sheet “Read data and IO control” is not present or not written correctly",
                                      ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0750_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0750 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0750: In the sheet “Read data and IO control” (or “mesures et commandes”), the column XXXX (to be indicated) is not present or not written correctly as in the document [DOC3]. "
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String,
                                      " In the sheet “Read data and IO control” (or “mesures et commandes”), the column XXXX (to be indicated) is not present or not written correctly as in the document [DOC3]. ",
                                      ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0760_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0760 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0760: The sheet “Not embedded diagnosis”  is not present or not written correctly"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String,
                                      " The sheet “Not embedded diagnosis”  is not present or not written correctly",
                                      ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0770_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0770 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0770: In the sheet “Not embedded diagnosis” (or “Read data and IO control”), the column XXXX (to be indicated) is not present or not written correctly as in the document [DOC3]."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String,
                                      "In the sheet “Not embedded diagnosis” (or “Read data and IO control”), the column XXXX (to be indicated) is not present or not written correctly as in the document [DOC3].",
                                      ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0780_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0780 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0780: The sheet “Customer effect” (or “Effets clients”) is not present or not written correctly"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String,
                                      "The sheet “Customer effect” (or “Effets clients”) is not present or not written correctly",
                                      ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0790_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0790 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0790: In the sheet “Customer effect” (or “Effets clients”), the column XXXX (to be indicated) is not present or not written correctly as in the document [DOC3]"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String,
                                      "In the sheet “Customer effect” (or “Effets clients”), the column XXXX (to be indicated) is not present or not written correctly as in the document [DOC3].",
                                      ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0800_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0800 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0800: The sheet “Feared events” (or “ER”)  is not present or not written correctly"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String,
                                      "The sheet “Feared events” (or “ER”)  is not present or not written correctly",
                                      ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0810_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0810 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0810: In the sheet “Feared events”, the column XXXX (to be indicated) is not present or not written correctly as in the document [DOC3].."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String,
                                      "In the sheet “Feared events”, the column XXXX (to be indicated) is not present or not written correctly as in the document [DOC3].",
                                      ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0820_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0820 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0820: The sheet “Parts” (or “Constituants”) is not present or not written correctly"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String,
                                      "The sheet “Parts” (or “Constituants”) is not present or not written correctly",
                                      ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0830_XLSX_XLSM(workBook)
        if testResult == 1:
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0830:  In the sheet “Parts” (or “Constituants”), the column XXXX (to be indicated) is not present or not written correctly as in the document [DOC3]."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String,
                                      " In the sheet “Parts” (or “Constituants”), the column XXXX (to be indicated) is not present or not written correctly as in the document [DOC3]",
                                      ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0840_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0840 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0840: The sheet “Situation” (or “situations de vie”) is not present or not written correctly"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String,
                                      "The sheet “Situation” (or “situations de vie”) is not present or not written correctly",
                                      ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0850_XLSX_XLSM(workBook)
        if testResult == 1:
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0850: In the sheet “Situation” (or “situations de vie”), the column XXXX (to be indicated) is not present or not written correctly as in the document [DOC3]. "
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String,
                                      "In the sheet “Situation” (or “situations de vie”), the column XXXX (to be indicated) is not present or not written correctly as in the document [DOC3]. ",
                                      ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0860_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0860 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0860: The sheet “Degraded mode” (or “Liste MDD”)  is not present or not written correctly."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String,
                                      "The sheet “Degraded mode” (or “Liste MDD”)  is not present or not written correctly.",
                                      ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0870_XLSX_XLSM(workBook)
        if testResult == 1:
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0870:In the sheet “Degraded mode” (or “Liste MDD”), the column XXXX (to be indicated) is not present or not written correctly as in the document [DOC3].."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String,
                                      "In the sheet “Degraded mode” (or “Liste MDD”), the column XXXX (to be indicated) is not present or not written correctly as in the document [DOC3].",
                                      ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0880_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0880 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0880: The sheet “Technical effect “ (or “Effets techniques”) is not present or not written correctly"
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String,
                                      "The sheet “Technical effect “ (or “Effets techniques”) is not present or not written correctly”.",
                                      ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0890_XLSX_XLSM(workBook)
        if testResult == 1:
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0890: In the sheet “Technical effect “ (or “Effets techniques”), the column XXXX (to be indicated) is not present or not written correctly as in the document [DOC3]."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String,
                                      " In the sheet “Technical effect “ (or “Effets techniques”), the column XXXX (to be indicated) is not present or not written correctly as in the document [DOC3].",
                                      ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0900_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0900 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0900: The sheet “Variant “ (or “Variantes”) is not present or not written correctly.."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String,
                                      "The sheet “Variant “ (or “Variantes”) is not present or not written correctly.",
                                      ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        str2 = str(stringInt + row)
        String = str1 + str2 + str3
        testResult = self.Test_02043_18_04939_STRUCT_0910_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0910 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", String, "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_0910: In the sheet “Variant “ (or “Variantes”), the column XXXX (to be indicated) is not present or not written correctly as in the document [DOC3]."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict[String][self.checkLevel], String,
                                      "In the sheet “Variant “ (or “Variantes”), the column XXXX (to be indicated) is not present or not written correctly as in the document [DOC3]. ",
                                      ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(
                    ["Good", "02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        elif testResult == 2:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060 NOT OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060",
                                      "The sheet “tableau” (or “table”) is not in the file", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            for key, value in testResult.items():
                if key == "0":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060: In the sheet “tableau” (or “table”), the column “Référence” is not completed."
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1000"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1000 ",
                                              "In the sheet “tableau” (or “table”), the column “Référence” is not completed.",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1000 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1000 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-1":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060: In the sheet “tableau” (or “table”), the column “Version” is not completed."
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1001"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1001 ",
                                              "In the sheet “tableau” (or “table”), the column “Version” is not completed.",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1001 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1001 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-2":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060: In the sheet “tableau” (or “table”), the columns “Constituant défaillant détecté” is not completed."
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1180"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1180 ",
                                              "In the sheet “tableau” (or “table”), the columns “Constituant défaillant détecté” is not completed",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1180 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1180 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-3":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060: In the sheet “tableau” (or “tableau”),  the columns “Défaillance constituant” is not completed"
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1190"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1190 ",
                                              "In the sheet “tableau” (or “tableau”),  the columns “Défaillance constituant” is not completed",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1190 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1190 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-4":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060: In the sheet “tableau” (or “tableau”), the columns “Situation de vie client” is not completed"
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1200"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1200 ",
                                              "In the sheet “tableau” (or “tableau”), the columns “Situation de vie client” is not completed ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1200 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1200 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-5":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060: In the sheet “tableau” (or “tableau”), the columns “Effet(s) client(s)” is not completed."
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1210"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1210 ",
                                              "In the sheet “tableau” (or “tableau”), the columns “Effet(s) client(s)” is not completed.",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1210 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1210 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-6":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060: In the sheet “tableau” (or “tableau”), the columns “Code défaut” is not completed"
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1220"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1220 ",
                                              "In the sheet “tableau” (or “tableau”), the columns “Code défaut” is not completed",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1220 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1220 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-7":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060: The project impact is not specified for the project in the sheet “tableau "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1060"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1060 ",
                                              "The project impact is not specified for the project in the sheet “tableau ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1060 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1060 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_WHOLENESS_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(
                    ["Good", "02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061", "",
                     ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        elif testResult == 2:
            row += 1
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061: NOT OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(
                    ["", "02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061",
                     "The sheet “codes défauts” is not present in the file", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            for key, value in testResult.items():
                if key == "0":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061:  In the sheet “codes défauts”, the column “Référence” is not completed "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1010"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1010",
                                              "In the sheet “codes défauts”, the column “Référence” is not completed. ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1010 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1010 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-1":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061: In the sheet “codes défauts”, the column “Version” is not completed. "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1011"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1011 ",
                                              "In the sheet “codes défauts”, the column “Version” is not completed. ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1011 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1011 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-2":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061:In the sheet “codes défauts”, the column “Code défaut” is not completed. "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1080"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1080 ",
                                              "In the sheet “codes défauts”, the column “Code défaut” is not completed. ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1080 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1080 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-3":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061:In the sheet “codes défauts”, the column “supporté par constituant (s)” is not completed. "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1090"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1090 ",
                                              "In the sheet “codes défauts”, the column “supporté par constituant (s)” is not completed. ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1090 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1090 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-4":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061:In the sheet “codes défauts”, the columns “libellé (signification)” is not completed "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1110"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1110 ",
                                              "In the sheet “codes défauts”, the columns “libellé (signification)” is not completed ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1110 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1110 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-5":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061:In the sheet “codes défauts”, the columns “Description de la strategie pour détecter le défaut” is not completed "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1120"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1120 ",
                                              "In the sheet “codes défauts”, the columns “Description de la strategie pour détecter le défaut” is not completed ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1120 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1120 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-6":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061:In the sheet “codes défauts”, the columns “Seuil de détection  /  valeur  du défaut” is not completed. "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1130"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1130 ",
                                              "In the sheet “codes défauts”, the columns “Seuil de détection  /  valeur  du défaut” is not completed.",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1130 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1130 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-7":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061:In the sheet “codes défauts”, the columns “Temps de confirmation du défaut” is not completed. "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1140"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1140 ",
                                              "In the sheet “codes défauts”, the columns “Temps de confirmation du défaut” is not completed. ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1140 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1140 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-8":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061:In the sheet “codes défauts”, the columns “Description de la strategie de disparition du défaut / Procedure à effectuer pour vérifier la disparition du défaut” is not completed "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1150"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1150 ",
                                              "In the sheet “codes défauts”, the columns “Description de la strategie de disparition du défaut / Procedure à effectuer pour vérifier la disparition du défaut” is not completed ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1150 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1150 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-9":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061:In the sheet “codes défauts”, the columns “Mode dégradé” is not completed. "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1160"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1160 ",
                                              "In the sheet “codes défauts”, the columns “Mode dégradé” is not completed. ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1160 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1160 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-10":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061:In the sheet “codes défauts”, the columns “Voyant” is not completed. "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1170"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1170 ",
                                              "In the sheet “codes défauts”, the columns “Voyant” is not completed. ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1170 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1170 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-11":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_STRUCT_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061: The project impact is not specified for the project in the sheet “codes défauts "
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1061"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1061 ",
                                              " The project impact is not specified for the project in the sheet “codes défauts ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1061 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1061 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_WHOLENESS_1020_1021_1100_1062_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_WHOLENESS_1020_1021_1100_1062 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1020_1021_1100_1062 OK", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        elif testResult == 2:
            row += 1
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_WHOLENESS_1020_1021_1100_1062 NOT OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["", "02043_18_04939_WHOLENESS_1020_1021_1100_1062",
                                      "The sheet “mesures et commandes” is not in the file", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        else:
            for key, value in testResult.items():
                if key == "0":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1020_1021_1100_1062: In the sheet “mesures et commandes”, the column “Référence” is not completed"
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1020"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1020 ",
                                              "In the sheet “mesures et commandes”, the column “Référence” is not completed ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1020 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1020 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-1":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1020_1021_1100_1062: In the sheet “mesures et commandes”, the column “Version” is not completed."
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1021"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1021 ",
                                              "In the sheet “mesures et commandes”, the column “Version” is not completed.",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENES_1021 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1021 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-2":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1020_1021_1100_1062: In the sheet “mesures et commandes”, the column “supporté par constituant (s)” is not completed."
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1100"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1100 ",
                                              "In the sheet “mesures et commandes”, the column “supporté par constituant (s)” is not completed. ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1100 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1100 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-3":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1020_1021_1100_1062: The project impact is not specified for the project in the sheet “mesures et commandes”"
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1062"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1062 ",
                                              "The project impact is not specified for the project in the sheet “mesures et commandes” ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1062 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1062 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_WHOLENESS_1030_1031_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_WHOLENESS_1030_1031 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1030_1031", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        elif testResult == 2:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_WHOLENESS_1030_1031 NOT OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(
                    ["", "02043_18_04939_WHOLENESS_1030_1031", "The sheet “Diagnostic débarqués” is not in the file",
                     ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        else:
            for key, value in testResult.items():
                if key == "0":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nIn the sheet “Diagnostic débarqués”, the column “Référence” is not completed."
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1030"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1030 ",
                                              "Diagnostic débarqués”, the column “Référence” is not completed. ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1030 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "002043_18_04939_WHOLENESS_1030 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-1":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nIn the sheet “Diagnostic débarqués”, the column “Version” is not completed"
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1031"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1031",
                                              "In the sheet “Diagnostic débarqués”, the column “Version” is not completed.",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1031 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1031 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_WHOLENESS_1040_1041_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_WHOLENESS_1040_1041 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1040_1041", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        elif testResult == 2:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_WHOLENESS_1040_1041 NOT OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(
                    ["Good", "02043_18_04939_WHOLENESS_1040_1041", "The sheet “Liste MDD” is not in the file", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            for key, value in testResult.items():
                if key == "0":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1040_1041: In the sheet “Liste MDD”, the column “Référence” is not completed."
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1040"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1040",
                                              "In the sheet “Liste MDD”, the column “Référence” is not completed. ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1040 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1040 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                if key == "-1":
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\nTest_02043_18_04939_WHOLENESS_1040_1041: In the sheet “Liste MDD”, the column “Version” is not completed."
                    self.tab1.textbox.setText(text)
                    flag = 0
                    for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1041"][self.checkLevel],
                                              "02043_18_04939_WHOLENESS_1041",
                                              "In the sheet “Liste MDD”, the column “Version” is not completed. ",
                                              str(value)]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name
                else:
                    row += 1
                    text = self.tab1.textbox.toPlainText()
                    text = text + "\n02043_18_04939_WHOLENESS_1041 OK"
                    self.tab1.textbox.setText(text)
                    for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1041 OK", "", ""]):
                        reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_WHOLENESS_1050_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_1050 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_1050", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        elif testResult == 2:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_1050 NOT OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_1050",
                                      "The “tableau”(or “table”) or “codes défauts” is not presetnt in the file", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_1050: The projects listed in the following sheets “tableau” and “codes défauts” are not identical."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1050"][self.checkLevel],
                                      "02043_18_04939_WHOLENESS_1050 ",
                                      " The projects listed in the following sheets “tableau” and “codes défauts” are not identical. ",
                                      ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_WHOLENESS_1055_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_1055 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_1055", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        elif testResult == 2:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_1055 NOT OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_STRUCT_1055",
                                      "The “tableau”(or “table”) or “mesures et commandes” is not presetnt in the file",
                                      ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_STRUCT_1055: The projects listed in the following sheets “tableau” and “mesures et commandes” are not identical. "
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS _1055"][self.checkLevel],
                                      "02043_18_04939_WHOLENESS _1055",
                                      "The projects listed in the following sheets “tableau” and “mesures et commandes” are not identical. ",
                                      ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_WHOLENESS_1240_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_WHOLENESS_1240 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_WHOLENESS_1240", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_WHOLENESS_1240: The reviewers have to be indicated in the paraph  “Liste de diffusion” or “Mailing list (the taking part)” of the sheet “Informations Générales” (or “General information”)."
            self.tab1.textbox.setText(text)
            flag = 0
            for i, name in enumerate([self.testReqDict["02043_18_04939_WHOLENESS_1240"][self.checkLevel],
                                      "02043_18_04939_WHOLENESS_1240",
                                      "The reviewers have to be indicated in the paraph  “Liste de diffusion” or “Mailing list (the taking part)” of the sheet “Informations Générales” (or “General information”).",
                                      ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name

        row += 1
        testResult = self.Test_02043_18_04939_COH_2080_XLSX_XLSM(workBook)
        if testResult == 1:
            text = self.tab1.textbox.toPlainText()
            text = text + "\nTest_02043_18_04939_COH_2080 OK"
            self.tab1.textbox.setText(text)
            for i, name in enumerate(["Good", "02043_18_04939_COH_2080", "", ""]):
                reportWorkSheet2.Cells(row, i + 1).Value = name
        else:
            if testResult == 2:
                text = self.tab1.textbox.toPlainText()
                text = text + "\nTest_02043_18_04939_COH_2080: One of the sheets “Effets clients”, “FR” or “GB” are not in the file  "

            else:
                text = self.tab1.textbox.toPlainText()
                text = text + "\nTest_02043_18_04939_COH_2080: The information specified in the column “Noms” of the “Effets clients” sheet is missing in the document [DOC7] in French or English. Please check it or ask ADRD people to add this new customer effect."
                self.tab1.textbox.setText(text)
                flag = 0
                for i, name in enumerate([self.testReqDict["02043_18_04939_COH_2080"][self.checkLevel],
                                          "02043_18_04939_COH_2080",
                                          " The information specified in the column “Noms” of the “Effets clients” sheet is missing in the document [DOC7] in French or English. Please check it or ask ADRD people to add this new customer effect.",
                                          ""]):
                    reportWorkSheet2.Cells(row, i + 1).Value = name

        reportWorkBook.Save()

    def GetTsdFileExtension(self):

        fileName = self.tab1.myTextBox1.toPlainText()
        tokens = fileName.split(".")
        self.tsdFileExtension = tokens[-1]

    def GetTsdFileWorkbook(self):

        fileName = self.tab1.myTextBox1.toPlainText()
        fileName = fileName.replace("\\", "/")
        if self.tsdFileExtension == "xls":
            return xlrd.open_workbook(fileName,  formatting_info=True)
        elif self.tsdFileExtension == "xlsx":
            with open(fileName, "rb") as fileReader:
                fileName = io.BytesIO(fileReader.read())
            return openpyxl.load_workbook(fileName, read_only=True)
        elif self.tsdFileExtension == "xlsm":
            with open(fileName, "rb") as fileReader:
                fileName = io.BytesIO(fileReader.read())
            return openpyxl.load_workbook(fileName, keep_vba=True, read_only=True)

    def TestTsdFile(self, path_Cesare, path_effect):

        self.GetTsdFileExtension()
        flag = True
        if self.tsdFileExtension:
            text = self.tab1.textbox.toPlainText()
            text = text + "\n\tTesting TSD FILE ----------------------------------\n"
            self.tab1.textbox.setText(text)
            workBook = self.GetTsdFileWorkbook()
            if self.tsdFileExtension == "xls":
                flag = self.TestGeneralStructureXLS_DOC3(workBook, self.tab1.myTextBox1.toPlainText(), path_Cesare, path_effect)
                self.TestGeneralStructure_DOC3_XLS(workBook)
                flag = flag and self.TestGeneralStructure_DOC3_XLS(workBook)
            else:
                flag = self.TestGeneralStructureXLSX_XLSM_DOC3(workBook, path_Cesare, path_effect)
                self.TestGeneralStructure_DOC3_XLSX_XLSM(workBook)
                flag = flag and self.TestGeneralStructure_DOC3_XLSX_XLSM(workBook)
            if flag == True:
                self.tab1.colorTextBox1.setStyleSheet('background-color: green')

            else:
                self.tab1.colorTextBox1.setStyleSheet('background-color: red')

        else:
            self.pbvalue = self.pbvalue + 0.8772*19
            self.tab1.pbar.setValue(self.pbvalue)
        return flag

    def GetTsdVehicleFunctionFileExtension(self):
        fileName = self.tab1.myTextBox2.toPlainText()
        tokens = fileName.split(".")
        self.tsdVehicleFunctionFileExtension = tokens[-1]

    def GetTsdVehicleFunctionFileWorkbook(self):

        fileName = self.tab1.myTextBox2.toPlainText()
        fileName = fileName.replace("\\", "/")
        if self.tsdVehicleFunctionFileExtension == "xls":
            return xlrd.open_workbook(fileName,  formatting_info=True)
        elif self.tsdVehicleFunctionFileExtension == "xlsx":
            with open(fileName, "rb") as fileReader:
                fileName = io.BytesIO(fileReader.read())
            return openpyxl.load_workbook(fileName, read_only=True)
        elif self.tsdVehicleFunctionFileExtension == "xlsm":
            with open(fileName, "rb") as fileReader:
                fileName = io.BytesIO(fileReader.read())
            return openpyxl.load_workbook(fileName, keep_vba=True, read_only=True)

    def TestTsdVehicleFunctionFile(self, path_Cesare, path_effect):

        self.GetTsdVehicleFunctionFileExtension()
        if self.tsdVehicleFunctionFileExtension:
            text = self.tab1.textbox.toPlainText()
            text = text + "\n\tTesting TSD Vehicle Function FILE --------------------------\n"
            self.tab1.textbox.setText(text)
            workBook = self.GetTsdVehicleFunctionFileWorkbook()
            if self.tsdVehicleFunctionFileExtension == "xls":
                flag = self.TestGeneralStructureXLS_DOC4(workBook, self.tab1.myTextBox2.toPlainText(), path_Cesare, path_effect)
                self.TestGeneralStructure_DOC4_XLS(workBook)
                flag = flag and self.TestGeneralStructure_DOC4_XLS(workBook)
            else:
                flag = self.TestGeneralStructureXLSX_XLSM_DOC4(workBook, path_Cesare, path_effect)
                self.TestGeneralStructure_DOC4_XLSX_XLSM(workBook)
                flag = flag and self.TestGeneralStructure_DOC4_XLSX_XLSM(workBook)
            if flag == True:
                self.tab1.colorTextBox2.setStyleSheet('background-color: green')
            else:
                self.tab1.colorTextBox2.setStyleSheet('background-color: red')
        else:
            self.pbvalue = self.pbvalue + 0.8772*19
            self.tab1.pbar.setValue(self.pbvalue)

    def GetTsdSystemFileExtension(self):
        fileName = self.tab1.myTextBox3.toPlainText()
        tokens = fileName.split(".")
        self.tsdSystemFileExtension = tokens[-1]

    def GetTsdSystemFileWorkbook(self):

        fileName = self.tab1.myTextBox3.toPlainText()
        if self.tsdSystemFileExtension == "xls":
            return xlrd.open_workbook(fileName,  formatting_info=True)
        elif self.tsdSystemFileExtension == "xlsx":
            with open(fileName, "rb") as fileReader:
                fileName = io.BytesIO(fileReader.read())
            return openpyxl.load_workbook(fileName, read_only=True)
        elif self.tsdSystemFileExtension == "xlsm":
            with open(fileName, "rb") as fileReader:
                fileName = io.BytesIO(fileReader.read())
            return openpyxl.load_workbook(fileName, keep_vba=True, read_only=True)

    def TestTsdSystemFile(self, path_Cesare, path_effect):

        self.GetTsdSystemFileExtension()
        if self.tsdSystemFileExtension:
            text = self.tab1.textbox.toPlainText()
            text = text + "\n\tTesting TSD System FILE ---------------------------------\n"
            self.tab1.textbox.setText(text)
            workBook = self.GetTsdSystemFileWorkbook()
            if self.tsdSystemFileExtension == "xls":
                flag = self.TestGeneralStructureXLS_DOC5(workBook, self.tab1.myTextBox3.toPlainText(), path_Cesare, path_effect)
                self.TestGeneralStructure_DOC5_XLS(workBook)
                flag = flag and self.TestGeneralStructure_DOC5_XLS(workBook)
            else:
                flag = self.TestGeneralStructureXLSX_XLSM_DOC5(workBook, path_Cesare, path_effect)
                self.TestGeneralStructure_DOC5_XLSX_XLSM(workBook)
                flag = flag and self.TestGeneralStructure_DOC5_XLSX_XLSM(workBook)
            if flag == True:
                self.tab1.colorTextBox3.setStyleSheet('background-color: green')
            else:
                self.tab1.colorTextBox3.setStyleSheet('background-color: red')
        else:
            self.pbvalue = self.pbvalue + 0.8772*19
            self.tab1.pbar.setValue(self.pbvalue)

    def GetAmdecFileExtension(self):
        fileName = self.tab1.myTextBox7.toPlainText()
        tokens = fileName.split(".")
        self.amdecFileExtension = tokens[-1]

    def GetAmdecFileWorkbook(self):

        fileName = self.tab1.myTextBox7.toPlainText()
        if self.amdecFileExtension == "xls":
            return xlrd.open_workbook(fileName,  formatting_info=True)
        elif self.amdecFileExtension == "xlsx":
            return openpyxl.load_workbook(fileName, read_only=True)
        elif self.amdecFileExtension == "xlsm":
            return openpyxl.load_workbook(fileName, keep_vba=True, read_only=True)

    def TestAmdecFile(self):

        self.GetAmdecFileExtension()
        if self.amdecFileExtension:
            text = self.tab1.textbox.toPlainText()
            text = text + "\n\tTesting AMDEC FILE -------------------------------------\n"
            self.tab1.textbox.setText(text)
            workBook = self.GetAmdecFileWorkbook()
            if self.amdecFileExtension == "xls":
                flag = self.TestGeneralStructureXLS(workBook, self.tab1.myTextBox7.toPlainText())
            else:
                flag = self.TestGeneralStructureXLSX_XLSM(workBook)
            if flag == True:
                self.tab1.colorTextBox4.setStyleSheet('background-color: green')
            else:
                self.tab1.colorTextBox4.setStyleSheet('background-color: red')
        else:
            self.pbvalue = self.pbvalue + 0.8772*19
            self.tab1.pbar.setValue(self.pbvalue)

    def GetExportMedialecMatriceFileExtension(self):
        fileName = self.tab1.myTextBox8.toPlainText()
        tokens = fileName.split(".")
        self.exportMedialecMatriceFileExtension = tokens[-1]

    def GetExportMedialecMatriceFileWorkbook(self):

        fileName = self.tab1.myTextBox8.toPlainText()
        if self.exportMedialecMatriceFileExtension == "xls":
            return xlrd.open_workbook(fileName,  formatting_info=True)
        elif self.exportMedialecMatriceFileExtension == "xlsx":
            return openpyxl.load_workbook(fileName, read_only=True)
        elif self.exportMedialecMatriceFileExtension == "xlsm":
            return openpyxl.load_workbook(fileName, keep_vba=True, read_only=True)

    def TestExportMedialecMatriceFile(self):

        self.GetExportMedialecMatriceFileExtension()
        if self.exportMedialecMatriceFileExtension:
            text = self.tab1.textbox.toPlainText()
            text = text + "\n\tTesting Export Medialec Matrice FILE --------------------------\n"
            self.tab1.textbox.setText(text)
            workBook = self.GetExportMedialecMatriceFileWorkbook()
            if self.exportMedialecMatriceFileExtension == "xls":
                flag = self.TestGeneralStructureXLS(workBook, self.tab1.myTextBox8.toPlainText())
            else:
                flag = self.TestGeneralStructureXLSX_XLSM(workBook)
            if flag == True:
                self.tab1.colorTextBox5.setStyleSheet('background-color: green')
            else:
                self.tab1.colorTextBox5.setStyleSheet('background-color: red')
        else:
            self.pbvalue = self.pbvalue + 0.8772*19
            self.tab1.pbar.setValue(self.pbvalue)

    def GetDiagnosticMatrixFileExtension(self):
        fileName = self.tab1.myTextBox10.toPlainText()
        tokens = fileName.split(".")
        self.diagnosticMatrixFileExtension = tokens[-1]

    def GetDiagnosticMatrixFileWorkbook(self):

        fileName = self.tab1.myTextBox10.toPlainText()
        if self.diagnosticMatrixFileExtension == "xls":
            return xlrd.open_workbook(fileName,  formatting_info=True)
        elif self.diagnosticMatrixFileExtension == "xlsx":
            return openpyxl.load_workbook(fileName, read_only=True)
        elif self.diagnosticMatrixFileExtension == "xlsm":
            return openpyxl.load_workbook(fileName, keep_vba=True, read_only=True)

    def TestDiagnosticMatrixFile(self):

        self.GetDiagnosticMatrixFileExtension()
        if self.diagnosticMatrixFileExtension:
            text = self.tab1.textbox.toPlainText()
            text = text + "\n\tTesting Diagnostic Matrix FILE ---------------------------------\n"
            self.tab1.textbox.setText(text)
            workBook = self.GetDiagnosticMatrixFileWorkbook()
            if self.diagnosticMatrixFileExtension == "xls":
                flag = self.TestGeneralStructureXLS(workBook, self.tab1.myTextBox10.toPlainText())
            else:
                flag = self.TestGeneralStructureXLSX_XLSM(workBook)
            if flag == True:
                self.tab1.colorTextBox6.setStyleSheet('background-color: green')
            else:
                self.tab1.colorTextBox6.setStyleSheet('background-color: red')
        else:
            self.pbvalue = self.pbvalue + 0.8772*19
            self.tab1.pbar.setValue(self.pbvalue)

#Requirements for General structure


    def Test_02043_18_04939_STRUCT_0000_XLS(self, workBook):

        sheetNames = [x.casefold() for x in workBook.sheet_names()]
        if "informations générales" in sheetNames or "general information" in sheetNames:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0000_XLSX_XLSM(self, workBook):

        sheetNames = [x.casefold() for x in workBook.sheetnames]
        if "informations générales" in sheetNames or "general information" in sheetNames:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0005_XLS(self, workBook):

        os.system("taskkill /f /im EXCEL.EXE")
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        wb = excel.Workbooks.Open(workBook)
        ws = wb.Worksheets(1)
        if ws.Cells(51,1).HasFormula is False:
            excel.Application.Quit()
            return 1

        else:
            excel.Application.Quit()
            return 0

    def Test_02043_18_04939_STRUCT_0005_XLSX_XLSM(self, workBook):

        workSheet = workBook.worksheets[0]
        if "=" in str(workSheet.cell(52,2).value):
            return 0
        else:
            return 1

    def Test_02043_18_04939_STRUCT_0010_XLS(self, workBook):
        workSheet = workBook.sheet_by_index(0)
        try:
            value = workSheet.cell_value(51, 1)
        except:
            return 0
        if isinstance(value, str) and value.strip():
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0010_XLSX_XLSM(self, workBook):

        workSheet = workBook.worksheets[0]
        if isinstance(workSheet.cell(52, 2).value, str) and workSheet.cell(52, 2).value and not workSheet.cell(52,2).value.isspace():
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0011_XLS(self, workBook):
        workSheet = workBook.sheet_by_index(0)
        try:
            value = workSheet.cell_value(51, 1)
        except:
            return 0
        if value in {"AEEV_IAEE07_0033", "02043_12_01665", "02043_12_01666"}:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0011_XLSX_XLSM(self, workBook):

        workSheet = workBook.worksheets[0]
        if workSheet.cell(52, 2).value in {"AEEV_IAEE07_0033", "02043_12_01665", "02043_12_01666"}:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0020_XLS(self, workBook):

        sheetNames = [x.casefold() for x in workBook.sheet_names()]
        if "suppression" in sheetNames:
            return 1
        else:
            return 0

    def  Test_02043_18_04939_STRUCT_0020_XLSX_XLSM(self, workBook):

         sheetNames = [x.casefold() for x in workBook.sheetnames]
         if "suppression" in sheetNames:
             return 1
         else:
             return 0

    def Test_02043_18_04939_STRUCT_0025_XLS(self, workBook):
        sheetNames = [x.casefold() for x in workBook.sheet_names()]
        try:
            index = sheetNames.index("suppression")
        except:
            return 0
        workSheet = workBook.sheet_by_index(index)
        try:
            row = workSheet.row(0)
        except:
            return 0
        for cell in row:
            if cell.value.casefold() in {"sheet", "onglet"}:
                return 1
        return 0

    def Test_02043_18_04939_STRUCT_0025_XLSX_XLSM(self, workBook):
        sheetNames = [x.casefold() for x in workBook.sheetnames]
        try:
            index = sheetNames.index("suppression")
        except:
            return 0
        workSheet = workBook.worksheets[index]

        row = workSheet.iter_rows(min_col=1, min_row=1, max_row=5, max_col=5)
        for cellTuple in row:
            for cell in cellTuple:
                if str(cell.value).casefold() in {"sheet", "onglet"}:
                    return 1
        return 0

    def Test_02043_18_04939_STRUCT_0030_XLS(self, workBook):
        sheetNames = [x.casefold() for x in workBook.sheet_names()]
        try:
            index = sheetNames.index("suppression")
        except:
            return 0
        workSheet = workBook.sheet_by_index(index)

        try:
            row = workSheet.row(0)
        except:
            return 0
        for cell in row:
            if cell.value.casefold() in {"référence de la ligne", "line number"}:
                return 1
        return 0

    def Test_02043_18_04939_STRUCT_0030_XLSX_XLSM(self, workBook):
        sheetNames = [x.casefold() for x in workBook.sheetnames]
        try:
            index = sheetNames.index("suppression")
        except:
            return 0
        workSheet = workBook.worksheets[index]

        row = workSheet.iter_rows(min_col=1, min_row=1, max_row=10, max_col=10)
        for cellTuple in row:
            for cell in cellTuple:
               if str(cell.value).casefold() in {"référence de la ligne", "line number"}:
                   return 1
        return 0

    def Test_02043_18_04939_STRUCT_0035_XLS(self, workBook):
        sheetNames = [x.casefold() for x in workBook.sheet_names()]
        try:
            index = sheetNames.index("suppression")
        except:
            return 0
        workSheet = workBook.sheet_by_index(index)

        try:
            row = workSheet.row(0)
        except:
            return 0
        for cell in row:
            if cell.value.casefold() in {"version du tsd", "version of the document"}:
                return 1
        return 0

    def Test_02043_18_04939_STRUCT_0035_XLSX_XLSM(self, workBook):
        sheetNames = [x.casefold() for x in workBook.sheetnames]
        try:
            index = sheetNames.index("suppression")
        except:
            return 0
        workSheet = workBook.worksheets[index]

        row = workSheet.iter_rows(min_col=1, min_row=1, max_row=10, max_col=10)
        for cellTuple in row:
            for cell in cellTuple:
                if str(cell.value).casefold() in {"version du tsd", "version of the document"}:
                   return 1
        return 0

    def Test_02043_18_04939_STRUCT_0040_XLS(self, workBook):
        sheetNames = [x.casefold() for x in workBook.sheet_names()]
        try:
            index = sheetNames.index("suppression")
        except:
            return 0
        workSheet = workBook.sheet_by_index(index)

        try:
            row = workSheet.row(0)
        except:
            return 0
        for cell in row:
            if str(cell.value).casefold() in {"justification de la modification", "change reason"}:
                return 1
        return 0

    def Test_02043_18_04939_STRUCT_0040_XLSX_XLSM(self, workBook):
        sheetNames = [x.casefold() for x in workBook.sheetnames]
        try:
            index = sheetNames.index("suppression")
        except:
            return 0
        workSheet = workBook.worksheets[index]

        row = workSheet.iter_rows(min_col=1, min_row=1, max_row=10, max_col=10)
        for cellTuple in row:
            for cell in cellTuple:
                 if str(cell.value).casefold() in {"justification de la modification", "change reason"}:
                    return 1
        return 0

    def Test_02043_18_04939_STRUCT_0051_XLS(self, workBook):
        sheetNames = [x.casefold() for x in workBook.sheet_names()]
        try:
            index = sheetNames.index("reference docs")
        except:
            return 0
        workSheet = workBook.sheet_by_index(index)
        indexCol = workSheet.ncols
        for index in range(0,indexCol):
            column = workSheet.col(index)
            for cell in column:
                if str(cell.value).strip().casefold() == "name":
                    colNameIndex = index
                if str(cell.value).strip().casefold() == "reference":
                    colReferenceIndex = index
        colName = workSheet.col(colNameIndex)
        for index in range(0,len(colName)):
            if str(colName[index].value).strip().casefold() in [x.casefold().strip() for x in [ "Vehicle Architecture Schematic", "Planche d'architecture véhicule"]]:
                if isinstance(workSheet.cell_value(index, colReferenceIndex), str) and workSheet.cell_value(index, colReferenceIndex).strip():
                    return 1
                else:
                    return 0
        return 0

    def Test_02043_18_04939_STRUCT_0051_XLSX_XLSM(self, workBook):
        sheetNames = [x.casefold() for x in workBook.sheetnames]
        try:
            index = sheetNames.index("reference docs")
        except:
            return 0
        workSheet = workBook.worksheets[index]
        rows = workSheet.iter_rows(min_row=1, max_row=10)
        for cellRowTuple in rows:
            for cell in cellRowTuple:
                if str(cell.value).strip().casefold() == "name":
                    colNameIndex = cell.column
                if str(cell.value).strip().casefold() == "reference":
                    colReferenceIndex = cell.column
        for rowIndex in range(1, workSheet.max_row):
            if str(workSheet.cell(row = rowIndex, column = colNameIndex).value).casefold().strip() in [x.casefold().strip() for x in ["Vehicle Architecture Schematic", "Planche d'architecture véhicule"]]:
                if isinstance(workSheet.cell(row = rowIndex, column = colReferenceIndex).value, str) and workSheet.cell(row = rowIndex, column = colReferenceIndex).value.strip():
                    return 1
                else:
                    return 0
        return 0

        """      colName = workSheet.iter_rows(min_col=colNameIndex, max_col=colNameIndex)
            for cellObjectTuple in colName:
                for cellObject in cellObjectTuple:
                    if str(cellObject.value).strip().casefold() == "vehicle architecture schematic" or str(cellObject.value).strip().casefold() == "planche d'architecture véhicule":
                        if isinstance(workSheet.cell(cellObject.row, colReferenceIndex).value, str) and workSheet.cell(cellObject.row, colReferenceIndex).value and not workSheet.cell(cellObject.row, colReferenceIndex).value.isspace():
                            return 1
                        else:
                            return 0
        """

    def Test_02043_18_04939_STRUCT_0052_XLS(self, workBook):
        sheetNames = [x.casefold() for x in workBook.sheet_names()]
        try:
            index = sheetNames.index("reference docs")
        except:
            return 0
        workSheet = workBook.sheet_by_index(index)
        indexCol = workSheet.ncols
        for index in range(0, indexCol):
            column = workSheet.col(index)
            for cell in column:
                if str(cell.value).strip().casefold() == "name":
                    colNameIndex = index
                if str(cell.value).strip().casefold() == "reference":
                    colReferenceIndex = index
        colName = workSheet.col(colNameIndex)
        for index in range(0, len(colName)):
            if str(colName[index].value).strip().casefold() in [x.casefold().strip() for x in["Diagnostic Matrix", "Matrice Diag"]]:
                if isinstance(workSheet.cell_value(index, colReferenceIndex), str) and workSheet.cell_value(index, colReferenceIndex).strip():
                    return 1
                else:
                    return 0
        return 0

    def Test_02043_18_04939_STRUCT_0052_XLSX_XLSM(self, workBook):
        sheetNames = [x.casefold() for x in workBook.sheetnames]
        try:
            index = sheetNames.index("reference docs")
        except:
            return 0
        workSheet = workBook.worksheets[index]
        rows = workSheet.iter_rows(min_row=1, max_row=10)
        for cellRowTuple in rows:
            for cell in cellRowTuple:
                if str(cell.value).strip().casefold() == "name":
                    colNameIndex = cell.column
                if str(cell.value).strip().casefold() == "reference":
                    colReferenceIndex = cell.column
        for rowIndex in range(1, workSheet.max_row):
            if str(workSheet.cell(row = rowIndex, column = colNameIndex).value).casefold().strip() in [x.casefold().strip() for x in ["Diagnostic Matrix", "Matrice Diag"]]:
                if isinstance(workSheet.cell(row = rowIndex, column = colReferenceIndex).value, str) and workSheet.cell(row = rowIndex, column = colReferenceIndex).value.strip():
                    return 1
                else:
                    return 0
        return 0

    def Test_02043_18_04939_STRUCT_0053_XLS(self, workBook):
        sheetNames = [x.casefold() for x in workBook.sheet_names()]
        try:
            index = sheetNames.index("reference docs")
        except:
            return 0
        workSheet = workBook.sheet_by_index(index)
        indexCol = workSheet.ncols
        for index in range(0, indexCol):
            column = workSheet.col(index)
            for cell in column:
                if str(cell.value).strip().casefold() == "name":
                    colNameIndex = index
                if str(cell.value).strip().casefold() == "reference":
                    colReferenceIndex = index
        colName = workSheet.col(colNameIndex)
        for index in range(0, len(colName)):
            if str(colName[index].value).strip().casefold() in [x.casefold().strip() for x in["Fault Tree", "AMDEC"]]:
                if isinstance(workSheet.cell_value(index, colReferenceIndex), str) and workSheet.cell_value(index, colReferenceIndex).strip():
                    return 1
                else:
                    return 0
        return 0

    def Test_02043_18_04939_STRUCT_0053_XLSX_XLSM(self, workBook):
        sheetNames = [x.casefold() for x in workBook.sheetnames]
        try:
            index = sheetNames.index("reference docs")
        except:
            return 0
        workSheet = workBook.worksheets[index]
        rows = workSheet.iter_rows(min_row=1, max_row=10)
        for cellRowTuple in rows:
            for cell in cellRowTuple:
                if str(cell.value).strip().casefold() == "name":
                    colNameIndex = cell.column
                if str(cell.value).strip().casefold() == "reference":
                    colReferenceIndex = cell.column
        for rowIndex in range(1, workSheet.max_row):
            if str(workSheet.cell(row = rowIndex, column = colNameIndex).value).casefold().strip() in [x.casefold().strip() for x in ["Fault Tree", "AMDEC"]]:
                if isinstance(workSheet.cell(row = rowIndex, column = colReferenceIndex).value, str) and workSheet.cell(row = rowIndex, column = colReferenceIndex).value.strip():
                    return 1
                else:
                    return 0
        return 0

    def Test_02043_18_04939_STRUCT_0054_XLS(self, workBook):
        sheetNames = [x.casefold() for x in workBook.sheet_names()]
        try:
            index = sheetNames.index("reference docs")
        except:
            return 0
        workSheet = workBook.sheet_by_index(index)
        indexCol = workSheet.ncols
        for index in range(0, indexCol):
            column = workSheet.col(index)
            for cell in column:
                if str(cell.value).strip().casefold() == "name":
                    colNameIndex = index
                if str(cell.value).strip().casefold() == "reference":
                    colReferenceIndex = index
        colName = workSheet.col(colNameIndex)
        for index in range(0, len(colName)):
            if str(colName[index].value).strip().casefold() in [x.casefold().strip() for x in["ECU schematic", "Synoptique ECU"]]:
                if isinstance(workSheet.cell_value(index, colReferenceIndex), str) and workSheet.cell_value(index, colReferenceIndex).strip():
                    return 1
                else:
                    return 0
        return 0

    def Test_02043_18_04939_STRUCT_0054_XLSX_XLSM(self, workBook):
        sheetNames = [x.casefold() for x in workBook.sheetnames]
        try:
            index = sheetNames.index("reference docs")
        except:
            return 0
        workSheet = workBook.worksheets[index]
        rows = workSheet.iter_rows(min_row=1, max_row=5)
        for cellRowTuple in rows:
            for cell in cellRowTuple:
                if str(cell.value).strip().casefold() == "name":
                    colNameIndex = cell.column
                if str(cell.value).strip().casefold() == "reference":
                    colReferenceIndex = cell.column
        for rowIndex in range(1, workSheet.max_row):
            if str(workSheet.cell(row = rowIndex, column = colNameIndex).value).casefold().strip() in [x.casefold().strip() for x in ["ECU schematic", "Synoptique ECU"]]:
                if isinstance(workSheet.cell(row = rowIndex, column = colReferenceIndex).value, str) and workSheet.cell(row = rowIndex, column = colReferenceIndex).value.strip():
                    return 1
                else:
                    return 0
        return 0

    def Test_02043_18_04939_STRUCT_0055_XLS(self, workBook):
        sheetNames = [x.casefold() for x in workBook.sheet_names()]
        try:
            index = sheetNames.index("reference docs")
        except:
            return 0
        workSheet = workBook.sheet_by_index(index)
        indexCol = workSheet.ncols
        for index in range(0, indexCol):
            column = workSheet.col(index)
            for cell in column:
                if str(cell.value).strip().casefold() == "name":
                    colNameIndex = index
                if str(cell.value).strip().casefold() == "reference":
                    colReferenceIndex = index
        colName = workSheet.col(colNameIndex)
        for index in range(0, len(colName)):
            if str(colName[index].value).strip().casefold() == "STD".casefold().strip():
                if isinstance(workSheet.cell_value(index, colReferenceIndex), str) and workSheet.cell_value(index, colReferenceIndex).strip():
                    return 1
                else:
                    return 0
        return 0

    def Test_02043_18_04939_STRUCT_0055_XLSX_XLSM(self, workBook):
        sheetNames = [x.casefold() for x in workBook.sheetnames]
        try:
            index = sheetNames.index("reference docs")
        except:
            return 0
        workSheet = workBook.worksheets[index]
        rows = workSheet.iter_rows(min_row=1, max_row=5)
        for cellRowTuple in rows:
            for cell in cellRowTuple:
                if str(cell.value).strip().casefold() == "name":
                    colNameIndex = cell.column
                if str(cell.value).strip().casefold() == "reference":
                    colReferenceIndex = cell.column
        for rowIndex in range(1, workSheet.max_row):
            if str(workSheet.cell(row = rowIndex, column = colNameIndex).value).casefold().strip() == "STD".casefold().strip():
                if isinstance(workSheet.cell(row = rowIndex, column = colReferenceIndex).value, str) and workSheet.cell(row = rowIndex, column = colReferenceIndex).value.strip():
                    return 1
                else:
                    return 0
        return 0

    def Test_02043_18_04939_STRUCT_0056_XLS(self, workBook):
        sheetNames = [x.casefold() for x in workBook.sheet_names()]
        try:
            index = sheetNames.index("reference docs")
        except:
            return 0
        workSheet = workBook.sheet_by_index(index)
        indexCol = workSheet.ncols
        for index in range(0, indexCol):
            column = workSheet.col(index)
            for cell in column:
                if str(cell.value).strip().casefold() == "name":
                    colNameIndex = index
                if str(cell.value).strip().casefold() == "reference":
                    colReferenceIndex = index
        colName = workSheet.col(colNameIndex)
        for index in range(0, len(colName)):
            if str(colName[index].value).strip().casefold() == "Complexity Matrix (Decli EE)".casefold().strip():
                if isinstance(workSheet.cell_value(index, colReferenceIndex), str) and workSheet.cell_value(index, colReferenceIndex).strip():
                    return 1
                else:
                    return 0
        return 0

    def Test_02043_18_04939_STRUCT_0056_XLSX_XLSM(self, workBook):
        sheetNames = [x.casefold() for x in workBook.sheetnames]
        try:
            index = sheetNames.index("reference docs")
        except:
            return 0
        workSheet = workBook.worksheets[index]
        rows = workSheet.iter_rows(min_row=1, max_row=5)
        for cellRowTuple in rows:
            for cell in cellRowTuple:
                if str(cell.value).strip().casefold() == "name":
                    colNameIndex = cell.column
                if str(cell.value).strip().casefold() == "reference":
                    colReferenceIndex = cell.column
        for rowIndex in range(1, workSheet.max_row):
            if str(workSheet.cell(row = rowIndex, column = colNameIndex).value).casefold().strip() == "Complexity Matrix (Decli EE)".casefold().strip():
                if isinstance(workSheet.cell(row = rowIndex, column = colReferenceIndex).value, str) and workSheet.cell(row = rowIndex, column = colReferenceIndex).value.strip():
                    return 1
                else:
                    return 0
        return 0

    def Test_02043_18_04939_STRUCT_0057_XLS(self, workBook):
        sheetNames = [x.casefold() for x in workBook.sheet_names()]
        try:
            index = sheetNames.index("reference docs")
        except:
            return 0
        workSheet = workBook.sheet_by_index(index)
        indexCol = workSheet.ncols
        for index in range(0, indexCol):
            column = workSheet.col(index)
            for cell in column:
                if str(cell.value).strip().casefold() == "name":
                    colNameIndex = index
                if str(cell.value).strip().casefold() == "reference":
                    colReferenceIndex = index
        colName = workSheet.col(colNameIndex)
        for index in range(0, len(colName)):
            if str(colName[index].value).strip().casefold() == "Décli".casefold().strip():
                if isinstance(workSheet.cell_value(index, colReferenceIndex), str) and workSheet.cell_value(index, colReferenceIndex).strip():
                    return 1
                else:
                    return 0
        return 0

    def Test_02043_18_04939_STRUCT_0057_XLSX_XLSM(self, workBook):
        sheetNames = [x.casefold() for x in workBook.sheetnames]
        try:
            index = sheetNames.index("reference docs")
        except:
            return 0
        workSheet = workBook.worksheets[index]
        rows = workSheet.iter_rows(min_row=1, max_row=5)
        for cellRowTuple in rows:
            for cell in cellRowTuple:
                if str(cell.value).strip().casefold() == "name":
                    colNameIndex = cell.column
                if str(cell.value).strip().casefold() == "reference":
                    colReferenceIndex = cell.column
        for rowIndex in range(1, workSheet.max_row):
            if str(workSheet.cell(row = rowIndex, column = colNameIndex).value).casefold().strip() == "Décli".casefold().strip():
                if isinstance(workSheet.cell(row = rowIndex, column = colReferenceIndex).value, str) and workSheet.cell(row = rowIndex, column = colReferenceIndex).value.strip():
                    return 1
                else:
                    return 0
        return 0

    def Test_02043_18_04939_STRUCT_0058_XLS(self, workBook):
        sheetNames = [x.casefold() for x in workBook.sheet_names()]
        try:
            index = sheetNames.index("reference docs")
        except:
            return 0
        workSheet = workBook.sheet_by_index(index)
        indexCol = workSheet.ncols
        for index in range(0, indexCol):
            column = workSheet.col(index)
            for cell in column:
                if str(cell.value).strip().casefold() == "name":
                    colNameIndex = index
                if str(cell.value).strip().casefold() == "reference":
                    colReferenceIndex = index
        colName = workSheet.col(colNameIndex)
        for index in range(0, len(colName)):
            if str(colName[index].value).strip().casefold() == "DCEE".casefold().strip():
                if isinstance(workSheet.cell_value(index, colReferenceIndex), str) and workSheet.cell_value(index, colReferenceIndex).strip():
                    return 1
                else:
                    return 0
        return 0

    def Test_02043_18_04939_STRUCT_0058_XLSX_XLSM(self, workBook):
        sheetNames = [x.casefold() for x in workBook.sheetnames]
        try:
            index = sheetNames.index("reference docs")
        except:
            return 0
        workSheet = workBook.worksheets[index]
        rows = workSheet.iter_rows(min_row=1, max_row=5)
        for cellRowTuple in rows:
            for cell in cellRowTuple:
                if str(cell.value).strip().casefold() == "name":
                    colNameIndex = cell.column
                if str(cell.value).strip().casefold() == "reference":
                    colReferenceIndex = cell.column
        for rowIndex in range(1, workSheet.max_row):
            if str(workSheet.cell(row = rowIndex, column = colNameIndex).value).casefold().strip() == "DCEE".casefold().strip():
                if isinstance(workSheet.cell(row = rowIndex, column = colReferenceIndex).value, str) and workSheet.cell(row = rowIndex, column = colReferenceIndex).value.strip():
                    return 1
                else:
                    return 0
        return 0

    def Test_02043_18_04939_STRUCT_0059_XLS(self, workBook):
        sheetNames = [x.casefold() for x in workBook.sheet_names()]
        try:
            index = sheetNames.index("reference docs")
        except:
            return 0
        workSheet = workBook.sheet_by_index(index)
        indexCol = workSheet.ncols
        for index in range(0, indexCol):
            column = workSheet.col(index)
            for cell in column:
                if str(cell.value).strip().casefold() == "name":
                    colNameIndex = index
                if str(cell.value).strip().casefold() == "reference":
                    colReferenceIndex = index
        colName = workSheet.col(colNameIndex)
        for index in range(0, len(colName)):
            if str(colName[index].value).strip().casefold() == "EEAD".casefold().strip():
                if isinstance(workSheet.cell_value(index, colReferenceIndex), str) and workSheet.cell_value(index, colReferenceIndex).strip():
                    return 1
                else:
                    return 0
        return 0

    def Test_02043_18_04939_STRUCT_0059_XLSX_XLSM(self, workBook):
        sheetNames = [x.casefold() for x in workBook.sheetnames]
        try:
            index = sheetNames.index("reference docs")
        except:
            return 0
        workSheet = workBook.worksheets[index]
        rows = workSheet.iter_rows(min_row=1, max_row=5)
        for cellRowTuple in rows:
            for cell in cellRowTuple:
                if str(cell.value).strip().casefold() == "name":
                    colNameIndex = cell.column
                if str(cell.value).strip().casefold() == "reference":
                    colReferenceIndex = cell.column
        for rowIndex in range(1, workSheet.max_row):
            if str(workSheet.cell(row = rowIndex, column = colNameIndex).value).casefold().strip() == "EEAD".casefold().strip():
                if isinstance(workSheet.cell(row = rowIndex, column = colReferenceIndex).value, str) and workSheet.cell(row = rowIndex, column = colReferenceIndex).value.strip():
                    return 1
                else:
                    return 0
        return 0

    def Test_02043_18_04939_STRUCT_0060_XLS(self, workBook):
        sheetNames = [x.casefold() for x in workBook.sheet_names()]
        try:
            index = sheetNames.index("reference docs")
        except:
            return 0
        workSheet = workBook.sheet_by_index(index)
        indexCol = workSheet.ncols
        for index in range(0, indexCol):
            column = workSheet.col(index)
            for cell in column:
                if str(cell.value).strip().casefold() == "name":
                    colNameIndex = index
                if str(cell.value).strip().casefold() == "reference":
                    colReferenceIndex = index
        colName = workSheet.col(colNameIndex)
        for index in range(0, len(colName)):
            if str(colName[index].value).strip().casefold() == "TFD".casefold().strip():
                if isinstance(workSheet.cell_value(index, colReferenceIndex), str) and workSheet.cell_value(index, colReferenceIndex).strip():
                    return 1
                else:
                    return 0
        return 0

    def Test_02043_18_04939_STRUCT_0060_XLSX_XLSM(self, workBook):
        sheetNames = [x.casefold() for x in workBook.sheetnames]
        try:
            index = sheetNames.index("reference docs")
        except:
            return 0
        workSheet = workBook.worksheets[index]
        rows = workSheet.iter_rows(min_row=1, max_row=5)
        for cellRowTuple in rows:
            for cell in cellRowTuple:
                if str(cell.value).strip().casefold() == "name":
                    colNameIndex = cell.column
                if str(cell.value).strip().casefold() == "reference":
                    colReferenceIndex = cell.column
        for rowIndex in range(1, workSheet.max_row):
            if str(workSheet.cell(row = rowIndex, column = colNameIndex).value).casefold().strip() == "TFD".casefold().strip():
                if isinstance(workSheet.cell(row = rowIndex, column = colReferenceIndex).value, str) and workSheet.cell(row = rowIndex, column = colReferenceIndex).value.strip():
                    return 1
                else:
                    return 0
        return 0

#Requirements for [DOC4]

    def Test_02043_18_04939_STRUCT_0400_XLS(self, workBook):

        sheetNames = [x.casefold() for x in workBook.sheet_names()]
        if "table" in sheetNames or "tableau" in sheetNames:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0400_XLSX_XLSM(self, workBook):

        sheetNames = [x.casefold() for x in workBook.sheetnames]
        if "table" in sheetNames or "tableau" in sheetNames:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0410_XLS(self, workBook):

        # get table sheet

        sheetNames = workBook.sheet_names()
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("table")
        except:
            try:
                index = sheetNames.index("tableau")
            except:
                return 0
        workSheet = workBook.sheet_by_index(index)

        list_test = list()
        errorColValueList = list()

        for i in range(2, 3):
            for j in range(0, 35):
                dict = {}
                dict['1'] = workSheet.cell(i - 2, j).value
                dict['2'] = workSheet.cell(i - 1, j).value
                dict['3'] = workSheet.cell(i, j).value
                list_test.append(dict)

        # load reference file

        fileName = self.download_DOC4(self.DOC4Link)

        wb_ref = xlrd.open_workbook(fileName)

        sheetNames = wb_ref.sheet_names()
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("table")
        except:
            index = sheetNames.index("tableau")

        workSheet = wb_ref.sheet_by_index(index)

        list_ref = list()

        for i in range(2, 3):
            for j in range(0, 35):
                dict = {}
                dict['1'] = workSheet.cell(i - 2, j).value
                dict['2'] = workSheet.cell(i - 1, j).value
                dict['3'] = workSheet.cell(i, j).value
                list_ref.append(dict)

        for element in list_ref:
            if element in list_test:
                pass
            else:
                errorColValueList.append(element['3'] + ", ")

        if errorColValueList:
            errorColValueString = "In the sheet “Table” (or “tableau”), the column(s) "
            errorColValueString = errorColValueString + "".join(errorColValueList)
            errorColValueString = errorColValueString + " is(are) not present or not written correctly as in the document [DOC3]."
            return errorColValueString
        else:
            return 1

    def Test_02043_18_04939_STRUCT_0410_XLSX_XLSM(self, workBook):


    # get table sheet

       sheetNames = workBook.sheetnames
       sheetNames = [x.casefold() for x in sheetNames]
       try:
           index = sheetNames.index("table")
       except:
           try:
               index = sheetNames.index("tableau")
           except:
               return 0

       workSheet =  workBook.worksheets[index]

       list_test = list()
       errorColValueList = list()

       for i in range(3, 4):
           for j in range(1, 36):
              dict = {}
              dict['1'] = workSheet.cell(i-2, j).value
              dict['2'] = workSheet.cell(i-1, j).value
              dict['3'] = workSheet.cell(i, j).value
              list_test.append(dict)

       for element in list_test:
         if element['1'] is None:
             element['1'] = ""
         if element['2'] is None:
             element['2'] = ""
         if element['3'] is None:
             element['3'] = ""

#load reference file

       fileName = self.download_DOC4(self.DOC4Link)

       wb_ref = xlrd.open_workbook(fileName)

       sheetNames = wb_ref.sheet_names()
       sheetNames = [x.casefold() for x in sheetNames]
       try:
          index = sheetNames.index("table")
       except:
          index = sheetNames.index("tableau")

       workSheet = wb_ref.sheet_by_index(index)

       list_ref = list()

       for i in range(2, 3):
          for j in range(0, 35):
             dict = {}
             dict['1'] = workSheet.cell(i - 2, j).value
             dict['2'] = workSheet.cell(i - 1, j).value
             dict['3'] = workSheet.cell(i, j).value
             list_ref.append(dict)

       for element in list_ref:
            if element in list_test:
                pass
            else:
                errorColValueList.append(element['3'] + ", ")

       if errorColValueList:
            errorColValueString = "In the sheet “Table” (or “tableau”), the column(s) "
            errorColValueString = errorColValueString + "".join(errorColValueList)
            errorColValueString = errorColValueString + " is(are) not present or not written correctly as in the document [DOC3]."
            return errorColValueString
       else:
            return 1

    def Test_02043_18_04939_STRUCT_0420_XLS(self, workBook):

        sheetNames = [x.casefold() for x in workBook.sheet_names()]
        if "diagnostic needs" in sheetNames:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0420_XLSX_XLSM(self, workBook):

        sheetNames = [x.casefold() for x in workBook.sheetnames]
        if "diagnostic needs" in sheetNames:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0430_XLS(self, workBook):

        cellNamesRow1 = ["Reference", "Version", "Label", "Description", "Situation during which the diagnosis is active",
                         "Technical Effect covers by the need", "Diversity", "Allocated to the system", "Upstream requirements",
                         "Taken into account", "comment"]

        # check if row 1 is OK

        sheetNames = workBook.sheet_names()
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("diagnostic needs")
        except:
            return 0
        workSheet = workBook.sheet_by_index(index)

        cellNamesRow1 = [x.strip().casefold() for x in cellNamesRow1]

        rowValueList = list()
        rowCellsList = workSheet.row(0)
        for cell in rowCellsList:
            rowValueList.append(str(cell.value).casefold().strip())
        errorColValueList = list()

        for value in cellNamesRow1:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            else:
                errorColValueList.append(value + ", ")

        if errorColValueList:
            errorColValueString = "In the sheet “diagnostic needs” the column(s) "
            errorColValueString = errorColValueString + "".join(errorColValueList)
            errorColValueString = errorColValueString + " is(are) not present or not written correctly as in the document [DOC3]."
            return errorColValueString
        else:
            return 1

    def Test_02043_18_04939_STRUCT_0430_XLSX_XLSM(self, workBook):

        cellNamesRow1 = ["Reference", "Version", "Label", "Description", "Situation during which the diagnosis is active",
                         "Technical Effect covers by the need", "Diversity", "Allocated to the system", "Upstream requirements",
                         "Taken into account", "comment"]

        # check if row 1 is OK

        sheetNames = workBook.sheetnames
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("diagnostic needs")
        except:
            return 0
        workSheet = workBook.worksheets[index]

        cellNamesRow1 = [x.strip().casefold() for x in cellNamesRow1]

        rowValueList = list()
        rowCellsGenrator = workSheet.iter_rows(min_row=1, max_row=1)
        for cellTuple in rowCellsGenrator:
            for cell in cellTuple:
                rowValueList.append(str(cell.value).casefold().strip())
        errorColValueList = list()

        for value in cellNamesRow1:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            else:
                errorColValueList.append(value + ", ")

        if errorColValueList:
            errorColValueString = "In the sheet “diagnostic needs”, the column(s) "
            errorColValueString = errorColValueString + "".join(errorColValueList)
            errorColValueString = errorColValueString + " is(are) not present or not written correctly as in the document [DOC3]."
            return errorColValueString
        else:
            return 1

    def Test_02043_18_04939_STRUCT_0440_XLS(self, workBook):

        sheetNames = [x.casefold() for x in workBook.sheet_names()]
        if "customer effects" in sheetNames:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0440_XLSX_XLSM(self, workBook):

        sheetNames = [x.casefold() for x in workBook.sheetnames]
        if "customer effects" in sheetNames:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0450_XLS(self, workBook):

        cellNamesRow1 = ["Name", "Taken into account", "Diagnosticability synthesis", "Comments"]

        # check if row 1 is OK

        sheetNames = workBook.sheet_names()
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("customer effects")
        except:
            return 0
        workSheet = workBook.sheet_by_index(index)

        cellNamesRow1 = [x.strip().casefold() for x in cellNamesRow1]

        rowValueList = list()
        rowCellsList = workSheet.row(0)
        for cell in rowCellsList:
            rowValueList.append(str(cell.value).casefold().strip())
        errorColValueList = list()

        for value in cellNamesRow1:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            else:
                errorColValueList.append(value + ", ")

        if errorColValueList:
            errorColValueString = "In the sheet “customer effects” the column(s) "
            errorColValueString = errorColValueString + "".join(errorColValueList)
            errorColValueString = errorColValueString + " is(are) not present or not written correctly as in the document [DOC3]."
            return errorColValueString
        else:
            return 1

    def Test_02043_18_04939_STRUCT_0450_XLSX_XLSM(self, workBook):

        cellNamesRow1 = ["Name", "Taken into account", "Diagnosticability synthesis", "Comments"]

        # check if row 1 is OK

        sheetNames = workBook.sheetnames
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("customer effects")
        except:
            return 0
        workSheet = workBook.worksheets[index]

        cellNamesRow1 = [x.strip().casefold() for x in cellNamesRow1]

        rowValueList = list()
        rowCellsGenrator = workSheet.iter_rows(min_row=1, max_row=1)
        for cellTuple in rowCellsGenrator:
            for cell in cellTuple:
                rowValueList.append(str(cell.value).casefold().strip())
        errorColValueList = list()

        for value in cellNamesRow1:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            else:
                errorColValueList.append(value + ", ")

        if errorColValueList:
            errorColValueString = "In the sheet “customer effects”, the column(s) "
            errorColValueString = errorColValueString + "".join(errorColValueList)
            errorColValueString = errorColValueString + " is(are) not present or not written correctly as in the document [DOC3]."
            return errorColValueString
        else:
            return 1

    def Test_02043_18_04939_STRUCT_0460_XLS(self, workBook):

        sheetNames = [x.casefold() for x in workBook.sheet_names()]
        if "feared events" in sheetNames or "er" in sheetNames:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0460_XLSX_XLSM(self, workBook):

        sheetNames = [x.casefold() for x in workBook.sheetnames]
        if "feared events" in sheetNames or "er" in sheetNames:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0470_XLS(self, workBook):

        cellNamesRow1 = ["Description", "Reference", "Severity", "Level", "Taken into account",
                         "Justification for not taking into account the dread Event", "Commentaire"]

        # check if row 1 is OK

        sheetNames = workBook.sheet_names()
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("feared events")
        except:
            try:
                index = sheetNames.index("er")
            except:
                return 0
        workSheet = workBook.sheet_by_index(index)

        cellNamesRow1 = [x.strip().casefold() for x in cellNamesRow1]

        rowValueList = list()
        rowCellsList = workSheet.row(0)
        for cell in rowCellsList:
            rowValueList.append(str(cell.value).casefold().strip())
        errorColValueList = list()

        for value in cellNamesRow1:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            else:
                errorColValueList.append(value + ", ")

        if errorColValueList:
            errorColValueString = "In the sheet “feared events” the column(s) "
            errorColValueString = errorColValueString + "".join(errorColValueList)
            errorColValueString = errorColValueString + " is(are) not present or not written correctly as in the document [DOC3]."
            return errorColValueString
        else:
            return 1

    def Test_02043_18_04939_STRUCT_0470_XLSX_XLSM(self, workBook):

        cellNamesRow1 = ["Description", "Reference", "Severity", "Level", "Taken into account",
                         "Justification for not taking into account the dread Event", "Commentaire"]

        # check if row 1 is OK

        sheetNames = workBook.sheetnames
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("feared events")
        except:
            try:
                index = sheetNames.index("er")
            except:
                return 0

        workSheet = workBook.worksheets[index]

        cellNamesRow1 = [x.strip().casefold() for x in cellNamesRow1]

        rowValueList = list()
        rowCellsGenrator = workSheet.iter_rows(min_row=1, max_row=1)
        for cellTuple in rowCellsGenrator:
            for cell in cellTuple:
                rowValueList.append(str(cell.value).casefold().strip())
        errorColValueList = list()

        for value in cellNamesRow1:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            else:
                errorColValueList.append(value + ", ")

        if errorColValueList:
            errorColValueString = "In the sheet “feared events”, the column(s) "
            errorColValueString = errorColValueString + "".join(errorColValueList)
            errorColValueString = errorColValueString + " is(are) not present or not written correctly as in the document [DOC3]."
            return errorColValueString
        else:
            return 1

    def Test_02043_18_04939_STRUCT_0480_XLS(self, workBook):

        sheetNames = [x.casefold() for x in workBook.sheet_names()]
        if "system" in sheetNames or "système" in sheetNames:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0480_XLSX_XLSM(self, workBook):

        sheetNames = [x.casefold() for x in workBook.sheetnames]
        if "system" in sheetNames or "système" in sheetNames:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0490_XLS(self, workBook):

        cellNamesRow1 = ["Description", "Taken into account"]

        # check if row 1 is OK

        sheetNames = workBook.sheet_names()
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("system")
        except:
            try:
                index = sheetNames.index("système")
            except:
                return 0

        workSheet = workBook.sheet_by_index(index)

        cellNamesRow1 = [x.strip().casefold() for x in cellNamesRow1]

        rowValueList = list()
        rowCellsList = workSheet.row(0)
        for cell in rowCellsList:
            rowValueList.append(str(cell.value).casefold().strip())
        errorColValueList = list()

        for value in cellNamesRow1:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            else:
                errorColValueList.append(value + ", ")

        if errorColValueList:
            errorColValueString = "In the sheet “system” the column(s) "
            errorColValueString = errorColValueString + "".join(errorColValueList)
            errorColValueString = errorColValueString + " is(are) not present or not written correctly as in the document [DOC3]."
            return errorColValueString
        else:
            return 1

    def Test_02043_18_04939_STRUCT_0490_XLSX_XLSM(self, workBook):

        cellNamesRow1 = ["Description", "Taken into account"]

        # check if row 1 is OK

        sheetNames = workBook.sheetnames
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("system")
        except:
            try:
                index = sheetNames.index("système")
            except:
                return 0

        workSheet = workBook.worksheets[index]

        cellNamesRow1 = [x.strip().casefold() for x in cellNamesRow1]

        rowValueList = list()
        rowCellsGenrator = workSheet.iter_rows(min_row=1, max_row=1)
        for cellTuple in rowCellsGenrator:
            for cell in cellTuple:
                rowValueList.append(str(cell.value).casefold().strip())
        errorColValueList = list()

        for value in cellNamesRow1:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            else:
                errorColValueList.append(value + ", ")

        if errorColValueList:
            errorColValueString = "In the sheet “system”, the column(s) "
            errorColValueString = errorColValueString + "".join(errorColValueList)
            errorColValueString = errorColValueString + " is(are) not present or not written correctly as in the document [DOC3]."
            return errorColValueString
        else:
            return 1

    def Test_02043_18_04939_STRUCT_0500_XLS(self, workBook):

        sheetNames = [x.casefold() for x in workBook.sheet_names()]
        if "operation situation" in sheetNames:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0500_XLSX_XLSM(self, workBook):

        sheetNames = [x.casefold() for x in workBook.sheetnames]
        if "operation situation" in sheetNames:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0510_XLS(self, workBook):

        cellNamesRow1 = ["Description", "Taken into account", "Comments"]

        # check if row 1 is OK

        sheetNames = workBook.sheet_names()
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("operation situation")
        except:
            return 0
        workSheet = workBook.sheet_by_index(index)

        cellNamesRow1 = [x.strip().casefold() for x in cellNamesRow1]

        rowValueList = list()
        rowCellsList = workSheet.row(0)
        for cell in rowCellsList:
            rowValueList.append(str(cell.value).casefold().strip())
        errorColValueList = list()

        for value in cellNamesRow1:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            else:
                errorColValueList.append(value + ", ")

        if errorColValueList:
            errorColValueString = "In the sheet “operation situation” the column(s) "
            errorColValueString = errorColValueString + "".join(errorColValueList)
            errorColValueString = errorColValueString + " is(are) not present or not written correctly as in the document [DOC3]."
            return errorColValueString
        else:
            return 1

    def Test_02043_18_04939_STRUCT_0510_XLSX_XLSM(self, workBook):

        cellNamesRow1 = ["Description", "Taken into account", "Comments"]

        # check if row 1 is OK

        sheetNames = workBook.sheetnames
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("operation situation")
        except:
            return 0
        workSheet = workBook.worksheets[index]

        cellNamesRow1 = [x.strip().casefold() for x in cellNamesRow1]

        rowValueList = list()
        rowCellsGenrator = workSheet.iter_rows(min_row=1, max_row=1)
        for cellTuple in rowCellsGenrator:
            for cell in cellTuple:
                rowValueList.append(str(cell.value).casefold().strip())
        errorColValueList = list()

        for value in cellNamesRow1:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            else:
                errorColValueList.append(value + ", ")

        if errorColValueList:
            errorColValueString = "In the sheet “operation situation”, the column(s) "
            errorColValueString = errorColValueString + "".join(errorColValueList)
            errorColValueString = errorColValueString + " is(are) not present or not written correctly as in the document [DOC3]."
            return errorColValueString
        else:
            return 1

    def Test_02043_18_04939_STRUCT_0520_XLS(self, workBook):

        sheetNames = [x.casefold() for x in workBook.sheet_names()]
        if "req. of tech. effects" in sheetNames:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0520_XLSX_XLSM(self, workBook):

        sheetNames = [x.casefold() for x in workBook.sheetnames]
        if "req. of tech. effects" in sheetNames:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0530_XLS(self, workBook):
        cellNamesRow1 = ["Reference", "version", "Description", "technical effect", "Allocated to", "Tracability with the TSD"]

        sheetNames = workBook.sheet_names()
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("req. of tech. effects")
        except:
            return 0
        workSheet = workBook.sheet_by_index(index)

        cellNamesRow1 = [x.strip().casefold() for x in cellNamesRow1]

        rowValueList = list()
        rowCellsList = workSheet.row(0)
        for cell in rowCellsList:
            rowValueList.append(str(cell.value).casefold().strip())
        errorColValueList = list()

        for value in cellNamesRow1:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            else:
                errorColValueList.append(value + ", ")

        if errorColValueList:
            errorColValueString = "In the sheet “req. of tech. effects” the column(s) "
            errorColValueString = errorColValueString + "".join(errorColValueList)
            errorColValueString = errorColValueString + " is(are) not present or not written correctly as in the document [DOC3]."
            return errorColValueString
        else:
            return 1

    def Test_02043_18_04939_STRUCT_0530_XLSX_XLSM(self, workBook):
        cellNamesRow1 = ["Reference", "version", "Description", "technical effect", "Allocated to", "Tracability with the TSD"]

        # check if row 1 is OK

        sheetNames = workBook.sheetnames
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("req. of tech. effects")
        except:
            return 0
        workSheet = workBook.worksheets[index]

        cellNamesRow1 = [x.strip().casefold() for x in cellNamesRow1]

        rowValueList = list()
        rowCellsGenrator = workSheet.iter_rows(min_row=1, max_row=1)
        for cellTuple in rowCellsGenrator:
            for cell in cellTuple:
                rowValueList.append(str(cell.value).casefold().strip())
        errorColValueList = list()

        for value in cellNamesRow1:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            else:
                errorColValueList.append(value + ", ")

        if errorColValueList:
            errorColValueString = "In the sheet “req. of tech. effects”, the column(s) "
            errorColValueString = errorColValueString + "".join(errorColValueList)
            errorColValueString = errorColValueString + " is(are) not present or not written correctly as in the document [DOC3]."
            return errorColValueString
        else:
            return 1

    # Requirements for [DOC3]

    def Test_02043_18_04939_STRUCT_0100_XLS(self, workBook):

        sheetNames = [x.casefold() for x in workBook.sheet_names()]
        if "table" in sheetNames or "tableau" in sheetNames:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0100_XLSX_XLSM(self, workBook):

        sheetNames = [x.casefold() for x in workBook.sheetnames]
        if "table" in sheetNames or "tableau" in sheetNames:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0110_XLS(self, workBook):

        # get table sheet

        sheetNames = workBook.sheet_names()
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("table")
        except:
            try:
                index = sheetNames.index("tableau")
            except:
                return 0

        workSheet = workBook.sheet_by_index(index)

        list_test = list()
        errorColValueList = list()

        for i in range(3, 4):
            for j in range(0, 35):
                dict = {}
                dict['1'] = workSheet.cell(i - 1, j).value
                dict['2'] = workSheet.cell(i, j).value
                list_test.append(dict)

        # load reference file
        fileName = self.download_DOC3(self.DOC3Link)

        wb_ref = xlrd.open_workbook(fileName)

        sheetNames = wb_ref.sheet_names()
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("table")
        except:
            index = sheetNames.index("tableau")

        workSheet = wb_ref.sheet_by_index(index)

        list_ref = list()

        for i in range(3, 4):
            for j in range(0, 35):
                dict = {}
                dict['1'] = workSheet.cell(i - 1, j).value
                dict['2'] = workSheet.cell(i, j).value

                list_ref.append(dict)

        for element in list_ref:
            if element in list_test:
                pass
            else:
                errorColValueList.append(element['2'] + ", ")

        if errorColValueList:
            errorColValueString = "In the sheet “Table” (or “tableau”), the column(s) "
            errorColValueString = errorColValueString + "".join(errorColValueList)
            errorColValueString = errorColValueString + " is(are) not present or not written correctly as in the document [DOC3]."
            return errorColValueString
        else:
            return 1

    def Test_02043_18_04939_STRUCT_0110_XLSX_XLSM(self, workBook):

            # get table sheet

        sheetNames = workBook.sheetnames
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("table")
        except:
            try:
                index = sheetNames.index("tableau")
            except:
                return 0

        workSheet = workBook.worksheets[index]

        list_test = list()
        errorColValueList = list()

        for i in range(4, 5):
            for j in range(1, 36):
                dict = {}
                dict['1'] = workSheet.cell(i - 1, j).value
                dict['2'] = workSheet.cell(i, j).value
                list_test.append(dict)

        for element in list_test:
            if element['1'] is None:
                element['1'] = ""
            if element['2'] is None:
                element['2'] = ""
            # load reference file

        fileName = self.download_DOC3(self.DOC3Link)

        wb_ref = xlrd.open_workbook(fileName)

        sheetNames = wb_ref.sheet_names()
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("table")
        except:
            index = sheetNames.index("tableau")

        workSheet = wb_ref.sheet_by_index(index)

        list_ref = list()

        for i in range(3, 4):
            for j in range(0, 35):
                dict = {}
                dict['1'] = workSheet.cell(i - 1, j).value
                dict['2'] = workSheet.cell(i, j).value
                list_ref.append(dict)

        for element in list_ref:
            if element in list_test:
                 pass
            else:
                errorColValueList.append(element['2'] + ", ")

        if errorColValueList:
            errorColValueString = "In the sheet “Table” (or “tableau”), the column(s) "
            errorColValueString = errorColValueString + "".join(errorColValueList)
            errorColValueString = errorColValueString + " is(are) not present or not written correctly as in the document [DOC3]."
            return errorColValueString
        else:
            return 1

    def Test_02043_18_04939_STRUCT_0120_XLS(self, workBook):

        sheetNames = [x.casefold() for x in workBook.sheet_names()]
        if "codes défauts" in sheetNames:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0120_XLSX_XLSM(self, workBook):

        sheetNames = [x.casefold() for x in workBook.sheetnames]
        if "codes défauts" in sheetNames:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0130_XLS(self, workBook):

        cellNamesRow2 = ["Référence", "Version", "Code défaut", "libellé (signification)", "Flux Fonctionnel", "Description de la strategie pour détecter le défaut",
                           "Seuil de détection  /  valeur  du défaut ", "Temps de confirmation du défaut",
                           "Description de la strategie de disparition du défaut / Procedure à effectuer pour vérifier la disparition du défaut",
                           "Situation de vie véhicule pour faire remonter le code défaut", "Mode dégradé", "Taux de remonté du code défaut",
                           "Voyant", "Accès scantool", "Groupe de contextes associés", "Diversité", "Applicabilité usine",
                           "condition d'applicabilité en usine", "supporté par constituant (s)", "se référer au document spécifiant DRD : (réf & version)",
                           "Référence amont", "Version de la référence amont", "Pris en compte", "Justification de la modification",
                           "Validation"]

        cellNamesRow1 = ["Liste des codes défauts", "Applicabilité projet"]

        # check if row 2 is OK

        sheetNames = workBook.sheet_names()
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("codes défauts")
        except:
            return 0
        workSheet = workBook.sheet_by_index(index)
        errorColValueList = list()

        # check row 2

        cellNamesRow2 = [x.strip().casefold() for x in cellNamesRow2]

        rowValueList = list()
        rowCellsList = workSheet.row(1)
        for cell in rowCellsList:
            rowValueList.append(str(cell.value).casefold().strip())

        for value in cellNamesRow2:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            else:
                errorColValueList.append(value + ", ")

        # check row 1

        cellNamesRow1 = [x.strip().casefold() for x in cellNamesRow1]

        rowValueList = list()
        rowCellsList = workSheet.row(0)
        for cell in rowCellsList:
            rowValueList.append(str(cell.value).casefold().strip())

        for value in cellNamesRow1:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            else:
                errorColValueList.append(value + ", ")

        if errorColValueList:
            errorColValueString = "In the sheet “codes défauts”, the column(s) "
            errorColValueString = errorColValueString + "".join(errorColValueList)
            errorColValueString = errorColValueString + " is(are) not present or not written correctly as in the document [DOC3]."
            return errorColValueString
        else:
            return 1

    def Test_02043_18_04939_STRUCT_0130_XLSX_XLSM(self, workBook):

        cellNamesRow2 = ["Référence", "Version", "Code défaut", "libellé (signification)", "Flux Fonctionnel", "Description de la strategie pour détecter le défaut",
                         "Seuil de détection  /  valeur  du défaut ", "Temps de confirmation du défaut",
                         "Description de la strategie de disparition du défaut / Procedure à effectuer pour vérifier la disparition du défaut",
                         "Situation de vie véhicule pour faire remonter le code défaut", "Mode dégradé", "Taux de remonté du code défaut",
                         "Voyant", "Accès scantool", "Groupe de contextes associés", "Diversité", "Applicabilité usine",
                         "condition d'applicabilité en usine", "supporté par constituant (s)", "se référer au document spécifiant DRD : (réf & version)",
                          "Référence amont", "Version de la référence amont", "Pris en compte", "Justification de la modification",
                          "Validation"]

        cellNamesRow1 = ["Liste des codes défauts", "Applicabilité projet"]



        sheetNames = workBook.sheetnames
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("codes défauts")
        except:
            return 0
        workSheet = workBook.worksheets[index]

        # check if row 2 is OK

        cellNamesRow2 = [x.strip().casefold() for x in cellNamesRow2]

        rowValueList = list()
        rowCellsGenrator = workSheet.iter_rows(min_row=2, max_row=2)
        for cellTuple in rowCellsGenrator:
            for cell in cellTuple:
                rowValueList.append(str(cell.value).casefold().strip())
        errorColValueList = list()

        for value in cellNamesRow2:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            else:
                errorColValueList.append(value + ", ")


        #check if row1 is OK

        cellNamesRow1 = [x.strip().casefold() for x in cellNamesRow1]

        rowValueList = list()
        rowCellsGenrator = workSheet.iter_rows(min_row=1, max_row=1)
        for cellTuple in rowCellsGenrator:
            for cell in cellTuple:
                rowValueList.append(str(cell.value).casefold().strip())
        errorColValueList = list()

        for value in cellNamesRow1:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            else:
                errorColValueList.append(value + ", ")


        if errorColValueList:
            errorColValueString = "In the sheet “codes défauts”, the column(s) "
            errorColValueString = errorColValueString + "".join(errorColValueList)
            errorColValueString = errorColValueString + " is(are) not present or not written correctly as in the document [DOC3]."
            return errorColValueString
        else:
            return 1

    def Test_02043_18_04939_STRUCT_0140_XLS(self, workBook):

        sheetNames = [x.casefold() for x in workBook.sheet_names()]
        if "mesures et commandes" in sheetNames:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0140_XLSX_XLSM(self, workBook):

        sheetNames = [x.casefold() for x in workBook.sheetnames]
        if "mesures et commandes" in sheetNames:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0150_XLS(self, workBook):

        cellNamesRow2 = ["Référence", "Version", "Type (choix par menu)", "libellé (signification)", "Description",
                         "Situation pendant laquelle la mesure ou commande est utilisable", "Statut",
                         "Taux de fiabilité du test (50%, 100%)", "Flux fonctionnel", "Uniquement \npour O Control\nlecture \nsortie effective /commande",
                         "Diversité", "Applicabilité usine", "condition d'applicabilité en usine",
                         "supporté par constituant (s)", "se référer au document spécifiant DRD : (réf & version)",
                         "Référence amont", "Version de la référence amont", "Pris en compte", "Justification de la modification", "Validation"]

        cellNamesRow1 = ["mesures et commandes", "Applicabilité projet"]

        # check if row 2 is OK

        sheetNames = workBook.sheet_names()
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("mesures et commandes")
        except:
            return 0
        workSheet = workBook.sheet_by_index(index)
        errorColValueList = list()

        #check row 2

        cellNamesRow2 = [x.strip().casefold() for x in cellNamesRow2]

        rowValueList = list()
        rowCellsList = workSheet.row(1)
        for cell in rowCellsList:
            rowValueList.append(str(cell.value).casefold().strip())

        for value in cellNamesRow2:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            else:
                errorColValueList.append(value + ", ")

        #check row 1

        cellNamesRow1 = [x.strip().casefold() for x in cellNamesRow1]

        rowValueList = list()
        rowCellsList = workSheet.row(0)
        for cell in rowCellsList:
            rowValueList.append(str(cell.value).casefold().strip())

        for value in cellNamesRow1:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            else:
                errorColValueList.append(value + ", ")

        if errorColValueList:
            errorColValueString = "In the sheet “mesures et commandes”, the column(s) "
            errorColValueString = errorColValueString + "".join(errorColValueList)
            errorColValueString = errorColValueString + " is(are) not present or not written correctly as in the document [DOC3]."
            return errorColValueString
        else:
            return 1

    def Test_02043_18_04939_STRUCT_0150_XLSX_XLSM(self, workBook):

        cellNamesRow2 = ["Référence", "Version", "Type (choix par menu)", "libellé (signification)", "Description",
                          "Situation pendant laquelle la mesure ou commande est utilisable", "Statut",
                          "Taux de fiabilité du test (50%, 100%)", "Flux fonctionnel", "Uniquement \npour O Control\nlecture \nsortie effective /commande",
                          "Diversité", "Applicabilité usine", "condition d'applicabilité en usine",
                          "supporté par constituant (s)", "se référer au document spécifiant DRD : (réf & version)",
                          "Référence amont", "Version de la référence amont", "Pris en compte",
                          "Justification de la modification", "Validation"]

        cellNamesRow1 = ["mesures et commandes", "Applicabilité projet"]

        # check if row 2 is OK


        sheetNames = workBook.sheetnames
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("mesures et commandes")
        except:
            return 0
        workSheet = workBook.worksheets[index]

        # check if row 2 is OK

        cellNamesRow2 = [x.strip().casefold() for x in cellNamesRow2]

        rowValueList = list()
        rowCellsGenrator = workSheet.iter_rows(min_row=2, max_row=2)
        for cellTuple in rowCellsGenrator:
            for cell in cellTuple:
                rowValueList.append(str(cell.value).casefold().strip())
        errorColValueList = list()

        for value in cellNamesRow2:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            else:
                errorColValueList.append(value + ", ")


        #check if row1 is OK

        cellNamesRow1 = [x.strip().casefold() for x in cellNamesRow1]

        rowValueList = list()
        rowCellsGenrator = workSheet.iter_rows(min_row=1, max_row=1)
        for cellTuple in rowCellsGenrator:
            for cell in cellTuple:
                rowValueList.append(str(cell.value).casefold().strip())
        errorColValueList = list()

        for value in cellNamesRow1:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            else:
                errorColValueList.append(value + ", ")


        if errorColValueList:
            errorColValueString = "In the sheet “mesures et commandes”, the column(s) "
            errorColValueString = errorColValueString + "".join(errorColValueList)
            errorColValueString = errorColValueString + " is(are) not present or not written correctly as in the document [DOC3]."
            return errorColValueString
        else:
            return 1

    def Test_02043_18_04939_STRUCT_0160_XLS(self, workBook):

        sheetNames = [x.casefold() for x in workBook.sheet_names()]
        if "diagnostic débarqués" in sheetNames:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0160_XLSX_XLSM(self, workBook):

        sheetNames = [x.casefold() for x in workBook.sheetnames]
        if "diagnostic débarqués" in sheetNames:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0170_XLS(self, workBook):

        cellNamesRow2 = ["Référence", "Version", "libellé (signification)", "Description", "Taux de fiabilité du test (50%, 100%)",
                          "Applicabilité Usine", "se référer au document spécifiant : (réf & version)",
                          "Pris en compte", "Justification de la modification", "Validation"]

        # check if row 2 is OK

        sheetNames = workBook.sheet_names()
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("diagnostic débarqués")
        except:
            return 0
        workSheet = workBook.sheet_by_index(index)

        cellNamesRow2 = [x.strip().casefold() for x in cellNamesRow2]

        rowValueList = list()
        rowCellsList = workSheet.row(1)
        for cell in rowCellsList:
            rowValueList.append(str(cell.value).casefold().strip())
        errorColValueList = list()

        for value in cellNamesRow2:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            else:
                errorColValueList.append(value + ", ")

        if errorColValueList:
            errorColValueString = "In the sheet “diagnostic débarqués” the column(s) "
            errorColValueString = errorColValueString + "".join(errorColValueList)
            errorColValueString = errorColValueString + " is(are) not present or not written correctly as in the document [DOC3]."
            return errorColValueString
        else:
            return 1

    def Test_02043_18_04939_STRUCT_0170_XLSX_XLSM(self, workBook):

        cellNamesRow1 = ["Référence", "Version", "libellé (signification)", "Description", "Taux de fiabilité du test (50%, 100%)",
                          "Applicabilité Usine", "se référer au document spécifiant : (réf & version)",
                          "Pris en compte", "Justification de la modification", "Validation"]


        # check if row 1 is OK

        sheetNames = workBook.sheetnames
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("diagnostic débarqués")
        except:
            return 0
        workSheet = workBook.worksheets[index]

        cellNamesRow1 = [x.strip().casefold() for x in cellNamesRow1]

        rowValueList = list()
        rowCellsGenrator = workSheet.iter_rows(min_row=2, max_row=2)
        for cellTuple in rowCellsGenrator:
            for cell in cellTuple:
                rowValueList.append(str(cell.value).casefold().strip())
        errorColValueList = list()

        for value in cellNamesRow1:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            else:
                errorColValueList.append(value + ", ")

        if errorColValueList:
            errorColValueString = "In the sheet “diagnostic débarqués”, the column(s) "
            errorColValueString = errorColValueString + "".join(errorColValueList)
            errorColValueString = errorColValueString + " is(are) not present or not written correctly as in the document [DOC3]."
            return errorColValueString
        else:
            return 1

    def Test_02043_18_04939_STRUCT_0180_XLS(self, workBook):

        sheetNames = [x.casefold() for x in workBook.sheet_names()]
        if "effets clients" in sheetNames:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0180_XLSX_XLSM(self, workBook):

        sheetNames = [x.casefold() for x in workBook.sheetnames]
        if "effets clients" in sheetNames:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0190_XLS(self, workBook):

        cellNamesRow2 = ["Noms", "Pris en compte", "Synthèse de la diagnosticabilité", "Justification de la modification"]

        cellNamesRow1= ["effets clients"]

        # check if row 2 is OK

        sheetNames = workBook.sheet_names()
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("effets clients")
        except:
            return 0
        workSheet = workBook.sheet_by_index(index)
        errorColValueList = list()

        # check row 2

        cellNamesRow2 = [x.strip().casefold() for x in cellNamesRow2]

        rowValueList = list()
        rowCellsList = workSheet.row(1)
        for cell in rowCellsList:
            rowValueList.append(str(cell.value).casefold().strip())

        for value in cellNamesRow2:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            else:
                errorColValueList.append(value + ", ")

        # check row 1

        cellNamesRow1 = [x.strip().casefold() for x in cellNamesRow1]

        rowValueList = list()
        rowCellsList = workSheet.row(0)
        for cell in rowCellsList:
            rowValueList.append(str(cell.value).casefold().strip())

        for value in cellNamesRow1:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            else:
                errorColValueList.append(value + ", ")

        if errorColValueList:
            errorColValueString = "In the sheet “effets clients” the column(s) "
            errorColValueString = errorColValueString + "".join(errorColValueList)
            errorColValueString = errorColValueString + " is(are) not present or not written correctly as in the document [DOC3]."
            return errorColValueString
        else:
            return 1

    def Test_02043_18_04939_STRUCT_0190_XLSX_XLSM(self, workBook):

        cellNamesRow2 = ["Noms", "Pris en compte", "Synthèse de la diagnosticabilité", "Justification de la modification"]

        cellNamesRow1 = ["effets clients"]

        # check if row 2 is OK

        sheetNames = workBook.sheetnames
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("effets clients")
        except:
            return 0
        workSheet = workBook.worksheets[index]

        # check if row 2 is OK

        cellNamesRow2 = [x.strip().casefold() for x in cellNamesRow2]

        rowValueList = list()
        rowCellsGenrator = workSheet.iter_rows(min_row=2, max_row=2)
        for cellTuple in rowCellsGenrator:
            for cell in cellTuple:
                rowValueList.append(str(cell.value).casefold().strip())
        errorColValueList = list()

        for value in cellNamesRow2:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            else:
                errorColValueList.append(value + ", ")

        # check if row1 is OK

        cellNamesRow1 = [x.strip().casefold() for x in cellNamesRow1]

        rowValueList = list()
        rowCellsGenrator = workSheet.iter_rows(min_row=1, max_row=1)
        for cellTuple in rowCellsGenrator:
            for cell in cellTuple:
                rowValueList.append(str(cell.value).casefold().strip())
        errorColValueList = list()

        for value in cellNamesRow1:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            else:
                errorColValueList.append(value + ", ")

        if errorColValueList:
            errorColValueString = "In the sheet “effets clients”, the column(s) "
            errorColValueString = errorColValueString + "".join(errorColValueList)
            errorColValueString = errorColValueString + " is(are) not present or not written correctly as in the document [DOC3]."
            return errorColValueString
        else:
            return 1

    def Test_02043_18_04939_STRUCT_0200_XLS(self, workBook):

        sheetNames = [x.casefold() for x in workBook.sheet_names()]
        if "er" in sheetNames:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0200_XLSX_XLSM(self, workBook):

        sheetNames = [x.casefold() for x in workBook.sheetnames]
        if "er" in sheetNames:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0210_XLS(self, workBook):

        cellNamesRow2 = ["nom", "désignation", "Gravité", "Pris en compte", "Justification de non prise en compte de l'ER",
                           "Justification de la modification"]

        cellNamesRow1 = ["liste des er"]


        # check if row 2 is OK


        sheetNames = workBook.sheet_names()
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("er")
        except:
            return 0
        workSheet = workBook.sheet_by_index(index)
        errorColValueList = list()

        # check row 2

        cellNamesRow2 = [x.strip().casefold() for x in cellNamesRow2]

        rowValueList = list()
        rowCellsList = workSheet.row(1)
        for cell in rowCellsList:
            rowValueList.append(str(cell.value).casefold().strip())

        for value in cellNamesRow2:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            else:
                errorColValueList.append(value + ", ")

        # check row 1

        cellNamesRow1 = [x.strip().casefold() for x in cellNamesRow1]

        rowValueList = list()
        rowCellsList = workSheet.row(0)
        for cell in rowCellsList:
            rowValueList.append(str(cell.value).casefold().strip())

        for value in cellNamesRow1:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            else:
                errorColValueList.append(value + ", ")

        if errorColValueList:
            errorColValueString = "In the sheet “er” the column(s) "
            errorColValueString = errorColValueString + "".join(errorColValueList)
            errorColValueString = errorColValueString + " is(are) not present or not written correctly as in the document [DOC3]."
            return errorColValueString
        else:
            return 1

    def Test_02043_18_04939_STRUCT_0210_XLSX_XLSM(self, workBook):

        cellNamesRow2 = ["nom", "désignation", "Gravité", "Pris en compte", "Justification de non prise en compte de l'ER",
                           "Justification de la modification"]

        cellNamesRow1 = ["liste des er"]

        # check if row 2 is OK

        sheetNames = workBook.sheetnames
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("er")
        except:
            return 0
        workSheet = workBook.worksheets[index]

        # check if row 2 is OK

        cellNamesRow2 = [x.strip().casefold() for x in cellNamesRow2]

        rowValueList = list()
        rowCellsGenrator = workSheet.iter_rows(min_row=2, max_row=2)
        for cellTuple in rowCellsGenrator:
            for cell in cellTuple:
                rowValueList.append(str(cell.value).casefold().strip())
        errorColValueList = list()

        for value in cellNamesRow2:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            else:
                errorColValueList.append(value + ", ")

        # check if row1 is OK

        cellNamesRow1 = [x.strip().casefold() for x in cellNamesRow1]

        rowValueList = list()
        rowCellsGenrator = workSheet.iter_rows(min_row=1, max_row=1)
        for cellTuple in rowCellsGenrator:
            for cell in cellTuple:
                rowValueList.append(str(cell.value).casefold().strip())
        errorColValueList = list()

        for value in cellNamesRow1:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            else:
                errorColValueList.append(value + ", ")

        if errorColValueList:
            errorColValueString = "In the sheet “er”, the column(s) "
            errorColValueString = errorColValueString + "".join(errorColValueList)
            errorColValueString = errorColValueString + " is(are) not present or not written correctly as in the document [DOC3]."
            return errorColValueString
        else:
            return 1

    def Test_02043_18_04939_STRUCT_0220_XLS(self, workBook):

        sheetNames = [x.casefold() for x in workBook.sheet_names()]
        if "constituants" in sheetNames:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0220_XLSX_XLSM(self, workBook):

        sheetNames = [x.casefold() for x in workBook.sheetnames]
        if "constituants" in sheetNames:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0230_XLS(self, workBook):

        cellNamesRow2 = ["Noms", "Description", "Taux de défaillance (en ppm)",
                          "Découpage PSA", "Pris en compte", "Justification de la modification"]

        #cellNamesRow1 = ["constituants"]

        # check if row 2 is OK


        sheetNames = workBook.sheet_names()
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("constituants")
        except:
            return 0
        workSheet = workBook.sheet_by_index(index)
        errorColValueList = list()

        # check row 2

        cellNamesRow2 = [x.strip().casefold() for x in cellNamesRow2]

        rowValueList = list()
        rowCellsList = workSheet.row(1)
        for cell in rowCellsList:
            rowValueList.append(str(cell.value).casefold().strip())

        for value in cellNamesRow2:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            else:
                errorColValueList.append(value + ", ")

        # check row 1
        #
        # cellNamesRow1 = [x.strip().casefold() for x in cellNamesRow1]
        #
        # rowValueList = list()
        # rowCellsList = workSheet.row(0)
        # for cell in rowCellsList:
        #     rowValueList.append(str(cell.value).casefold().strip())
        #
        # for value in cellNamesRow1:
        #     if value in rowValueList and rowValueList.count(value) == 1:
        #         pass
        #     else:
        #         errorColValueList.append(value + ", ")

        if errorColValueList:
            errorColValueString = "In the sheet “constituants” the column(s) "
            errorColValueString = errorColValueString + "".join(errorColValueList)
            errorColValueString = errorColValueString + " is(are) not present or not written correctly as in the document [DOC3]."
            return errorColValueString
        else:
            return 1

    def Test_02043_18_04939_STRUCT_0230_XLSX_XLSM(self, workBook):

        cellNamesRow2 = ["Noms", "Description", "Taux de défaillance (en ppm)",
                          "Découpage PSA", "Pris en compte", "Justification de la modification"]

        #cellNamesRow1 = ["constituants"]

        # check if row 2 is OK

        sheetNames = workBook.sheetnames
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("constituants")
        except:
            return 0
        workSheet = workBook.worksheets[index]

        # check if row 2 is OK

        cellNamesRow2 = [x.strip().casefold() for x in cellNamesRow2]

        rowValueList = list()
        rowCellsGenrator = workSheet.iter_rows(min_row=2, max_row=2)
        for cellTuple in rowCellsGenrator:
            for cell in cellTuple:
                rowValueList.append(str(cell.value).casefold().strip())
        errorColValueList = list()

        for value in cellNamesRow2:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            else:
                errorColValueList.append(value + ", ")

        # check if row1 is OK
        #
        # cellNamesRow1 = [x.strip().casefold() for x in cellNamesRow1]
        #
        # rowValueList = list()
        # rowCellsGenrator = workSheet.iter_rows(min_row=1, max_row=1)
        # for cellTuple in rowCellsGenrator:
        #     for cell in cellTuple:
        #         rowValueList.append(str(cell.value).casefold().strip())
        # errorColValueList = list()
        #
        # for value in cellNamesRow1:
        #     if value in rowValueList and rowValueList.count(value) == 1:
        #         pass
        #     else:
        #         errorColValueList.append(value + ", ")

        if errorColValueList:
            errorColValueString = "In the sheet “constituants”, the column(s) "
            errorColValueString = errorColValueString + "".join(errorColValueList)
            errorColValueString = errorColValueString + " is(are) not present or not written correctly as in the document [DOC3]."
            return errorColValueString
        else:
            return 1

    def Test_02043_18_04939_STRUCT_0240_XLS(self, workBook):

        sheetNames = [x.casefold() for x in workBook.sheet_names()]
        if "situations de vie" in sheetNames:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0240_XLSX_XLSM(self, workBook):

        sheetNames = [x.casefold() for x in workBook.sheetnames]
        if "situations de vie" in sheetNames:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0250_XLS(self, workBook):

        cellNamesRow2 = ["Situations de vie", "Justification de la modification"]

        cellNamesRow1 = ["situations de vie de la fonction ou du système:"]

        # check if row 2 is OK

        sheetNames = workBook.sheet_names()
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("situations de vie")
        except:
            return 0
        workSheet = workBook.sheet_by_index(index)
        errorColValueList = list()

        # check row 2

        cellNamesRow2 = [x.strip().casefold() for x in cellNamesRow2]

        rowValueList = list()
        rowCellsList = workSheet.row(1)
        for cell in rowCellsList:
            rowValueList.append(str(cell.value).casefold().strip())

        for value in cellNamesRow2:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            else:
                errorColValueList.append(value + ", ")

        # check row 1

        cellNamesRow1 = [x.strip().casefold() for x in cellNamesRow1]

        rowValueList = list()
        rowCellsList = workSheet.row(0)
        for cell in rowCellsList:
            rowValueList.append(str(cell.value).casefold().strip())

        for value in cellNamesRow1:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            else:
                errorColValueList.append(value + ", ")

        if errorColValueList:
            errorColValueString = "In the sheet “situations de vie” the column(s) "
            errorColValueString = errorColValueString + "".join(errorColValueList)
            errorColValueString = errorColValueString + " is(are) not present or not written correctly as in the document [DOC3]."
            return errorColValueString
        else:
            return 1

    def Test_02043_18_04939_STRUCT_0250_XLSX_XLSM(self, workBook):

        cellNamesRow2 = ["Situations de vie", "Justification de la modification"]

        cellNamesRow1 = ["situations de vie de la fonction ou du système:"]

        # check if row 2 is OK

        sheetNames = workBook.sheetnames
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("situations de vie")
        except:
            return 0
        workSheet = workBook.worksheets[index]

        # check if row 2 is OK

        cellNamesRow2 = [x.strip().casefold() for x in cellNamesRow2]

        rowValueList = list()
        rowCellsGenrator = workSheet.iter_rows(min_row=2, max_row=2)
        for cellTuple in rowCellsGenrator:
            for cell in cellTuple:
                rowValueList.append(str(cell.value).casefold().strip())
        errorColValueList = list()

        for value in cellNamesRow2:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            else:
                errorColValueList.append(value + ", ")

        # check if row1 is OK

        cellNamesRow1 = [x.strip().casefold() for x in cellNamesRow1]

        rowValueList = list()
        rowCellsGenrator = workSheet.iter_rows(min_row=1, max_row=1)
        for cellTuple in rowCellsGenrator:
            for cell in cellTuple:
                rowValueList.append(str(cell.value).casefold().strip())
        errorColValueList = list()

        for value in cellNamesRow1:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            else:
                errorColValueList.append(value + ", ")

        if errorColValueList:
            errorColValueString = "In the sheet “situations de vie”, the column(s) "
            errorColValueString = errorColValueString + "".join(errorColValueList)
            errorColValueString = errorColValueString + " is(are) not present or not written correctly as in the document [DOC3]."
            return errorColValueString
        else:
            return 1

    def Test_02043_18_04939_STRUCT_0260_XLS(self, workBook):

        sheetNames = [x.casefold() for x in workBook.sheet_names()]
        if "liste mdd" in sheetNames:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0260_XLSX_XLSM(self, workBook):

        sheetNames = [x.casefold() for x in workBook.sheetnames]
        if "liste mdd" in sheetNames:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0270_XLS(self, workBook):

        cellNamesRow2 = ["Modes dégradés:", "Justification de la modification"]

        cellNamesRow1 = ["modes dégradés:"]


        # check if row 2 is OK

        sheetNames = workBook.sheet_names()
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("liste mdd")
        except:
            return 0
        workSheet = workBook.sheet_by_index(index)
        errorColValueList = list()

        # check row 2

        cellNamesRow2 = [x.strip().casefold() for x in cellNamesRow2]

        rowValueList = list()
        rowCellsList = workSheet.row(1)
        for cell in rowCellsList:
            rowValueList.append(str(cell.value).casefold().strip())

        for value in cellNamesRow2:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            else:
                errorColValueList.append(value + ", ")

        # check row 1

        cellNamesRow1 = [x.strip().casefold() for x in cellNamesRow1]

        rowValueList = list()
        rowCellsList = workSheet.row(0)
        for cell in rowCellsList:
            rowValueList.append(str(cell.value).casefold().strip())

        for value in cellNamesRow1:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            else:
                errorColValueList.append(value + ", ")

        if errorColValueList:
            errorColValueString = "In the sheet “liste mdd” the column(s) "
            errorColValueString = errorColValueString + "".join(errorColValueList)
            errorColValueString = errorColValueString + " is(are) not present or not written correctly as in the document [DOC3]."
            return errorColValueString
        else:
            return 1

    def Test_02043_18_04939_STRUCT_0270_XLSX_XLSM(self, workBook):

        cellNamesRow2 = ["Modes dégradés:", "Justification de la modification"]

        cellNamesRow1 = ["modes dégradés:"]

        # check if row 2 is OK

        sheetNames = workBook.sheetnames
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("liste mdd")
        except:
            return 0
        workSheet = workBook.worksheets[index]

        # check if row 2 is OK

        cellNamesRow2 = [x.strip().casefold() for x in cellNamesRow2]

        rowValueList = list()
        rowCellsGenrator = workSheet.iter_rows(min_row=2, max_row=2)
        for cellTuple in rowCellsGenrator:
            for cell in cellTuple:
                rowValueList.append(str(cell.value).casefold().strip())
        errorColValueList = list()

        for value in cellNamesRow2:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            else:
                errorColValueList.append(value + ", ")

        # check if row1 is OK

        cellNamesRow1 = [x.strip().casefold() for x in cellNamesRow1]

        rowValueList = list()
        rowCellsGenrator = workSheet.iter_rows(min_row=1, max_row=1)
        for cellTuple in rowCellsGenrator:
            for cell in cellTuple:
                rowValueList.append(str(cell.value).casefold().strip())
        errorColValueList = list()

        for value in cellNamesRow1:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            else:
                errorColValueList.append(value + ", ")

        if errorColValueList:
            errorColValueString = "In the sheet “liste mdd”, the column(s) "
            errorColValueString = errorColValueString + "".join(errorColValueList)
            errorColValueString = errorColValueString + " is(are) not present or not written correctly as in the document [DOC3]."
            return errorColValueString
        else:
            return 1


# Requirements for [DOC5]

    def Test_02043_18_04939_STRUCT_0700_XLS(self, workBook):
        sheetNames = [x.casefold() for x in workBook.sheet_names()]
        if "table" in sheetNames or "tableau" in sheetNames:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0700_XLSX_XLSM(self, workBook):

        sheetNames = [x.casefold() for x in workBook.sheetnames]
        if "table" in sheetNames or "tableau" in sheetNames:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0710_XLS(self, workBook):

        cellNamesRow3 = ["Reference", "Version", "Document of reference", "Variant/\noption",
                         "Sub-function of the system incriminated",
                         "Module / Group of parts", "Defective part", "Contribution to fonctionnality",
                         "Logical failure mode",
                         "Physical failure mode", "Weight", "Situation", "Detailed situation",
                         "Link to another DST",
                         "Technical effect", "Customer effect", "Feared events", "Degraded mode",
                         "HMI\n(Indicator lights/messages)", "Data Trouble code", "Mislead Data trouble code",
                         "Read data or I/O control", "decision criterion", "Non-embedded diagnosis",
                         "decision criterion", "Action on the incriminated part", "to do list / Comments",
                         "FMEA reference"]

        cellHeaderFailureAnalysis = ["Document of reference", "Variant/\noption",
                                     "Sub-function of the system incriminated",
                                     "Module / Group of parts", "Defective part", "Contribution to fonctionnality",
                                     "Logical failure mode",
                                     "Physical failure mode", "Weight", "Situation", "Detailed situation",
                                     "Link to another DST"]
        cellHeaderFailureAnalysisPosition = []
        cellHeaderCustomerEffects = ["Technical effect", "Customer effect", "Feared events", "Degraded mode",
                                     "HMI\n(Indicator lights/messages)"]
        cellHeaderCustomerEffectsPosition = []
        cellHeaderElementLeading = ["Data Trouble code", "Mislead Data trouble code",
                                    "Read data or I/O control", "decision criterion", "Non-embedded diagnosis"]
        cellHeaderElementLeadingPosition = []
        cellHeaderFollowUp = ["Action on the incriminated part", "to do list / Comments",
                              "FMEA reference"]
        cellHeaderFollowUpPosition = []

        cellNamesRow3 = [x.strip().casefold() for x in cellNamesRow3]
        cellHeaderFailureAnalysis = [x.strip().casefold() for x in cellHeaderFailureAnalysis]
        cellHeaderCustomerEffects = [x.strip().casefold() for x in cellHeaderCustomerEffects]
        cellHeaderElementLeading = [x.strip().casefold() for x in cellHeaderElementLeading]
        cellHeaderFollowUp = [x.strip().casefold() for x in cellHeaderFollowUp]

        # get table sheet

        sheetNames = workBook.sheet_names()
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("table")
        except:
            try:
                index = sheetNames.index("tableau")
            except:
                return 0

        workSheet = workBook.sheet_by_index(index)

        # check if row 3 is ok

        rowValueList = list()
        rowCellsList = workSheet.row(1)
        for cell in rowCellsList:
            rowValueList.append(str(cell.value).casefold().strip())
       # rowValueList = [str(x).casefold().strip() for x in rowCellsList.value]
        errorColValueList = list()

        for value in cellNamesRow3:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            elif value == "decision criterion" and value in rowValueList and rowValueList.count(value) == 2:
                pass
            else:
                errorColValueList.append(value + ", ")

        rowCellsList = workSheet.row(0)
        tempList = list()
        for cell in rowCellsList:
            tempList.append(str(cell.value).casefold().strip())
        if not "Project applicability".casefold().strip() in tempList:
            errorColValueList.append("Project applicability")

        if errorColValueList:
            errorColValueString = "In the sheet “Table” (or “tableau”, the column(s) "
            errorColValueString = errorColValueString + "".join(errorColValueList)
            errorColValueString = errorColValueString + " is(are) not present or not written correctly as in the document [DOC3]."
            return errorColValueString

        # get index of different headers in sheet

        for value in cellHeaderFailureAnalysis:
            cellHeaderFailureAnalysisPosition.append(rowValueList.index(value))
        for value in cellHeaderCustomerEffects:
            cellHeaderCustomerEffectsPosition.append(rowValueList.index(value))
        for value in cellHeaderElementLeading:
            if value == "decision criterion":
                for index, string in enumerate(rowValueList):
                    if string == "decision criterion":
                        cellHeaderElementLeadingPosition.append(index)
            else:
                cellHeaderElementLeadingPosition.append(rowValueList.index(value))
        for value in cellHeaderFollowUp:
            cellHeaderFollowUpPosition.append(rowValueList.index(value))

        # sort index

        cellHeaderFailureAnalysisPosition.sort()
        cellHeaderCustomerEffectsPosition.sort()
        cellHeaderElementLeadingPosition.sort()
        cellHeaderFollowUpPosition.sort()

        # see if subcells of headers are together

        for listIndex in range(1, len(cellHeaderFailureAnalysisPosition)):
            if cellHeaderFailureAnalysisPosition[listIndex] - cellHeaderFailureAnalysisPosition[listIndex - 1] > 1:
                errorColValueString = "In the sheet “Table” (or “tableau”), the column(s) below Failure Analysis: "
                errorColValueString =[errorColValueString + x + ", " for x in cellHeaderFailureAnalysis]
                errorColValueString = errorColValueString + "are not adjacent"
                return errorColValueString

        for listIndex in range(1, len(cellHeaderCustomerEffectsPosition)):
            if cellHeaderCustomerEffectsPosition[listIndex] - cellHeaderCustomerEffectsPosition[listIndex - 1] > 1:
                errorColValueString = "In the sheet “Table” (or “tableau”), the column(s) below Customer Effects: "
                errorColValueString =[errorColValueString + x + ", " for x in cellHeaderCustomerEffects]
                errorColValueString = errorColValueString + "are not adjacent"
                return errorColValueString

        for listIndex in range(1, len(cellHeaderElementLeadingPosition)):
            if cellHeaderElementLeadingPosition[listIndex] - cellHeaderElementLeadingPosition[listIndex - 1] > 1:
                errorColValueString = "In the sheet “Table” (or “tableau”), the column(s) below ELEMENT LEADING TO THE DEFECTIVE PART: "
                errorColValueString = [errorColValueString + x + ", " for x in cellHeaderElementLeading]
                errorColValueString = errorColValueString + "are not adjacent"
                return errorColValueString

        for listIndex in range(1, len(cellHeaderFollowUpPosition)):
            if cellHeaderFollowUpPosition[listIndex] - cellHeaderFollowUpPosition[listIndex - 1] > 1:
                errorColValueString = "In the sheet “Table” (or “tableau”), the column(s) below FOLLOW-UP: "
                errorColValueString = [errorColValueString + x + ", " for x in cellHeaderFollowUp]
                errorColValueString = errorColValueString + "are not adjacent"
                return errorColValueString


        #check headers


        rowCellsList = workSheet.row(0)

        tempindex = cellHeaderFailureAnalysisPosition[0]
        tempcell = rowCellsList[tempindex]
        tempstring = str(tempcell.value)

        if str(rowCellsList[cellHeaderFailureAnalysisPosition[0]].value).casefold().strip() != "FAILURE ANALYSIS".casefold().strip():
            errorColValueList.append("FAILURE ANALYSIS")

        if str(rowCellsList[cellHeaderCustomerEffectsPosition[0]].value).casefold().strip() != "CUSTOMER EFFECTS".casefold().strip():
            errorColValueList.append("CUSTOMER EFFECTS")


        if str(rowCellsList[cellHeaderElementLeadingPosition[0]].value).casefold().strip() != "ELEMENT LEADING TO THE DEFECTIVE PART".casefold().strip():
            errorColValueList.append("ELEMENT LEADING TO THE DEFECTIVE PART")


        if str(rowCellsList[cellHeaderFollowUpPosition[0]].value).casefold().strip() != "FOLLOW-UP".casefold().strip():
            errorColValueList.append("FOLLOW-UP")


        if errorColValueList:
            errorColValueString = "In the sheet “Table” (or “tableau”), the column(s) "
            errorColValueString = errorColValueString + "".join(errorColValueList)
            errorColValueString = errorColValueString + " is(are) not present or not written correctly as in the document [DOC3]."
            return errorColValueString
        else:
            return 1

    def Test_02043_18_04939_STRUCT_0710_XLSX_XLSM(self, workBook):


        cellNamesRow3 = ["Reference", "Version", "Document of reference", "Variant/\noption", "Sub-function of the system incriminated",
                         "Module / Group of parts", "Defective part", "Contribution to fonctionnality", "Logical failure mode",
                         "Physical failure mode", "Weight", "Situation", "Detailed situation", "Link to another DST", "Technical effect", "Customer effect", "Feared events", "Degraded mode",
                         "HMI\n(Indicator lights/messages)", "Data Trouble code", "Mislead Data trouble code",
                         "Read data or I/O control", "decision criterion", "Non-embedded diagnosis",
                         "decision criterion", "Action on the incriminated part", "to do list / Comments", "FMEA reference"]

        cellHeaderFailureAnalysis = ["Document of reference", "Variant/\noption",
                                     "Sub-function of the system incriminated",
                                     "Module / Group of parts", "Defective part", "Contribution to fonctionnality",
                                     "Logical failure mode",
                                     "Physical failure mode", "Weight", "Situation", "Detailed situation",
                                     "Link to another DST"]
        cellHeaderFailureAnalysisPosition = []
        cellHeaderCustomerEffects = ["Technical effect", "Customer effect", "Feared events", "Degraded mode",
                                     "HMI\n(Indicator lights/messages)"]
        cellHeaderCustomerEffectsPosition = []
        cellHeaderElementLeading = ["Data Trouble code", "Mislead Data trouble code",
                                    "Read data or I/O control", "decision criterion", "Non-embedded diagnosis"]
        cellHeaderElementLeadingPosition = []
        cellHeaderFollowUp = ["Action on the incriminated part", "to do list / Comments",
                              "FMEA reference"]
        cellHeaderFollowUpPosition = []

        cellNamesRow3 = [x.strip().casefold() for x in cellNamesRow3]
        cellHeaderFailureAnalysis = [x.strip().casefold() for x in cellHeaderFailureAnalysis]
        cellHeaderCustomerEffects = [x.strip().casefold() for x in cellHeaderCustomerEffects]
        cellHeaderElementLeading = [x.strip().casefold() for x in cellHeaderElementLeading]
        cellHeaderFollowUp = [x.strip().casefold() for x in cellHeaderFollowUp]

        # get table sheet

        sheetNames = workBook.sheetnames
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("table")
        except:
            try:
                index = sheetNames.index("tableau")
            except:
                return 0

        workSheet = workBook.worksheets[index]

        # check if row 3 is ok

        rowValueList = list()
        rowCellsGenrator = workSheet.iter_rows(min_row=2, max_row=2)
        for cellTuple in rowCellsGenrator:
            for cell in cellTuple:
                rowValueList.append(str(cell.value).casefold().strip())
        errorColValueList = list()

        for value in cellNamesRow3:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            elif value == "decision criterion" and value in rowValueList and rowValueList.count(value) == 2:
                pass
            else:
                errorColValueList.append(value + ", ")

        rowCellsGenrator = workSheet.iter_rows(min_row=1, max_row=1)
        tempList = list()
        for cellTuple in rowCellsGenrator:
            for cell in cellTuple:
                tempList.append(str(cell.value).casefold().strip())
        if not "Project applicability".casefold().strip() in tempList:
            errorColValueList.append("Project applicability")

        if errorColValueList:
            errorColValueString = "In the sheet “Table” (or “tableau”), the column(s) "
            errorColValueString = errorColValueString + "".join(errorColValueList)
            errorColValueString = errorColValueString + " is(are) not present or not written correctly as in the document [DOC3]."
            return errorColValueString

        # get index of different headers in sheet

        for value in cellHeaderFailureAnalysis:
            cellHeaderFailureAnalysisPosition.append(rowValueList.index(value))
        for value in cellHeaderCustomerEffects:
            cellHeaderCustomerEffectsPosition.append(rowValueList.index(value))
        for value in cellHeaderElementLeading:
            if value == "decision criterion":
                for index, string in enumerate(rowValueList):
                    if string == "decision criterion":
                        cellHeaderElementLeadingPosition.append(index)
            else:
                cellHeaderElementLeadingPosition.append(rowValueList.index(value))
        for value in cellHeaderFollowUp:
            cellHeaderFollowUpPosition.append(rowValueList.index(value))

        # sort index

        cellHeaderFailureAnalysisPosition.sort()
        cellHeaderCustomerEffectsPosition.sort()
        cellHeaderElementLeadingPosition.sort()
        cellHeaderFollowUpPosition.sort()

        # see if subcells of headers are together

        for listIndex in range(1, len(cellHeaderFailureAnalysisPosition)):
            if cellHeaderFailureAnalysisPosition[listIndex] - cellHeaderFailureAnalysisPosition[listIndex - 1] > 1:
                errorColValueString = "In the sheet “Table” (or “tableau”), the column(s) below Failure Analysis: "
                errorColValueString = [errorColValueString + x + ", " for x in cellHeaderFailureAnalysis]
                errorColValueString = errorColValueString + "are not adjacent"
                return errorColValueString

        for listIndex in range(1, len(cellHeaderCustomerEffectsPosition)):
            if cellHeaderCustomerEffectsPosition[listIndex] - cellHeaderCustomerEffectsPosition[listIndex - 1] > 1:
                errorColValueString = "In the sheet “Table” (or “tableau”), the column(s) below Customer Effects: "
                errorColValueString = [errorColValueString + x + ", " for x in cellHeaderCustomerEffects]
                errorColValueString = errorColValueString + "are not adjacent"
                return errorColValueString

        for listIndex in range(1, len(cellHeaderElementLeadingPosition)):
            if cellHeaderElementLeadingPosition[listIndex] - cellHeaderElementLeadingPosition[listIndex - 1] > 1:
                errorColValueString = "In the sheet “Table” (or “tableau”), the column(s) below ELEMENT LEADING TO THE DEFECTIVE PART: "
                errorColValueString = [errorColValueString + x + ", " for x in cellHeaderElementLeading]
                errorColValueString = errorColValueString + "are not adjacent"
                return errorColValueString

        for listIndex in range(1, len(cellHeaderFollowUpPosition)):
            if cellHeaderFollowUpPosition[listIndex] - cellHeaderFollowUpPosition[listIndex - 1] > 1:
                errorColValueString = "In the sheet “Table” (or “tableau”), the column(s) below FOLLOW-UP: "
                errorColValueString = [errorColValueString + x + ", " for x in cellHeaderFollowUp]
                errorColValueString = errorColValueString + "are not adjacent"
                return errorColValueString

        # check headers


        rowCellsGenrator = workSheet.iter_rows(min_row=1, max_row=1)
        rowValueList = list()
        for cellTuple in rowCellsGenrator:
            for cell in cellTuple:
                rowValueList.append(str(cell.value).casefold().strip())

        if str(rowValueList[cellHeaderFailureAnalysisPosition[0]]).casefold().strip() != "FAILURE ANALYSIS".casefold().strip():
            errorColValueList.append("FAILURE ANALYSIS")

        if str(rowValueList[cellHeaderCustomerEffectsPosition[0]]).casefold().strip() != "CUSTOMER EFFECTS".casefold().strip():
            errorColValueList.append("CUSTOMER EFFECTS")

        if str(rowValueList[cellHeaderElementLeadingPosition[0]]).casefold().strip() != "ELEMENT LEADING TO THE DEFECTIVE PART".casefold().strip():
            errorColValueList.append("ELEMENT LEADING TO THE DEFECTIVE PART")

        if str(rowValueList[cellHeaderFollowUpPosition[0]]).casefold().strip() != "FOLLOW-UP".casefold().strip():
            errorColValueList.append("FOLLOW-UP")

        if errorColValueList:
            errorColValueString = "In the sheet “Table” (or “tableau”), the column(s) "
            errorColValueString = errorColValueString + "".join(errorColValueList)
            errorColValueString = errorColValueString + " is(are) not present or not written correctly as in the document [DOC3]."
            return errorColValueString
        else:
            return 1

    def Test_02043_18_04939_STRUCT_0720_XLS(self, workBook):
        sheetNames = [x.casefold() for x in workBook.sheet_names()]
        if "data trouble codes" in sheetNames or "codes défauts" in sheetNames:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0720_XLSX_XLSM(self, workBook):

        sheetNames = [x.casefold() for x in workBook.sheetnames]
        if "data trouble codes" in sheetNames or "codes défauts" in sheetNames:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0730_XLS(self, workBook):

        cellNamesRow1 = ["Reference", "Version", "Data trouble code", "Label", "Description of the qualification conditions",
                           "Detection threshold", "Qualification time","Description of the dequalification conditions / Operation to do to check if the defect disappeared",
                           "Conditions of the diagnostic activation", "Degraded mode", "Failure detection rate",
                           "Indicateur light", "Visibility of the failure with the Scantool", "Freeze Frame Class",
                            "Diversity", "Stored by the ECU", "Upstream requirements", "Taken into account", "B78", "DV", "projet X", "Projet Y"]

        # check if row 1 is OK

        # get table sheet

        sheetNames = workBook.sheet_names()
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("data trouble codes")
        except:
            try:
                index = sheetNames.index("codes défauts")
            except:
                return 0

        workSheet = workBook.sheet_by_index(index)

        cellNamesRow1 = [x.strip().casefold() for x in cellNamesRow1]

        rowValueList = list()
        rowCellsList = workSheet.row(0)
        for cell in rowCellsList:
            rowValueList.append(str(cell.value).casefold().strip())
        errorColValueList = list()

        for value in cellNamesRow1:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            else:
                errorColValueList.append(value + ", ")

        if errorColValueList:
            errorColValueString = "In the sheet “data trouble codes” (or “codes défauts”), the column(s) "
            errorColValueString = errorColValueString + "".join(errorColValueList)
            errorColValueString = errorColValueString + " is(are) not present or not written correctly as in the document [DOC3]."
            return errorColValueString
        else:
            return 1

    def Test_02043_18_04939_STRUCT_0730_XLSX_XLSM(self, workBook):

        cellNamesRow1 = ["Reference", "Version", "Data trouble code", "Label", "Description of the qualification conditions",
                           "Detection threshold", "Qualification time",
                           "Description of the dequalification conditions / Operation to do to check if the defect disappeared",
                           "Conditions of the diagnostic activation", "Degraded mode", "Failure detection rate",
                           "Indicateur light", "Visibility of the failure with the Scantool", "Freeze Frame Class",
                            "Diversity", "Stored by the ECU", "Upstream requirements", "Taken into account",
                           "B78", "DV", "projet X", "Projet Y"]

        # check if row 1 is OK

        sheetNames = workBook.sheetnames
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("data trouble codes")
        except:
            try:
                index = sheetNames.index("codes défauts")
            except:
                return 0

        workSheet = workBook.worksheets[index]

        cellNamesRow1 = [x.strip().casefold() for x in cellNamesRow1]

        rowValueList = list()
        rowCellsGenrator = workSheet.iter_rows(min_row=1, max_row=1)
        for cellTuple in rowCellsGenrator:
            for cell in cellTuple:
                rowValueList.append(str(cell.value).casefold().strip())
        errorColValueList = list()


        for value in cellNamesRow1:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            else:
                errorColValueList.append(value + ", ")


        if errorColValueList:
            errorColValueString = "In the sheet “data trouble codes” (or “codes défauts”), the column(s) "
            errorColValueString = errorColValueString + "".join(errorColValueList)
            errorColValueString = errorColValueString + " is(are) not present or not written correctly as in the document [DOC3]."
            return errorColValueString
        else:
            return 1

    def Test_02043_18_04939_STRUCT_0740_XLS(self, workBook):
        sheetNames = [x.casefold() for x in workBook.sheet_names()]
        if "read data and io control" in sheetNames  or "mesures et commandes" in sheetNames:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0740_XLSX_XLSM(self, workBook):

        sheetNames = [x.casefold() for x in workBook.sheetnames]
        if "read data and io control" in sheetNames or "mesures et commandes" in sheetNames:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0750_XLS(self, workBook):

        cellNamesRow1 = ["Reference", "Version", "Type of diagnosis", "Label", "Description", "Conditions of the diagnostic activation",
                           "Status", "Diversity", "Stored by the ECU", "Upstream requirements", "Taken into account",
                           "B78", "DV", "projet X", "Projet Y"]

        # check if row 1 is OK

        sheetNames = workBook.sheet_names()
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("read data and io control")
        except:
            try:
                index = sheetNames.index("mesures et commandes")
            except:
                return 0

        workSheet = workBook.sheet_by_index(index)

        cellNamesRow1 = [x.strip().casefold() for x in cellNamesRow1]

        rowValueList = list()
        rowCellsList = workSheet.row(0)
        for cell in rowCellsList:
            rowValueList.append(str(cell.value).casefold().strip())
        errorColValueList = list()

        for value in cellNamesRow1:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            else:
                errorColValueList.append(value + ", ")

        if errorColValueList:
            errorColValueString = "In the sheet “Read data and IO control” (or “mesures et commandes”), the column(s) "
            errorColValueString = errorColValueString + "".join(errorColValueList)
            errorColValueString = errorColValueString + " is(are) not present or not written correctly as in the document [DOC3]."
            return errorColValueString
        else:
            return 1

    def Test_02043_18_04939_STRUCT_0750_XLSX_XLSM(self, workBook):

        cellNamesRow1 = ["Reference", "Version", "Type of diagnosis", "Label", "Description", "Conditions of the diagnostic activation",
                           "Status", "Diversity", "Stored by the ECU", "Upstream requirements", "Taken into account",
                           "B78", "DV", "projet X", "Projet Y"]

        # check if row 1 is OK

        sheetNames = workBook.sheetnames
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("read data and io control")
        except:
            try:
                index = sheetNames.index("mesures et commandes")
            except:
                return 0

        workSheet = workBook.worksheets[index]

        cellNamesRow1 = [x.strip().casefold() for x in cellNamesRow1]

        rowValueList = list()
        rowCellsGenrator = workSheet.iter_rows(min_row=1, max_row=1)
        for cellTuple in rowCellsGenrator:
            for cell in cellTuple:
                rowValueList.append(str(cell.value).casefold().strip())
        errorColValueList = list()


        for value in cellNamesRow1:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            else:
                errorColValueList.append(value + ", ")


        if errorColValueList:
            errorColValueString = "In the sheet “Read data and IO control” (or “mesures et commandes”), the column(s) "
            errorColValueString = errorColValueString + "".join(errorColValueList)
            errorColValueString = errorColValueString + " is(are) not present or not written correctly as in the document [DOC3]."
            return errorColValueString
        else:
            return 1

    def Test_02043_18_04939_STRUCT_0760_XLS(self, workBook):
        sheetNames = [x.casefold() for x in workBook.sheet_names()]
        if "not embedded diagnosis" in sheetNames:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0760_XLSX_XLSM(self, workBook):

        sheetNames = [x.casefold() for x in workBook.sheetnames]
        if "not embedded diagnosis" in sheetNames:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0770_XLS(self, workBook):

        cellNamesRow1 = ["Reference", "Version", "Label", "Description", "Upstream requirements",
                           "Taken into account", "B78", "DV", "projet X", "Projet Y"]

        # check if row 1 is OK

        sheetNames = workBook.sheet_names()
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("not embedded diagnosis")
        except:
            return 0

        workSheet = workBook.sheet_by_index(index)

        cellNamesRow1 = [x.strip().casefold() for x in cellNamesRow1]

        rowValueList = list()
        rowCellsList = workSheet.row(0)
        for cell in rowCellsList:
            rowValueList.append(str(cell.value).casefold().strip())
        errorColValueList = list()

        for value in cellNamesRow1:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            else:
                errorColValueList.append(value + ", ")

        if errorColValueList:
            errorColValueString = "In the sheet “Not embedded diagnosis“, the column(s) "
            errorColValueString = errorColValueString + "".join(errorColValueList)
            errorColValueString = errorColValueString + " is(are) not present or not written correctly as in the document [DOC3]."
            return errorColValueString
        else:
            return 1

    def Test_02043_18_04939_STRUCT_0770_XLSX_XLSM(self, workBook):

        cellNamesRow1 = ["Reference", "Version", "Label", "Description", "Upstream requirements",
                           "Taken into account", "B78", "DV", "projet X", "Projet Y"]

        # check if row 1 is OK

        sheetNames = workBook.sheetnames
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("not embedded diagnosis")
        except:
            return 0

        workSheet = workBook.worksheets[index]

        cellNamesRow1 = [x.strip().casefold() for x in cellNamesRow1]

        rowValueList = list()
        rowCellsGenrator = workSheet.iter_rows(min_row=1, max_row=1)
        for cellTuple in rowCellsGenrator:
            for cell in cellTuple:
                rowValueList.append(str(cell.value).casefold().strip())
        errorColValueList = list()

        for value in cellNamesRow1:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            else:
                errorColValueList.append(value + ", ")

        if errorColValueList:
            errorColValueString = "In the sheet “Not embedded diagnosis“, the column(s) "
            errorColValueString = errorColValueString + "".join(errorColValueList)
            errorColValueString = errorColValueString + " is(are) not present or not written correctly as in the document [DOC3]."
            return errorColValueString
        else:
            return 1

    def Test_02043_18_04939_STRUCT_0780_XLS(self, workBook):
        sheetNames = [x.casefold() for x in workBook.sheet_names()]
        if "customer effect" in sheetNames or "effets clients" in sheetNames:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0780_XLSX_XLSM(self, workBook):

        sheetNames = [x.casefold() for x in workBook.sheetnames]
        if "customer effect" in sheetNames or "effets clients" in sheetNames:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0790_XLS(self, workBook):

        cellNamesRow1 = ["Name", "Taken into account", "Diagnosticability synthesis"]

        # check if row 1 is OK

        sheetNames = workBook.sheet_names()
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("customer effect")
        except:
            try:
                index = sheetNames.index("effets clients")
            except:
                return 0

        workSheet = workBook.sheet_by_index(index)

        cellNamesRow1 = [x.strip().casefold() for x in cellNamesRow1]

        rowValueList = list()
        rowCellsList = workSheet.row(0)
        for cell in rowCellsList:
            rowValueList.append(str(cell.value).casefold().strip())
        errorColValueList = list()

        for value in cellNamesRow1:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            else:
                errorColValueList.append(value + ", ")

        if errorColValueList:
            errorColValueString = "In the sheet “Customer effect” (or “Effets clients”), the column(s) "
            errorColValueString = errorColValueString + "".join(errorColValueList)
            errorColValueString = errorColValueString + " is(are) not present or not written correctly as in the document [DOC3]."
            return errorColValueString
        else:
            return 1

    def Test_02043_18_04939_STRUCT_0790_XLSX_XLSM(self, workBook):

        cellNamesRow1 = ["Name", "Taken into account", "Diagnosticability synthesis"]

        # check if row 1 is OK

        sheetNames = workBook.sheetnames
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("customer effect")
        except:
            try:
                index = sheetNames.index("effets clients")
            except:
                return 0

        workSheet = workBook.worksheets[index]

        cellNamesRow1 = [x.strip().casefold() for x in cellNamesRow1]

        rowValueList = list()
        rowCellsGenrator = workSheet.iter_rows(min_row=1, max_row=1)
        for cellTuple in rowCellsGenrator:
            for cell in cellTuple:
                rowValueList.append(str(cell.value).casefold().strip())
        errorColValueList = list()

        for value in cellNamesRow1:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            else:
                errorColValueList.append(value + ", ")

        if errorColValueList:
            errorColValueString = "In the sheet “Customer effect” (or “Effets clients”), the column(s) "
            errorColValueString = errorColValueString + "".join(errorColValueList)
            errorColValueString = errorColValueString + " is(are) not present or not written correctly as in the document [DOC3]."
            return errorColValueString
        else:
            return 1

    def Test_02043_18_04939_STRUCT_0800_XLS(self, workBook):
        sheetNames = [x.casefold() for x in workBook.sheet_names()]
        if "feared events" in sheetNames or "er" in sheetNames:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0800_XLSX_XLSM(self, workBook):

        sheetNames = [x.casefold() for x in workBook.sheetnames]
        if "feared events" in sheetNames or "er" in sheetNames:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0810_XLS(self, workBook):

        cellNamesRow1 = ["Description", "Description", "Severity", "Taken into account",
                           "Justification for not taking into account the dread Event"]

        # check if row 1 is OK

        sheetNames = workBook.sheet_names()
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("feared events")
        except:
            try:
                index = sheetNames.index("er")
            except:
                return 0

        workSheet = workBook.sheet_by_index(index)

        cellNamesRow1 = [x.strip().casefold() for x in cellNamesRow1]

        rowValueList = list()
        rowCellsList = workSheet.row(0)
        for cell in rowCellsList:
            rowValueList.append(str(cell.value).casefold().strip())
        errorColValueList = list()

        for value in cellNamesRow1:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            else:
                errorColValueList.append(value + ", ")

        if errorColValueList:
            errorColValueString = "In the sheet “Feared events” (or “ER”), the column(s) "
            errorColValueString = errorColValueString + "".join(errorColValueList)
            errorColValueString = errorColValueString + " is(are) not present or not written correctly as in the document [DOC3]."
            return errorColValueString
        else:
            return 1

    def Test_02043_18_04939_STRUCT_0810_XLSX_XLSM(self, workBook):

        cellNamesRow1 = ["Description", "Description", "Severity", "Taken into account",
                           "Justification for not taking into account the dread Event"]

        # check if row 1 is OK

        sheetNames = workBook.sheetnames
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("feared events")
        except:
            try:
                index = sheetNames.index("er")
            except:
                return 0

        workSheet = workBook.worksheets[index]
        cellNamesRow1 = [x.strip().casefold() for x in cellNamesRow1]

        rowValueList = list()
        rowCellsGenrator = workSheet.iter_rows(min_row=1, max_row=1)
        for cellTuple in rowCellsGenrator:
            for cell in cellTuple:
                rowValueList.append(str(cell.value).casefold().strip())
        errorColValueList = list()

        for value in cellNamesRow1:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            else:
                errorColValueList.append(value + ", ")

        if errorColValueList:
            errorColValueString = "In the sheet “Feared events” (or “ER”), the column(s) "
            errorColValueString = errorColValueString + "".join(errorColValueList)
            errorColValueString = errorColValueString + " is(are) not present or not written correctly as in the document [DOC3]."
            return errorColValueString
        else:
            return 1

    def Test_02043_18_04939_STRUCT_0820_XLS(self, workBook):
        sheetNames = [x.casefold() for x in workBook.sheet_names()]
        if "parts" in sheetNames or "constituants" in sheetNames:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0820_XLSX_XLSM(self, workBook):

        sheetNames = [x.casefold() for x in workBook.sheetnames]
        if "parts" in sheetNames or "constituants" in sheetNames:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0830_XLS(self, workBook):

        cellNamesRow1 = ["Name", "Description", "Taken into account"]

        # check if row 1 is OK

        sheetNames = workBook.sheet_names()
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("parts")
        except:
            try:
                index = sheetNames.index("constituants")
            except:
                return 0

        workSheet = workBook.sheet_by_index(index)

        cellNamesRow1 = [x.strip().casefold() for x in cellNamesRow1]

        rowValueList = list()
        rowCellsList = workSheet.row(0)
        for cell in rowCellsList:
            rowValueList.append(str(cell.value).casefold().strip())
        errorColValueList = list()

        for value in cellNamesRow1:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            else:
                errorColValueList.append(value + ", ")

        if errorColValueList:
            errorColValueString = "In the sheet “Parts” (or “Constituants”), the column(s) "
            errorColValueString = errorColValueString + "".join(errorColValueList)
            errorColValueString = errorColValueString + " is(are) not present or not written correctly as in the document [DOC3]."
            return errorColValueString
        else:
            return 1

    def Test_02043_18_04939_STRUCT_0830_XLSX_XLSM(self, workBook):

        cellNamesRow1 = ["Name", "Description", "Taken into account"]

        # check if row 1 is OK

        sheetNames = workBook.sheetnames
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("parts")
        except:
            try:
                index = sheetNames.index("constituants")
            except:
                return 0

        workSheet = workBook.worksheets[index]

        cellNamesRow1 = [x.strip().casefold() for x in cellNamesRow1]

        rowValueList = list()
        rowCellsGenrator = workSheet.iter_rows(min_row=1, max_row=1)
        for cellTuple in rowCellsGenrator:
            for cell in cellTuple:
                rowValueList.append(str(cell.value).casefold().strip())
        errorColValueList = list()

        for value in cellNamesRow1:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            else:
                errorColValueList.append(value + ", ")

        if errorColValueList:
            errorColValueString = "In the sheet “Parts” (or “Constituants”), the column(s) "
            errorColValueString = errorColValueString + "".join(errorColValueList)
            errorColValueString = errorColValueString + " is(are) not present or not written correctly as in the document [DOC3]."
            return errorColValueString
        else:
            return 1

    def Test_02043_18_04939_STRUCT_0840_XLS(self, workBook):
        sheetNames = [x.casefold() for x in workBook.sheet_names()]
        if "situation" in sheetNames or "situations de vie" in sheetNames:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0840_XLSX_XLSM(self, workBook):

        sheetNames = [x.casefold() for x in workBook.sheetnames]
        if "situation" in sheetNames or "situations de vie" in sheetNames:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0850_XLS(self, workBook):

        cellNamesRow1 = ["Description", "Taken into account", "Comments"]

        # check if row 1 is OK

        sheetNames = workBook.sheet_names()
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("situation")
        except:
            try:
                index = sheetNames.index("situations de vie")
            except:
                return 0

        workSheet = workBook.sheet_by_index(index)

        cellNamesRow1 = [x.strip().casefold() for x in cellNamesRow1]

        rowValueList = list()
        rowCellsList = workSheet.row(0)
        for cell in rowCellsList:
            rowValueList.append(str(cell.value).casefold().strip())
        errorColValueList = list()

        for value in cellNamesRow1:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            else:
                errorColValueList.append(value + ", ")

        if errorColValueList:
            errorColValueString = "In the sheet “Situation” (or “situations de vie”), the column(s) "
            errorColValueString = errorColValueString + "".join(errorColValueList)
            errorColValueString = errorColValueString + " is(are) not present or not written correctly as in the document [DOC3]."
            return errorColValueString
        else:
            return 1

    def Test_02043_18_04939_STRUCT_0850_XLSX_XLSM(self, workBook):

        cellNamesRow1 = ["Description", "Taken into account", "Comments"]

        # check if row 1 is OK

        sheetNames = workBook.sheetnames
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("situation")
        except:
            try:
                index = sheetNames.index("situations de vie")
            except:
                return 0

        workSheet = workBook.worksheets[index]

        cellNamesRow1 = [x.strip().casefold() for x in cellNamesRow1]

        rowValueList = list()
        rowCellsGenrator = workSheet.iter_rows(min_row=1, max_row=1)
        for cellTuple in rowCellsGenrator:
            for cell in cellTuple:
                rowValueList.append(str(cell.value).casefold().strip())
        errorColValueList = list()

        for value in cellNamesRow1:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            else:
                errorColValueList.append(value + ", ")

        if errorColValueList:
            errorColValueString = "In the sheet “Situation” (or “situations de vie”), the column(s) "
            errorColValueString = errorColValueString + "".join(errorColValueList)
            errorColValueString = errorColValueString + " is(are) not present or not written correctly as in the document [DOC3]."
            return errorColValueString
        else:
            return 1

    def Test_02043_18_04939_STRUCT_0860_XLS(self, workBook):
        sheetNames = [x.casefold() for x in workBook.sheet_names()]
        if "degraded mode" in sheetNames or "liste mdd" in sheetNames:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0860_XLSX_XLSM(self, workBook):

        sheetNames = [x.casefold() for x in workBook.sheetnames]
        if "degraded mode" in sheetNames or "liste mdd" in sheetNames:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0870_XLS(self, workBook):

        cellNamesRow1 = ["Modes dégradés:", "Taken into account"]

        # check if row 1 is OK

        sheetNames = workBook.sheet_names()
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("degraded mode")
        except:
            try:
                index = sheetNames.index("liste mdd")
            except:
                return 0

        workSheet = workBook.sheet_by_index(index)

        cellNamesRow1 = [x.strip().casefold() for x in cellNamesRow1]

        rowValueList = list()
        rowCellsList = workSheet.row(0)
        for cell in rowCellsList:
            rowValueList.append(str(cell.value).casefold().strip())
        errorColValueList = list()

        for value in cellNamesRow1:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            else:
                errorColValueList.append(value + ", ")

        if errorColValueList:
            errorColValueString = "In the sheet “Degraded mode” (or “Liste MDD”), the column(s) "
            errorColValueString = errorColValueString + "".join(errorColValueList)
            errorColValueString = errorColValueString + " is(are) not present or not written correctly as in the document [DOC3]."
            return errorColValueString
        else:
            return 1

    def Test_02043_18_04939_STRUCT_0870_XLSX_XLSM(self, workBook):

        cellNamesRow1 = ["Modes dégradés:", "Taken into account"]

        # check if row 1 is OK

        sheetNames = workBook.sheetnames
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("degraded mode")
        except:
            try:
                index = sheetNames.index("liste mdd")
            except:
                return 0


        workSheet = workBook.worksheets[index]

        cellNamesRow1 = [x.strip().casefold() for x in cellNamesRow1]

        rowValueList = list()
        rowCellsGenrator = workSheet.iter_rows(min_row=1, max_row=1)
        for cellTuple in rowCellsGenrator:
            for cell in cellTuple:
                rowValueList.append(str(cell.value).casefold().strip())
        errorColValueList = list()

        for value in cellNamesRow1:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            else:
                errorColValueList.append(value + ", ")

        if errorColValueList:
            errorColValueString = "In the sheet “Degraded mode” (or “Liste MDD”), the column(s) "
            errorColValueString = errorColValueString + "".join(errorColValueList)
            errorColValueString = errorColValueString + " is(are) not present or not written correctly as in the document [DOC3]."
            return errorColValueString
        else:
            return 1

    def Test_02043_18_04939_STRUCT_0880_XLS(self, workBook):
        sheetNames = [x.casefold() for x in workBook.sheet_names()]
        if "technical effect" in sheetNames or "effets techniques" in sheetNames:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0880_XLSX_XLSM(self, workBook):

        sheetNames = [x.casefold() for x in workBook.sheetnames]
        if "technical effect" in sheetNames or "effets techniques" in sheetNames:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0890_XLS(self, workBook):

        cellNamesRow1 = ["Name", "Taken into account", "Upstream requirements"]

        # check if row 1 is OK

        sheetNames = workBook.sheet_names()
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("technical effect")
        except:
            try:
                index = sheetNames.index("effets techniques")
            except:
                return 0


        workSheet = workBook.sheet_by_index(index)

        cellNamesRow1 = [x.strip().casefold() for x in cellNamesRow1]

        rowValueList = list()
        rowCellsList = workSheet.row(0)
        for cell in rowCellsList:
            rowValueList.append(str(cell.value).casefold().strip())
        errorColValueList = list()

        for value in cellNamesRow1:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            else:
                errorColValueList.append(value + ", ")

        if errorColValueList:
            errorColValueString = "In the sheet “Technical effect “ (or “Effets techniques”), the column(s) "
            errorColValueString = errorColValueString + "".join(errorColValueList)
            errorColValueString = errorColValueString + " is(are) not present or not written correctly as in the document [DOC3]."
            return errorColValueString
        else:
            return 1

    def Test_02043_18_04939_STRUCT_0890_XLSX_XLSM(self, workBook):

        cellNamesRow1 = ["Name", "Taken into account", "Upstream requirements"]

        # check if row 1 is OK

        sheetNames = workBook.sheetnames
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("technical effect")
        except:
            try:
                index = sheetNames.index("effets techniques")
            except:
                return 0

        workSheet = workBook.worksheets[index]

        cellNamesRow1 = [x.strip().casefold() for x in cellNamesRow1]

        rowValueList = list()
        rowCellsGenrator = workSheet.iter_rows(min_row=1, max_row=1)
        for cellTuple in rowCellsGenrator:
            for cell in cellTuple:
                rowValueList.append(str(cell.value).casefold().strip())
        errorColValueList = list()

        for value in cellNamesRow1:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            else:
                errorColValueList.append(value + ", ")

        if errorColValueList:
            errorColValueString = "In the sheet “Technical effect “ (or “Effets techniques”), the column(s) "
            errorColValueString = errorColValueString + "".join(errorColValueList)
            errorColValueString = errorColValueString + " is(are) not present or not written correctly as in the document [DOC3]."
            return errorColValueString
        else:
            return 1

    def Test_02043_18_04939_STRUCT_0900_XLS(self, workBook):
        sheetNames = [x.casefold() for x in workBook.sheet_names()]
        if "variant" in sheetNames or "variantes" in sheetNames:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0900_XLSX_XLSM(self, workBook):

        sheetNames = [x.casefold() for x in workBook.sheetnames]
        if "variant" in sheetNames or "variantes" in sheetNames:
            return 1
        else:
            return 0

    def Test_02043_18_04939_STRUCT_0910_XLS(self, workBook):

        cellNamesRow1 = ["Name", "Description", "Taken into account"]

        # check if row 1 is OK

        sheetNames = workBook.sheet_names()
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("variant")
        except:
            try:
                index = sheetNames.index("variantes")
            except:
                return 0

        workSheet = workBook.sheet_by_index(index)

        cellNamesRow1 = [x.strip().casefold() for x in cellNamesRow1]

        rowValueList = list()
        rowCellsList = workSheet.row(0)
        for cell in rowCellsList:
            rowValueList.append(str(cell.value).casefold().strip())
        errorColValueList = list()

        for value in cellNamesRow1:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            else:
                errorColValueList.append(value + ", ")

        if errorColValueList:
            errorColValueString = "In the sheet “Variant “ (or “Variantes”), the column(s) "
            errorColValueString = errorColValueString + "".join(errorColValueList)
            errorColValueString = errorColValueString + " is(are) not present or not written correctly as in the document [DOC3]."
            return errorColValueString
        else:
            return 1

    def Test_02043_18_04939_STRUCT_0910_XLSX_XLSM(self, workBook):

        cellNamesRow1 = ["Name", "Description", "Taken into account"]

        # check if row 1 is OK

        sheetNames = workBook.sheetnames
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("variant")
        except:
            try:
                index = sheetNames.index("variantes")
            except:
                return 0

        workSheet = workBook.worksheets[index]

        cellNamesRow1 = [x.strip().casefold() for x in cellNamesRow1]

        rowValueList = list()
        rowCellsGenrator = workSheet.iter_rows(min_row=1, max_row=1)
        for cellTuple in rowCellsGenrator:
            for cell in cellTuple:
                rowValueList.append(str(cell.value).casefold().strip())
        errorColValueList = list()

        for value in cellNamesRow1:
            if value in rowValueList and rowValueList.count(value) == 1:
                pass
            else:
                errorColValueList.append(value + ", ")

        if errorColValueList:
            errorColValueString = "In the sheet “Variant “ (or “Variantes”), the column(s) "
            errorColValueString = errorColValueString + "".join(errorColValueList)
            errorColValueString = errorColValueString + " is(are) not present or not written correctly as in the document [DOC3]."
            return errorColValueString
        else:
            return 1

 #Wholeness

    def Test_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060_XLS(self, workBook):

        # get table sheet

        sheetNames = workBook.sheet_names()
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("table")
        except:
            try:
                index = sheetNames.index("tableau")
            except:
                return 2

        workSheet = workBook.sheet_by_index(index)
        rows = workSheet.get_rows()

        row_index = 0
        row_start = 10
        for row in rows:
            if row_index >= row_start:
                testflag = False
                for cell in row:
                    if cell.value:
                        testflag = True
                        break
                    else:
                        pass
                if testflag == True:
                    dictResult = {}
                    if not str(row[ref_col].value):
                        dictResult["0"] = str(colnum_string(ref_col + 1)) + str(row_index + 1)
                    if not str(row[ver_col].value):
                        dictResult["-1"] = str(colnum_string(ver_col + 1)) + str(row_index + 1)
                    if not str(row[defDet_col].value):
                        dictResult["-2"] = str(colnum_string(defDet_col + 1)) + str(row_index + 1)
                    if not str(row[defConst_col].value):
                        dictResult["-3"] = str(colnum_string(defConst_col + 1)) + str(row_index + 1)
                    if not str(row[sit_col].value):
                        dictResult["-4"] = str(colnum_string(sit_col + 1)) + str(row_index + 1)
                    if not str(row[eff_col].value):
                        dictResult["-5"] = str(colnum_string(eff_col + 1)) + str(row_index + 1)
                    if not str(row[code_col].value):
                        dictResult["-6"] = str(colnum_string(code_col + 1)) + str(row_index + 1)
                    if str(row[app_col].value) in {"NA", "X"}:
                        dictResult["-7"] = str(colnum_string(app_col + 1)) + str(row_index + 1)
                    else:
                        pass
                    if dictResult:
                        return dictResult
                else:
                    break
            if row_start == 10:
                for colindex,cell in enumerate(row):
                    if str(cell.value).casefold() in {"référence", "reference"}:
                        ref_col = colindex
                        ref_tuple = (row_index, row_index+2, ref_col, ref_col +1)
                        if ref_tuple in workSheet.merged_cells:
                            row_start = row_index +2
                        else:
                            row_start = row_index + 1
                    if str(cell.value).casefold() in {"version"}:
                        ver_col = colindex
                    if str(cell.value).casefold() in {"constituant défaillant détecté"}:
                        defDet_col = colindex
                    if str(cell.value).casefold() in {"défaillance constituant"}:
                        defConst_col = colindex
                    if str(cell.value).casefold() in {"situation de vie client"}:
                        sit_col = colindex
                    if str(cell.value).casefold() in {"effet(s) client(s)"}:
                        eff_col = colindex
                    if str(cell.value).casefold() in {"code défaut"}:
                        code_col = colindex
                    if str(cell.value).casefold() in {"applicabilité projet"}:
                        app_col = colindex

            row_index = row_index + 1

        return 1

    def Test_02043_18_04939_WHOLENESS_1000_1001_1180_1190_1200_1210_1220_1230_1060_XLSX_XLSM(self, workBook):

        # get table sheet

        sheetNames = workBook.sheetnames
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("table")
        except:
            try:
                index = sheetNames.index("tableau")
            except:
                return 2

        workSheet = workBook.worksheets[index]
        rows = workSheet.iter_rows()
        row_index = 1
        row_start = 10
        for row in rows:
            if row_index >= row_start:
                testflag = False
                for cell in row:
                    if cell.value:
                        testflag = True
                        break
                    else:
                        pass
                if testflag == True:
                    dictResult = {}
                    if not str(row[ref_col].value):
                        dictResult["0"] = str(colnum_string(ref_col + 1)) + str(row_index + 1)
                    if not str(row[ver_col].value):
                        dictResult["-1"] = str(colnum_string(ver_col + 1)) + str(row_index + 1)
                    if not str(row[defDet_col].value):
                        dictResult["-2"] = str(colnum_string(defDet_col + 1)) + str(row_index + 1)
                    if not str(row[defConst_col].value):
                        dictResult["-3"] = str(colnum_string(defConst_col + 1)) + str(row_index + 1)
                    if not str(row[sit_col].value):
                        dictResult["-4"] = str(colnum_string(sit_col + 1)) + str(row_index + 1)
                    if not str(row[eff_col].value):
                        dictResult["-5"] = str(colnum_string(eff_col + 1)) + str(row_index + 1)
                    if not str(row[code_col].value):
                        dictResult["-6"] = str(colnum_string(code_col + 1)) + str(row_index + 1)
                    if str(row[app_col].value) in {"NA", "X"}:
                        dictResult["-7"] = str(colnum_string(app_col + 1)) + str(row_index + 1)
                    else:
                        pass
                    if dictResult:
                        return dictResult
                else:
                    break
            if row_start == 10:
                for cell in row:
                    if str(cell.value).casefold() in {"référence", "reference"}:
                        ref_col = cell.column
                        if workSheet.cell(row_index+1,ref_col).value:
                            row_start = row_index + 1
                        else:
                            row_start = row_index + 2
                    if str(cell.value).casefold() in {"version"}:
                        ver_col = cell.column
                    if str(cell.value).casefold() in {"constituant défaillant détecté"}:
                        defDet_col = cell.column
                    if str(cell.value).casefold() in {"défaillance constituant"}:
                        defConst_col = cell.column
                    if str(cell.value).casefold() in {"situation de vie client"}:
                        sit_col = cell.column
                    if str(cell.value).casefold() in {"effet(s) client(s)"}:
                        eff_col = cell.column
                    if str(cell.value).casefold() in {"code défaut"}:
                        code_col = cell.column
                    if str(cell.value).casefold() in {"applicabilité projet"}:
                        app_col = cell.column
            row_index = row_index + 1

        return 1

    def Test_02043_18_04939_WHOLENESS_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061_XLS(self, workBook):

        # get table sheet

        sheetNames = workBook.sheet_names()
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("codes défauts")
        except:
            return 2

        workSheet = workBook.sheet_by_index(index)
        rows = workSheet.get_rows()

        row_index = 0
        row_start = 10
        for row in rows:
            if row_index >= row_start:
                testflag = False
                for cell in row:
                    if cell.value:
                        testflag = True
                        break
                    else:
                        pass
                if testflag == True:
                    dictResult = {}
                    if not str(row[ref_col].value):
                        dictResult["0"] = str(colnum_string(ref_col + 1)) + str(row_index + 1)
                    if not str(row[ver_col].value):
                        dictResult["-1"] = str(colnum_string(ver_col + 1)) + str(row_index + 1)
                    if not str(row[codeDef_col].value):
                        dictResult["-2"] = str(colnum_string(codeDef_col + 1)) + str(row_index + 1)
                    if not str(row[supp_col].value):
                        dictResult["-3"] = str(colnum_string(supp_col + 1)) + str(row_index + 1)
                    if not str(row[lib_col].value):
                        dictResult["-4"] = str(colnum_string(lib_col + 1)) + str(row_index + 1)
                    if not str(row[strat_col].value):
                        dictResult["-5"] = str(colnum_string(strat_col + 1)) + str(row_index + 1)
                    if not str(row[det_col].value):
                        dictResult["-6"] = str(colnum_string(det_col + 1)) + str(row_index + 1)
                    if not str(row[temps_col].value):
                        dictResult["-7"] = str(colnum_string(temps_col + 1)) + str(row_index + 1)
                    if not str(row[disp_col].value):
                        dictResult["-8"] = str(colnum_string(disp_col + 1)) + str(row_index + 1)
                    if not str(row[mode_col].value):
                        dictResult["-9"] = str(colnum_string(mode_col + 1)) + str(row_index + 1)
                    if not str(row[voy_col].value):
                        dictResult["-10"] = str(colnum_string(voy_col + 1)) + str(row_index + 1)
                    if str(row[app_col].value) in {"NA", "X"}:
                        dictResult["-11"] = str(colnum_string(app_col + 1)) + str(row_index + 1)
                    else:
                        pass
                    if dictResult:
                        return dictResult
                else:
                    break
            if row_start == 10:
                for colindex,cell in enumerate(row):
                    if str(cell.value).casefold() in {"référence", "reference"}:
                        ref_col = colindex
                        ref_tuple = (row_index, row_index+2, ref_col, ref_col +1)
                        if ref_tuple in workSheet.merged_cells:
                            row_start = row_index +2
                        else:
                            row_start = row_index + 1
                    if str(cell.value).casefold() in {"version"}:
                        ver_col = colindex
                    if str(cell.value).casefold() in {"code défaut"}:
                        codeDef_col = colindex
                    if str(cell.value).casefold() in {"supporté par constituant (s)"}:
                        supp_col = colindex
                    if str(cell.value).casefold() in {"libellé (signification)"}:
                        lib_col = colindex
                    if str(cell.value).casefold() in {"description de la strategie pour détecter le défaut"}:
                        strat_col = colindex
                    if str(cell.value).casefold() in {"seuil de détection  /  valeur  du défaut "}:
                        det_col = colindex
                    if str(cell.value).casefold() in {"temps de confirmation du défaut"}:
                        temps_col = colindex
                    if str(cell.value).casefold() in {"description de la strategie de disparition du défaut / procedure à effectuer pour vérifier la disparition du défaut"}:
                        disp_col = colindex
                    if str(cell.value).casefold() in {"mode dégradé"}:
                        mode_col = colindex
                    if str(cell.value).casefold() in {"voyant"}:
                        voy_col = colindex
                    if str(cell.value).casefold() in {"applicabilité projet"}:
                        app_col = colindex
            row_index = row_index + 1

        return 1

    def Test_02043_18_04939_WHOLENESS_1010_1011_1080_1090_1110_1120_1130_1140_1150_1160_1170_1061_XLSX_XLSM(self, workBook):

        # get table sheet

        sheetNames = workBook.sheetnames
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("codes défauts")
        except:
            return 2

        workSheet = workBook.worksheets[index]
        rows = workSheet.iter_rows()
        row_index = 1
        row_start = 10
        for row in rows:
            if row_index >= row_start:
                testflag = False
                for cell in row:
                    if cell.value:
                        testflag = True
                        break
                    else:
                        pass
                if testflag == True:
                    dictResult = {}
                    if not str(row[ref_col].value):
                        dictResult["0"] = str(colnum_string(ref_col + 1)) + str(row_index + 1)
                    if not str(row[ver_col].value):
                        dictResult["-1"] = str(colnum_string(ver_col + 1)) + str(row_index + 1)
                    if not str(row[codeDef_col].value):
                        dictResult["-2"] = str(colnum_string(codeDef_col + 1)) + str(row_index + 1)
                    if not str(row[supp_col].value):
                        dictResult["-3"] = str(colnum_string(supp_col + 1)) + str(row_index + 1)
                    if not str(row[lib_col].value):
                        dictResult["-4"] = str(colnum_string(lib_col + 1)) + str(row_index + 1)
                    if not str(row[strat_col].value):
                        dictResult["-5"] = str(colnum_string(strat_col + 1)) + str(row_index + 1)
                    if not str(row[det_col].value):
                        dictResult["-6"] = str(colnum_string(det_col + 1)) + str(row_index + 1)
                    if not str(row[temps_col].value):
                        dictResult["-7"] = str(colnum_string(temps_col + 1)) + str(row_index + 1)
                    if not str(row[disp_col].value):
                        dictResult["-8"] = str(colnum_string(disp_col + 1)) + str(row_index + 1)
                    if not str(row[mode_col].value):
                        dictResult["-9"] = str(colnum_string(mode_col + 1)) + str(row_index + 1)
                    if not str(row[voy_col].value):
                        dictResult["-10"] = str(colnum_string(voy_col + 1)) + str(row_index + 1)
                    if str(row[app_col].value) in {"NA", "X"}:
                        dictResult["-11"] = str(colnum_string(app_col + 1)) + str(row_index + 1)
                    else:
                        pass
                    if dictResult:
                        return dictResult
                else:
                    break
            if row_start == 10:
                for cell in row:
                    if str(cell.value).casefold() in {"référence", "reference"}:
                        ref_col = cell.column
                        if workSheet.cell(row_index+1,ref_col).value:
                            row_start = row_index + 1
                        else:
                            row_start = row_index + 2
                    if str(cell.value).casefold() in {"version"}:
                        ver_col = cell.column
                    if str(cell.value).casefold() in {"code défaut"}:
                        codeDef_col = cell.column
                    if str(cell.value).casefold() in {"supporté par constituant (s)"}:
                        supp_col = cell.column
                    if str(cell.value).casefold() in {"libellé (signification)"}:
                        lib_col = cell.column
                    if str(cell.value).casefold() in {"description de la strategie pour détecter le défaut"}:
                        strat_col = cell.column
                    if str(cell.value).casefold() in {"seuil de détection  /  valeur  du défaut"}:
                        det_col = cell.column
                    if str(cell.value).casefold() in {"temps de confirmation du défaut"}:
                        temps_col = cell.column
                    if str(cell.value).casefold() in {"description de la strategie de disparition du défaut / Procedure à effectuer pour vérifier la disparition du défaut"}:
                        disp_col = cell.column
                    if str(cell.value).casefold() in {"mode dégradé"}:
                        mode_col = cell.column
                    if str(cell.value).casefold() in {"Voyant"}:
                        voy_col = cell.column
                    if str(cell.value).casefold() in {"applicabilité projet"}:
                        app_col = cell.column
            row_index = row_index + 1

        return 1

    def Test_02043_18_04939_WHOLENESS_1020_1021_1100_1062_XLS(self, workBook):

        # get table sheet

        sheetNames = workBook.sheet_names()
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("mesures et commandes")
        except:
            return 2

        workSheet = workBook.sheet_by_index(index)
        rows = workSheet.get_rows()

        row_index = 0
        row_start = 10
        for row in rows:
            if row_index >= row_start:
                testflag = False
                for cell in row:
                    if cell.value:
                        testflag = True
                        break
                    else:
                        pass
                if testflag == True:
                    dictResult = {}
                    if not str(row[ref_col].value):
                        dictResult["0"] = str(colnum_string(ref_col + 1)) + str(row_index + 1)
                    if not str(row[ver_col].value):
                        dictResult["-1"] = str(colnum_string(ver_col + 1)) + str(row_index + 1)
                    if not str(row[supp_col].value):
                        dictResult["-2"] = str(colnum_string(supp_col + 1)) + str(row_index + 1)
                    if str(row[app_col].value) in {"NA", "X"}:
                        dictResult["-3"] = str(colnum_string(app_col + 1)) + str(row_index + 1)
                    else:
                        pass
                    if dictResult:
                        return dictResult
                else:
                    break
            if row_start == 10:
                for colindex, cell in enumerate(row):
                    if str(cell.value).casefold() in {"référence", "reference"}:
                        ref_col = colindex
                        ref_tuple = (row_index, row_index+2, ref_col, ref_col +1)
                        if ref_tuple in workSheet.merged_cells:
                            row_start = row_index + 2
                        else:
                            row_start = row_index + 1
                    if str(cell.value).casefold() in {"version"}:
                        ver_col = colindex
                    if str(cell.value).casefold() in {"supporté par constituant (s)"}:
                        supp_col = colindex
                    if str(cell.value).casefold() in {"applicabilité projet"}:
                        app_col = colindex
            row_index = row_index + 1


        return 1

    def Test_02043_18_04939_WHOLENESS_1020_1021_1100_1062_XLSX_XLSM(self, workBook):

        # get table sheet

        sheetNames = workBook.sheetnames
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("mesures et commandes")
        except:
            return 2

        workSheet = workBook.worksheets[index]
        rows = workSheet.iter_rows()
        row_index = 1
        row_start = 10
        for row in rows:
            if row_index >= row_start:
                testflag = False
                for cell in row:
                    if cell.value:
                        testflag = True
                        break
                    else:
                        pass
                if testflag == True:
                    dictResult = {}
                    if not str(row[ref_col].value):
                        dictResult["0"] = str(colnum_string(ref_col + 1)) + str(row_index + 1)
                    if not str(row[ver_col].value):
                        dictResult["-1"] = str(colnum_string(ver_col + 1)) + str(row_index + 1)
                    if not str(row[supp_col].value):
                        dictResult["-2"] = str(colnum_string(supp_col + 1)) + str(row_index + 1)
                    if str(row[app_col].value) in {"NA", "X"}:
                        dictResult["-3"] = str(colnum_string(app_col + 1)) + str(row_index + 1)
                    else:
                        pass
                    if dictResult:
                        return dictResult
                else:
                    break
            if row_start == 10:
                for cell in row:
                    if str(cell.value).casefold() in {"référence", "reference"}:
                        ref_col = cell.column
                        if workSheet.cell(row_index+1,ref_col).value:
                            row_start = row_index + 1
                        else:
                            row_start = row_index + 2
                    if str(cell.value).casefold() in {"version"}:
                        ver_col = cell.column
                    if str(cell.value).casefold() in {"supporté par constituant (s)"}:
                        supp_col = cell.column
                    if str(cell.value).casefold() in {"applicabilité projet"}:
                        app_col = cell.column
            row_index = row_index + 1


        return 1

    def Test_02043_18_04939_WHOLENESS_1030_1031_XLS(self, workBook):

        # get table sheet

        sheetNames = workBook.sheet_names()
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("diagnostic débarqués")
        except:
            return 2

        workSheet = workBook.sheet_by_index(index)
        rows = workSheet.get_rows()

        row_index = 0
        row_start = 10
        for row in rows:
            if row_index >= row_start:
                testflag = False
                for cell in row:
                    if cell.value:
                        testflag = True
                        break
                    else:
                        pass
                if testflag == True:
                    dictResult = {}
                    if not str(row[ref_col].value):
                        dictResult["0"] = str(colnum_string(ref_col + 1)) + str(row_index + 1)
                    if not str(row[ver_col].value):
                        dictResult["-1"] = str(colnum_string(ver_col + 1)) + str(row_index + 1)
                    if dictResult:
                        return dictResult
                else:
                    break
            if row_start == 10:
                for colindex,cell in enumerate(row):
                    if str(cell.value).casefold() in {"référence", "reference"}:
                        ref_col = colindex
                        ref_tuple = (row_index, row_index+2, ref_col, ref_col +1)
                        if ref_tuple in workSheet.merged_cells:
                            row_start = row_index +2
                        else:
                            row_start = row_index + 1
                    if str(cell.value).casefold() in {"version"}:
                        ver_col = colindex
            row_index = row_index + 1

        return 1

    def Test_02043_18_04939_WHOLENESS_1030_1031_XLSX_XLSM(self, workBook):

        # get table sheet

        sheetNames = workBook.sheetnames
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("Diagnostic débarqués")
        except:
            return 2

        workSheet = workBook.worksheets[index]
        rows = workSheet.iter_rows()
        row_index = 1
        row_start = 10
        for row in rows:
            if row_index >= row_start:
                testflag = False
                for cell in row:
                    if cell.value:
                        testflag = True
                        break
                    else:
                        pass
                if testflag == True:
                    dictResult = {}
                    if not str(row[ref_col].value):
                        dictResult["0"] = str(colnum_string(ref_col + 1)) + str(row_index + 1)
                    if not str(row[ver_col].value):
                        dictResult["-1"] = str(colnum_string(ver_col + 1)) + str(row_index + 1)
                    if dictResult:
                        return dictResult
                else:
                    break
            if row_start == 10:
                for cell in row:
                    if str(cell.value).casefold() in {"référence", "reference"}:
                        ref_col = cell.column
                        if workSheet.cell(row_index+1,ref_col).value:
                            row_start = row_index + 1
                        else:
                            row_start = row_index + 2
                    if str(cell.value).casefold() in {"version"}:
                        ver_col = cell.column
            row_index = row_index + 1

        return 1

    def Test_02043_18_04939_WHOLENESS_1040_1041_XLS(self, workBook):

        # get table sheet

        sheetNames = workBook.sheet_names()
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("Liste MDD")
        except:
            return 2

        workSheet = workBook.sheet_by_index(index)
        rows = workSheet.get_rows()

        row_index = 0
        row_start = 10
        for row in rows:
            if row_index >= row_start:
                testflag = False
                for cell in row:
                    if cell.value:
                        testflag = True
                        break
                    else:
                        pass
                if testflag == True:
                    dictResult = {}
                    if not str(row[ref_col].value):
                        dictResult["0"] = str(colnum_string(ref_col + 1)) + str(row_index + 1)
                    if not str(row[ver_col].value):
                        dictResult["-1"] = str(colnum_string(ver_col + 1)) + str(row_index + 1)
                    if dictResult:
                        return dictResult
                else:
                    break
            if row_start == 10:
                for colindex,cell in enumerate(row):
                    if str(cell.value).casefold() in {"référence", "reference"}:
                        ref_col = colindex
                        ref_tuple = (row_index, row_index+2, ref_col, ref_col +1)
                        if ref_tuple in workSheet.merged_cells:
                            row_start = row_index +2
                        else:
                            row_start = row_index + 1
                    if str(cell.value).casefold() in {"version"}:
                        ver_col = colindex
            row_index = row_index + 1

        return 1

    def Test_02043_18_04939_WHOLENESS_1040_1041_XLSX_XLSM(self, workBook):

        # get table sheet

        sheetNames = workBook.sheetnames
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("Liste MDD")
        except:
            return 2

        workSheet = workBook.worksheets[index]
        rows = workSheet.iter_rows()
        row_index = 1
        row_start = 10
        for row in rows:
            if row_index >= row_start:
                testflag = False
                for cell in row:
                    if cell.value:
                        testflag = True
                        break
                    else:
                        pass
                if testflag == True:
                    dictResult = {}
                    if not str(row[ref_col].value):
                        dictResult["0"] = str(colnum_string(ref_col + 1)) + str(row_index + 1)
                    if not str(row[ver_col].value):
                        dictResult["-1"] = str(colnum_string(ver_col + 1)) + str(row_index + 1)
                    if dictResult:
                        return dictResult
                else:
                    break
            if row_start == 10:
                for cell in row:
                    if str(cell.value).casefold() in {"référence", "reference"}:
                        ref_col = cell.column
                        if workSheet.cell(row_index+1,ref_col).value:
                            row_start = row_index + 1
                        else:
                            row_start = row_index + 2
                    if str(cell.value).casefold() in {"version"}:
                        ver_col = cell.column
            row_index = row_index + 1

        return 1

    def Test_02043_18_04939_WHOLENESS_1050_XLS(self, workBook):

        sheetNames = workBook.sheet_names()
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index_table = sheetNames.index("table")
        except:
            try:
                index_table = sheetNames.index("tableau")
            except:
                return 2

        workSheetTable = workBook.sheet_by_index(index_table)
        rows_table = workSheetTable.get_rows()

        try:
            index_codes = sheetNames.index("codes défauts")
        except:
            try:
                index_codes = sheetNames.index("data trouble codes")
            except:
                return 0

        workSheetCodes = workBook.sheet_by_index(index_codes)
        rows_codes = workSheetCodes.get_rows()
        rowValueListTable  = list()
        rowValueListCodes = list()

        row_index = 0
        row_start = 10
        for row_table in rows_table:
            if row_index >= row_start:
                for i in range(0, 12):
                    app_col_table = app_col_table + 1
                    rowValueListTable.append(str(row_table[app_col_table-1].value).casefold().strip())
                    row_index = 0

            if row_start == 10:
                for colindex, cell in enumerate(row_table):
                    if str(cell.value).casefold() in {"applicabilité projet", "project applicability"}:
                        app_col_table = colindex
                        app_tuple = (row_index, row_index + 2, app_col_table, app_col_table + 12)
                        if app_tuple in workSheetTable.merged_cells:
                            row_start = row_index + 2
                        else:
                            row_start = row_index + 1
                row_index = row_index + 1


        row_index_codes = 0
        row_start_codes = 10
        for row_codes in rows_codes:
            if row_index_codes >= row_start_codes:
                for i in range(0, 8):
                    app_col_code = app_col_code + 1
                    rowValueListCodes.append(str(row_codes[app_col_code-1].value).casefold().strip())
                    row_index_codes = 0
            if row_start_codes == 10:
                for colindex, cell in enumerate(row_codes):
                    if str(cell.value).casefold() in {"applicabilité projet", "project applicability"}:
                        app_col_code = colindex
                        app_tuple = (row_index_codes, row_index_codes + 1, app_col_code, app_col_code + 8)
                        if app_tuple in workSheetCodes.merged_cells:
                            row_start_codes = row_index_codes + 1
                row_index_codes = row_index_codes + 1

        check = 0
        if str(rowValueListTable) == '[]'.casefold():
            check += 1
        else:
            for element in rowValueListTable:
                if element in rowValueListCodes:
                    pass
                else:
                   check += 1

        if check == 0:
            return 1
        else:
            return 0

    def Test_02043_18_04939_WHOLENESS_1050_XLSX_XLSM(self, workBook):

        sheetNames = workBook.sheetnames
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index_table = sheetNames.index("table")
        except:
            try:
                index_table = sheetNames.index("tableau")
            except:
                return 2

        workSheetTable = workBook.worksheets[index_table]
        rows_table = workSheetTable.iter_rows()

        try:
            index_codes = sheetNames.index("codes défauts")
        except:
            try:
                index_codes = sheetNames.index("data trouble codes")
            except:
                return 0

        workSheetCodes = workBook.worksheets[index_codes]
        rows_codes = workSheetCodes.iter_rows()
        rowValueListTable  = list()
        rowValueListCodes = list()

        row_index = 1
        row_start = 10
        for row_table in rows_table:
            if row_index >= row_start:
                for i in range(0, 12):
                    rowValueListTable.append(str(row_table[app_col_table - 1].value).casefold().strip())
                    app_col_table = app_col_table + 1
                    row_index = 0

            if row_start == 10:
                for  cell in row_table:
                    if str(cell.value).casefold() in {"applicabilité projet", "project applicability"}:
                        app_col_table = cell.column
                        if workSheetTable.cell(row_index + 1, app_col_table).value:
                            row_start = row_index + 1
                        else:
                            row_start = row_index + 2
                row_index = row_index + 1

        row_index_codes = 1
        row_start_codes = 10
        for row_codes in rows_codes:
            if row_index_codes >= row_start_codes:
                    for i in range(0, 8):
                        rowValueListCodes.append(str(row_codes[app_col_code - 1].value).casefold().strip())
                        app_col_code = app_col_code + 1
                        row_index_codes = 1

            if row_start_codes == 10:
                for cell in row_codes:
                    if str(cell.value).casefold() in {"applicabilité projet", "project applicability"}:
                        app_col_code = cell.column
                        if workSheetCodes.cell(row_index_codes + 1, app_col_code).value:
                            row_start_codes = row_index_codes + 1
                        else:
                            row_start_codes = row_index_codes + 2
                row_index_codes = row_index_codes + 1

        check = 0
        if str(rowValueListTable) == '[]'.casefold():
            check += 1
        else:
            for element in rowValueListTable:
                if element in rowValueListCodes:
                    pass
                else:
                   check = check + 1

        if check == 0:
            return 1
        else:
            return 0

    def Test_02043_18_04939_WHOLENESS_1055_XLS(self, workBook):

        sheetNames = workBook.sheet_names()
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index_table = sheetNames.index("table")
        except:
            try:
                index_table = sheetNames.index("tableau")
            except:
                return 2

        workSheetTable = workBook.sheet_by_index(index_table)
        rows_table = workSheetTable.get_rows()

        try:
            index_measures = sheetNames.index("mesures et commandes")
        except:
            return 0

        workSheetMeasures = workBook.sheet_by_index(index_measures)
        rows_measures = workSheetMeasures.get_rows()
        rowValueListTable  = list()
        rowValueListMeasures = list()

        row_index = 0
        row_start = 10
        for row_table in rows_table:
            if row_index >= row_start:
                for i in range(0, 12):
                    app_col_table = app_col_table + 1
                    rowValueListTable.append(str(row_table[app_col_table-1].value).casefold().strip())
                    row_index = 0

            if row_start == 10:
                for colindex, cell in enumerate(row_table):
                    if str(cell.value).casefold() in {"applicabilité projet", "project applicability"}:
                        app_col_table = colindex
                        app_tuple = (row_index, row_index + 2, app_col_table, app_col_table + 12)
                        if app_tuple in workSheetTable.merged_cells:
                            row_start = row_index + 2
                        else:
                            row_start = row_index + 1
                row_index = row_index + 1


        row_index_measures = 0
        row_start_measures = 10
        for row_measures in rows_measures:
            if row_index_measures >= row_start_measures:
                for i in range(0, 21):
                    app_col_measures = app_col_measures + 1
                    rowValueListMeasures.append(str(row_measures[app_col_measures-1].value).casefold().strip())
                    row_index_measures = 0

            if row_start_measures == 10:
                for colindex, cell in enumerate(row_measures):
                    if str(cell.value).casefold() in {"applicabilité projet"}:
                        app_col_measures = colindex
                        row_start_measures = row_index_measures + 1
                row_index_measures = row_index_measures + 1

        check = 0
        for element in rowValueListTable:
            if element in rowValueListMeasures:
                pass
            else:
               check = check + 1

        if check == 0:
            return 1
        else:
            return 0

    def Test_02043_18_04939_WHOLENESS_1055_XLSX_XLSM(self, workBook):

        sheetNames = workBook.sheetnames
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index_table = sheetNames.index("table")
        except:
            try:
                index_table = sheetNames.index("tableau")
            except:
                return 2

        workSheetTable = workBook.worksheets[index_table]
        rows_table = workSheetTable.iter_rows()

        try:
            index_measures = sheetNames.index("mesures et commandes")
        except:
            return 0

        workSheetMeasures = workBook.worksheets[index_measures]
        rows_measures = workSheetMeasures.iter_rows()
        rowValueListTable  = list()
        rowValueListMeasures = list()

        row_index = 1
        row_start = 10
        for row_table in rows_table:
            if row_index >= row_start:
                for i in range(0, 12):
                    rowValueListTable.append(str(row_table[app_col_table - 1].value).casefold().strip())
                    app_col_table = app_col_table + 1
                    row_index = 1

            if row_start == 10:
                for  cell in row_table:
                    if str(cell.value).casefold() in {"applicabilité projet", "project applicability"}:
                        app_col_table = cell.column
                        if workSheetTable.cell(row_index + 1, app_col_table).value:
                            row_start = row_index + 1
                        else:
                            row_start = row_index + 2
                row_index = row_index + 1

        row_index_measures = 1
        row_start_measures = 10
        for row_measures in rows_measures:
            if row_index_measures >= row_start_measures:
                testflag = False
                for cell in row_measures:
                    if cell.value:
                        testflag = True
                        break
                    else:
                        pass
                if testflag == True:
                    for i in range(0, 21):
                        rowValueListMeasures.append(str(row_measures[app_col_measures - 1].value).casefold().strip())
                        app_col_measures = app_col_measures + 1
                else:
                    break
            if row_start_measures == 10:
                for cell in row_measures:
                    if str(cell.value).casefold() in {"applicabilité projet", "project applicability"}:
                        app_col_measures = cell.column
                        if workSheetMeasures.cell(row_index_measures + 1, app_col_measures).value:
                            row_start_measures = row_index_measures + 1
                        else:
                            row_start_measures = row_index_measures + 2
            row_index_measures = row_index_measures + 1

        check = 0
        for element in rowValueListTable:
            if element in rowValueListMeasures:
                pass
            else:
               check = check + 1

        if check == 0:
            return 1
        else:
            return 0

    def Test_02043_18_04939_WHOLENESS_1240_XLS(self, workBook):

        sheetNames = workBook.sheet_names()
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("informations générales")
        except:
            try:
                index = sheetNames.index("general information")
            except:
                return 2

        workSheet = workBook.sheet_by_index(index)
        rows = workSheet.get_rows()

        row_index = 1
        for row in rows:
            for cell in row:
                if str(cell.value).casefold().strip() in {"liste de diffusion (participant)", "mailing list (the taking part)"}:
                    row_start = row_index + 3
                if str(cell.value).casefold().strip() in {"copie", "copy"}:
                    row_stop = row_index - 3
            row_index = row_index + 1

        test_rows = workSheet._cell_values[row_start:row_stop]

        for row in test_rows:
             if not str(row[0]):
                 return 0
             else:
                 pass
        return 1

    def Test_02043_18_04939_WHOLENESS_1240_XLSX_XLSM(self, workBook):

        sheetNames = workBook.sheetnames
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("informations générales")
        except:
            try:
                index = sheetNames.index("general information")
            except:
                return 2

        workSheet = workBook.worksheets[index]
        rows = workSheet.iter_rows()

        row_index = 1
        for row in rows:
            for cell in row:
                if str(cell.value).casefold().strip() in {"liste de diffusion (participant)", "mailing list (the taking part)"}:
                    row_start = row_index + 3
                if str(cell.value).casefold().strip() in {"copie", "copy"}:
                    row_stop = row_index - 3
            row_index = row_index + 1

        test_rows = workSheet.iter_rows(min_row= row_start, max_row= row_stop)

        for row in test_rows:
             if str(row[0].value) == "None":
                return 0
             else:
                 pass
        return 1




#Coherence checks requirements

    def Test_02043_18_04939_COH_2000_XLS(self, workBook):

        # get table sheet

        sheetNames = workBook.sheet_names()
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index_table = sheetNames.index("table")
        except:
            try:
                index_table = sheetNames.index("tableau")
            except:
                return 2

        try:
            index_mesures = sheetNames.index("mesures et commandes")
        except:
            return 2

        workSheetTable = workBook.sheet_by_index(index_table)
        rows_table = workSheetTable.get_rows()
        rowValueListTable = list()
        flag_table = 0
        workSheetMesures = workBook.sheet_by_index(index_mesures)
        rows_mesures = workSheetMesures.get_rows()
        rowValueListMesures = list()
        flag_mesures = 0
        row_index_table = 0
        row_index_mesures = 0

        for row_table in rows_table:
            if flag_table == 1:
                if not str(row_table[table_col]):
                    rowValueListTable.append(str(row_table[table_col].value).casefold().strip())
                else:
                    break

            if flag_table == 0:
                for colindex, cell in enumerate(row_table):
                    if str(cell.value).casefold().strip() == "mesures et commandes (Mesure Parametre et Test Actionneur) / Tests de cohérence".casefold():
                        table_col = colindex
                        table_tuple = (row_index_table, row_index_table + 2, table_col, table_col + 1)
                        if table_tuple in workSheetTable.merged_cells:
                            row_index_table = row_index_table + 2
                            flag_table = 1
                        else:
                            row_index_table = row_index_table + 1
                            flag_table = 1
                        break

            row_index_table = row_index_table + 1

        for row_mesures in rows_mesures:
            if flag_mesures == 1:
                if not str(row_mesures[mesures_col]):
                    rowValueListMesures.append(str(row_mesures[mesures_col].value).casefold().strip())
                else:
                    break

            if flag_mesures == 0:
                for colindex, cell in enumerate(row_mesures):
                    if str(cell.value).casefold() == "libellé (signification)".casefold():
                        mesures_col = colindex
                        row_index_mesures = row_index_mesures + 1
                        flag_mesures = 1
            row_index_mesures = row_index_mesures + 1




        check = 0
        for element in rowValueListTable:
            if element in rowValueListMesures:
                pass
            else:
                check = check + 1

        if check == 0:
            return 1
        else:
            return 0

    def Test_02043_18_04939_COH_2000_XLSX_XLSM(self, workBook):

        # get table sheet

        sheetNames = workBook.sheetnames
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index_table = sheetNames.index("table")
        except:
            try:
                index_table = sheetNames.index("tableau")
            except:
                return 2

        try:
            index_mesures = sheetNames.index("mesures et commandes")
        except:
            return 2


        workSheetTable = workBook.worksheets[index_table]
        rows_table = workSheetTable.iter_rows()
        rowValueListTable = list()
        flag_table = 0
        workSheetMesures = workBook.worksheets[index_mesures]
        rows_mesures = workSheetMesures.iter_rows()
        rowValueListMesures = list()
        flag_mesures = 0
        row_index_table = 1
        row_index_mesures = 1

        for row_table in rows_table:
            if flag_table == 1:
                if str(row_table[table_col].value) == "None":
                    break
                else:
                    rowValueListTable.append(str(row_table[table_col].value).casefold().strip())

            if flag_table == 0:
                for cell in row_table:
                    if str(cell.value).casefold().strip() == "mesures et commandes (Mesure Parametre et Test Actionneur) / Tests de cohérence".casefold():
                        table_col = cell.column
                        if workSheetTable.cell(row_index_table + 1, table_col).value:
                            row_index_table = row_index_table + 1
                            flag_table = 1
                        else:
                            row_index_table = row_index_table + 2
                            flag_table = 1
                        break

            row_index_table = row_index_table + 1

        for row_mesures in rows_mesures:
            if flag_mesures == 1:
                if str(row_mesures[mesures_col].value) == "None":
                    break
                else:
                    rowValueListMesures.append(str(row_mesures[mesures_col].value).casefold().strip())

            if flag_mesures == 0:
                for cell in row_mesures:
                    if str(cell.value).casefold() == "libellé (signification)".casefold():
                        mesures_col = cell.column
                        flag_mesures = 1
            row_index_mesures = row_index_mesures + 1

        check = 0
        for element in rowValueListTable:
            if element in rowValueListMesures:
                pass
            else:
                check = check + 1

        if check == 0:
            return 1
        else:
            return 0

    def Test_02043_18_04939_COH_2005_XLS(self, workBook):
        # get table sheet

        sheetNames = workBook.sheet_names()
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("codes défauts")
        except:
            return 2

        workSheet = workBook.sheet_by_index(index)
        rows = workSheet.get_rows()
        flag = 0
        row_index = 0
        listValues = list()
        famillyList = list()
        codeList = list()
        localisation = str()
        ok = 1
        for row in rows:
            if flag == 1:
                if str(row[code_col]).count('-') != 2:
                   localisation =(str(colnum_string(code_col + 1)) + str(row_index + 1))
                   ok = 0
                   break
                else:
                    listValues = str(row[code_col].value).casefold().split('-')
                    if not listValues[0].isascii():
                        ok = 0
                    if not listValues[1][0].isalpha():
                        ok = 0
                    try:
                        int(listValues[1][1:], 16)
                    except:
                        ok = 0
                    try:
                        int(listValues[2], 16)
                    except:
                        ok = 0
                    if ok == 1:
                        tempDict = dict()
                        tempDict["value"] = listValues[0]
                        tempDict["codenr"] = listValues[1]
                        tempDict["row"] = str(row_index + 1)
                        tempDict["col"] = str(colnum_string(code_col + 1))
                        famillyList.append(dict(tempDict))
                    else:
                        localisation = (str(colnum_string(code_col + 1)) + str(row_index + 1))
                        break

            if flag == 0:
                for colindex, cell in enumerate(row):
                    if str(cell.value).casefold().strip() == "code défaut".casefold():
                        code_col = colindex
                        flag = 1
                        break
                    else:
                        pass

            row_index = row_index + 1

        if ok == 0:
             return (0,famillyList,localisation)
        else:
             return (1, famillyList, localisation)

    def Test_02043_18_04939_COH_2005_XLSX_XLSM(self, workBook):
        # get table sheet

        sheetNames = workBook.sheetnames
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("codes défauts")
        except:
            return 2

        workSheet = workBook.worksheets[index]
        rows = workSheet.iter_rows()
        flag = 0
        listValues = list()
        row_index = 1
        localisation = str()
        ok = 1
        famillyList = list()
        for row in rows:
            if flag == 1:
                if str(row[code_col - 1].value).count('-') != 2:
                    localisation = (str(colnum_string(code_col)) + str(row_index + 1))
                    ok = 0
                    break
                else:
                    listValues = str(row[code_col -1].value).casefold().split('-')
                    if not listValues[0].isascii():
                        ok = 0
                    if not listValues[1][0].isalpha():
                        ok = 0
                    try:
                        int(listValues[1][1:], 16)
                    except:
                        ok = 0
                    try:
                        int(listValues[2], 16)
                    except:
                        ok = 0
                    if ok == 1:
                        tempDict = dict()
                        tempDict["value"] = listValues[0]
                        tempDict["codenr"] = listValues[1]
                        tempDict["row"] = str(row_index + 1)
                        tempDict["col"] = str(colnum_string(code_col))
                        famillyList.append(dict(tempDict))
                    else:
                        localisation = (str(colnum_string(code_col + 1)) + str(row_index + 1))
                        break

            if flag == 0:
                for cell in row:
                    if str(cell.value).casefold().strip() == "code défaut".casefold():
                        code_col = cell.column
                        flag = 1
                        break
                    else:
                        pass

            row_index = row_index + 1

        if ok == 0:
             return (0,famillyList,localisation)
        else:
             return (1, famillyList,localisation)

    def Test_02043_18_04939_COH_2006_XLS(self, workBook, famillyList):
        # get table sheet

        sheetNames = workBook.sheet_names()
        sheetNames = [x.casefold() for x in sheetNames]

        try:
            index_codes = sheetNames.index("codes défauts")
        except:
            return 2

        count = 0
        returnList = list()

        if self.tab2.myTextBox5.toPlainText():
            fileName = self.tab2.myTextBox5.toPlainText()
            win32.gencache.EnsureDispatch('Excel.Application')
            workBook = excel.Workbooks.Open(fileName)
            workSheet = workBook.Sheets(1)
            col = 0
        else:
             fileName = self.path_Cesare
             excel = win32.gencache.EnsureDispatch('Excel.Application')
             workBook = excel.Workbooks.Open(fileName)
             workSheet = workBook.Sheets(1)
             col = 0

        for row in workSheet.Rows:
             for cell in row.Cells:
                 if str(cell.Value).strip().casefold() == " Nom de la sous famille ".strip().casefold():
                     col = cell.Column
                     break
             if col != 0:
                 break

        CesareList = list()
        for row in workSheet.Rows:
            if row.Row > 2:
                if workSheet.Cells(row.Row, col).Value is None:
                    break
                else:
                    CesareList.append(str(workSheet.Cells(row.Row, col).Value))

        for element in famillyList:
            if element["value"] in CesareList:
                pass
            else:
                count +=1
                returnList.append(dict(element))


        if count == 0:
             return 1
        else:
            return (0, returnList)

    def Test_02043_18_04939_COH_2006_XLSX_XLSM(self, workBook, famillyList):
        # get table sheet

        sheetNames = workBook.sheetnames
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index = sheetNames.index("codes défauts")
        except:
            return 2

        count = 0
        returnList = list()

        if self.tab2.myTextBox5.toPlainText():
            fileName = self.tab2.myTextBox5.toPlainText()
            win32.gencache.EnsureDispatch('Excel.Application')
            workBook = excel.Workbooks.Open(fileName)
            workSheet = workBook.Sheets(1)
            col = 0
        else:
             fileName = self.path_Cesare
             excel = win32.gencache.EnsureDispatch('Excel.Application')
             workBook = excel.Workbooks.Open(fileName)
             workSheet = workBook.Sheets(1)
             col = 0

        for row in workSheet.Rows:
             for cell in row.Cells:
                 if str(cell.Value).strip().casefold() == " Nom de la sous famille ".strip().casefold():
                     col = cell.Column
                     break
             if col != 0:
                 break

        CesareList = list()
        for row in workSheet.Rows:
            if row.Row > 2:
                if workSheet.Cells(row.Row, col).Value is None:
                    break
                else:
                    CesareList.append(str(workSheet.Cells(row.Row, col).Value))

        for element in famillyList:
            if element["value"] in CesareList:
                pass
            else:
                count +=1
                returnList.append(dict(element))


        if count == 0:
             return 1
        else:
            return (0, returnList)

    def Test_02043_18_04939_COH_2007_XLS(self, workBook, famillyList):

        sheetNames = workBook.sheet_names()
        sheetNames = [x.casefold() for x in sheetNames]
        count = 0

        try:
            index_codes = sheetNames.index("codes défauts")
        except:
            return 2

        fileName = self.path_DOC14
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        workBook = excel.Workbooks.Open(fileName)
        workSheet = workBook.Sheets("Matrix")
        returnList = list()
        troubleCodes = list()

        for row in workSheet.Rows(2):
             for cell in row.Cells:
                 if str(cell.Value).strip().casefold() == "Data Trouble Code (DTC)".strip().casefold():
                     col = cell.Column
                     break
                 if cell.Value is None:
                    break
             if col != 0:
                 break

        for row in workSheet.Rows:
            if row.Row > 2:
                if workSheet.Cells(row.Row, col).Value is None:
                    break
                else:
                    troubleCodes.append(str(workSheet.Cells(row.Row, col).Value).casefold().strip())

        for element in famillyList:
            if element["codenr"] in troubleCodes:
                pass
            else:
                count += 1
                returnList.append(dict(element))

        if count == 0:
             return 1
        else:
            return (0, returnList)

    def Test_02043_18_04939_COH_2007_XLSX_XLSM(self, workBook, famillyList):

        sheetNames = workBook.sheetnames
        sheetNames = [x.casefold() for x in sheetNames]
        count = 0

        try:
            index = sheetNames.index("codes défauts")
        except:
            return 2

        fileName = self.path_DOC14
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        workBook = excel.Workbooks.Open(fileName)
        workSheet = workBook.Sheets("Matrix")
        returnList = list()
        troubleCodes = list()

        for row in workSheet.Rows(2):
             for cell in row.Cells:
                 if str(cell.Value).strip().casefold() == "Data Trouble Code (DTC)".strip().casefold():
                     col = cell.Column
                     break
                 if cell.Value is None:
                    break
             if col != 0:
                 break

        for row in workSheet.Rows:
            if row.Row > 2:
                if workSheet.Cells(row.Row, col).Value is None:
                    break
                else:
                    troubleCodes.append(str(workSheet.Cells(row.Row, col).Value))

        for element in famillyList:
            if element["codenr"] in troubleCodes:
                pass
            else:
                count +=1
                returnList.append(dict(element))

        if count == 0:
             return 1
        else:
            return (0, returnList)

    def Test_02043_18_04939_COH_2010_XLS(self, workBook):

        sheetNames = workBook.sheet_names()
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index_table = sheetNames.index("table")
        except:
            try:
                index_table = sheetNames.index("tableau")
            except:
                return 2

        try:
            index_codes = sheetNames.index("codes défauts")
        except:
            return 2

        workSheetTable = workBook.sheet_by_index(index_table)
        rows_table = workSheetTable.get_rows()
        rowValueListTable = list()
        workSheetCodes = workBook.sheet_by_index(index_codes)
        rows_codes = workSheetCodes.get_rows()
        rowValueListCodes = list()
        row_index_table = 0
        row_index_codes = 0
        flag_table = 0
        flag_codes = 0
        check = 0

        for row_table in rows_table:
            if flag_table == 1:
                if not str(row_table[table_col]):
                    if str(row_table[table_col]).casefold().strip() != "No DTC".casefold().strip():
                        rowValueListTable.append(str(row_table[table_col].value).casefold().strip())
                    else:
                        pass
                else:
                    break

            if flag_table == 0:
                for colindex, cell in enumerate(row_table):
                    if str(cell.value).casefold().strip() == "code défaut".casefold():
                        table_col = colindex
                        table_tuple = (row_index_table, row_index_table + 2, table_col, table_col + 1)
                        if table_tuple in workSheetTable.merged_cells:
                            row_index_table = row_index_table + 2
                            flag_table = 1
                        else:
                            row_index_table = row_index_table + 1
                            flag_table = 1
                        break

            row_index_table = row_index_table + 1

        for row_codes in rows_codes:
            if flag_codes == 1:
                if not str(row_codes[codes_col]):
                    rowValueListCodes.append(str(row_codes[codes_col].value).casefold().strip())
                else:
                    break

            if flag_codes == 0:
                for colindex, cell in enumerate(row_codes):
                    if str(cell.value).casefold() == "Code défaut".casefold():
                        codes_col = colindex
                        row_index_codes = row_index_codes + 1
                        flag_codes = 1
            row_index_codes = row_index_codes + 1


        if str(rowValueListTable) == '[]'.casefold():
            check += 1
        else:
            for element in rowValueListTable:
                if element in rowValueListCodes:
                    pass
                else:
                    check = check + 1

        if check == 0:
            return 1
        else:
            return 0

    def Test_02043_18_04939_COH_2010_XLSX_XLSM(self, workBook):

        sheetNames = workBook.sheetnames
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index_table = sheetNames.index("table")
        except:
            try:
                index_table = sheetNames.index("tableau")
            except:
                return 2

        try:
            index_codes = sheetNames.index("codes défauts")
        except:
            return 2

        workSheetTable = workBook.worksheets[index_table]
        rows_table = workSheetTable.iter_rows()
        rowValueListTable = list()
        workSheetCodes = workBook.worksheets[index_codes]
        rows_codes = workSheetCodes.iter_rows()
        rowValueListCodes = list()
        row_index_table = 0
        row_index_codes = 0
        flag_table = 0
        flag_codes = 0
        check = 0

        for row_table in rows_table:
            if flag_table == 1:
                if str(row_table[table_col].value) == "None":
                    break
                else:
                    if str(row_table[table_col]).casefold().strip() != "No DTC".casefold().strip():
                        rowValueListTable.append(str(row_table[table_col].value).casefold().strip())
                    else:
                        pass

            if flag_table == 0:
                for colindex, cell in enumerate(row_table):
                    if str(cell.value).casefold().strip() == "code défaut".casefold():
                        table_col = cell.column
                        if workSheetTable.cell(row_index_table + 1, table_col).value:
                            row_index_table = row_index_table + 1
                            flag_table = 1
                        else:
                            row_index_table = row_index_table + 2
                            flag_table = 1
                        break

            row_index_table = row_index_table + 1

        for row_codes in rows_codes:
            if flag_codes == 1:
                if str(row_codes[codes_col].value) == "None":
                    break
                else:
                    rowValueListCodes.append(str(row_codes[codes_col].value).casefold().strip())

            if flag_codes == 0:
                for colindex, cell in enumerate(row_codes):
                    if str(cell.value).casefold() == "Code défaut".casefold():
                        codes_col = cell.column
                        row_index_codes = row_index_codes + 1
                        flag_codes = 1
            row_index_codes = row_index_codes + 1


        if str(rowValueListTable) == '[]'.casefold():
            check += 1
        else:
            for element in rowValueListTable:
                if element in rowValueListCodes:
                    pass
                else:
                    check = check + 1

        if check == 0:
            return 1
        else:
            return 0

    def Test_02043_18_04939_COH_2020_XLS(self, workBook):

        sheetNames = workBook.sheet_names()
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index_table = sheetNames.index("table")
        except:
            try:
                index_table = sheetNames.index("tableau")
            except:
                return 2

        try:
            index_constituants = sheetNames.index("constituants")
        except:
            return 2

        workSheetTable = workBook.sheet_by_index(index_table)
        rows_table = workSheetTable.get_rows()
        rowValueListTable = list()
        workSheetConstituants = workBook.sheet_by_index(index_constituants)
        rows_constituants = workSheetConstituants.get_rows()
        rowValueListConstituants = list()
        row_index_table = 0
        row_index_constituants = 0
        flag_table = 0
        flag_constituants = 0
        check = 0

        for row_table in rows_table:
            if flag_table == 1:
                if not str(row_table[table_col]):
                    rowValueListTable.append(str(row_table[table_col].value).casefold().strip())
                else:
                    check += 1
                    break

            if flag_table == 0:
                for colindex, cell in enumerate(row_table):
                    if str(cell.value).casefold().strip() == "constituant défaillant détecté".casefold():
                        table_col = colindex
                        table_tuple = (row_index_table, row_index_table + 2, table_col, table_col + 1)
                        if table_tuple in workSheetTable.merged_cells:
                            row_index_table = row_index_table + 2
                            flag_table = 1
                        else:
                            row_index_table = row_index_table + 1
                            flag_table = 1
                        break

            row_index_table = row_index_table + 1

        for row_constituants in rows_constituants:
            if flag_constituants == 1:
                if not str(row_constituants[constituants_col]):
                    rowValueListConstituants.append(str(row_constituants[constituants_col].value).casefold().strip())
                else:
                    break

            if flag_constituants == 0:
                for colindex, cell in enumerate(row_constituants):
                    if str(cell.value).casefold() == "noms".casefold():
                        constituants_col = colindex
                        row_index_constituants = row_index_constituants + 1
                        flag_constituants = 1
            row_index_constituants = row_index_constituants + 1


        if str(rowValueListTable) == '[]'.casefold():
            check += 1
        else:
            for element in rowValueListTable:
                if element in rowValueListConstituants:
                    pass
                else:
                    check = check + 1

        if check == 0:
            return 1
        else:
            return 0

    def Test_02043_18_04939_COH_2020_XLSX_XLSM(self, workBook):

        sheetNames = workBook.sheetnames
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index_table = sheetNames.index("table")
        except:
            try:
                index_table = sheetNames.index("tableau")
            except:
                return 2

        try:
            index_constituants = sheetNames.index("constituants")
        except:
            return 2

        workSheetTable = workBook.worksheets[index_table]
        rows_table = workSheetTable.iter_rows()
        rowValueListTable = list()
        workSheetConstituants = workBook.worksheets[index_constituants]
        rows_constituants = workSheetConstituants.iter_rows()
        rowValueListConstituants = list()
        row_index_table = 0
        row_index_constituants = 0
        flag_table = 0
        flag_constituants = 0
        check = 0

        for row_table in rows_table:
            if flag_table == 1:
                if str(row_table[table_col].value) == "None":
                    break
                else:
                    rowValueListTable.append(str(row_table[table_col].value).casefold().strip())
            if flag_table == 0:
                for colindex, cell in enumerate(row_table):
                    if str(cell.value).casefold().strip() == "constituant défaillant détecté".casefold():
                        table_col = cell.column
                        if workSheetTable.cell(row_index_table + 1, table_col).value:
                            row_index_table = row_index_table + 1
                            flag_table = 1
                        else:
                            row_index_table = row_index_table + 2
                            flag_table = 1
                        break

            row_index_table = row_index_table + 1

        for row_constituants in rows_constituants:
            if flag_constituants == 1:
                if str(row_constituants[constituants_col].value) == "None":
                    break
                else:
                    rowValueListConstituants.append(str(row_constituants[constituants_col].value).casefold().strip())

            if flag_constituants == 0:
                for colindex, cell in enumerate(row_constituants):
                    if str(cell.value).casefold() == "noms".casefold():
                        constituants_col = cell.column
                        row_index_constituants = row_index_constituants + 1
                        flag_constituants = 1
            row_index_constituants = row_index_constituants + 1


        if str(rowValueListTable) == '[]'.casefold():
            check += 1
        else:
            for element in rowValueListTable:
                if element in rowValueListConstituants:
                    pass
                else:
                    check = check + 1

        if check == 0:
            return 1
        else:
            return 0

    def Test_02043_18_04939_COH_2030_XLS(self, workBook):

        sheetNames = workBook.sheet_names()
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index_table = sheetNames.index("table")
        except:
            try:
                index_table = sheetNames.index("tableau")
            except:
                return 2

        try:
            index_effet = sheetNames.index("effets clients")
        except:
            return 2

        workSheetTable = workBook.sheet_by_index(index_table)
        rows_table = workSheetTable.get_rows()
        rowValueListTable = list()
        workSheetEffet = workBook.sheet_by_index(index_effet)
        rows_effet = workSheetEffet.get_rows()
        rowValueListEffet = list()
        row_index_table = 0
        row_index_effet = 0
        flag_table = 0
        flag_effet = 0
        check = 0

        for row_table in rows_table:
            if flag_table == 1:
                if not str(row_table[table_col]):
                    rowValueListTable.append(str(row_table[table_col].value).casefold().strip())
                else:
                    check += 1
                    break

            if flag_table == 0:
                for colindex, cell in enumerate(row_table):
                    if str(cell.value).casefold().strip() == "effet(s) client(s)".casefold():
                        table_col = colindex
                        table_tuple = (row_index_table, row_index_table + 2, table_col, table_col + 1)
                        if table_tuple in workSheetTable.merged_cells:
                            row_index_table = row_index_table + 2
                            flag_table = 1
                        else:
                            row_index_table = row_index_table + 1
                            flag_table = 1
                        break

            row_index_table = row_index_table + 1

        for row_effet in rows_effet:
            if flag_effet == 1:
                if not str(row_effet[effet_col]):
                    rowValueListEffet.append(str(row_effet[effet_col].value).casefold().strip())
                else:
                    break

            if flag_effet == 0:
                for colindex, cell in enumerate(row_effet):
                    if str(cell.value).casefold() == "noms".casefold():
                        effet_col = colindex
                        row_index_effet = row_index_effet + 1
                        flag_effet = 1
            row_index_effet = row_index_effet + 1

        if str(rowValueListTable) == '[]'.casefold():
            check += 1
        else:
            for element in rowValueListTable:
                if element in rowValueListEffet:
                    pass
                else:
                    check = check + 1

        if check == 0:
            return 1
        else:
            return 0

    def Test_02043_18_04939_COH_2030_XLSX_XLSM(self, workBook):

        sheetNames = workBook.sheetnames
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index_table = sheetNames.index("table")
        except:
            try:
                index_table = sheetNames.index("tableau")
            except:
                return 2

        try:
            index_effet = sheetNames.index("effets clients")
        except:
            return 2

        workSheetTable = workBook.worksheets[index_table]
        rows_table = workSheetTable.iter_rows()
        rowValueListTable = list()
        workSheetEffet = workBook.worksheets[index_effet]
        rows_effet = workSheetEffet.iter_rows()
        rowValueListEffet = list()
        row_index_table = 0
        row_index_effet = 0
        flag_table = 0
        flag_effet = 0
        check = 0

        for row_table in rows_table:
            if flag_table == 1:
                if str(row_table[table_col].value) == "None":
                    break
                else:
                    rowValueListTable.append(str(row_table[table_col].value).casefold().strip())
            if flag_table == 0:
                for colindex, cell in enumerate(row_table):
                    if str(cell.value).casefold().strip() == "constituant défaillant détecté".casefold():
                        table_col = cell.column
                        if workSheetTable.cell(row_index_table + 1, table_col).value:
                            row_index_table = row_index_table + 1
                            flag_table = 1
                        else:
                            row_index_table = row_index_table + 2
                            flag_table = 1
                        break

            row_index_table = row_index_table + 1

        for row_effet in rows_effet:
            if flag_effet == 1:
                if str(row_effet[effet_col].value) == "None":
                    break
                else:
                    rowValueListEffet.append(str(row_effet[effet_col].value).casefold().strip())

            if flag_effet == 0:
                for colindex, cell in enumerate(row_effet):
                    if str(cell.value).casefold() == "noms".casefold():
                        effet_col = cell.column
                        row_index_effet = row_index_effet + 1
                        flag_effet = 1
            row_index_effet = row_index_effet + 1


        if str(rowValueListTable) == '[]'.casefold():
            check += 1
        else:
            for element in rowValueListTable:
                if element in rowValueListEffet:
                    pass
                else:
                    check = check + 1

        if check == 0:
            return 1
        else:
            return 0

    def Test_02043_18_04939_COH_2040_XLS(self, workBook):

        sheetNames = workBook.sheet_names()
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index_table = sheetNames.index("table")
        except:
            try:
                index_table = sheetNames.index("tableau")
            except:
                return 2

        try:
            index_diag = sheetNames.index("diagnostic débarqués")
        except:
            return 2

        workSheetTable = workBook.sheet_by_index(index_table)
        rows_table = workSheetTable.get_rows()
        rowValueListTable = list()
        workSheetDiag = workBook.sheet_by_index(index_diag)
        rows_diag = workSheetDiag.get_rows()
        rowValueListDiag = list()
        row_index_table = 0
        row_index_diag = 0
        flag_table = 0
        flag_diag = 0
        check = 0

        for row_table in rows_table:
            if flag_table == 1:
                if not str(row_table[table_col]):
                    if str(row_table[table_col]).casefold().strip() != "N/A".casefold().strip():
                        rowValueListTable.append(str(row_table[table_col].value).casefold().strip())
                    else:
                        pass
                else:
                    break

            if flag_table == 0:
                for colindex, cell in enumerate(row_table):
                    if str(cell.value).casefold().strip() == "DIAGNOSTIC DEBARQUE".casefold():
                        table_col = colindex
                        table_tuple = (row_index_table, row_index_table + 2, table_col, table_col + 1)
                        if table_tuple in workSheetTable.merged_cells:
                            row_index_table = row_index_table + 2
                            flag_table = 1
                        else:
                            row_index_table = row_index_table + 1
                            flag_table = 1
                        break

            row_index_table = row_index_table + 1

        for row_diag in rows_diag:
            if flag_diag == 1:
                if not str(row_diag[diag_col]):
                    rowValueListDiag.append(str(row_diag[diag_col].value).casefold().strip())
                else:
                    break

            if flag_diag == 0:
                for colindex, cell in enumerate(row_diag):
                    if str(cell.value).casefold() == "libellé (signification)".casefold():
                        diag_col = colindex
                        row_index_diag = row_index_diag + 1
                        flag_diag = 1
            row_index_diag = row_index_diag + 1

        if str(rowValueListTable) == '[]'.casefold():
            check += 1
        else:
            for element in rowValueListTable:
                if element in rowValueListDiag:
                    pass
                else:
                    check = check + 1

        if check == 0:
            return 1
        else:
            return 0

    def Test_02043_18_04939_COH_2040_XLSX_XLSM(self, workBook):

        sheetNames = workBook.sheetnames
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index_table = sheetNames.index("table")
        except:
            try:
                index_table = sheetNames.index("tableau")
            except:
                return 2

        try:
            index_diag = sheetNames.index("diagnostic débarqués")
        except:
            return 2

        workSheetTable = workBook.worksheets[index_table]
        rows_table = workSheetTable.iter_rows()
        rowValueListTable = list()
        workSheetDiag = workBook.worksheets[index_diag]
        rows_diag = workSheetDiag.iter_rows()
        rowValueListDiag = list()
        row_index_table = 0
        row_index_diag = 0
        flag_table = 0
        flag_diag = 0
        check = 0

        for row_table in rows_table:
            if flag_table == 1:
                if str(row_table[table_col].value) == "None":
                    break
                else:
                    if str(row_table[table_col]).casefold().strip() != "N/A".casefold().strip():
                        rowValueListTable.append(str(row_table[table_col].value).casefold().strip())
                    else:
                        pass
            if flag_table == 0:
                for colindex, cell in enumerate(row_table):
                    if str(cell.value).casefold().strip() == "DIAGNOSTIC DEBARQUE".casefold():
                        table_col = cell.column
                        if workSheetTable.cell(row_index_table + 1, table_col).value:
                            row_index_table = row_index_table + 1
                            flag_table = 1
                        else:
                            row_index_table = row_index_table + 2
                            flag_table = 1
                        break

            row_index_table = row_index_table + 1

        for row_diag in rows_diag:
            if flag_diag == 1:
                if str(row_diag[diag_col].value) == "None":
                    break
                else:
                    rowValueListDiag.append(str(row_diag[diag_col].value).casefold().strip())

            if flag_diag == 0:
                for colindex, cell in enumerate(row_diag):
                    if str(cell.value).casefold() == "libellé (signification)".casefold():
                        diag_col = cell.column
                        row_index_diag = row_index_diag + 1
                        flag_diag = 1
            row_index_diag = row_index_diag + 1


        if str(rowValueListTable) == '[]'.casefold():
            check += 1
        else:
            for element in rowValueListTable:
                if element in rowValueListDiag:
                    pass
                else:
                    check = check + 1

        if check == 0:
            return 1
        else:
            return 0

    def Test_02043_18_04939_COH_2050_XLS(self, workBook):

        sheetNames = workBook.sheet_names()
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index_table = sheetNames.index("table")
        except:
            try:
                index_table = sheetNames.index("tableau")
            except:
                return 2

        try:
            index_er = sheetNames.index("er")
        except:
            return 2

        workSheetTable = workBook.sheet_by_index(index_table)
        rows_table = workSheetTable.get_rows()
        rowValueListTable = list()
        workSheetEr = workBook.sheet_by_index(index_er)
        rows_er = workSheetEr.get_rows()
        rowValueListEr = list()
        row_index_table = 0
        row_index_er = 0
        flag_table = 0
        flag_er = 0
        check = 0

        for row_table in rows_table:
            if flag_table == 1:
                if not str(row_table[table_col]):
                    rowValueListTable.append(str(row_table[table_col].value).casefold().strip())
                else:
                    check += 1
                    break

            if flag_table == 0:
                for colindex, cell in enumerate(row_table):
                    if str(cell.value).casefold().strip() == "evenement(s) redouté(s) (ER)".casefold():
                        table_col = colindex
                        table_tuple = (row_index_table, row_index_table + 2, table_col, table_col + 1)
                        if table_tuple in workSheetTable.merged_cells:
                            row_index_table = row_index_table + 2
                            flag_table = 1
                        else:
                            row_index_table = row_index_table + 1
                            flag_table = 1
                        break

            row_index_table = row_index_table + 1

        for row_er in rows_er:
            if flag_er == 1:
                if not str(row_er[er_col]):
                    rowValueListEr.append(str(row_er[er_col].value).casefold().strip())
                else:
                    break

            if flag_er == 0:
                for colindex, cell in enumerate(row_er):
                    if str(cell.value).casefold() == "nom".casefold():
                        er_col = colindex
                        row_index_er = row_index_er + 1
                        flag_er = 1
            row_index_er = row_index_er + 1

        if str(rowValueListTable) == '[]'.casefold():
            check += 1
        else:
            for element in rowValueListTable:
                if element in rowValueListEr:
                    pass
                else:
                    check = check + 1

        if check == 0:
            return 1
        else:
            return 0

    def Test_02043_18_04939_COH_2050_XLSX_XLSM(self, workBook):

        sheetNames = workBook.sheetnames
        sheetNames = [x.casefold() for x in sheetNames]
        try:
            index_table = sheetNames.index("table")
        except:
            try:
                index_table = sheetNames.index("tableau")
            except:
                return 2

        try:
            index_er = sheetNames.index("effets clients")
        except:
            return 2

        workSheetTable = workBook.worksheets[index_table]
        rows_table = workSheetTable.iter_rows()
        rowValueListTable = list()
        workSheetEr = workBook.worksheets[index_er]
        rows_er = workSheetEr.iter_rows()
        rowValueListEr = list()
        row_index_table = 0
        row_index_er = 0
        flag_table = 0
        flag_er = 0
        check = 0

        for row_table in rows_table:
            if flag_table == 1:
                if str(row_table[table_col].value) == "None":
                    break
                else:
                    rowValueListTable.append(str(row_table[table_col].value).casefold().strip())
            if flag_table == 0:
                for cell in row_table:
                    if str(cell.value).casefold().strip() == "constituant défaillant détecté".casefold():
                        table_col = cell.column
                        if workSheetTable.cell(row_index_table + 1, table_col).value:
                            row_index_table = row_index_table + 1
                            flag_table = 1
                        else:
                            row_index_table = row_index_table + 2
                            flag_table = 1
                        break

            row_index_table = row_index_table + 1

        for row_er in rows_er:
            if flag_er == 1:
                if str(row_er[er_col].value) == "None":
                    break
                else:
                    rowValueListEr.append(str(row_er[er_col].value).casefold().strip())

            if flag_er == 0:
                for colindex, cell in enumerate(row_er):
                    if str(cell.value).casefold() == "noms".casefold():
                        er_col = cell.column
                        row_index_er = row_index_er + 1
                        flag_er = 1
            row_index_er = row_index_er + 1


        if str(rowValueListTable) == '[]'.casefold():
            check += 1
        else:
            for element in rowValueListTable:
                if element in rowValueListEr:
                    pass
                else:
                    check = check + 1

        if check == 0:
            return 1
        else:
            return 0

    def Test_02043_18_04939_COH_2060_XLS(self, workBook):

        sheetNames = workBook.sheet_names()
        sheetNames = [x.casefold() for x in sheetNames]

        try:
            index_effet = sheetNames.index("effets clients")
        except:
             return 2

        workSheetEffet = workBook.sheet_by_index(index_effet)
        rows_effet = workSheetEffet.get_rows()
        rowValueListEffet = list()
        row_index_effet = 0
        flag_effet = 0
        count  = 0

        for row_effet in rows_effet:
            if flag_effet == 1:
                if not str(row_effet[effet_col]):
                    rowValueListEffet.append(str(row_effet[effet_col].value).casefold().strip())
                else:
                    break

            if flag_effet == 0:
                for colindex, cell in enumerate(row_effet):
                    if str(cell.value).casefold() == "noms".casefold():
                        effet_col = colindex
                        row_index_effet = row_index_effet + 1
                        flag_effet = 1
            row_index_effet = row_index_effet + 1


        if self.tab2.myTextBox6.toPlainText():
            fileName = self.tab2.myTextBox6.toPlainText()
            excel = win32.gencache.EnsureDispatch('Excel.Application')
            workBook = excel.Workbooks.Open(fileName)
            workSheetFR = workBook.Sheets("FR")
            workSheetGB = workBook.Sheets("GB")
            col = 0
        else:
            fileName = self.path_effect
            excel = win32.gencache.EnsureDispatch('Excel.Application')
            workBook = excel.Workbooks.Open(fileName)
            workSheetFR = workBook.Sheets("FR")
            workSheetGB = workBook.Sheets("GB")
            col = 0

        for row in workSheetFR.Rows:
             for cell in row.Cells:
                 if str(cell.Value).strip().casefold() == "Libellé N1".strip().casefold():
                     colN1_FR = cell.Column
                     col += 1
                 if str(cell.Value).strip().casefold() == "Libellé N2".strip().casefold():
                     colN2_FR = cell.Column
                     col += 1
                 if str(cell.Value).strip().casefold() == "Libellé N3".strip().casefold():
                     colN3_FR = cell.Column
                     col += 1
                 if col == 3:
                     break
             if col == 3:
                 break

        col = 0
        for row in workSheetGB.Rows:
             for cell in row.Cells:
                 if str(cell.Value).strip().casefold() == "Libellé N1".strip().casefold():
                     colN1_GB = cell.Column
                     col += 1
                 if str(cell.Value).strip().casefold() == "Libellé N2".strip().casefold():
                     colN2_GB = cell.Column
                     col += 1
                 if str(cell.Value).strip().casefold() == "Libellé N3".strip().casefold():
                     colN3_GB = cell.Column
                     col += 1
                 if col == 3:
                     break
             if col == 3:
                 break


        EffetListFR = list()
        for row in workSheetFR.Rows:
            if workSheetFR.Cells(row.Row, colN1_FR).Value is not None:
                EffetListFR.append(str(workSheetFR.Cells(row.Row, colN1_FR).Value))
            elif workSheetFR.Cells(row.Row, colN2_FR).Value is not None:
                EffetListFR.append(str(workSheetFR.Cells(row.Row, colN2_FR).Value))
            elif workSheetFR.Cells(row.Row, colN3_FR).Value is not None:
                EffetListFR.append(str(workSheetFR.Cells(row.Row, colN3_FR).Value))
            else:
                break

        EffetListGB = list()
        for row in workSheetGB.Rows:
            if workSheetGB.Cells(row.Row, colN1_GB).Value is not None:
                EffetListGB.append(str(workSheetGB.Cells(row.Row, colN1_GB).Value))
            elif workSheetGB.Cells(row.Row, colN2_GB).Value is not None:
                EffetListGB.append(str(workSheetGB.Cells(row.Row, colN2_GB).Value))
            elif workSheetGB.Cells(row.Row, colN3_GB).Value is not None:
                EffetListGB.append(str(workSheetGB.Cells(row.Row, colN3_GB).Value))
            else:
                break

        if str(rowValueListEffet) == '[]'.casefold():
            count += 1
        else:
            for element in rowValueListEffet:
                if element in EffetListFR or element in EffetListGB:
                    pass
                else:
                    count +=1


        if count == 0:
             return 1
        else:
            return 0

    def Test_02043_18_04939_COH_2060_XLSX_XLSM(self, workBook):

        sheetNames = workBook.sheetnames
        sheetNames = [x.casefold() for x in sheetNames]

        try:
            index_effet = sheetNames.index("effets clients")
        except:
             return 2

        workSheetEffet = workBook.worksheets[index_effet]
        rows_effet = workSheetEffet.iter_rows()
        rowValueListEffet = list()
        row_index_effet = 0
        flag_effet = 0
        count = 0

        for row_effet in rows_effet:
            if flag_effet == 1:
                if not str(row_effet[effet_col]):
                    rowValueListEffet.append(str(row_effet[effet_col].value).casefold().strip())
                else:
                    break

            if flag_effet == 0:
                for cell in row_effet:
                    if str(cell.value).casefold() == "noms".casefold():
                        effet_col = cell.column
                        row_index_effet = row_index_effet + 1
                        flag_effet = 1
            row_index_effet = row_index_effet + 1


        if self.tab2.myTextBox6.toPlainText():
            fileName = self.tab2.myTextBox6.toPlainText()
            excel = win32.gencache.EnsureDispatch('Excel.Application')
            workBook = excel.Workbooks.Open(fileName)
            workSheetFR = workBook.Sheets("FR")
            workSheetGB = workBook.Sheets("GB")
            col = 0
        else:
            fileName = self.path_effect
            excel = win32.gencache.EnsureDispatch('Excel.Application')
            workBook = excel.Workbooks.Open(fileName)
            workSheetFR = workBook.Sheets("FR")
            workSheetGB = workBook.Sheets("GB")
            col = 0

        for row in workSheetFR.Rows:
             for cell in row.Cells:
                 if str(cell.Value).strip().casefold() == "Libellé N1".strip().casefold():
                     colN1_FR = cell.Column
                     col += 1
                 if str(cell.Value).strip().casefold() == "Libellé N2".strip().casefold():
                     colN2_FR = cell.Column
                     col += 1
                 if str(cell.Value).strip().casefold() == "Libellé N3".strip().casefold():
                     colN3_FR = cell.Column
                     col += 1
                 if col == 3:
                     break
             if col == 3:
                 break

        col = 0
        for row in workSheetGB.Rows:
             for cell in row.Cells:
                 if str(cell.Value).strip().casefold() == "Libellé N1".strip().casefold():
                     colN1_GB = cell.Column
                     col += 1
                 if str(cell.Value).strip().casefold() == "Libellé N2".strip().casefold():
                     colN2_GB = cell.Column
                     col += 1
                 if str(cell.Value).strip().casefold() == "Libellé N3".strip().casefold():
                     colN3_GB = cell.Column
                     col += 1
                 if col == 3:
                     break
             if col == 3:
                 break


        EffetListFR = list()
        for row in workSheetFR.Rows:
            if workSheetFR.Cells(row.Row, colN1_FR).Value is not None:
                EffetListFR.append(str(workSheetFR.Cells(row.Row, colN1_FR).Value))
            elif workSheetFR.Cells(row.Row, colN2_FR).Value is not None:
                EffetListFR.append(str(workSheetFR.Cells(row.Row, colN2_FR).Value))
            elif workSheetFR.Cells(row.Row, colN3_FR).Value is not None:
                EffetListFR.append(str(workSheetFR.Cells(row.Row, colN3_FR).Value))
            else:
                break

        EffetListGB = list()
        for row in workSheetGB.Rows:
            if workSheetGB.Cells(row.Row, colN1_GB).Value is not None:
                EffetListGB.append(str(workSheetGB.Cells(row.Row, colN1_GB).Value))
            elif workSheetGB.Cells(row.Row, colN2_GB).Value is not None:
                EffetListGB.append(str(workSheetGB.Cells(row.Row, colN2_GB).Value))
            elif workSheetGB.Cells(row.Row, colN3_GB).Value is not None:
                EffetListGB.append(str(workSheetGB.Cells(row.Row, colN3_GB).Value))
            else:
                break


        if str(rowValueListEffet) == '[]'.casefold():
            count += 1
        else:
            for element in rowValueListEffet:
                if element in EffetListFR or element in EffetListGB:
                    pass
                else:
                    count +=1


        if count == 0:
             return 1
        else:
            return 0

    def Test_02043_18_04939_COH_2070_XLS(self, workBook):

        sheetNames = workBook.sheet_names()
        sheetNames = [x.casefold() for x in sheetNames]

        try:
            index_effet = sheetNames.index("customer effects")
        except:
             return 2

        workSheetEffet = workBook.sheet_by_index(index_effet)
        rows_effet = workSheetEffet.get_rows()
        rowValueListEffet = list()
        row_index_effet = 0
        flag_effet = 0
        count  = 0

        for row_effet in rows_effet:
            if flag_effet == 1:
                if not str(row_effet[effet_col]):
                    rowValueListEffet.append(str(row_effet[effet_col].value).casefold().strip())
                else:
                    break

            if flag_effet == 0:
                for colindex, cell in enumerate(row_effet):
                    if str(cell.value).casefold() == "name".casefold():
                        effet_col = colindex
                        row_index_effet = row_index_effet + 1
                        flag_effet = 1
            row_index_effet = row_index_effet + 1


        if self.tab2.myTextBox6.toPlainText():
            fileName = self.tab2.myTextBox6.toPlainText()
            excel = win32.gencache.EnsureDispatch('Excel.Application')
            workBook = excel.Workbooks.Open(fileName)
            workSheetFR = workBook.Sheets("FR")
            workSheetGB = workBook.Sheets("GB")
            col = 0
        else:
            fileName = self.path_effect
            excel = win32.gencache.EnsureDispatch('Excel.Application')
            workBook = excel.Workbooks.Open(fileName)
            workSheetFR = workBook.Sheets("FR")
            workSheetGB = workBook.Sheets("GB")
            col = 0

        for row in workSheetFR.Rows:
             for cell in row.Cells:
                 if str(cell.Value).strip().casefold() == "Libellé N1".strip().casefold():
                     colN1_FR = cell.Column
                     col += 1
                 if str(cell.Value).strip().casefold() == "Libellé N2".strip().casefold():
                     colN2_FR = cell.Column
                     col += 1
                 if str(cell.Value).strip().casefold() == "Libellé N3".strip().casefold():
                     colN3_FR = cell.Column
                     col += 1
                 if col == 3:
                     break
             if col == 3:
                 break

        col = 0
        for row in workSheetGB.Rows:
             for cell in row.Cells:
                 if str(cell.Value).strip().casefold() == "Libellé N1".strip().casefold():
                     colN1_GB = cell.Column
                     col += 1
                 if str(cell.Value).strip().casefold() == "Libellé N2".strip().casefold():
                     colN2_GB = cell.Column
                     col += 1
                 if str(cell.Value).strip().casefold() == "Libellé N3".strip().casefold():
                     colN3_GB = cell.Column
                     col += 1
                 if col == 3:
                     break
             if col == 3:
                 break


        EffetListFR = list()
        for row in workSheetFR.Rows:
            if workSheetFR.Cells(row.Row, colN1_FR).Value is not None:
                EffetListFR.append(str(workSheetFR.Cells(row.Row, colN1_FR).Value))
            elif workSheetFR.Cells(row.Row, colN2_FR).Value is not None:
                EffetListFR.append(str(workSheetFR.Cells(row.Row, colN2_FR).Value))
            elif workSheetFR.Cells(row.Row, colN3_FR).Value is not None:
                EffetListFR.append(str(workSheetFR.Cells(row.Row, colN3_FR).Value))
            else:
                break

        EffetListGB = list()
        for row in workSheetGB.Rows:
            if workSheetGB.Cells(row.Row, colN1_GB).Value is not None:
                EffetListGB.append(str(workSheetGB.Cells(row.Row, colN1_GB).Value))
            elif workSheetGB.Cells(row.Row, colN2_GB).Value is not None:
                EffetListGB.append(str(workSheetGB.Cells(row.Row, colN2_GB).Value))
            elif workSheetGB.Cells(row.Row, colN3_GB).Value is not None:
                EffetListGB.append(str(workSheetGB.Cells(row.Row, colN3_GB).Value))
            else:
                break

        if str(rowValueListEffet) == '[]'.casefold():
            count += 1
        else:
            for element in rowValueListEffet:
                if element in EffetListFR or element in EffetListGB:
                    pass
                else:
                    count +=1


        if count == 0:
             return 1
        else:
            return 0

    def Test_02043_18_04939_COH_2070_XLSX_XLSM(self, workBook):

        sheetNames = workBook.sheetnames
        sheetNames = [x.casefold() for x in sheetNames]

        try:
            index_effet = sheetNames.index("customer effects")
        except:
            return 2

        workSheetEffet = workBook.worksheets[index_effet]
        rows_effet = workSheetEffet.iter_rows()
        rowValueListEffet = list()
        row_index_effet = 0
        flag_effet = 0
        count = 0

        for row_effet in rows_effet:
            if flag_effet == 1:
                if not str(row_effet[effet_col]):
                    rowValueListEffet.append(str(row_effet[effet_col].value).casefold().strip())
                else:
                    break

            if flag_effet == 0:
                for cell in row_effet:
                    if str(cell.value).casefold() == "name".casefold():
                        effet_col = cell.column
                        row_index_effet = row_index_effet + 1
                        flag_effet = 1
            row_index_effet = row_index_effet + 1

        if self.tab2.myTextBox6.toPlainText():
            fileName = self.tab2.myTextBox6.toPlainText()
            excel = win32.gencache.EnsureDispatch('Excel.Application')
            workBook = excel.Workbooks.Open(fileName)
            workSheetFR = workBook.Sheets("FR")
            workSheetGB = workBook.Sheets("GB")
            col = 0
        else:
            fileName = self.path_effect
            excel = win32.gencache.EnsureDispatch('Excel.Application')
            workBook = excel.Workbooks.Open(fileName)
            workSheetFR = workBook.Sheets("FR")
            workSheetGB = workBook.Sheets("GB")
            col = 0

        for row in workSheetFR.Rows:
            for cell in row.Cells:
                if str(cell.Value).strip().casefold() == "Libellé N1".strip().casefold():
                    colN1_FR = cell.Column
                    col += 1
                if str(cell.Value).strip().casefold() == "Libellé N2".strip().casefold():
                    colN2_FR = cell.Column
                    col += 1
                if str(cell.Value).strip().casefold() == "Libellé N3".strip().casefold():
                    colN3_FR = cell.Column
                    col += 1
                if col == 3:
                    break
            if col == 3:
                break

        col = 0
        for row in workSheetGB.Rows:
            for cell in row.Cells:
                if str(cell.Value).strip().casefold() == "Libellé N1".strip().casefold():
                    colN1_GB = cell.Column
                    col += 1
                if str(cell.Value).strip().casefold() == "Libellé N2".strip().casefold():
                    colN2_GB = cell.Column
                    col += 1
                if str(cell.Value).strip().casefold() == "Libellé N3".strip().casefold():
                    colN3_GB = cell.Column
                    col += 1
                if col == 3:
                    break
            if col == 3:
                break

        EffetListFR = list()
        for row in workSheetFR.Rows:
            if workSheetFR.Cells(row.Row, colN1_FR).Value is not None:
                EffetListFR.append(str(workSheetFR.Cells(row.Row, colN1_FR).Value))
            elif workSheetFR.Cells(row.Row, colN2_FR).Value is not None:
                EffetListFR.append(str(workSheetFR.Cells(row.Row, colN2_FR).Value))
            elif workSheetFR.Cells(row.Row, colN3_FR).Value is not None:
                EffetListFR.append(str(workSheetFR.Cells(row.Row, colN3_FR).Value))
            else:
                break

        EffetListGB = list()
        for row in workSheetGB.Rows:
            if workSheetGB.Cells(row.Row, colN1_GB).Value is not None:
                EffetListGB.append(str(workSheetGB.Cells(row.Row, colN1_GB).Value))
            elif workSheetGB.Cells(row.Row, colN2_GB).Value is not None:
                EffetListGB.append(str(workSheetGB.Cells(row.Row, colN2_GB).Value))
            elif workSheetGB.Cells(row.Row, colN3_GB).Value is not None:
                EffetListGB.append(str(workSheetGB.Cells(row.Row, colN3_GB).Value))
            else:
                break

        if str(rowValueListEffet) == '[]'.casefold():
            count += 1
        else:
            for element in rowValueListEffet:
                if element in EffetListFR or element in EffetListGB:
                    pass
                else:
                    count += 1

        if count == 0:
            return 1
        else:
            return 0

    def Test_02043_18_04939_COH_2080_XLS(self, workBook):

        sheetNames = workBook.sheet_names()
        sheetNames = [x.casefold() for x in sheetNames]

        try:
            index_effet = sheetNames.index("effets clients")
        except:
             return 2

        workSheetEffet = workBook.sheet_by_index(index_effet)
        rows_effet = workSheetEffet.get_rows()
        rowValueListEffet = list()
        row_index_effet = 0
        flag_effet = 0
        count  = 0

        for row_effet in rows_effet:
            if flag_effet == 1:
                if not str(row_effet[effet_col]):
                    rowValueListEffet.append(str(row_effet[effet_col].value).casefold().strip())
                else:
                    break

            if flag_effet == 0:
                for colindex, cell in enumerate(row_effet):
                    if str(cell.value).casefold() == "noms".casefold():
                        effet_col = colindex
                        row_index_effet = row_index_effet + 1
                        flag_effet = 1
            row_index_effet = row_index_effet + 1


        if self.tab2.myTextBox6.toPlainText():
            fileName = self.tab2.myTextBox6.toPlainText()
            excel = win32.gencache.EnsureDispatch('Excel.Application')
            workBook = excel.Workbooks.Open(fileName)
            workSheetFR = workBook.Sheets("FR")
            workSheetGB = workBook.Sheets("GB")
            col = 0
        else:
            fileName = self.path_effect
            excel = win32.gencache.EnsureDispatch('Excel.Application')
            workBook = excel.Workbooks.Open(fileName)
            workSheetFR = workBook.Sheets("FR")
            workSheetGB = workBook.Sheets("GB")
            col = 0

        for row in workSheetFR.Rows:
             for cell in row.Cells:
                 if str(cell.Value).strip().casefold() == "Libellé N1".strip().casefold():
                     colN1_FR = cell.Column
                     col += 1
                 if str(cell.Value).strip().casefold() == "Libellé N2".strip().casefold():
                     colN2_FR = cell.Column
                     col += 1
                 if str(cell.Value).strip().casefold() == "Libellé N3".strip().casefold():
                     colN3_FR = cell.Column
                     col += 1
                 if col == 3:
                     break
             if col == 3:
                 break

        col = 0
        for row in workSheetGB.Rows:
             for cell in row.Cells:
                 if str(cell.Value).strip().casefold() == "Libellé N1".strip().casefold():
                     colN1_GB = cell.Column
                     col += 1
                 if str(cell.Value).strip().casefold() == "Libellé N2".strip().casefold():
                     colN2_GB = cell.Column
                     col += 1
                 if str(cell.Value).strip().casefold() == "Libellé N3".strip().casefold():
                     colN3_GB = cell.Column
                     col += 1
                 if col == 3:
                     break
             if col == 3:
                 break


        EffetListFR = list()
        for row in workSheetFR.Rows:
            if workSheetFR.Cells(row.Row, colN1_FR).Value is not None:
                EffetListFR.append(str(workSheetFR.Cells(row.Row, colN1_FR).Value))
            elif workSheetFR.Cells(row.Row, colN2_FR).Value is not None:
                EffetListFR.append(str(workSheetFR.Cells(row.Row, colN2_FR).Value))
            elif workSheetFR.Cells(row.Row, colN3_FR).Value is not None:
                EffetListFR.append(str(workSheetFR.Cells(row.Row, colN3_FR).Value))
            else:
                break

        EffetListGB = list()
        for row in workSheetGB.Rows:
            if workSheetGB.Cells(row.Row, colN1_GB).Value is not None:
                EffetListGB.append(str(workSheetGB.Cells(row.Row, colN1_GB).Value))
            elif workSheetGB.Cells(row.Row, colN2_GB).Value is not None:
                EffetListGB.append(str(workSheetGB.Cells(row.Row, colN2_GB).Value))
            elif workSheetGB.Cells(row.Row, colN3_GB).Value is not None:
                EffetListGB.append(str(workSheetGB.Cells(row.Row, colN3_GB).Value))
            else:
                break

        if str(rowValueListEffet) == '[]'.casefold():
            count += 1
        else:
            for element in rowValueListEffet:
                if element in EffetListFR or element in EffetListGB:
                    pass
                else:
                    count +=1


        if count == 0:
             return 1
        else:
            return 0

    def Test_02043_18_04939_COH_2080_XLSX_XLSM(self, workBook):

        sheetNames = workBook.sheetnames
        sheetNames = [x.casefold() for x in sheetNames]

        try:
            index_effet = sheetNames.index("effets clients")
        except:
            return 2

        workSheetEffet = workBook.worksheets[index_effet]
        rows_effet = workSheetEffet.iter_rows()
        rowValueListEffet = list()
        row_index_effet = 0
        flag_effet = 0
        count = 0

        for row_effet in rows_effet:
            if flag_effet == 1:
                if not str(row_effet[effet_col]):
                    rowValueListEffet.append(str(row_effet[effet_col].value).casefold().strip())
                else:
                    break

            if flag_effet == 0:
                for cell in row_effet:
                    if str(cell.value).casefold() == "noms".casefold():
                        effet_col = cell.column
                        row_index_effet = row_index_effet + 1
                        flag_effet = 1
            row_index_effet = row_index_effet + 1

        if self.tab2.myTextBox6.toPlainText():
            fileName = self.tab2.myTextBox6.toPlainText()
            excel = win32.gencache.EnsureDispatch('Excel.Application')
            workBook = excel.Workbooks.Open(fileName)
            workSheetFR = workBook.Sheets("FR")
            workSheetGB = workBook.Sheets("GB")
            col = 0
        else:
            fileName = self.path_effect
            excel = win32.gencache.EnsureDispatch('Excel.Application')
            workBook = excel.Workbooks.Open(fileName)
            workSheetFR = workBook.Sheets("FR")
            workSheetGB = workBook.Sheets("GB")
            col = 0

        for row in workSheetFR.Rows:
            for cell in row.Cells:
                if str(cell.Value).strip().casefold() == "Libellé N1".strip().casefold():
                    colN1_FR = cell.Column
                    col += 1
                if str(cell.Value).strip().casefold() == "Libellé N2".strip().casefold():
                    colN2_FR = cell.Column
                    col += 1
                if str(cell.Value).strip().casefold() == "Libellé N3".strip().casefold():
                    colN3_FR = cell.Column
                    col += 1
                if col == 3:
                    break
            if col == 3:
                break

        col = 0
        for row in workSheetGB.Rows:
            for cell in row.Cells:
                if str(cell.Value).strip().casefold() == "Libellé N1".strip().casefold():
                    colN1_GB = cell.Column
                    col += 1
                if str(cell.Value).strip().casefold() == "Libellé N2".strip().casefold():
                    colN2_GB = cell.Column
                    col += 1
                if str(cell.Value).strip().casefold() == "Libellé N3".strip().casefold():
                    colN3_GB = cell.Column
                    col += 1
                if col == 3:
                    break
            if col == 3:
                break

        EffetListFR = list()
        for row in workSheetFR.Rows:
            if workSheetFR.Cells(row.Row, colN1_FR).Value is not None:
                EffetListFR.append(str(workSheetFR.Cells(row.Row, colN1_FR).Value))
            elif workSheetFR.Cells(row.Row, colN2_FR).Value is not None:
                EffetListFR.append(str(workSheetFR.Cells(row.Row, colN2_FR).Value))
            elif workSheetFR.Cells(row.Row, colN3_FR).Value is not None:
                EffetListFR.append(str(workSheetFR.Cells(row.Row, colN3_FR).Value))
            else:
                break

        EffetListGB = list()
        for row in workSheetGB.Rows:
            if workSheetGB.Cells(row.Row, colN1_GB).Value is not None:
                EffetListGB.append(str(workSheetGB.Cells(row.Row, colN1_GB).Value))
            elif workSheetGB.Cells(row.Row, colN2_GB).Value is not None:
                EffetListGB.append(str(workSheetGB.Cells(row.Row, colN2_GB).Value))
            elif workSheetGB.Cells(row.Row, colN3_GB).Value is not None:
                EffetListGB.append(str(workSheetGB.Cells(row.Row, colN3_GB).Value))
            else:
                break

        if str(rowValueListEffet) == '[]'.casefold():
            count += 1
        else:
            for element in rowValueListEffet:
                if element in EffetListFR or element in EffetListGB:
                    pass
                else:
                    count += 1

        if count == 0:
            return 1
        else:
            return 0

    def Test_02043_18_04939_COH_2100_XLS(self, workBook):

        sheetNames = workBook.sheet_names()
        sheetNames = [x.casefold() for x in sheetNames]

        try:
            index = sheetNames.index("codes défauts")
        except:
             return 2

        workSheetCodes = workBook.sheet_by_index(index)
        rows_codes = workSheetCodes.get_rows()
        rowValueListCodes = list()
        row_index_codes = 0
        flag_codes = 0
        count = 0

        for row_codes in rows_codes:
            if flag_codes == 1:
                if not str(row_codes[codes_col]):
                    rowValueListCodes.append(str(row_codes[codes_col].value).casefold().strip())
                else:
                    break

            if flag_codes == 0:
                for colindex, cell in enumerate(row_codes):
                    if str(cell.value).casefold() == "supporté par constituant (s)".casefold():
                        codes_col = colindex
                        row_index_codes = row_index_codes + 1
                        flag_codes = 1
            row_index_codes = row_index_codes + 1

            fileName = self.path_DOC8
            excel = win32.gencache.EnsureDispatch('Excel.Application')
            workBook = excel.Workbooks.Open(fileName)
            workSheet = workBook.Sheets("sous familles Cesare")


        for row in workSheet.Rows:
             for cell in row.Cells:
                 if str(cell.Value).strip().casefold() == "Nom de la sous famille".strip().casefold():
                     col = cell.Column
             if col != 0:
                 break


        CodeList = list()
        for row in workSheet.Rows:
            if workSheet.Cells(row.Row, col).Value is not None:
                CodeList.append(str(workSheet.Cells(row.Row, col).Value))
            else:
                break


        if str(rowValueListCodes) == '[]'.casefold():
            count += 1
        else:
            for element in rowValueListCodes:
                if element in CodeList :
                    pass
                else:
                    count +=1


        if count == 0:
             return 1
        else:
            return 0




    def buttonClicked(self):

        self.tab1.colorTextBox1.setStyleSheet('background-color: grey')
        self.tab1.colorTextBox2.setStyleSheet('background-color: grey')
        self.tab1.colorTextBox3.setStyleSheet('background-color: grey')

        os.system("taskkill /f /im EXCEL.EXE")
        self.tab1.textbox.setText("")
        self.tab1.pbar.setValue(0)
        self.pbvalue = 0
        if not self.tab2.myTextBox5.toPlainText():
            self.path_Cesare = self.download_file(self.CesareLink)
        else:
            self.path_Cesare = self.tab2.myTextBox5.toPlainText()
        if not self.tab2.myTextBox4.toPlainText():
            self.path_config = self.download_file(self.TSDConfigLink)
        if not self.tab2.myTextBox6.toPlainText():
            self.path_effect = self.download_file(self.CustomerEffectLink)
        else:
            self.path_effect = self.tab2.myTextBox6.toPlainText()
        if not self.tab2.myTextBox9.toPlainText():
            self.path_diversity = self.download_file(self.DiversityLink)
        if self.tab1.myTextBox1.toPlainText():
            self.result = self.download_DOC3(self.DOC3Link)
        if self.tab1.myTextBox2.toPlainText():
            self.result = self.download_DOC4(self.DOC4Link)
        self.path_DOC14 = self.download_file(self.DOC14Link)
        self.path_DOC8 = self.download_file(self.DOC8Link)
        if self.path_Cesare == "Error":
            self.tab1.textbox.setText("Check connection type")
            return
        if self.path_Cesare == "False":
            return

        filePath = self.download_doc9(self.DOC9Link)
        self.testReqDict = dict()
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        workBook = excel.Workbooks.Open(filePath)
        workSheet = workBook.Sheets(1)
        for row in workSheet.Rows:
            if workSheet.Cells(row.Row, 1).Value == "Requirements":
                rowStartIndex = row.Row
            if workSheet.Cells(row.Row, 1).Value is None:
                if workSheet.Cells(row.Row + 1, 1).Value is None and workSheet.Cells(row.Row + 2,1).Value is None and workSheet.Cells(row.Row + 3, 1).Value is None and workSheet.Cells(row.Row + 4, 1).Value is None:
                    rowStopIndex = row.Row - 1
                    break

        cellRange = workSheet.Range(workSheet.Cells(rowStartIndex + 1, 1), workSheet.Cells(rowStopIndex, 4))
        for row in cellRange.Rows:
            dictTemp = dict()
            dictTemp["Previsional"] = workSheet.Cells(row.Row, 2).Value
            dictTemp["Consolidated"] = workSheet.Cells(row.Row, 3).Value
            dictTemp["Validated"] = workSheet.Cells(row.Row, 4).Value
            self.testReqDict[str(workSheet.Cells(row.Row, 1).Value).strip()] = dictTemp

        self.checkLevel = str(self.tab1.combo.currentText()).strip()

        self.TestTsdFile(self.path_Cesare, self.path_effect)
        self.TestTsdVehicleFunctionFile(self.path_Cesare, self.path_effect)
        self.TestTsdSystemFile(self.path_Cesare, self.path_effect)
        self.TestAmdecFile()
        self.TestExportMedialecMatriceFile()
        self.TestDiagnosticMatrixFile()

        os.system("taskkill /f /im EXCEL.EXE")

if __name__ == '__main__':


    try:
        FindWindow(None, "TSD Checker  V0.5")
        windll.user32.MessageBoxW(0, "Application already running", "Warning", 0|48)

    except:
        app = QApplication(sys.argv)
        apel = Test()
        myQLabel = QLabel()
        sys.exit(app.exec_())



